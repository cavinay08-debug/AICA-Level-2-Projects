"""
tally_converter.py
==================
Converts a standard Tally Trial Balance Excel/CSV export to the flat columnar
format required by the Schedule III application import engine.

Tally TB structure (Excel):
  Rows 1-N : Company header / date range (skip)
  Header row: Particulars | Opening Balance | Debit | Credit | Closing Balance
  Data rows : Group headers (bold / indent 0) + Ledger rows (indented / not bold)
  Last row  : Grand Total (end marker)

Output format (matches generate_sample_excel_file):
  Particulars / Ledger Name | Parent Tally Group |
  Opening Balance | Debit Amount | Credit Amount |
  Closing Balance | Prior Year Closing Balance
"""

import re
import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Amount Parser ────────────────────────────────────────────────────────────

def _parse_amount(val) -> float:
    """
    Parse a Tally amount value.
    Handles:
      • "12,34,567.89 Dr"  →  +1234567.89
      • "12,34,567.89 Cr"  →  -1234567.89
      • "(12345.00)"       →  -12345.00
      • plain float/int   →  as-is
      • None / blank      →  0.0
    """
    if val is None:
        return 0.0
    s = str(val).strip().replace(',', '')
    if not s or s in ['-', '—', '–', '']:
        return 0.0

    # Check Cr / Dr suffix
    sl = s.lower()
    is_cr = sl.endswith('cr')
    is_dr = sl.endswith('dr')

    # Strip text
    clean = re.sub(r'[a-zA-Z\s]', '', s).strip()

    # Handle parentheses (accounting negative)
    if clean.startswith('(') and clean.endswith(')'):
        clean = clean[1:-1]
        is_cr = True

    try:
        num = float(clean)
    except ValueError:
        return 0.0

    if is_cr:
        return -num
    elif is_dr:
        return num
    else:
        return num   # plain number — sign from value itself


def _fmt_amount(val: float, decimals: int = 2) -> str:
    """Format a signed float as Tally-style string (e.g. '1,23,456.78 Dr')."""
    if val == 0.0:
        return f"0.{'0'*decimals} Dr"
    suffix = 'Cr' if val < 0 else 'Dr'
    abs_val = abs(val)
    # Indian number formatting
    formatted = f"{abs_val:,.{decimals}f}"
    return f"{formatted} {suffix}"


# ─── Indent Detection ─────────────────────────────────────────────────────────

def _get_indent(cell, name: str) -> int:
    """Return the logical indent level of a Tally TB row."""
    # 1. openpyxl alignment.indent (works for proper Excel exports)
    try:
        if cell.alignment and cell.alignment.indent:
            return int(cell.alignment.indent)
    except Exception:
        pass

    # 2. Leading spaces in the cell string value
    if isinstance(name, str):
        raw = name
        stripped = name.lstrip()
        spaces = len(raw) - len(stripped)
        if spaces > 0:
            return max(1, spaces // 2)

    return 0


def _is_bold(cell) -> bool:
    """Return True if the cell has bold font (Tally uses bold for group headers)."""
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


# ─── Header Row Detection ─────────────────────────────────────────────────────

def _find_header_row(all_rows):
    """
    Find the row index containing 'Particulars'/'Ledger'/'Name' and
    return (header_row_index, col_map dict).

    Also searches the row immediately below the found row for
    'Debit' / 'Credit' in case Tally uses a two-row header with a merged
    'Transactions' cell above Debit/Credit.
    """
    col_map = {}
    header_idx = None

    for i, row in enumerate(all_rows):
        vals = [str(c.value or '').strip().lower() for c in row]
        combined = ' '.join(vals)
        if 'particulars' in combined or 'ledger' in combined:
            header_idx = i

            for j, cell in enumerate(row):
                v = str(cell.value or '').strip().lower()
                if any(k in v for k in ['particular', 'ledger', 'account']) and j <= 2:
                    col_map.setdefault('particulars', j)
                elif 'opening' in v:
                    col_map.setdefault('opening', j)
                elif v in ('debit', 'dr') or ('debit' in v and 'credit' not in v):
                    col_map.setdefault('debit', j)
                elif v in ('credit', 'cr') or ('credit' in v and 'debit' not in v):
                    col_map.setdefault('credit', j)
                elif 'closing' in v or ('balance' in v and j > 2):
                    col_map.setdefault('closing', j)

            # Check the next row too (two-row header)
            if i + 1 < len(all_rows):
                next_row = all_rows[i + 1]
                for j, cell in enumerate(next_row):
                    v = str(cell.value or '').strip().lower()
                    if v in ('debit', 'dr') or ('debit' in v and 'credit' not in v):
                        col_map.setdefault('debit', j)
                    elif v in ('credit', 'cr') or ('credit' in v and 'debit' not in v):
                        col_map.setdefault('credit', j)

            # Positional fallbacks
            n = len(row)
            col_map.setdefault('particulars', 0)
            if n > 1: col_map.setdefault('opening', 1)
            if n > 2: col_map.setdefault('debit', 2)
            if n > 3: col_map.setdefault('credit', 3)
            if n > 4: col_map.setdefault('closing', 4)

            break

    return header_idx, col_map


# ─── Main Converter ───────────────────────────────────────────────────────────

def convert_tally_excel(file_bytes_or_path) -> tuple:
    """
    Convert a Tally Trial Balance Excel file to a list of flat ledger dicts.

    Parameters
    ----------
    file_bytes_or_path : bytes | str | BytesIO

    Returns
    -------
    (ledgers: list[dict], stats: dict, errors: list[str])

    Each ledger dict has keys:
      ledger_name, tally_group,
      opening_dr, opening_cr,
      debit, credit,
      closing_dr, closing_cr,
      closing_net, prior_closing_net
    """
    errors = []

    # ── Open workbook ─────────────────────────────────────────────────────────
    if isinstance(file_bytes_or_path, bytes):
        file_bytes_or_path = io.BytesIO(file_bytes_or_path)
    try:
        wb = openpyxl.load_workbook(file_bytes_or_path, data_only=True)
    except Exception as e:
        return [], {}, [f"Cannot open Excel file: {e}. Ensure the file is a valid .xlsx or .xls export from Tally."]

    ws = wb.active
    all_rows = list(ws.iter_rows())

    # ── Find header row ───────────────────────────────────────────────────────
    header_idx, col_map = _find_header_row(all_rows)
    if header_idx is None:
        return [], {}, [
            "Could not find the column header row (expected 'Particulars' or 'Ledger Name'). "
            "Please export the Trial Balance from Tally with standard column headers."
        ]

    # ── Parse data rows ───────────────────────────────────────────────────────
    SKIP_NAMES = {'', 'total', 'grand total', 'particulars', 'ledger name',
                  'nil', 'debit', 'credit', 'dr', 'cr', 'opening balance',
                  'closing balance', 'transactions'}

    raw_rows = []
    for row in all_rows[header_idx + 1:]:
        if col_map['particulars'] >= len(row):
            continue
        cell_a = row[col_map['particulars']]
        raw_name = cell_a.value
        name = str(raw_name or '').strip()
        # Keep the ORIGINAL unstripped string for indent detection:
        # Tally sometimes indents via leading spaces in the cell value.
        # If we pass the stripped `name`, all space counts become 0 and
        # every row appears at indent level 0 (breaking sub-group detection).
        raw_name_str = str(raw_name or '')
        name_clean = name.lower()

        if not name or name_clean in SKIP_NAMES:
            continue
        if 'grand total' in name_clean or 'net balance' in name_clean:
            break

        def _get_col(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return 0.0
            return _parse_amount(row[idx].value)

        indent    = _get_indent(cell_a, raw_name_str)  # ← unstripped name preserves leading spaces
        bold      = _is_bold(cell_a)
        opening   = _get_col('opening')
        debit     = abs(_get_col('debit'))
        credit    = abs(_get_col('credit'))
        closing   = _get_col('closing')

        has_amounts = (
            abs(closing) > 0.001 or debit > 0.001 or
            credit > 0.001 or abs(opening) > 0.001
        )

        raw_rows.append({
            'name': name,
            'indent': indent,
            'bold': bold,
            'has_amounts': has_amounts,
            'opening': opening,
            'debit': debit,
            'credit': credit,
            'closing': closing,
            'row_ref': row,   # keep reference for closing-cell sign check
        })

    if not raw_rows:
        return [], {}, ["No data rows found after the header. Please check the file."]

    # ── Master Tally Groups & Sub-Groups ───────────────────────────────────────
    TOP_LEVEL_GROUPS = {
        'capital account', 'current liabilities', 'fixed assets', 'current assets',
        'sales accounts', 'indirect expenses', 'direct expenses', 'purchase accounts',
        'direct incomes', 'indirect incomes', 'investments', 'loans (liability)',
        'branch / divisions', 'suspense a/c', 'misc. expenses (asset)', 'primary'
    }

    ALWAYS_SUBGROUPS = {
        'duties & taxes', 'bank accounts', 'cash-in-hand', 'provisions',
        'reserves & surplus', 'bank od a/c', 'secured loans', 'unsecured loans',
        'deposits (asset)', 'loans & advances (asset)', 'stock-in-hand'
    }

    def _is_group_row(i):
        entry = raw_rows[i]
        name_clean = entry['name'].lower()
        if entry['bold'] or name_clean in TOP_LEVEL_GROUPS or name_clean in ALWAYS_SUBGROUPS:
            return True
        if not entry['has_amounts']:
            return True

        # Party groups (Sundry Debtors / Sundry Creditors):
        # A row named Sundry Debtors or Sundry Creditors is a Sub-Group if and only if
        # subsequent child party rows sum up to its total.
        if name_clean in {'sundry debtors', 'sundry creditors'}:
            target = max(entry['debit'], entry['credit'], abs(entry['closing']))
            if target <= 0.01:
                return True
            acc = 0.0
            for j in range(i + 1, len(raw_rows)):
                nxt = raw_rows[j]
                nxt_name_clean = nxt['name'].lower()
                if nxt['bold'] or nxt_name_clean in TOP_LEVEL_GROUPS or nxt_name_clean in ALWAYS_SUBGROUPS:
                    break
                nxt_amt = max(nxt['debit'], nxt['credit'], abs(nxt['closing']))
                acc += nxt_amt
                if abs(acc - target) <= 1.0:
                    return True
                if acc > target + 1.0:
                    break
            return False

        return False

    # ── Classify rows and extract leaf ledgers ────────────────────────────────
    ledgers = []
    current_top_group = 'Primary'
    current_sub_group = None
    sub_group_target = 0.0
    sub_group_acc = 0.0
    groups_seen = set()
    ledger_count = 0

    for i, entry in enumerate(raw_rows):
        name = entry['name'].strip()
        name_clean = name.lower()

        if _is_group_row(i):
            groups_seen.add(name)
            if entry['bold'] or name_clean in TOP_LEVEL_GROUPS:
                current_top_group = name
                current_sub_group = None
                sub_group_target = 0.0
                sub_group_acc = 0.0
            else:
                current_sub_group = name
                sub_group_target = max(entry['debit'], entry['credit'], abs(entry['closing']))
                sub_group_acc = 0.0
            continue

        # ── Leaf ledger ───────────────────────────────────────────────────────
        parent = current_sub_group if current_sub_group else current_top_group

        # Check if sub-group has accumulated its full total
        if current_sub_group:
            child_amt = max(entry['debit'], entry['credit'], abs(entry['closing']))
            sub_group_acc += child_amt
            if sub_group_target > 0 and abs(sub_group_acc - sub_group_target) <= 1.0:
                current_sub_group = None  # Subgroup scope fulfilled

        # Determine closing balance sign:
        # If the closing cell has an explicit Dr/Cr suffix → trust it.
        # If it's a plain unsigned number → use the double-entry formula:
        #   closing = opening(signed) + debit(abs) - credit(abs)
        row_ref = entry['row_ref']
        raw_closing_cell = row_ref[col_map['closing']].value if col_map.get('closing') is not None and col_map['closing'] < len(row_ref) else None
        closing_raw_str = str(raw_closing_cell or '').strip().lower()
        closing_has_sign = closing_raw_str.endswith('dr') or closing_raw_str.endswith('cr')

        closing = entry['closing']
        if not closing_has_sign:
            closing = entry['opening'] + entry['debit'] - entry['credit']

        opening = entry['opening']

        ledgers.append({
            'ledger_name'      : name,
            'tally_group'      : parent,
            'opening_dr'       : opening if opening >= 0 else 0.0,
            'opening_cr'       : abs(opening) if opening < 0 else 0.0,
            'debit'            : entry['debit'],
            'credit'           : entry['credit'],
            'closing_dr'       : closing if closing >= 0 else 0.0,
            'closing_cr'       : abs(closing) if closing < 0 else 0.0,
            'closing_net'      : closing,
            'prior_closing_net': 0.0,
        })
        ledger_count += 1

    # ── Deduplicate by ledger_name (keep last occurrence) ────────────────────
    seen_ledgers = {}
    for l in ledgers:
        seen_ledgers[l['ledger_name']] = l
    ledgers = list(seen_ledgers.values())

    stats = {
        'total_ledgers'   : len(ledgers),
        'groups_detected' : len(groups_seen),
        'total_rows_read' : len(raw_rows),
    }



    if ledger_count == 0:
        errors.append(
            "No ledger rows could be extracted. This may happen if the Tally Excel uses "
            "a non-standard format or all rows are classified as group headers (bold). "
            "Tip: Export from Tally with 'Show Ledger Details' enabled."
        )

    return ledgers, stats, errors


# ─── Output Excel Generator ───────────────────────────────────────────────────

def generate_app_format_excel(ledgers: list, decimals: int = 2) -> io.BytesIO:
    """
    Generate an Excel workbook in the application's required import format.

    Columns:
      Particulars / Ledger Name | Parent Tally Group |
      Opening Balance | Debit Amount | Credit Amount |
      Closing Balance | Prior Year Closing Balance
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tally Trial Balance"

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    grp_font  = Font(name='Calibri', bold=True, color='1F497D', size=10)
    grp_fill  = PatternFill(start_color='EDF2FB', end_color='EDF2FB', fill_type='solid')

    reg_font  = Font(name='Calibri', size=10)
    num_align = Alignment(horizontal='right')
    thin      = Side(style='thin', color='D9D9D9')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'Particulars / Ledger Name',
        'Parent Tally Group',
        'Opening Balance',
        'Debit Amount',
        'Credit Amount',
        'Closing Balance',
        'Prior Year Closing Balance',
    ]

    ws.append(headers)
    for col_i, cell in enumerate(ws[1], 1):
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
    ws.row_dimensions[1].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    prev_group = None
    row_num = 2

    for l in ledgers:
        group = l['tally_group']

        # Insert group separator row when group changes.
        # Prefix with '—' so that parse_tally_excel will skip it on re-import.
        if group != prev_group:
            ws.append([f'— {group} —', '', '', '', '', '', ''])
            for col_i in range(1, 8):
                c = ws.cell(row=row_num, column=col_i)
                c.font   = grp_font
                c.fill   = grp_fill
                c.border = border
            row_num += 1
            prev_group = group

        opening_net = l['opening_dr'] - l['opening_cr']
        closing_net = l['closing_net']

        ws.append([
            l['ledger_name'],
            group,
            _fmt_amount(opening_net, decimals),
            round(l['debit'],   decimals),
            round(l['credit'],  decimals),
            _fmt_amount(closing_net, decimals),
            _fmt_amount(l['prior_closing_net'], decimals),
        ])

        for col_i in range(1, 8):
            c = ws.cell(row=row_num, column=col_i)
            c.font   = reg_font
            c.border = border
            if col_i >= 3:
                c.alignment = num_align

        row_num += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [45, 28, 18, 16, 16, 18, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Info sheet ────────────────────────────────────────────────────────────
    ws_info = wb.create_sheet(title='Conversion Info')
    ws_info.append(['Field', 'Value'])
    ws_info.append(['Total Ledgers Converted', len(ledgers)])
    ws_info.append(['Unique Tally Groups', len({l['tally_group'] for l in ledgers})])
    ws_info.append(['File Format', 'Application Import Format (Schedule III Builder)'])
    ws_info.append(['Note', 'Upload the "Tally Trial Balance" sheet via Step 2 → Import TB'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── CSV Parser (bonus support) ──────────────────────────────────────────────

def convert_tally_csv(file_bytes: bytes) -> tuple:
    """
    Convert Tally Trial Balance exported as CSV.
    Assumes structure similar to Excel export but uses leading spaces for indentation.
    """
    import csv

    content = file_bytes.decode('utf-8-sig', errors='ignore')
    reader = csv.reader(io.StringIO(content))
    rows = list(reader)

    # Find header row
    header_idx = None
    col_map = {}
    for i, row in enumerate(rows):
        vals = [v.strip().lower() for v in row]
        if 'particulars' in vals or 'ledger' in ' '.join(vals):
            header_idx = i
            for j, v in enumerate(vals):
                if 'particular' in v or 'ledger' in v:
                    col_map.setdefault('particulars', j)
                elif 'opening' in v:
                    col_map.setdefault('opening', j)
                elif 'debit' in v or v == 'dr':
                    col_map.setdefault('debit', j)
                elif 'credit' in v or v == 'cr':
                    col_map.setdefault('credit', j)
                elif 'closing' in v or ('balance' in v and j > 2):
                    col_map.setdefault('closing', j)
            col_map.setdefault('particulars', 0)
            col_map.setdefault('opening', 1)
            col_map.setdefault('debit', 2)
            col_map.setdefault('credit', 3)
            col_map.setdefault('closing', 4)
            break

    if header_idx is None:
        return [], {}, ["Could not find header row in CSV."]

    ledgers = []
    group_stack = []

    SKIP = {'', 'total', 'grand total', 'particulars', 'nil'}

    for row in rows[header_idx + 1:]:
        if not row:
            continue
        name = row[col_map.get('particulars', 0)].strip() if row else ''
        if not name or name.lower() in SKIP:
            continue
        if 'grand total' in name.lower():
            break

        # Detect indent from leading spaces
        raw = row[col_map.get('particulars', 0)]
        spaces = len(raw) - len(raw.lstrip())
        indent = max(0, spaces // 2)

        def safe(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return 0.0
            return _parse_amount(row[idx])

        opening = safe('opening')
        debit   = abs(safe('debit'))
        credit  = abs(safe('credit'))
        closing = safe('closing')
        has_amt = abs(closing) > 0.001 or debit > 0.001 or credit > 0.001

        while group_stack and group_stack[-1][1] >= indent:
            group_stack.pop()

        if not has_amt:
            group_stack.append((name, indent))
            continue

        parent  = group_stack[-1][0] if group_stack else 'Primary'
        # Trust closing as-is; only estimate if both closing AND opening are absent
        if abs(closing) < 0.001 and abs(opening) < 0.001 and (debit > 0.001 or credit > 0.001):
            closing = debit - credit

        ledgers.append({
            'ledger_name'      : name,
            'tally_group'      : parent,
            'opening_dr'       : opening if opening >= 0 else 0.0,
            'opening_cr'       : abs(opening) if opening < 0 else 0.0,
            'debit'            : debit,
            'credit'           : credit,
            'closing_dr'       : closing if closing >= 0 else 0.0,
            'closing_cr'       : abs(closing) if closing < 0 else 0.0,
            'closing_net'      : closing,
            'prior_closing_net': 0.0,
        })

    stats = {'total_ledgers': len(ledgers),
             'groups_detected': len({l['tally_group'] for l in ledgers})}
    return ledgers, stats, []

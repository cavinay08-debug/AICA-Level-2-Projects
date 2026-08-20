import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.database import get_settings, save_invoice
from app.utils import validate_gstin, validate_pan, validate_pin

def create_sample_excel_template(output_path):
    """Creates a pre-formatted sample Excel template with guidelines and examples."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices Import Template"
    
    # Columns definition
    headers = [
        "Invoice Number", "Date (YYYY-MM-DD)", "Invoice Type (Goods/Services)", 
        "GST Treatment (Auto-Detect/Intrastate/Interstate)", "Place of Supply", 
        "Reverse Charge (RCM) (Yes/No)", "Customer Name", "Customer Mobile", 
        "Customer Address", "Customer GSTIN", "Customer PAN", "Customer PIN", 
        "Item Name", "HSN/SAC", "Qty", "Rate / Amount", "GST %", "Cess %"
    ]
    
    # Formatting styles
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Border styles
    thin = Side(border_style="thin", color="CBD5E1")
    grid_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    # Set headers
    ws.append(headers)
    ws.row_dimensions[1].height = 28
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    # Sample rows
    samples = [
        [
            "INV-IM-001", "2026-08-12", "Goods", "Auto-Detect", "Maharashtra (27)", 
            "No", "Acme Corp", "9876543210", "123 Mumbai Rd", "27AAAAA1111A1Z1", 
            "AAAAA1111A", "400001", "Business Laptop", "84713010", 5, 65000.0, 18, 0.0
        ],
        [
            "INV-IM-001", "2026-08-12", "Goods", "Auto-Detect", "Maharashtra (27)", 
            "No", "Acme Corp", "9876543210", "123 Mumbai Rd", "27AAAAA1111A1Z1", 
            "AAAAA1111A", "400001", "Wireless Mouse", "84716040", 10, 850.0, 18, 0.0
        ],
        [
            "INV-IM-002", "2026-08-12", "Services", "Auto-Detect", "Delhi (07)", 
            "Yes", "Beta Tech Ltd", "", "456 Delhi St", "07BBBBB2222B2Z2", 
            "BBBBB2222B", "110001", "Software Development", "998313", 1, 150000.0, 18, 1.0
        ]
    ]
    
    for row in samples:
        ws.append(row)
        
    # Apply standard fonts & borders to data rows
    for row_idx in range(2, 5):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Segoe UI", size=9.5)
            cell.border = grid_border
            if col_idx in [2, 14, 15, 17, 18]:
                cell.alignment = align_center
            elif col_idx in [16]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = align_left
                
    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output_path)
    return True

def parse_row(row_values):
    """Parses a row of strings into standard types safely."""
    try:
        inv_no = str(row_values[0]).strip()
        date = str(row_values[1]).strip()
        inv_type = "Services" if "service" in str(row_values[2]).lower() else "Goods"
        
        treatment_str = str(row_values[3]).strip().lower()
        if "intra" in treatment_str:
            treatment = "CGST_SGST"
        elif "inter" in treatment_str:
            treatment = "IGST"
        else:
            treatment = "Auto"
            
        place_of_supply = str(row_values[4]).strip()
        rcm = True if str(row_values[5]).strip().lower() in ["yes", "true", "y"] else False
        
        cust_name = str(row_values[6]).strip()
        cust_mobile = str(row_values[7]).strip() if row_values[7] is not None else ""
        cust_address = str(row_values[8]).strip() if row_values[8] is not None else ""
        cust_gst = str(row_values[9]).strip().upper() if row_values[9] is not None else ""
        cust_pan = str(row_values[10]).strip().upper() if row_values[10] is not None else ""
        cust_pin = str(row_values[11]).strip() if row_values[11] is not None else ""
        
        item_name = str(row_values[12]).strip()
        hsn = str(row_values[13]).strip() if row_values[13] is not None else ""
        
        qty = float(row_values[14]) if row_values[14] is not None else 1.0
        rate = float(row_values[15]) if row_values[15] is not None else 0.0
        gst_rate = float(row_values[16]) if row_values[16] is not None else 0.0
        cess_rate = float(row_values[17]) if row_values[17] is not None else 0.0
        
        return {
            "invoice_number": inv_no, "date": date, "invoice_type": inv_type,
            "gst_treatment": treatment, "place_of_supply": place_of_supply, "rcm": rcm,
            "cust_name": cust_name, "cust_mobile": cust_mobile, "cust_address": cust_address,
            "cust_gst": cust_gst, "cust_pan": cust_pan, "cust_pin": cust_pin,
            "item_name": item_name, "hsn": hsn, "qty": qty, "rate": rate,
            "gst_rate": gst_rate, "cess_rate": cess_rate
        }
    except Exception as e:
        raise ValueError(f"Formatting parsing error: {e}")

def import_invoices_from_file(file_path):
    """Parses Excel or CSV invoice rows, aggregates them, runs tax math, and saves."""
    rows = []
    
    # Read rows based on file extension
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xlsx":
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            # Iterate rows skipping header
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Ignore fully empty rows
                if not any(row):
                    continue
                rows.append(row)
        except Exception as e:
            return 0, [f"Failed to open Excel workbook: {e}"]
    elif ext == ".csv":
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader, None) # Skip header
                for row in reader:
                    if not any(row):
                        continue
                    # Pad row if columns are missing
                    while len(row) < 18:
                        row.append(None)
                    rows.append(row)
        except Exception as e:
            return 0, [f"Failed to open CSV file: {e}"]
    else:
        return 0, ["Unsupported file format. Please import an Excel (.xlsx) or CSV file."]

    if not rows:
        return 0, ["The selected import file contains no data rows."]

    grouped_invoices = {}
    errors = []
    
    # First pass: parse rows and group items by Invoice Number
    for idx, raw_row in enumerate(rows, 2):
        try:
            parsed = parse_row(raw_row)
            inv_no = parsed["invoice_number"]
            if not inv_no:
                errors.append(f"Row {idx}: Invoice Number is empty.")
                continue
            if not parsed["cust_name"]:
                errors.append(f"Row {idx}: Customer Name is empty.")
                continue
            if not parsed["item_name"]:
                errors.append(f"Row {idx}: Item Name is empty.")
                continue
                
            # Validations
            if parsed["cust_gst"] and not validate_gstin(parsed["cust_gst"]):
                errors.append(f"Row {idx}: Customer GSTIN '{parsed['cust_gst']}' has invalid format.")
                continue
            if parsed["cust_pan"] and not validate_pan(parsed["cust_pan"]):
                errors.append(f"Row {idx}: Customer PAN '{parsed['cust_pan']}' has invalid format.")
                continue
            if parsed["cust_pin"] and not validate_pin(parsed["cust_pin"]):
                errors.append(f"Row {idx}: PIN Code '{parsed['cust_pin']}' must be exactly 6 digits.")
                continue
                
            if inv_no not in grouped_invoices:
                grouped_invoices[inv_no] = {
                    "header": parsed,
                    "items": []
                }
            grouped_invoices[inv_no]["items"].append(parsed)
        except Exception as e:
            errors.append(f"Row {idx}: {e}")

    success_count = 0
    settings = get_settings()
    seller = settings.get("seller", {})
    seller_gst = seller.get("gstin", "").strip()

    # Second pass: calculate taxes, assemble JSON schemas, and save
    for inv_no, data in grouped_invoices.items():
        header = data["header"]
        items_raw = data["items"]
        
        # Determine GST treatment
        treatment = header["gst_treatment"]
        if treatment == "Auto":
            buyer_gst = header["cust_gst"].strip()
            if len(seller_gst) >= 2 and len(buyer_gst) >= 2:
                if seller_gst[:2] == buyer_gst[:2]:
                    treatment = "CGST_SGST"
                else:
                    treatment = "IGST"
            else:
                treatment = "CGST_SGST"
                
        is_intrastate = (treatment == "CGST_SGST")
        rcm_applicable = header["rcm"]
        
        items_list = []
        subtotal = 0.0
        cgst_total = 0.0
        sgst_total = 0.0
        igst_total = 0.0
        cess_total = 0.0
        
        for item in items_raw:
            qty = item["qty"]
            rate = item["rate"]
            gst_rate = item["gst_rate"]
            cess_rate = item["cess_rate"]
            
            taxable = qty * rate
            
            cgst_amt = 0.0
            sgst_amt = 0.0
            igst_amt = 0.0
            
            if is_intrastate:
                cgst_amt = taxable * ((gst_rate / 2.0) / 100.0)
                sgst_amt = taxable * ((gst_rate / 2.0) / 100.0)
            else:
                igst_amt = taxable * (gst_rate / 100.0)
                
            cess_amt = taxable * (cess_rate / 100.0)
            
            # If RCM, GST is not added to item total billing amount
            if rcm_applicable:
                total_amt = taxable + cess_amt
            else:
                total_amt = taxable + cgst_amt + sgst_amt + igst_amt + cess_amt
                
            items_list.append({
                "name": item["item_name"],
                "hsn": item["hsn"],
                "qty": qty,
                "rate": rate,
                "gst_rate": gst_rate,
                "cess_rate": cess_rate,
                "taxable_amount": taxable,
                "cgst_amount": cgst_amt,
                "sgst_amount": sgst_amt,
                "igst_amount": igst_amt,
                "cess_amount": cess_amt,
                "total_amount": total_amt
            })
            
            subtotal += taxable
            cgst_total += cgst_amt
            sgst_total += sgst_amt
            igst_total += igst_amt
            cess_total += cess_amt
            
        # Invoice totals
        if rcm_applicable:
            raw_grand = subtotal + cess_total
        else:
            raw_grand = subtotal + cgst_total + sgst_total + igst_total + cess_total
            
        grand_total = round(raw_grand)
        round_off = grand_total - raw_grand
        
        # Assemble invoice dict
        invoice_json = {
            "invoice_number": inv_no,
            "date": header["date"],
            "invoice_type": header["invoice_type"],
            "gst_treatment": treatment,
            "place_of_supply": header["place_of_supply"],
            "rcm": rcm_applicable,
            "customer": {
                "name": header["cust_name"],
                "mobile": header["cust_mobile"],
                "address": header["cust_address"],
                "gstin": header["cust_gst"],
                "pan": header["cust_pan"],
                "pin": header["cust_pin"]
            },
            "ship_to": {}, # Bulk importer defaults to empty ship-to
            "seller": seller,
            "items": items_list,
            "summary": {
                "subtotal": subtotal,
                "cgst_total": cgst_total,
                "sgst_total": sgst_total,
                "igst_total": igst_total,
                "cess_total": cess_total,
                "round_off": round_off,
                "grand_total": grand_total
            }
        }
        
        if save_invoice(invoice_json):
            success_count += 1
        else:
            errors.append(f"Invoice {inv_no}: Failed to save record to disk.")

    return success_count, errors

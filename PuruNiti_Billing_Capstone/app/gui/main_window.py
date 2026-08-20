import os
import sys
import shutil
import sqlite3
import calendar
import subprocess
from datetime import datetime

from PyQt5.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QSplitter, QListWidget, QListWidgetItem, QLineEdit, 
                             QPushButton, QLabel, QGroupBox, QFormLayout, 
                             QDateEdit, QTableWidget, QTableWidgetItem, QHeaderView,
                             QMessageBox, QDoubleSpinBox, QComboBox, QScrollArea,
                             QFileDialog, QApplication, QCheckBox, QInputDialog,
                             QRadioButton, QButtonGroup, QTabWidget, QGridLayout,
                             QSpinBox, QStackedWidget, QMenu, QDialog, QAbstractItemView)
from PyQt5.QtCore import Qt, QDate, QThread, pyqtSignal, QRectF
from PyQt5.QtGui import QPainter, QColor, QFont, QPen, QBrush

from app.database import (get_settings, get_next_invoice_number, save_invoice, 
                           delete_invoice, get_invoice, search_invoices, get_all_invoices,
                           get_database_dir, update_database_path, save_settings,
                           migrate_json_to_sqlite, get_db_connection,
                           add_payment, delete_payment, get_payments_for_invoice,
                           get_invoice_payment_summary, get_aging_report, get_payments_by_period)
from app.pdf_generator import generate_pdf
from app.utils import (num_to_words_indian, validate_gstin, validate_pan, validate_pin)
from app.gui.seller_dialog import SellerDialog
from app.gui.styles import MAIN_STYLESHEET
from app.gui.auth_dialogs import ChangePassDialog, UserSettingsDialog, RecordPaymentDialog
from app.license_manager import check_license_status, reset_license, get_base_dir
from app.gst_reports import (generate_gstr1_excel, generate_gstr1_pdf,
                              generate_dashboard_excel, generate_dashboard_pdf,
                              generate_dynamic_dashboard_excel, generate_dynamic_dashboard_pdf)

# GST State codes mapping
STATE_CODES = {
    "01": "Jammu & Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh",
    "05": "Uttarakhand", "06": "Haryana", "07": "Delhi", "08": "Rajasthan", "09": "Uttar Pradesh",
    "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur",
    "15": "Mizoram", "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal",
    "20": "Jharkhand", "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat",
    "25": "Daman & Diu", "26": "Dadra & Nagar Haveli", "27": "Maharashtra", "28": "Andhra Pradesh",
    "29": "Karnataka", "30": "Goa", "31": "Lakshadweep", "32": "Kerala", "33": "Tamil Nadu",
    "34": "Puducherry", "35": "Andaman & Nicobar Islands", "36": "Telangana", "37": "Andhra Pradesh (New)"
}


# --- Custom Painter-Based Charting Widget ---
class SimpleChartWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.data = [] # List of tuples: (label, value)
        self.metric_name = "Value"
        self.setMinimumHeight(400)
        
    def setData(self, data, metric_name="Value"):
        self.data = data
        self.metric_name = metric_name
        self.update() # Trigger paint event redraw
        
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        # Draw background container
        painter.fillRect(self.rect(), QColor("#FFFFFF"))
        
        # Border
        painter.setPen(QPen(QColor("#E2E8F0"), 1))
        painter.setBrush(Qt.NoBrush)
        painter.drawRoundedRect(QRectF(0, 0, self.width() - 1, self.height() - 1), 6, 6)
        
        if not self.data:
            painter.setPen(QColor("#64748B"))
            painter.setFont(QFont("Segoe UI", 11))
            painter.drawText(self.rect(), Qt.AlignCenter, "No query results loaded to represent.")
            return
            
        margin_left = 170
        margin_right = 110
        margin_top = 50
        margin_bottom = 20
        
        w = self.width() - margin_left - margin_right
        h = self.height() - margin_top - margin_bottom
        
        max_val = max(val for label, val in self.data)
        if max_val == 0:
            max_val = 1.0
            
        row_height = h / len(self.data)
        bar_height = min(row_height * 0.55, 26)
        
        # Title text
        painter.setPen(QColor("#0F172A"))
        painter.setFont(QFont("Segoe UI", 10, QFont.Bold))
        painter.drawText(15, 30, f"Analysis Graph: {self.metric_name} Breakdown")
        
        # Color spectrum
        colors_list = [
            QColor("#1E3A8A"), # Navy
            QColor("#0D9488"), # Teal
            QColor("#B45309"), # Amber
            QColor("#DC2626"), # Red
            QColor("#6366F1"), # Indigo
            QColor("#0284C7"), # Sky Blue
            QColor("#16A34A"), # Green
            QColor("#7C3AED")  # Purple
        ]
        
        for idx, (label, val) in enumerate(self.data):
            y = margin_top + (idx * row_height) + (row_height - bar_height) / 2
            
            # Label
            painter.setPen(QColor("#475569"))
            painter.setFont(QFont("Segoe UI", 9))
            label_text = label if label else "N/A"
            if len(label_text) > 22:
                label_text = label_text[:19] + "..."
                
            label_rect = QRectF(10, y, margin_left - 20, bar_height)
            painter.drawText(label_rect, Qt.AlignRight | Qt.AlignVCenter, label_text)
            
            # Bar width
            bar_w = (val / max_val) * w
            if bar_w < 2:
                bar_w = 2 # Draw tiny sliver for non-zero values
                
            bar_rect = QRectF(margin_left, y, bar_w, bar_height)
            
            color = colors_list[idx % len(colors_list)]
            painter.setBrush(QBrush(color))
            painter.setPen(Qt.NoPen)
            painter.drawRoundedRect(bar_rect, 3, 3)
            
            # Metric Value
            painter.setPen(QColor("#0F172A"))
            painter.setFont(QFont("Segoe UI", 9, QFont.Bold))
            val_text = f"₹ {val:,.2f}"
            val_rect = QRectF(margin_left + bar_w + 10, y, margin_right - 15, bar_height)
            painter.drawText(val_rect, Qt.AlignLeft | Qt.AlignVCenter, val_text)


# --- PyQT Background Threads ---
class PDFWorker(QThread):
    finished = pyqtSignal(str)
    error = pyqtSignal(str)
    
    def __init__(self, invoice_data, pdf_path):
        super().__init__()
        self.invoice_data = invoice_data
        self.pdf_path = pdf_path
        
    def run(self):
        try:
            generate_pdf(self.invoice_data, self.pdf_path)
            self.finished.emit(self.pdf_path)
        except Exception as e:
            self.error.emit(str(e))

class ImportWorker(QThread):
    finished = pyqtSignal(int, list)
    
    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path
        
    def run(self):
        from app.bulk_importer import import_invoices_from_file
        try:
            success, errors = import_invoices_from_file(self.file_path)
            self.finished.emit(success, errors)
        except Exception as e:
            self.finished.emit(0, [f"Import thread error: {e}"])

class ReportWorker(QThread):
    finished = pyqtSignal(bool, str)
    
    def __init__(self, report_type, start_date, end_date, output_path, format_type, extra_args=None):
        super().__init__()
        self.report_type = report_type # "gstr1", "dash", "dash_dyn"
        self.start_date = start_date
        self.end_date = end_date
        self.output_path = output_path
        self.format_type = format_type # "excel" or "pdf"
        self.extra_args = extra_args or {} # For dynamic reports
        
    def run(self):
        try:
            if self.report_type == "gstr1":
                if self.format_type == "excel":
                    generate_gstr1_excel(self.start_date, self.end_date, self.output_path)
                else:
                    generate_gstr1_pdf(self.start_date, self.end_date, self.output_path)
            elif self.report_type == "dash":
                if self.format_type == "excel":
                    generate_dashboard_excel(self.start_date, self.end_date, self.output_path)
                else:
                    generate_dashboard_pdf(self.start_date, self.end_date, self.output_path)
            else: # "dash_dyn"
                dim_name = self.extra_args.get("dim")
                meas_name = self.extra_args.get("meas")
                data = self.extra_args.get("data")
                if self.format_type == "excel":
                    generate_dynamic_dashboard_excel(dim_name, meas_name, self.start_date, self.end_date, data, self.output_path)
                else:
                    generate_dynamic_dashboard_pdf(dim_name, meas_name, self.start_date, self.end_date, data, self.output_path)
            self.finished.emit(True, self.output_path)
        except Exception as e:
            self.finished.emit(False, str(e))


class MainWindow(QMainWindow):
    def __init__(self, current_user="admin", current_role="admin"):
        super().__init__()
        self.current_user = current_user
        self.current_role = current_role
        self.logout_requested = False
        self.setWindowTitle("PuruNiti Smart Billing system")
        self.resize(1400, 930)
        self.setStyleSheet(MAIN_STYLESHEET)
        
        self.current_invoice_id = None
        self.active_dyn_data = [] # Stores active queried values
        
        # Run startup legacy database migrations
        migrate_json_to_sqlite()
        
        self.init_ui()
        self.load_seller_profiles_list()
        self.refresh_invoice_list()
        self.new_invoice()  
        self.update_license_display()
        self.auto_create_import_template()
        self.refresh_dashboard_data() # Load dashboard
        self.refresh_payments_tab()

    def auto_create_import_template(self):
        template_path = self.get_safe_dialog_path("Invoice_Import_Template.xlsx")
        if not os.path.exists(template_path):
            from app.bulk_importer import create_sample_excel_template
            try:
                create_sample_excel_template(template_path)
            except Exception as e:
                print(f"Failed to auto-create Excel import template: {e}")

    def get_safe_dialog_path(self, default_filename):
        documents = os.path.expanduser("~/Documents")
        if os.path.exists(documents):
            return os.path.join(documents, default_filename)
        desktop = os.path.expanduser("~/Desktop")
        if os.path.exists(desktop):
            return os.path.join(desktop, default_filename)
        return default_filename

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        outer_layout = QVBoxLayout(central_widget)
        outer_layout.setContentsMargins(10, 10, 10, 10)
        outer_layout.setSpacing(10)
        
        # --- HEADER ROW (Branding & License status) ---
        header_row = QHBoxLayout()
        header_row.setContentsMargins(5, 5, 5, 5)
        
        app_title = QLabel("PuruNiti Smart Billing system")
        app_title.setObjectName("titleLabel")
        header_row.addWidget(app_title)
        
        header_row.addStretch()
        
        # User authentication info display
        self.lbl_user_info = QLabel(f"User: {self.current_user} ({self.current_role})")
        self.lbl_user_info.setStyleSheet("font-weight: bold; font-size: 13px; color: #1e3a8a; margin-right: 10px;")
        header_row.addWidget(self.lbl_user_info)
        
        self.btn_change_password = QPushButton("Change Password")
        self.btn_change_password.setStyleSheet("font-size: 11px; padding: 4px 8px; margin-right: 5px;")
        self.btn_change_password.clicked.connect(self.open_change_password)
        header_row.addWidget(self.btn_change_password)
        
        self.btn_logout = QPushButton("Switch User")
        self.btn_logout.setStyleSheet("font-size: 11px; padding: 4px 8px; margin-right: 5px;")
        self.btn_logout.clicked.connect(self.handle_logout_action)
        header_row.addWidget(self.btn_logout)
        
        self.btn_manage_users = QPushButton("Manage Accounts")
        self.btn_manage_users.setStyleSheet("font-size: 11px; padding: 4px 8px; margin-right: 15px;")
        self.btn_manage_users.clicked.connect(self.open_user_settings)
        self.btn_manage_users.setVisible(self.current_role == "admin")
        header_row.addWidget(self.btn_manage_users)
        
        self.lbl_license_status = QLabel("License: Checking...")
        self.lbl_license_status.setStyleSheet("font-weight: bold; font-size: 13px; color: #15803D; margin-right: 10px;")
        header_row.addWidget(self.lbl_license_status)
        
        self.btn_reset_license = QPushButton("Reset License")
        self.btn_reset_license.setObjectName("btnDanger")
        self.btn_reset_license.setProperty("danger", True)
        self.btn_reset_license.setStyleSheet("font-size: 11px; padding: 4px 8px;")
        self.btn_reset_license.clicked.connect(self.reset_license_key)
        header_row.addWidget(self.btn_reset_license)
        
        outer_layout.addLayout(header_row)
        
        # --- TAB CONTROLS ---
        self.tab_widget = QTabWidget()
        outer_layout.addWidget(self.tab_widget)
        
        # TAB 1: Billing Form
        tab1_widget = QWidget()
        tab1_layout = QHBoxLayout(tab1_widget)
        tab1_layout.setContentsMargins(0, 0, 0, 0)
        self.tab_widget.addTab(tab1_widget, "Billing & Outward Invoices")
        
        # Splitter inside Tab 1
        splitter = QSplitter(Qt.Horizontal)
        tab1_layout.addWidget(splitter)
        
        # Sidebar layout
        sidebar_widget = QWidget()
        sidebar_layout = QVBoxLayout(sidebar_widget)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)
        sidebar_layout.setSpacing(10)
        
        sidebar_subtitle = QLabel("Saved Invoices")
        sidebar_subtitle.setStyleSheet("font-weight: bold; font-size: 14px; color: #1A365D;")
        sidebar_layout.addWidget(sidebar_subtitle)
        
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Search #, customer, mobile...")
        self.txt_search.textChanged.connect(self.refresh_invoice_list)
        sidebar_layout.addWidget(self.txt_search)
        
        self.lst_invoices = QListWidget()
        self.lst_invoices.itemDoubleClicked.connect(self.load_selected_invoice)
        sidebar_layout.addWidget(self.lst_invoices)
        
        self.btn_sidebar_new = QPushButton("New Invoice")
        self.btn_sidebar_new.setObjectName("btnSuccess")
        self.btn_sidebar_new.setProperty("success", True)
        self.btn_sidebar_new.clicked.connect(self.new_invoice)
        sidebar_layout.addWidget(self.btn_sidebar_new)
        
        self.btn_sidebar_import = QPushButton("Bulk Import (Excel/CSV)")
        self.btn_sidebar_import.setObjectName("btnPrimary")
        self.btn_sidebar_import.setProperty("primary", True)
        self.btn_sidebar_import.clicked.connect(self.import_bulk_invoices)
        sidebar_layout.addWidget(self.btn_sidebar_import)
        
        db_config_group = QGroupBox("Database Storage Path")
        db_config_layout = QVBoxLayout(db_config_group)
        db_config_layout.setContentsMargins(10, 10, 10, 10)
        db_config_layout.setSpacing(6)
        
        self.txt_db_path = QLineEdit()
        self.txt_db_path.setReadOnly(True)
        self.txt_db_path.setStyleSheet("font-size: 11px; background-color: #F8FAFC; color: #64748B;")
        self.txt_db_path.setText(get_database_dir())
        db_config_layout.addWidget(self.txt_db_path)
        
        self.btn_change_db = QPushButton("Change Location")
        self.btn_change_db.setStyleSheet("font-size: 11px; padding: 4px;")
        self.btn_change_db.clicked.connect(self.change_database_folder)
        self.btn_change_db.setEnabled(self.current_role == "admin")
        db_config_layout.addWidget(self.btn_change_db)
        sidebar_layout.addWidget(db_config_group)
        
        splitter.addWidget(sidebar_widget)
        
        # Right billing form
        form_scroll = QScrollArea()
        form_scroll.setWidgetResizable(True)
        form_scroll.setStyleSheet("QScrollArea { border: none; }")
        
        form_widget = QWidget()
        form_layout = QVBoxLayout(form_widget)
        form_layout.setContentsMargins(10, 0, 10, 0)
        form_layout.setSpacing(12)
        
        # 1. Seller Info Block (Multi-Supplier Dropdown)
        seller_group = QGroupBox("Active Supplier / Seller Profile")
        seller_group_layout = QHBoxLayout(seller_group)
        seller_group_layout.setContentsMargins(15, 10, 15, 10)
        
        self.cmb_seller = QComboBox()
        self.cmb_seller.currentTextChanged.connect(self.on_active_seller_changed)
        seller_group_layout.addWidget(self.cmb_seller, 3)
        
        self.btn_edit_seller = QPushButton("Manage Supplier Profiles")
        self.btn_edit_seller.clicked.connect(self.open_seller_dialog)
        seller_group_layout.addWidget(self.btn_edit_seller, 1)
        form_layout.addWidget(seller_group)
        
        # 2. Customer & Invoice Meta Details Block (2 columns)
        meta_layout = QHBoxLayout()
        meta_layout.setSpacing(15)
        
        # Customer Info Group
        cust_group = QGroupBox("Customer Details")
        cust_form = QFormLayout(cust_group)
        cust_form.setSpacing(8)
        cust_form.setContentsMargins(15, 15, 15, 15)
        
        cust_name_layout = QHBoxLayout()
        self.txt_cust_name = QLineEdit()
        self.txt_cust_name.setPlaceholderText("Customer Trade Name")
        cust_name_layout.addWidget(self.txt_cust_name, 3)
        
        self.btn_client_directory = QPushButton("Contacts Directory")
        self.btn_client_directory.setStyleSheet("font-size: 11px; padding: 4px 8px;")
        self.btn_client_directory.clicked.connect(self.open_client_directory)
        cust_name_layout.addWidget(self.btn_client_directory, 1)
        
        self.txt_cust_mobile = QLineEdit()
        self.txt_cust_mobile.setPlaceholderText("Phone Number (Optional)")
        
        self.txt_cust_address = QLineEdit()
        self.txt_cust_address.setPlaceholderText("Billing Address")
        
        self.txt_cust_gst = QLineEdit()
        self.txt_cust_gst.setPlaceholderText("Customer GSTIN (15-digit)")
        self.txt_cust_gst.textChanged.connect(self.on_customer_gstin_changed)
        
        self.txt_cust_pan = QLineEdit()
        self.txt_cust_pan.setPlaceholderText("Customer PAN (10-character)")
        self.txt_cust_pan.textChanged.connect(self.on_customer_pan_changed)
        
        self.txt_cust_pin = QLineEdit()
        self.txt_cust_pin.setPlaceholderText("6-digit PIN Code")
        self.txt_cust_pin.textChanged.connect(self.on_customer_pin_changed)
        
        self.txt_place_of_supply = QLineEdit()
        self.txt_place_of_supply.setPlaceholderText("e.g. Maharashtra (27)")
        
        self.chk_ship_to = QCheckBox("Ship to Different Address")
        self.chk_ship_to.stateChanged.connect(self.on_ship_to_toggled)
        
        # Shipping fields container
        self.ship_to_widget = QWidget()
        ship_to_layout = QFormLayout(self.ship_to_widget)
        ship_to_layout.setContentsMargins(0, 5, 0, 0)
        ship_to_layout.setSpacing(6)
        
        self.txt_ship_name = QLineEdit()
        self.txt_ship_name.setPlaceholderText("Recipient Name")
        self.txt_ship_mobile = QLineEdit()
        self.txt_ship_mobile.setPlaceholderText("Recipient Phone")
        self.txt_ship_address = QLineEdit()
        self.txt_ship_address.setPlaceholderText("Delivery/Shipping Destination")
        self.txt_ship_gstin = QLineEdit()
        self.txt_ship_gstin.setPlaceholderText("Shipping GSTIN (optional)")
        
        ship_to_layout.addRow("Ship-To Name", self.txt_ship_name)
        ship_to_layout.addRow("Ship-To Mobile", self.txt_ship_mobile)
        ship_to_layout.addRow("Ship-To Address", self.txt_ship_address)
        ship_to_layout.addRow("Ship-To GSTIN", self.txt_ship_gstin)
        self.ship_to_widget.setVisible(False)
        
        cust_form.addRow("Customer Name *", cust_name_layout)
        cust_form.addRow("Mobile No.", self.txt_cust_mobile)
        cust_form.addRow("Address", self.txt_cust_address)
        cust_form.addRow("GSTIN (Optional)", self.txt_cust_gst)
        cust_form.addRow("PAN (Optional)", self.txt_cust_pan)
        cust_form.addRow("PIN Code (Optional)", self.txt_cust_pin)
        cust_form.addRow("Place of Supply *", self.txt_place_of_supply)
        cust_form.addRow("", self.chk_ship_to)
        cust_form.addRow(self.ship_to_widget)
        
        meta_layout.addWidget(cust_group, 3)
        
        # Invoice details group
        inv_group = QGroupBox("Invoice details")
        inv_form = QFormLayout(inv_group)
        inv_form.setSpacing(8)
        inv_form.setContentsMargins(15, 15, 15, 15)
        
        self.txt_inv_number = QLineEdit()
        self.txt_inv_number.setPlaceholderText("Auto-generated")
        
        self.dt_inv_date = QDateEdit(QDate.currentDate())
        self.dt_inv_date.setCalendarPopup(True)
        self.dt_inv_date.setDisplayFormat("yyyy-MM-dd")
        
        self.rad_goods = QRadioButton("Goods (Qty/Rate)")
        self.rad_services = QRadioButton("Services (Amount Only)")
        self.rad_goods.setChecked(True)
        
        self.invoice_type_group = QButtonGroup(self)
        self.invoice_type_group.addButton(self.rad_goods)
        self.invoice_type_group.addButton(self.rad_services)
        
        self.rad_goods.toggled.connect(self.on_invoice_type_changed)
        self.rad_services.toggled.connect(self.on_invoice_type_changed)
        
        type_widget = QWidget()
        type_layout = QHBoxLayout(type_widget)
        type_layout.setContentsMargins(0, 0, 0, 0)
        type_layout.addWidget(self.rad_goods)
        type_layout.addWidget(self.rad_services)
        
        self.cmb_gst_treatment = QComboBox()
        self.cmb_gst_treatment.addItems([
            "Auto-Detect (Based on GSTIN)", 
            "Intrastate (CGST + SGST)", 
            "Interstate (IGST)"
        ])
        self.cmb_gst_treatment.currentTextChanged.connect(lambda: self.calculate_invoice_totals())
        
        self.chk_rcm = QCheckBox("Reverse Charge (RCM) Applicable")
        self.chk_rcm.stateChanged.connect(lambda _: self.calculate_invoice_totals())
        
        inv_form.addRow("Invoice No. *", self.txt_inv_number)
        inv_form.addRow("Invoice Date *", self.dt_inv_date)
        inv_form.addRow("Invoice Type *", type_widget)
        inv_form.addRow("GST Treatment *", self.cmb_gst_treatment)
        self.cmb_template_style = QComboBox()
        self.cmb_template_style.addItems(["Classic Elegant", "Modern Minimalist", "Professional Compact"])
        inv_form.addRow("Invoice Style", self.cmb_template_style)
        inv_form.addRow("", self.chk_rcm)
        
        meta_layout.addWidget(inv_group, 2)
        form_layout.addLayout(meta_layout)
        
        # 3. Line Items Grid (13 columns)
        items_group = QGroupBox("Line Items")
        items_layout = QVBoxLayout(items_group)
        items_layout.setContentsMargins(10, 15, 10, 10)
        
        self.tbl_items = QTableWidget(0, 13)
        self.tbl_items.setHorizontalHeaderLabels([
            "Item Name", "HSN/SAC", "Qty", "Rate", "GST %", "Cess %", "Taxable Value", 
            "CGST", "SGST", "IGST", "Cess Amt", "Total", "Action"
        ])
        
        header = self.tbl_items.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        for i in range(1, 12):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(12, QHeaderView.Fixed)
        self.tbl_items.setColumnWidth(12, 70)
        self.tbl_items.setFixedHeight(220)
        
        self.tbl_items.itemChanged.connect(self.on_table_item_changed)
        items_layout.addWidget(self.tbl_items)
        
        quick_entry_layout = QHBoxLayout()
        quick_entry_layout.setSpacing(6)
        
        self.quick_item_name = QLineEdit()
        self.quick_item_name.setPlaceholderText("Enter Item Name")
        self.quick_item_hsn = QLineEdit()
        self.quick_item_hsn.setPlaceholderText("HSN/SAC")
        self.quick_item_hsn.setMaximumWidth(70)
        
        self.lbl_quick_qty = QLabel("Qty:")
        self.quick_item_qty = QDoubleSpinBox()
        self.quick_item_qty.setRange(0.001, 999999999999.0)
        self.quick_item_qty.setValue(1.0)
        self.quick_item_qty.setDecimals(2)
        self.quick_item_qty.setMaximumWidth(90)
        
        self.lbl_quick_rate_amt = QLabel("Rate:")
        self.quick_item_rate = QDoubleSpinBox()
        self.quick_item_rate.setRange(0.00, 999999999999.0)
        self.quick_item_rate.setDecimals(2)
        
        self.quick_item_gst = QComboBox()
        self.quick_item_gst.addItems(["0", "5", "12", "18", "28"])
        self.quick_item_gst.setCurrentText("18")
        
        self.quick_item_cess = QDoubleSpinBox()
        self.quick_item_cess.setRange(0.00, 99.0)
        self.quick_item_cess.setDecimals(2)
        self.quick_item_cess.setSuffix("%")
        self.quick_item_cess.setMaximumWidth(70)
        
        self.btn_add_item = QPushButton("Add Item")
        self.btn_add_item.setObjectName("btnSuccess")
        self.btn_add_item.setProperty("success", True)
        self.btn_add_item.clicked.connect(self.quick_add_item)
        
        quick_entry_layout.addWidget(self.quick_item_name, 3)
        quick_entry_layout.addWidget(self.quick_item_hsn, 1)
        quick_entry_layout.addWidget(self.lbl_quick_qty, 0)
        quick_entry_layout.addWidget(self.quick_item_qty, 1)
        quick_entry_layout.addWidget(self.lbl_quick_rate_amt, 0)
        quick_entry_layout.addWidget(self.quick_item_rate, 2)
        quick_entry_layout.addWidget(QLabel("GST:"), 0)
        quick_entry_layout.addWidget(self.quick_item_gst, 1)
        self.lbl_quick_cess = QLabel("Cess:")
        quick_entry_layout.addWidget(self.lbl_quick_cess, 0)
        quick_entry_layout.addWidget(self.quick_item_cess, 1)
        quick_entry_layout.addWidget(self.btn_add_item, 1)
        items_layout.addLayout(quick_entry_layout)
        form_layout.addWidget(items_group)
        
        # Calculation Summary (Bottom Layout)
        bottom_summary_layout = QHBoxLayout()
        bottom_summary_layout.setSpacing(20)
        
        words_group = QGroupBox("Invoice Info")
        words_layout = QVBoxLayout(words_group)
        self.lbl_amount_words = QLabel("Rupees Zero Only")
        self.lbl_amount_words.setWordWrap(True)
        self.lbl_amount_words.setStyleSheet("font-size: 13px; font-weight: bold; color: #475569; font-style: italic;")
        words_layout.addWidget(self.lbl_amount_words)
        
        self.lbl_rcm_summary = QLabel("")
        self.lbl_rcm_summary.setWordWrap(True)
        self.lbl_rcm_summary.setStyleSheet("font-size: 12px; font-weight: bold; color: #B91C1C;")
        words_layout.addWidget(self.lbl_rcm_summary)
        bottom_summary_layout.addWidget(words_group, 3)
        
        totals_group = QGroupBox("Calculation Summary")
        self.totals_form = QFormLayout(totals_group)
        self.totals_form.setSpacing(8)
        self.totals_form.setContentsMargins(15, 15, 15, 15)
        
        self.lbl_subtotal_val = QLabel("₹ 0.00")
        self.lbl_subtotal_val.setObjectName("subTotalLabel")
        self.lbl_cgst_val = QLabel("₹ 0.00")
        self.lbl_cgst_val.setObjectName("taxLabel")
        self.lbl_sgst_val = QLabel("₹ 0.00")
        self.lbl_sgst_val.setObjectName("taxLabel")
        self.lbl_igst_val = QLabel("₹ 0.00")
        self.lbl_igst_val.setObjectName("taxLabel")
        self.lbl_cess_val = QLabel("₹ 0.00")
        self.lbl_cess_val.setObjectName("taxLabel")
        self.lbl_round_val = QLabel("₹ 0.00")
        self.lbl_round_val.setObjectName("roundOffLabel")
        self.lbl_grand_total_val = QLabel("₹ 0.00")
        self.lbl_grand_total_val.setObjectName("grandTotalValue")
        
        self.totals_form.addRow("Taxable Value (Subtotal):", self.lbl_subtotal_val)
        self.totals_form.addRow("CGST Total:", self.lbl_cgst_val)
        self.totals_form.addRow("SGST Total:", self.lbl_sgst_val)
        self.totals_form.addRow("IGST Total:", self.lbl_igst_val)
        self.lbl_cess_title = QLabel("Cess Total:")
        self.totals_form.addRow(self.lbl_cess_title, self.lbl_cess_val)
        self.totals_form.addRow("Round Off:", self.lbl_round_val)
        self.totals_form.addRow("Grand Total:", self.lbl_grand_total_val)
        bottom_summary_layout.addWidget(totals_group, 2)
        form_layout.addLayout(bottom_summary_layout)
        
        # Core Action buttons toolbar
        actions_layout = QHBoxLayout()
        actions_layout.setSpacing(10)
        
        self.btn_action_new = QPushButton("New (Clear)")
        self.btn_action_new.clicked.connect(self.new_invoice)
        
        self.btn_action_save = QPushButton("Save Invoice")
        self.btn_action_save.setObjectName("btnPrimary")
        self.btn_action_save.setProperty("primary", True)
        self.btn_action_save.clicked.connect(self.save_invoice_data)
        
        self.btn_action_preview = QPushButton("Preview PDF")
        self.btn_action_preview.clicked.connect(self.preview_invoice_pdf)
        
        self.btn_action_export = QPushButton("Export PDF")
        self.btn_action_export.clicked.connect(self.export_invoice_pdf)
        
        self.btn_action_print = QPushButton("Print Invoice")
        self.btn_action_print.clicked.connect(self.print_invoice_pdf)
        
        self.btn_action_account = QPushButton("Accounting Export")
        self.btn_action_account.setObjectName("btnSuccess")
        self.btn_action_account.setProperty("success", True)
        self.btn_action_account.clicked.connect(self.export_accounting_data)
        
        self.btn_action_delete = QPushButton("Delete Invoice")
        self.btn_action_delete.setObjectName("btnDanger")
        self.btn_action_delete.setProperty("danger", True)
        self.btn_action_delete.clicked.connect(self.delete_current_invoice)
        self.btn_action_delete.setEnabled(False)
        self.btn_action_delete.setVisible(self.current_role == "admin")
        
        actions_layout.addWidget(self.btn_action_new)
        actions_layout.addWidget(self.btn_action_save)
        actions_layout.addWidget(self.btn_action_preview)
        actions_layout.addWidget(self.btn_action_export)
        actions_layout.addWidget(self.btn_action_print)
        actions_layout.addWidget(self.btn_action_account)
        actions_layout.addWidget(self.btn_action_delete)
        
        form_layout.addLayout(actions_layout)
        form_scroll.setWidget(form_widget)
        splitter.addWidget(form_scroll)
        splitter.setSizes([260, 940])
        
        # TAB 2: Dynamic Analytics Dashboard
        if self.current_role == "admin":
            tab2_widget = QWidget()
            self.init_dashboard_ui(tab2_widget)
            self.tab_widget.addTab(tab2_widget, "Dynamic Analytics")
            
            # TAB 3: GST Returns & Reports
            tab3_widget = QWidget()
            self.init_reports_ui(tab3_widget)
            self.tab_widget.addTab(tab3_widget, "GST Filing & Presets")
            
        # TAB 4: Timesheets & Billing
        tab4_widget = QWidget()
        self.init_timesheets_ui(tab4_widget)
        self.tab_widget.addTab(tab4_widget, "Timesheets & Billing")

        # TAB 5: Payments & Aging Outstanding
        tab5_widget = QWidget()
        self.init_payments_ui(tab5_widget)
        self.tab_widget.addTab(tab5_widget, "Payments & Aging Outstanding")

    def init_dashboard_ui(self, parent_widget):
        layout = QVBoxLayout(parent_widget)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(12)
        
        # Dynamic Query Builder Control Panel
        filter_box = QGroupBox("Dynamic Query Builder Controls")
        filter_layout = QGridLayout(filter_box)
        filter_layout.setSpacing(10)
        
        # Row 0: Dimension, Metric, Limits
        filter_layout.addWidget(QLabel("<b>Analyze By (Group By):</b>"), 0, 0)
        self.cmb_dash_group = QComboBox()
        self.cmb_dash_group.addItems(["Client/Customer", "Product/Service", "HSN/SAC Code", "Place of Supply", "Month"])
        self.cmb_dash_group.currentTextChanged.connect(lambda _: self.refresh_dashboard_data())
        filter_layout.addWidget(self.cmb_dash_group, 0, 1)
        
        filter_layout.addWidget(QLabel("<b>Measure Metric:</b>"), 0, 2)
        self.cmb_dash_measure = QComboBox()
        self.cmb_dash_measure.addItems([
            "Taxable Value (Subtotal)", 
            "Taxes (CGST+SGST+IGST)", 
            "Cess Collected", 
            "Net Grand Total"
        ])
        self.cmb_dash_measure.currentTextChanged.connect(lambda _: self.refresh_dashboard_data())
        filter_layout.addWidget(self.cmb_dash_measure, 0, 3)
        
        filter_layout.addWidget(QLabel("<b>Filter Count:</b>"), 0, 4)
        self.cmb_dash_limit = QComboBox()
        self.cmb_dash_limit.addItems(["Top 5", "Top 10", "Top 25", "All"])
        self.cmb_dash_limit.currentTextChanged.connect(lambda _: self.refresh_dashboard_data())
        filter_layout.addWidget(self.cmb_dash_limit, 0, 5)
        
        # Row 1: Period, Date parameters, and Action Buttons
        filter_layout.addWidget(QLabel("<b>Date Filter:</b>"), 1, 0)
        self.cmb_dash_period = QComboBox()
        self.cmb_dash_period.addItems(["This Month", "Last 30 Days", "Financial Year", "Custom Date Range"])
        self.cmb_dash_period.currentTextChanged.connect(self.on_dash_filter_changed)
        filter_layout.addWidget(self.cmb_dash_period, 1, 1)
        
        self.lbl_dash_start = QLabel("Start Date:")
        self.dt_dash_start = QDateEdit(QDate.currentDate().addDays(-30))
        self.dt_dash_start.setCalendarPopup(True)
        self.dt_dash_start.setDisplayFormat("yyyy-MM-dd")
        self.dt_dash_start.setEnabled(False)
        self.dt_dash_start.dateChanged.connect(lambda _: self.refresh_dashboard_data())
        
        self.lbl_dash_end = QLabel("End Date:")
        self.dt_dash_end = QDateEdit(QDate.currentDate())
        self.dt_dash_end.setCalendarPopup(True)
        self.dt_dash_end.setDisplayFormat("yyyy-MM-dd")
        self.dt_dash_end.setEnabled(False)
        self.dt_dash_end.dateChanged.connect(lambda _: self.refresh_dashboard_data())
        
        filter_layout.addWidget(self.lbl_dash_start, 1, 2, Qt.AlignRight | Qt.AlignVCenter)
        filter_layout.addWidget(self.dt_dash_start, 1, 3)
        filter_layout.addWidget(self.lbl_dash_end, 1, 4, Qt.AlignRight | Qt.AlignVCenter)
        filter_layout.addWidget(self.dt_dash_end, 1, 5)
        
        # Button Controls inside filter
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(6)
        
        self.btn_dash_refresh = QPushButton("Refresh")
        self.btn_dash_refresh.setObjectName("btnPrimary")
        self.btn_dash_refresh.setProperty("primary", True)
        self.btn_dash_refresh.clicked.connect(self.refresh_dashboard_data)
        btn_layout.addWidget(self.btn_dash_refresh)
        
        self.btn_dash_export_xls = QPushButton("Download Excel")
        self.btn_dash_export_xls.setObjectName("btnSuccess")
        self.btn_dash_export_xls.setProperty("success", True)
        self.btn_dash_export_xls.clicked.connect(self.export_dashboard_excel)
        btn_layout.addWidget(self.btn_dash_export_xls)
        
        self.btn_dash_export_pdf = QPushButton("Download PDF")
        self.btn_dash_export_pdf.clicked.connect(self.export_dashboard_pdf)
        btn_layout.addWidget(self.btn_dash_export_pdf)
        
        filter_layout.addLayout(btn_layout, 1, 6)
        layout.addWidget(filter_box)
        
        # Financial Cards Overview
        cards_group = QGroupBox("Overall Financial KPI Summary")
        cards_layout = QGridLayout(cards_group)
        cards_layout.setSpacing(15)
        
        self.dash_card_subtotal = QLabel("₹ 0.00")
        self.dash_card_subtotal.setStyleSheet("font-size: 20px; font-weight: bold; color: #1E3A8A;")
        self.dash_card_tax = QLabel("₹ 0.00")
        self.dash_card_tax.setStyleSheet("font-size: 20px; font-weight: bold; color: #0D9488;")
        self.dash_card_cess = QLabel("₹ 0.00")
        self.dash_card_cess.setStyleSheet("font-size: 20px; font-weight: bold; color: #B45309;")
        self.dash_card_grand = QLabel("₹ 0.00")
        self.dash_card_grand.setStyleSheet("font-size: 24px; font-weight: bold; color: #15803D;")
        
        cards_layout.addWidget(QLabel("<b>TAXABLE VALUE (SUBTOTAL)</b>"), 0, 0)
        cards_layout.addWidget(self.dash_card_subtotal, 1, 0)
        cards_layout.addWidget(QLabel("<b>GST TAX LIABILITY</b>"), 0, 1)
        cards_layout.addWidget(self.dash_card_tax, 1, 1)
        cards_layout.addWidget(QLabel("<b>CESS COLLECTED</b>"), 0, 2)
        cards_layout.addWidget(self.dash_card_cess, 1, 2)
        cards_layout.addWidget(QLabel("<b>NET GROSS REVENUE (GRAND TOTAL)</b>"), 0, 3)
        cards_layout.addWidget(self.dash_card_grand, 1, 3)
        layout.addWidget(cards_group)
        
        # Spacious Data Table + Custom Painter-Based Charting side-by-side splitter
        data_splitter = QSplitter(Qt.Horizontal)
        
        # Left Grid
        grid_box = QGroupBox("Tabulated Analysis Results")
        grid_layout = QVBoxLayout(grid_box)
        self.tbl_dash_results = QTableWidget(0, 3)
        self.tbl_dash_results.setHorizontalHeaderLabels(["Group By Value", "Measure (INR)", "Share %"])
        self.tbl_dash_results.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tbl_dash_results.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.tbl_dash_results.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        grid_layout.addWidget(self.tbl_dash_results)
        data_splitter.addWidget(grid_box)
        
        # Right Chart
        self.chart_widget = SimpleChartWidget()
        data_splitter.addWidget(self.chart_widget)
        
        data_splitter.setSizes([600, 700])
        layout.addWidget(data_splitter)
        layout.setStretch(2, 3)

    def init_reports_ui(self, parent_widget):
        layout = QVBoxLayout(parent_widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        title = QLabel("GSTR-1 Outward Sales Return Exporter")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #1A365D;")
        layout.addWidget(title)
        
        desc = QLabel("Prepare returns file for uploading outward supply invoices. The system will compile and "
                      "format all schedules (B2B, B2CS, HSN code splits, and Document Summary) for the selected period.")
        desc.setWordWrap(True)
        desc.setStyleSheet("font-size: 13px; color: #475569; margin-bottom: 10px;")
        layout.addWidget(desc)
        
        # Filing Period presets
        preset_group = QGroupBox("Filing Frequency & Dates Selection")
        preset_layout = QFormLayout(preset_group)
        preset_layout.setSpacing(12)
        preset_layout.setContentsMargins(15, 15, 15, 15)
        
        self.cmb_report_freq = QComboBox()
        self.cmb_report_freq.addItems(["Monthly", "Quarterly", "Custom Range"])
        self.cmb_report_freq.currentTextChanged.connect(self.on_report_frequency_changed)
        preset_layout.addRow("Filing Frequency:", self.cmb_report_freq)
        
        # Stacked widgets to hold dynamic controls
        self.report_period_stack = QStackedWidget()
        
        # 1. Monthly widget
        monthly_w = QWidget()
        monthly_layout = QHBoxLayout(monthly_w)
        monthly_layout.setContentsMargins(0, 0, 0, 0)
        self.cmb_report_month = QComboBox()
        self.cmb_report_month.addItems([
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ])
        # Set default to current month
        self.cmb_report_month.setCurrentIndex(datetime.now().month - 1)
        
        self.spn_report_month_year = QSpinBox()
        self.spn_report_month_year.setRange(2020, 2040)
        self.spn_report_month_year.setValue(datetime.now().year)
        
        monthly_layout.addWidget(self.cmb_report_month, 2)
        monthly_layout.addWidget(QLabel("Year:"), 0)
        monthly_layout.addWidget(self.spn_report_month_year, 1)
        self.report_period_stack.addWidget(monthly_w)
        
        # 2. Quarterly widget
        quarter_w = QWidget()
        quarter_layout = QHBoxLayout(quarter_w)
        quarter_layout.setContentsMargins(0, 0, 0, 0)
        self.cmb_report_quarter = QComboBox()
        self.cmb_report_quarter.addItems([
            "Q1 (Apr - Jun)", 
            "Q2 (Jul - Sep)", 
            "Q3 (Oct - Dec)", 
            "Q4 (Jan - Mar)"
        ])
        
        self.spn_report_quarter_year = QSpinBox()
        self.spn_report_quarter_year.setRange(2020, 2040)
        self.spn_report_quarter_year.setValue(datetime.now().year)
        
        quarter_layout.addWidget(self.cmb_report_quarter, 2)
        quarter_layout.addWidget(QLabel("Year:"), 0)
        quarter_layout.addWidget(self.spn_report_quarter_year, 1)
        self.report_period_stack.addWidget(quarter_w)
        
        # 3. Custom Range widget
        custom_w = QWidget()
        custom_layout = QHBoxLayout(custom_w)
        custom_layout.setContentsMargins(0, 0, 0, 0)
        self.dt_report_start = QDateEdit(QDate.currentDate().addDays(-30))
        self.dt_report_start.setCalendarPopup(True)
        self.dt_report_start.setDisplayFormat("yyyy-MM-dd")
        
        self.dt_report_end = QDateEdit(QDate.currentDate())
        self.dt_report_end.setCalendarPopup(True)
        self.dt_report_end.setDisplayFormat("yyyy-MM-dd")
        
        custom_layout.addWidget(QLabel("From:"))
        custom_layout.addWidget(self.dt_report_start)
        custom_layout.addWidget(QLabel("To:"))
        custom_layout.addWidget(self.dt_report_end)
        self.report_period_stack.addWidget(custom_w)
        
        preset_layout.addRow("Filing Period:", self.report_period_stack)
        layout.addWidget(preset_group)
        
        # GSTR Outward actions group
        action_group = QGroupBox("Generate outward returns (GSTR-1)")
        action_layout = QHBoxLayout(action_group)
        action_layout.setContentsMargins(15, 15, 15, 15)
        action_layout.setSpacing(20)
        
        self.btn_export_gstr_xls = QPushButton("Download GSTR-1 Excel")
        self.btn_export_gstr_xls.setObjectName("btnSuccess")
        self.btn_export_gstr_xls.setProperty("success", True)
        self.btn_export_gstr_xls.setFixedHeight(45)
        self.btn_export_gstr_xls.clicked.connect(self.export_gstr1_excel)
        action_layout.addWidget(self.btn_export_gstr_xls)
        
        self.btn_export_gstr_pdf = QPushButton("Download GSTR-1 PDF")
        self.btn_export_gstr_pdf.setFixedHeight(45)
        self.btn_export_gstr_pdf.clicked.connect(self.export_gstr1_pdf)
        action_layout.addWidget(self.btn_export_gstr_pdf)
        
        layout.addWidget(action_group)
        layout.addStretch()

    def on_report_frequency_changed(self, text):
        if text == "Monthly":
            self.report_period_stack.setCurrentIndex(0)
        elif text == "Quarterly":
            self.report_period_stack.setCurrentIndex(1)
        else: # Custom
            self.report_period_stack.setCurrentIndex(2)

    def get_report_dates(self):
        freq = self.cmb_report_freq.currentText()
        if freq == "Monthly":
            month_idx = self.cmb_report_month.currentIndex() + 1
            year = self.spn_report_month_year.value()
            last_day = calendar.monthrange(year, month_idx)[1]
            start = QDate(year, month_idx, 1)
            end = QDate(year, month_idx, last_day)
        elif freq == "Quarterly":
            q_idx = self.cmb_report_quarter.currentIndex()
            year = self.spn_report_quarter_year.value()
            if q_idx == 0: # Q1: Apr-Jun
                start = QDate(year, 4, 1)
                end = QDate(year, 6, 30)
            elif q_idx == 1: # Q2: Jul-Sep
                start = QDate(year, 7, 1)
                end = QDate(year, 9, 30)
            elif q_idx == 2: # Q3: Oct-Dec
                start = QDate(year, 10, 1)
                end = QDate(year, 12, 31)
            else: # Q4: Jan-Mar
                start = QDate(year, 1, 1)
                end = QDate(year, 3, 31)
        else: # Custom
            start = self.dt_report_start.date()
            end = self.dt_report_end.date()
            
        return start.toString("yyyy-MM-dd"), end.toString("yyyy-MM-dd")

    def get_dashboard_dates(self):
        period = self.cmb_dash_period.currentText()
        today = QDate.currentDate()
        
        if period == "This Month":
            start = QDate(today.year(), today.month(), 1)
            end = today
        elif period == "Last 30 Days":
            start = today.addDays(-30)
            end = today
        elif period == "Financial Year":
            year = today.year()
            if today.month() < 4:
                year -= 1
            start = QDate(year, 4, 1)
            end = today
        else: # Custom
            start = self.dt_dash_start.date()
            end = self.dt_dash_end.date()
            
        return start.toString("yyyy-MM-dd"), end.toString("yyyy-MM-dd")

    def on_dash_filter_changed(self, text):
        is_custom = (text == "Custom Date Range")
        self.dt_dash_start.setEnabled(is_custom)
        self.dt_dash_end.setEnabled(is_custom)
        self.refresh_dashboard_data()

    def refresh_dashboard_data(self):
        """Queries SQLite database dynamically based on dimension and metric selections."""
        if self.current_role != "admin":
            return
        start_str, end_str = self.get_dashboard_dates()
        
        group_by = self.cmb_dash_group.currentText()
        measure = self.cmb_dash_measure.currentText()
        limit_text = self.cmb_dash_limit.currentText()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # 1. KPI Aggregates (Always updated)
            cursor.execute("""
                SELECT 
                    SUM(subtotal) as sub, SUM(cgst_total) as cgst, SUM(sgst_total) as sgst,
                    SUM(igst_total) as igst, SUM(cess_total) as cess, SUM(grand_total) as grand
                FROM invoices 
                WHERE date BETWEEN ? AND ?
            """, (start_str, end_str))
            kpi = cursor.fetchone()
            
            sub = kpi["sub"] or 0.0
            tax = (kpi["cgst"] or 0.0) + (kpi["sgst"] or 0.0) + (kpi["igst"] or 0.0)
            cess = kpi["cess"] or 0.0
            grand = kpi["grand"] or 0.0
            
            self.dash_card_subtotal.setText(f"₹ {sub:,.2f}")
            self.dash_card_tax.setText(f"₹ {tax:,.2f}")
            self.dash_card_cess.setText(f"₹ {cess:,.2f}")
            self.dash_card_grand.setText(f"₹ {grand:,.2f}")
            
            # 2. Build Dynamic SQL Query
            dim_field = ""
            is_items_join = False
            
            if group_by == "Client/Customer":
                dim_field = "customer_name"
            elif group_by == "Product/Service":
                dim_field = "item.name"
                is_items_join = True
            elif group_by == "HSN/SAC Code":
                dim_field = "item.hsn"
                is_items_join = True
            elif group_by == "Place of Supply":
                dim_field = "place_of_supply"
            else: # Month
                dim_field = "SUBSTR(date, 1, 7)"
                
            meas_field = ""
            if is_items_join:
                if measure == "Taxable Value (Subtotal)":
                    meas_field = "SUM(item.taxable_amount)"
                elif measure == "Taxes (CGST+SGST+IGST)":
                    meas_field = "SUM(item.cgst_amount+item.sgst_amount+item.igst_amount)"
                elif measure == "Cess Collected":
                    meas_field = "SUM(item.cess_amount)"
                else: # Grand Total
                    meas_field = "SUM(item.total_amount)"
            else:
                if measure == "Taxable Value (Subtotal)":
                    meas_field = "SUM(subtotal)"
                elif measure == "Taxes (CGST+SGST+IGST)":
                    meas_field = "SUM(cgst_total+sgst_total+igst_total)"
                elif measure == "Cess Collected":
                    meas_field = "SUM(cess_total)"
                else: # Grand Total
                    meas_field = "SUM(grand_total)"
                    
            limit_clause = ""
            if limit_text == "Top 5":
                limit_clause = "LIMIT 5"
            elif limit_text == "Top 10":
                limit_clause = "LIMIT 10"
            elif limit_text == "Top 25":
                limit_clause = "LIMIT 25"
                
            if is_items_join:
                sql = f"""
                    SELECT {dim_field} AS label, {meas_field} AS val
                    FROM invoice_items item
                    JOIN invoices inv ON item.invoice_number = inv.invoice_number
                    WHERE inv.date BETWEEN ? AND ?
                    GROUP BY label
                    ORDER BY val DESC
                    {limit_clause}
                """
            else:
                sql = f"""
                    SELECT {dim_field} AS label, {meas_field} AS val
                    FROM invoices
                    WHERE date BETWEEN ? AND ?
                    GROUP BY label
                    ORDER BY val DESC
                    {limit_clause}
                """
                
            cursor.execute(sql, (start_str, end_str))
            rows = cursor.fetchall()
            
            self.active_dyn_data = []
            self.tbl_dash_results.setRowCount(0)
            
            total_val = sum(r["val"] for r in rows if r["val"])
            if total_val == 0:
                total_val = 1.0
                
            for idx, r in enumerate(rows):
                label = r["label"] if r["label"] else "N/A"
                val = r["val"] if r["val"] else 0.0
                share = (val / total_val) * 100.0
                
                self.active_dyn_data.append((label, val))
                
                self.tbl_dash_results.insertRow(idx)
                self.tbl_dash_results.setItem(idx, 0, QTableWidgetItem(label))
                self.tbl_dash_results.setItem(idx, 1, QTableWidgetItem(f"₹ {val:,.2f}"))
                self.tbl_dash_results.setItem(idx, 2, QTableWidgetItem(f"{share:.2f}%"))
                
            # Update custom chart widget
            self.chart_widget.setData(self.active_dyn_data, measure)
            
        except Exception as e:
            print(f"Error loading dynamic dashboard stats: {e}")
        finally:
            conn.close()

    def export_dashboard_excel(self):
        start, end = self.get_dashboard_dates()
        dim = self.cmb_dash_group.currentText()
        meas = self.cmb_dash_measure.currentText()
        
        default_name = self.get_safe_dialog_path(f"Analysis_{dim}_{start}_to_{end}.xlsx")
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Export Dynamic Dashboard Excel", default_name, "Excel Files (*.xlsx)"
        )
        if save_path:
            self.statusBar().showMessage("Exporting Dynamic Analysis Excel...")
            extra_args = {"dim": dim, "meas": meas, "data": self.active_dyn_data}
            self.dash_worker = ReportWorker("dash_dyn", start, end, save_path, "excel", extra_args)
            self.dash_worker.finished.connect(self.on_report_export_finished)
            self.dash_worker.start()

    def export_dashboard_pdf(self):
        start, end = self.get_dashboard_dates()
        dim = self.cmb_dash_group.currentText()
        meas = self.cmb_dash_measure.currentText()
        
        default_name = self.get_safe_dialog_path(f"Analysis_{dim}_{start}_to_{end}.pdf")
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Export Dynamic Dashboard PDF", default_name, "PDF Files (*.pdf)"
        )
        if save_path:
            self.statusBar().showMessage("Exporting Dynamic Analysis PDF...")
            extra_args = {"dim": dim, "meas": meas, "data": self.active_dyn_data}
            self.dash_worker = ReportWorker("dash_dyn", start, end, save_path, "pdf", extra_args)
            self.dash_worker.finished.connect(self.on_report_export_finished)
            self.dash_worker.start()

    def export_gstr1_excel(self):
        start, end = self.get_report_dates()
        default_name = self.get_safe_dialog_path(f"GSTR1_Return_{start}_to_{end}.xlsx")
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Export GSTR-1 Excel Outward Returns", default_name, "Excel Files (*.xlsx)"
        )
        if save_path:
            self.statusBar().showMessage("Exporting GSTR-1 Excel Return file...")
            self.gstr_worker = ReportWorker("gstr1", start, end, save_path, "excel")
            self.gstr_worker.finished.connect(self.on_report_export_finished)
            self.gstr_worker.start()

    def export_gstr1_pdf(self):
        start, end = self.get_report_dates()
        default_name = self.get_safe_dialog_path(f"GSTR1_Return_Summary_{start}_to_{end}.pdf")
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Export GSTR-1 PDF Outward Return Summary", default_name, "PDF Files (*.pdf)"
        )
        if save_path:
            self.statusBar().showMessage("Exporting GSTR-1 PDF Return Summary...")
            self.gstr_worker = ReportWorker("gstr1", start, end, save_path, "pdf")
            self.gstr_worker.finished.connect(self.on_report_export_finished)
            self.gstr_worker.start()

    def on_report_export_finished(self, success, result_msg):
        self.statusBar().clearMessage()
        if success:
            QMessageBox.information(self, "Export Success", f"Report generated and saved successfully at:\n{result_msg}")
        else:
            QMessageBox.critical(self, "Export Failed", f"Failed to generate report:\n{result_msg}")

    def load_seller_profiles_list(self):
        settings = get_settings()
        sellers = settings.get("sellers", {})
        active = settings.get("active_seller_id", "")
        
        self.cmb_seller.blockSignals(True)
        self.cmb_seller.clear()
        for key in sellers.keys():
            self.cmb_seller.addItem(key)
        if active in sellers:
            self.cmb_seller.setCurrentText(active)
        self.cmb_seller.blockSignals(False)

    def on_active_seller_changed(self, trade_name):
        if not trade_name:
            return
        settings = get_settings()
        if trade_name in settings.get("sellers", {}):
            settings["active_seller_id"] = trade_name
            settings["seller"] = settings["sellers"][trade_name]
            save_settings(settings)
            
            safe_p = "".join([c if c.isalnum() else "_" for c in trade_name])
            logo_target = os.path.join(get_database_dir(), f"{safe_p}_logo.png")
            std_logo = os.path.join(get_database_dir(), "logo.png")
            
            if os.path.exists(logo_target):
                try:
                    shutil.copy2(logo_target, std_logo)
                except Exception:
                    pass
            else:
                try:
                    if os.path.exists(std_logo):
                        os.remove(std_logo)
                except Exception:
                    pass
            self.calculate_invoice_totals()

    def open_client_directory(self):
        from app.gui.customer_dialog import CustomerDialog
        dialog = CustomerDialog(self)
        if dialog.exec_() == CustomerDialog.Accepted:
            cust = dialog.selected_customer
            if cust:
                self.txt_cust_name.setText(cust.get("name", ""))
                self.txt_cust_mobile.setText(cust.get("mobile", ""))
                self.txt_cust_address.setText(cust.get("address", ""))
                self.txt_cust_gst.setText(cust.get("gstin", ""))
                self.txt_cust_pan.setText(cust.get("pan", ""))
                self.txt_cust_pin.setText(cust.get("pin", ""))
                self.txt_place_of_supply.setText(cust.get("place_of_supply", ""))

    def on_ship_to_toggled(self, state):
        self.ship_to_widget.setVisible(state == Qt.Checked)

    def on_customer_gstin_changed(self, text):
        gstin = text.strip().upper()
        if not gstin:
            self.txt_cust_gst.setStyleSheet("")
        elif validate_gstin(gstin):
            self.txt_cust_gst.setStyleSheet("border: 1.5px solid #16A34A;")
            if len(gstin) >= 2:
                code = gstin[:2]
                if code in STATE_CODES:
                    self.txt_place_of_supply.setText(f"{STATE_CODES[code]} ({code})")
        else:
            self.txt_cust_gst.setStyleSheet("border: 1.5px solid #DC2626;")
        self.calculate_invoice_totals()

    def on_customer_pan_changed(self, text):
        pan = text.strip().upper()
        if not pan:
            self.txt_cust_pan.setStyleSheet("")
        elif validate_pan(pan):
            self.txt_cust_pan.setStyleSheet("border: 1.5px solid #16A34A;")
        else:
            self.txt_cust_pan.setStyleSheet("border: 1.5px solid #DC2626;")

    def on_customer_pin_changed(self, text):
        pin = text.strip()
        if not pin:
            self.txt_cust_pin.setStyleSheet("")
        elif validate_pin(pin):
            self.txt_cust_pin.setStyleSheet("border: 1.5px solid #16A34A;")
        else:
            self.txt_cust_pin.setStyleSheet("border: 1.5px solid #DC2626;")

    def get_effective_gst_treatment(self):
        selection = self.cmb_gst_treatment.currentText()
        if "Intrastate" in selection:
            return "CGST_SGST"
        elif "Interstate" in selection:
            return "IGST"
            
        settings = get_settings()
        seller_gstin = settings.get("seller", {}).get("gstin", "").strip()
        buyer_gstin = self.txt_cust_gst.text().strip()
        
        if len(seller_gstin) >= 2 and len(buyer_gstin) >= 2:
            seller_state = seller_gstin[:2]
            buyer_state = buyer_gstin[:2]
            if seller_state.isdigit() and buyer_state.isdigit():
                if seller_state == buyer_state:
                    return "CGST_SGST"
                else:
                    return "IGST"
        return "CGST_SGST"

    def on_invoice_type_changed(self):
        is_services = self.rad_services.isChecked()
        self.tbl_items.setColumnHidden(2, is_services)
        self.tbl_items.setColumnHidden(3, is_services)
        self.tbl_items.setColumnHidden(5, is_services)
        self.tbl_items.setColumnHidden(10, is_services)
        if hasattr(self, 'lbl_cess_title') and self.lbl_cess_title:
            self.lbl_cess_title.setVisible(not is_services)
        if hasattr(self, 'lbl_cess_val') and self.lbl_cess_val:
            self.lbl_cess_val.setVisible(not is_services)
        
        header_item = self.tbl_items.horizontalHeaderItem(6)
        if header_item:
            header_item.setText("Amount" if is_services else "Taxable Value")
            
        self.chk_ship_to.setVisible(not is_services)
        if is_services:
            self.chk_ship_to.setChecked(False)
            
        self.lbl_quick_qty.setVisible(not is_services)
        self.quick_item_qty.setVisible(not is_services)
        self.lbl_quick_cess.setVisible(not is_services)
        self.quick_item_cess.setVisible(not is_services)
        if is_services:
            self.quick_item_cess.setValue(0.0)
        self.lbl_quick_rate_amt.setText("Amount:" if is_services else "Rate:")
        self.quick_item_name.setPlaceholderText("Enter Service Description" if is_services else "Enter Item Name")
        self.quick_item_hsn.setPlaceholderText("SAC" if is_services else "HSN")
        
        for r in range(self.tbl_items.rowCount()):
            cell = self.tbl_items.item(r, 6)
            if cell:
                if is_services:
                    cell.setFlags(cell.flags() | Qt.ItemIsEditable)
                else:
                    cell.setFlags(cell.flags() & ~Qt.ItemIsEditable)
                    
        self.calculate_invoice_totals()

    def update_license_display(self):
        valid, msg, remaining = check_license_status()
        self.lbl_license_status.setText(f"License Status: {msg}")
        if not valid:
            QMessageBox.critical(
                self, "License Expired / Invalid",
                f"Validation Alert:\n{msg}\n\nThe application will now terminate."
            )
            sys.exit(0)
            
        if remaining > 0 and remaining <= 7:
            self.lbl_license_status.setStyleSheet("font-weight: bold; font-size: 13px; color: #B91C1C; margin-right: 10px;")
        elif remaining == -1:
            self.lbl_license_status.setStyleSheet("font-weight: bold; font-size: 13px; color: #1E3A8A; margin-right: 10px;")
        else:
            self.lbl_license_status.setStyleSheet("font-weight: bold; font-size: 13px; color: #15803D; margin-right: 10px;")

    def reset_license_key(self):
        confirm = QMessageBox.question(
            self, "Reset License",
            "Are you sure you want to clear your current license registration?\n\n"
            "This will close the application and prompt you to input a new key on next start.",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm == QMessageBox.Yes:
            reset_license()
            QMessageBox.information(self, "License Reset", "License key removed successfully. Re-launching application...")
            subprocess.Popen([sys.executable] + sys.argv)
            QApplication.quit()

    def change_database_folder(self):
        current_dir = get_database_dir()
        selected_dir = QFileDialog.getExistingDirectory(
            self, "Select Database Directory", current_dir
        )
        if selected_dir:
            selected_dir = os.path.abspath(selected_dir)
            if update_database_path(selected_dir):
                self.txt_db_path.setText(selected_dir)
                self.load_seller_profiles_list()
                self.refresh_invoice_list()
                self.new_invoice()
                QMessageBox.information(
                    self, "Database Changed",
                    f"Database location changed successfully!\nFiles have been ported to:\n{selected_dir}"
                )
            else:
                QMessageBox.critical(
                    self, "Error", "Failed to update and transfer files to the new database path."
                )

    def open_seller_dialog(self):
        dialog = SellerDialog(self)
        if dialog.exec_() == SellerDialog.Accepted:
            self.load_seller_profiles_list()
            self.calculate_invoice_totals()

    def refresh_invoice_list(self):
        query = self.txt_search.text().strip()
        invoices = search_invoices(query)
        self.lst_invoices.clear()
        for inv in invoices:
            inv_no = inv.get("invoice_number", "")
            customer = inv.get("customer", {}).get("name", "Unknown Customer")
            date = inv.get("date", "")
            grand_total = inv.get("summary", {}).get("grand_total", 0.0)
            display_text = f"{inv_no} - {customer}\nDate: {date} | Total: ₹{grand_total:,.2f}"
            
            item = QListWidgetItem(display_text)
            item.setData(Qt.UserRole, inv_no)
            self.lst_invoices.addItem(item)

    def new_invoice(self):
        self.current_invoice_id = None
        self.txt_cust_name.clear()
        self.txt_cust_mobile.clear()
        self.txt_cust_address.clear()
        self.txt_cust_gst.clear()
        self.txt_cust_pan.clear()
        self.txt_cust_pin.clear()
        self.txt_place_of_supply.clear()
        
        self.txt_cust_gst.setStyleSheet("")
        self.txt_cust_pan.setStyleSheet("")
        self.txt_cust_pin.setStyleSheet("")
        
        self.chk_ship_to.setChecked(False)
        self.txt_ship_name.clear()
        self.txt_ship_mobile.clear()
        self.txt_ship_address.clear()
        self.txt_ship_gstin.clear()
        
        next_no = get_next_invoice_number()
        self.txt_inv_number.setText(next_no)
        self.dt_inv_date.setDate(QDate.currentDate())
        self.rad_goods.setChecked(True)
        self.cmb_gst_treatment.setCurrentIndex(0)
        self.chk_rcm.setChecked(False)
        
        self.tbl_items.itemChanged.disconnect(self.on_table_item_changed)
        self.tbl_items.setRowCount(0)
        self.tbl_items.itemChanged.connect(self.on_table_item_changed)
        
        self.quick_item_name.clear()
        self.quick_item_hsn.clear()
        self.quick_item_qty.setValue(1.0)
        self.quick_item_rate.setValue(0.0)
        self.quick_item_gst.setCurrentText("18")
        self.quick_item_cess.setValue(0.0)
        
        self.btn_action_delete.setEnabled(False)
        self.calculate_invoice_totals()

    def quick_add_item(self):
        is_services = self.rad_services.isChecked()
        name = self.quick_item_name.text().strip()
        hsn = self.quick_item_hsn.text().strip()
        gst = float(self.quick_item_gst.currentText())
        cess = 0.0 if is_services else self.quick_item_cess.value()
        
        if not name:
            QMessageBox.warning(self, "Validation Error", "Item/Service Description is required.")
            self.quick_item_name.setFocus()
            return
            
        if is_services:
            amount = self.quick_item_rate.value()
            qty = 1.0
            rate = amount
        else:
            qty = self.quick_item_qty.value()
            rate = self.quick_item_rate.value()
            if qty <= 0:
                QMessageBox.warning(self, "Validation Error", "Quantity must be greater than zero.")
                self.quick_item_qty.setFocus()
                return
                
        row_idx = self.tbl_items.rowCount()
        self.tbl_items.itemChanged.disconnect(self.on_table_item_changed)
        self.tbl_items.insertRow(row_idx)
        
        self.tbl_items.setItem(row_idx, 0, QTableWidgetItem(name))
        self.tbl_items.setItem(row_idx, 1, QTableWidgetItem(hsn))
        
        qty_cell = QTableWidgetItem(f"{qty:g}")
        qty_cell.setTextAlignment(Qt.AlignCenter)
        self.tbl_items.setItem(row_idx, 2, qty_cell)
        
        rate_cell = QTableWidgetItem(f"{rate:.2f}")
        rate_cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.tbl_items.setItem(row_idx, 3, rate_cell)
        
        gst_cell = QTableWidgetItem(f"{gst:g}")
        gst_cell.setTextAlignment(Qt.AlignCenter)
        self.tbl_items.setItem(row_idx, 4, gst_cell)
        
        cess_cell = QTableWidgetItem(f"{cess:g}")
        cess_cell.setTextAlignment(Qt.AlignCenter)
        self.tbl_items.setItem(row_idx, 5, cess_cell)
        
        taxable_val = qty * rate
        taxable_item = QTableWidgetItem(f"{taxable_val:.2f}")
        taxable_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        if is_services:
            taxable_item.setFlags(taxable_item.flags() | Qt.ItemIsEditable)
        else:
            taxable_item.setFlags(taxable_item.flags() & ~Qt.ItemIsEditable)
        self.tbl_items.setItem(row_idx, 6, taxable_item)
        
        for col in range(7, 12):
            cell = QTableWidgetItem("0.00")
            cell.setFlags(cell.flags() & ~Qt.ItemIsEditable)
            cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_items.setItem(row_idx, col, cell)
            
        btn_del = QPushButton("Delete")
        btn_del.setObjectName("btnDanger")
        btn_del.setProperty("danger", True)
        btn_del.setStyleSheet("padding: 2px 5px; font-size: 11px;")
        btn_del.clicked.connect(lambda _, r=row_idx: self.remove_table_row(r))
        self.tbl_items.setCellWidget(row_idx, 12, btn_del)
        
        self.tbl_items.itemChanged.connect(self.on_table_item_changed)
        self.update_table_row_calculations(row_idx)
        
        self.quick_item_name.clear()
        self.quick_item_hsn.clear()
        self.quick_item_qty.setValue(1.0)
        self.quick_item_rate.setValue(0.0)
        self.quick_item_gst.setCurrentText("18")
        self.quick_item_cess.setValue(0.0)
        self.quick_item_name.setFocus()

    def remove_table_row(self, row_idx):
        button = self.sender()
        if button:
            for r in range(self.tbl_items.rowCount()):
                if self.tbl_items.cellWidget(r, 12) == button:
                    self.tbl_items.removeRow(r)
                    break
        self.calculate_invoice_totals()

    def on_table_item_changed(self, item):
        row = item.row()
        col = item.column()
        is_services = self.rad_services.isChecked()
        trigger_cols = [1, 4, 5, 6] if is_services else [1, 2, 3, 4, 5]
        if col in trigger_cols:
            self.tbl_items.itemChanged.disconnect(self.on_table_item_changed)
            self.update_table_row_calculations(row)
            self.tbl_items.itemChanged.connect(self.on_table_item_changed)

    def update_table_row_calculations(self, row):
        is_services = self.rad_services.isChecked()
        gst_treatment = self.get_effective_gst_treatment()
        is_intrastate = (gst_treatment == "CGST_SGST")
        is_rcm = self.chk_rcm.isChecked()
        
        try:
            gst_val = float(self.tbl_items.item(row, 4).text())
        except (ValueError, AttributeError):
            gst_val = 0.0
            self.tbl_items.item(row, 4).setText("0")
            
        try:
            cess_val = float(self.tbl_items.item(row, 5).text())
        except (ValueError, AttributeError):
            cess_val = 0.0
            self.tbl_items.item(row, 5).setText("0")

        if is_services:
            try:
                taxable = float(self.tbl_items.item(row, 6).text())
            except (ValueError, AttributeError):
                taxable = 0.0
                self.tbl_items.item(row, 6).setText("0.00")
            
            self.tbl_items.item(row, 2).setText("1")
            self.tbl_items.item(row, 3).setText(f"{taxable:.2f}")
        else:
            try:
                qty = float(self.tbl_items.item(row, 2).text())
            except (ValueError, AttributeError):
                qty = 0.0
                self.tbl_items.item(row, 2).setText("0")
                
            try:
                rate = float(self.tbl_items.item(row, 3).text())
            except (ValueError, AttributeError):
                rate = 0.0
                self.tbl_items.item(row, 3).setText("0.00")
                
            taxable = qty * rate
            self.tbl_items.item(row, 6).setText(f"{taxable:.2f}")

        cgst_amt = 0.0
        sgst_amt = 0.0
        igst_amt = 0.0
        if is_intrastate:
            cgst_amt = taxable * ((gst_val / 2.0) / 100.0)
            sgst_amt = taxable * ((gst_val / 2.0) / 100.0)
        else:
            igst_amt = taxable * (gst_val / 100.0)
            
        cess_amt = taxable * (cess_val / 100.0)
        
        self.tbl_items.item(row, 7).setText(f"{cgst_amt:.2f}")
        self.tbl_items.item(row, 8).setText(f"{sgst_amt:.2f}")
        self.tbl_items.item(row, 9).setText(f"{igst_amt:.2f}")
        self.tbl_items.item(row, 10).setText(f"{cess_amt:.2f}")
        
        if is_rcm:
            total_amt = taxable + cess_amt
        else:
            total_amt = taxable + cgst_amt + sgst_amt + igst_amt + cess_amt
        self.tbl_items.item(row, 11).setText(f"{total_amt:.2f}")
        self.calculate_invoice_totals()

    def calculate_invoice_totals(self):
        subtotal = 0.0
        cgst_total = 0.0
        sgst_total = 0.0
        igst_total = 0.0
        cess_total = 0.0
        
        gst_treatment = self.get_effective_gst_treatment()
        is_intrastate = (gst_treatment == "CGST_SGST")
        is_rcm = self.chk_rcm.isChecked()
        
        self.tbl_items.setColumnHidden(7, not is_intrastate)
        self.tbl_items.setColumnHidden(8, not is_intrastate)
        self.tbl_items.setColumnHidden(9, is_intrastate)
        
        for r in range(self.tbl_items.rowCount()):
            try:
                taxable = float(self.tbl_items.item(r, 6).text())
                cgst = float(self.tbl_items.item(r, 7).text())
                sgst = float(self.tbl_items.item(r, 8).text())
                igst = float(self.tbl_items.item(r, 9).text())
                cess = float(self.tbl_items.item(r, 10).text())
                
                subtotal += taxable
                cgst_total += cgst
                sgst_total += sgst
                igst_total += igst
                cess_total += cess
            except (ValueError, AttributeError):
                pass
                
        if is_rcm:
            raw_grand_total = subtotal + cess_total
            rcm_tax = cgst_total + sgst_total + igst_total
            self.lbl_rcm_summary.setText(f"* TAX PAYABLE UNDER RCM: ₹ {rcm_tax:,.2f} (Excluded from Grand Total)")
        else:
            raw_grand_total = subtotal + cgst_total + sgst_total + igst_total + cess_total
            self.lbl_rcm_summary.setText("")
            
        grand_total = round(raw_grand_total)
        round_off = grand_total - raw_grand_total
        
        self.lbl_subtotal_val.setText(f"₹ {subtotal:,.2f}")
        self.lbl_cgst_val.setText(f"₹ {cgst_total:,.2f}")
        self.lbl_sgst_val.setText(f"₹ {sgst_total:,.2f}")
        self.lbl_igst_val.setText(f"₹ {igst_total:,.2f}")
        self.lbl_cess_val.setText(f"₹ {cess_total:,.2f}")
        self.lbl_round_val.setText(f"₹ {round_off:+.2f}")
        self.lbl_grand_total_val.setText(f"₹ {grand_total:,.2f}")
        
        words = num_to_words_indian(grand_total)
        self.lbl_amount_words.setText(words)

    def extract_invoice_data(self):
        inv_no = self.txt_inv_number.text().strip()
        inv_date = self.dt_inv_date.date().toString("yyyy-MM-dd")
        invoice_type = "Services" if self.rad_services.isChecked() else "Goods"
        gst_treatment = self.get_effective_gst_treatment()
        place_of_supply = self.txt_place_of_supply.text().strip()
        rcm_applicable = self.chk_rcm.isChecked()
        
        cust_name = self.txt_cust_name.text().strip()
        cust_mobile = self.txt_cust_mobile.text().strip()
        cust_address = self.txt_cust_address.text().strip()
        cust_gst = self.txt_cust_gst.text().strip().upper()
        cust_pan = self.txt_cust_pan.text().strip().upper()
        cust_pin = self.txt_cust_pin.text().strip()
        
        if not inv_no:
            QMessageBox.warning(self, "Validation Error", "Invoice Number is required.")
            self.txt_inv_number.setFocus()
            return None
        if not cust_name:
            QMessageBox.warning(self, "Validation Error", "Customer Name is required.")
            self.txt_cust_name.setFocus()
            return None
        if not place_of_supply:
            QMessageBox.warning(self, "Validation Error", "Place of Supply is required.")
            self.txt_place_of_supply.setFocus()
            return None
            
        if cust_gst and not validate_gstin(cust_gst):
            QMessageBox.critical(self, "Validation Error", "Invalid Customer GSTIN format.")
            self.txt_cust_gst.setFocus()
            return None
        if cust_pan and not validate_pan(cust_pan):
            QMessageBox.critical(self, "Validation Error", "Invalid Customer PAN format.")
            self.txt_cust_pan.setFocus()
            return None
        if cust_pin and not validate_pin(cust_pin):
            QMessageBox.critical(self, "Validation Error", "Invalid PIN Code format.")
            self.txt_cust_pin.setFocus()
            return None
            
        row_count = self.tbl_items.rowCount()
        if row_count == 0:
            QMessageBox.warning(self, "Validation Error", "Please add at least one line item.")
            return None
            
        items = []
        for r in range(row_count):
            try:
                name_cell = self.tbl_items.item(r, 0)
                name = name_cell.text().strip() if name_cell else ""
                
                hsn_cell = self.tbl_items.item(r, 1)
                hsn = hsn_cell.text().strip() if hsn_cell else ""
                
                qty_cell = self.tbl_items.item(r, 2)
                qty = float(qty_cell.text()) if qty_cell and qty_cell.text() else 0.0
                
                rate_cell = self.tbl_items.item(r, 3)
                rate = float(rate_cell.text()) if rate_cell and rate_cell.text() else 0.0
                
                gst_cell = self.tbl_items.item(r, 4)
                gst_rate = float(gst_cell.text()) if gst_cell and gst_cell.text() else 0.0
                
                cess_cell = self.tbl_items.item(r, 5)
                cess_rate = float(cess_cell.text()) if cess_cell and cess_cell.text() else 0.0
                
                taxable_cell = self.tbl_items.item(r, 6)
                taxable = float(taxable_cell.text()) if taxable_cell and taxable_cell.text() else 0.0
                
                cgst_cell = self.tbl_items.item(r, 7)
                cgst_amt = float(cgst_cell.text()) if cgst_cell and cgst_cell.text() else 0.0
                
                sgst_cell = self.tbl_items.item(r, 8)
                sgst_amt = float(sgst_cell.text()) if sgst_cell and sgst_cell.text() else 0.0
                
                igst_cell = self.tbl_items.item(r, 9)
                igst_amt = float(igst_cell.text()) if igst_cell and igst_cell.text() else 0.0
                
                cess_amt_cell = self.tbl_items.item(r, 10)
                cess_amt = float(cess_amt_cell.text()) if cess_amt_cell and cess_amt_cell.text() else 0.0
                
                total_cell = self.tbl_items.item(r, 11)
                total = float(total_cell.text()) if total_cell and total_cell.text() else 0.0
            except (ValueError, AttributeError):
                QMessageBox.warning(
                    self, "Validation Error", 
                    f"Line item {r+1} has invalid or blank fields. Please double check all numbers."
                )
                return None
            
            if not name:
                QMessageBox.warning(self, "Validation Error", f"Line item {r+1} description cannot be empty.")
                return None
                
            items.append({
                "name": name,
                "hsn": hsn,
                "qty": qty,
                "rate": rate,
                "gst_rate": gst_rate,
                "cess_rate": cess_rate,
                "taxable_amount": taxable,
                "cgst_amount": cgst_amt,
                "sgst_amount": sgst_amt,
                "igst_amount": igst_amt,
                "cess_amount": cess_amt,
                "total_amount": total
            })
            
        ship_to = {}
        if invoice_type == "Goods" and self.chk_ship_to.isChecked():
            ship_to = {
                "name": self.txt_ship_name.text().strip(),
                "mobile": self.txt_ship_mobile.text().strip(),
                "address": self.txt_ship_address.text().strip(),
                "gstin": self.txt_ship_gstin.text().strip().upper()
            }
            if not ship_to["name"] or not ship_to["address"]:
                QMessageBox.warning(self, "Validation Error", "Ship-To Name and Address are required when shipping is active.")
                return None
                
        settings = get_settings()
        active_seller_key = self.cmb_seller.currentText()
        seller = settings.get("sellers", {}).get(active_seller_key, settings.get("seller", {}))
        
        subtotal = sum(item["taxable_amount"] for item in items)
        cgst_total = sum(item["cgst_amount"] for item in items)
        sgst_total = sum(item["sgst_amount"] for item in items)
        igst_total = sum(item["igst_amount"] for item in items)
        cess_total = sum(item["cess_amount"] for item in items)
        
        if rcm_applicable:
            raw_grand = subtotal + cess_total
        else:
            raw_grand = subtotal + cgst_total + sgst_total + igst_total + cess_total
            
        grand_total = round(raw_grand)
        round_off = grand_total - raw_grand
        
        invoice_data = {
            "invoice_number": inv_no,
            "date": inv_date,
            "invoice_type": invoice_type,
            "template_style": self.cmb_template_style.currentText(),
            "gst_treatment": gst_treatment,
            "place_of_supply": place_of_supply,
            "rcm": rcm_applicable,
            "customer": {
                "name": cust_name,
                "mobile": cust_mobile,
                "address": cust_address,
                "gstin": cust_gst,
                "pan": cust_pan,
                "pin": cust_pin
            },
            "ship_to": ship_to,
            "seller": seller,
            "items": items,
            "summary": {
                "subtotal": subtotal,
                "cgst_total": cgst_total,
                "sgst_total": sgst_total,
                "igst_total": igst_total,
                "cess_total": cess_total,
                "round_off": round_off,
                "grand_total": grand_total
            }
        }
        return invoice_data

    def save_invoice_data(self):
        invoice_data = self.extract_invoice_data()
        if not invoice_data:
            return False
            
        # Restrict staff from overwriting existing invoices
        invoice_number = invoice_data["invoice_number"]
        existing = get_invoice(invoice_number)
        if existing and self.current_role != "admin":
            QMessageBox.critical(self, "Access Denied", "Staff users do not have permissions to modify existing invoices.")
            return False
            
        if save_invoice(invoice_data):
            self.current_invoice_id = invoice_data["invoice_number"]
            self.btn_action_delete.setEnabled(True)
            self.refresh_invoice_list()
            self.refresh_dashboard_data()
            QMessageBox.information(self, "Success", f"Invoice {self.current_invoice_id} saved successfully!")
            return True
        else:
            QMessageBox.critical(self, "Database Error", "Failed to save invoice record.")
            return False

    def load_selected_invoice(self, item):
        invoice_id = item.data(Qt.UserRole)
        invoice_data = get_invoice(invoice_id)
        
        if not invoice_data:
            QMessageBox.critical(self, "Error", f"Failed to load invoice {invoice_id}.")
            return
            
        self.current_invoice_id = invoice_id
        self.txt_inv_number.setText(invoice_data.get("invoice_number", ""))
        self.dt_inv_date.setDate(QDate.fromString(invoice_data.get("date", ""), "yyyy-MM-dd"))
        
        inv_type = invoice_data.get("invoice_type", "Goods")
        if inv_type == "Services":
            self.rad_services.setChecked(True)
        else:
            self.rad_goods.setChecked(True)
            
        treatment = invoice_data.get("gst_treatment", "CGST_SGST")
        treatment_map = {
            "CGST_SGST": "Intrastate (CGST + SGST)",
            "IGST": "Interstate (IGST)"
        }
        self.cmb_gst_treatment.setCurrentText(treatment_map.get(treatment, "Auto-Detect (Based on GSTIN)"))
        self.chk_rcm.setChecked(invoice_data.get("rcm", False))
        self.txt_place_of_supply.setText(invoice_data.get("place_of_supply", ""))
        self.cmb_template_style.setCurrentText(invoice_data.get("template_style", "Classic Elegant"))
        
        seller_trade = invoice_data.get("seller", {}).get("trade_name", "")
        if seller_trade and self.cmb_seller.findText(seller_trade) != -1:
            self.cmb_seller.setCurrentText(seller_trade)
            
        customer = invoice_data.get("customer", {})
        self.txt_cust_name.setText(customer.get("name", ""))
        self.txt_cust_mobile.setText(customer.get("mobile", ""))
        self.txt_cust_address.setText(customer.get("address", ""))
        self.txt_cust_gst.setText(customer.get("gstin", ""))
        self.txt_cust_pan.setText(customer.get("pan", ""))
        self.txt_cust_pin.setText(customer.get("pin", ""))
        
        ship_to = invoice_data.get("ship_to", {})
        if ship_to and ship_to.get("name") and inv_type == "Goods":
            self.chk_ship_to.setChecked(True)
            self.txt_ship_name.setText(ship_to.get("name", ""))
            self.txt_ship_mobile.setText(ship_to.get("mobile", ""))
            self.txt_ship_address.setText(ship_to.get("address", ""))
            self.txt_ship_gstin.setText(ship_to.get("gstin", ""))
        else:
            self.chk_ship_to.setChecked(False)
            self.txt_ship_name.clear()
            self.txt_ship_mobile.clear()
            self.txt_ship_address.clear()
            self.txt_ship_gstin.clear()
            
        self.tbl_items.itemChanged.disconnect(self.on_table_item_changed)
        items = invoice_data.get("items", [])
        self.tbl_items.setRowCount(0)
        
        for r_idx, item_data in enumerate(items):
            self.tbl_items.insertRow(r_idx)
            self.tbl_items.setItem(r_idx, 0, QTableWidgetItem(item_data.get("name", "")))
            self.tbl_items.setItem(r_idx, 1, QTableWidgetItem(item_data.get("hsn", "")))
            
            qty_cell = QTableWidgetItem(f"{item_data.get('qty', 0):g}")
            qty_cell.setTextAlignment(Qt.AlignCenter)
            self.tbl_items.setItem(r_idx, 2, qty_cell)
            
            rate_cell = QTableWidgetItem(f"{item_data.get('rate', 0):.2f}")
            rate_cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_items.setItem(r_idx, 3, rate_cell)
            
            gst_cell = QTableWidgetItem(f"{item_data.get('gst_rate', 0):g}")
            gst_cell.setTextAlignment(Qt.AlignCenter)
            self.tbl_items.setItem(r_idx, 4, gst_cell)
            
            cess_cell = QTableWidgetItem(f"{item_data.get('cess_rate', 0):g}")
            cess_cell.setTextAlignment(Qt.AlignCenter)
            self.tbl_items.setItem(r_idx, 5, cess_cell)
            
            taxable_cell = QTableWidgetItem(f"{item_data.get('taxable_amount', 0):.2f}")
            taxable_cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if inv_type == "Services":
                taxable_cell.setFlags(taxable_cell.flags() | Qt.ItemIsEditable)
            else:
                taxable_cell.setFlags(taxable_cell.flags() & ~Qt.ItemIsEditable)
            self.tbl_items.setItem(r_idx, 6, taxable_cell)
            
            cgst_c = QTableWidgetItem(f"{item_data.get('cgst_amount', 0):.2f}")
            sgst_c = QTableWidgetItem(f"{item_data.get('sgst_amount', 0):.2f}")
            igst_c = QTableWidgetItem(f"{item_data.get('igst_amount', 0):.2f}")
            cess_c = QTableWidgetItem(f"{item_data.get('cess_amount', 0):.2f}")
            tot_c = QTableWidgetItem(f"{item_data.get('total_amount', 0):.2f}")
            
            for c, cell_w in [(7, cgst_c), (8, sgst_c), (9, igst_c), (10, cess_c), (11, tot_c)]:
                cell_w.setFlags(cell_w.flags() & ~Qt.ItemIsEditable)
                cell_w.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tbl_items.setItem(r_idx, c, cell_w)
                
            btn_del = QPushButton("Delete")
            btn_del.setObjectName("btnDanger")
            btn_del.setProperty("danger", True)
            btn_del.setStyleSheet("padding: 2px 5px; font-size: 11px;")
            btn_del.clicked.connect(lambda _, r=r_idx: self.remove_table_row(r))
            self.tbl_items.setCellWidget(r_idx, 12, btn_del)
            
        self.tbl_items.itemChanged.connect(self.on_table_item_changed)
        self.btn_action_delete.setEnabled(True)
        self.calculate_invoice_totals()
        
        self.tab_widget.setCurrentIndex(0)

    def delete_current_invoice(self):
        if not self.current_invoice_id:
            return
            
        confirm = QMessageBox.question(
            self, "Confirm Delete", 
            f"Are you sure you want to permanently delete Invoice {self.current_invoice_id}?",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm == QMessageBox.Yes:
            if delete_invoice(self.current_invoice_id):
                QMessageBox.information(self, "Deleted", "Invoice deleted successfully.")
                self.new_invoice()
                self.refresh_invoice_list()
                self.refresh_dashboard_data()
            else:
                QMessageBox.critical(self, "Error", "Failed to delete invoice from disk.")

    def preview_invoice_pdf(self):
        if hasattr(self, "pdf_worker") and self.pdf_worker.isRunning():
            QMessageBox.warning(self, "Task In Progress", "A PDF generation task is currently running. Please wait...")
            return
            
        invoice_data = self.extract_invoice_data()
        if not invoice_data:
            return
            
        pdf_dir = os.path.join(get_base_dir(), "invoices_pdf")
        os.makedirs(pdf_dir, exist_ok=True)
        sanitized_no = self.txt_inv_number.text().strip().replace("/", "_").replace("\\", "_")
        pdf_path = os.path.join(pdf_dir, f"{sanitized_no}.pdf")
        
        self.statusBar().showMessage("Generating PDF preview...")
        self.pdf_worker = PDFWorker(invoice_data, pdf_path)
        self.pdf_worker.finished.connect(self.on_pdf_preview_finished)
        self.pdf_worker.error.connect(self.on_pdf_error)
        self.pdf_worker.start()

    def on_pdf_preview_finished(self, pdf_path):
        self.statusBar().clearMessage()
        try:
            os.startfile(pdf_path)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to launch PDF viewer:\n{e}")

    def on_pdf_error(self, err_msg):
        self.statusBar().clearMessage()
        QMessageBox.critical(self, "PDF Compilation Error", f"Failed to compile PDF document:\n{err_msg}")

    def export_invoice_pdf(self):
        if hasattr(self, "pdf_worker") and self.pdf_worker.isRunning():
            QMessageBox.warning(self, "Task In Progress", "A PDF generation task is currently running. Please wait...")
            return
            
        if not self.save_invoice_data():
            return
            
        invoice_data = self.extract_invoice_data()
        if not invoice_data:
            return
            
        file_name = f"{invoice_data['invoice_number'].replace('/', '_')}.pdf"
        default_name = self.get_safe_dialog_path(file_name)
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Export PDF Invoice", default_name, "PDF Files (*.pdf)"
        )
        
        if save_path:
            self.statusBar().showMessage("Exporting PDF invoice...")
            self.pdf_worker = PDFWorker(invoice_data, save_path)
            self.pdf_worker.finished.connect(self.on_pdf_export_finished)
            self.pdf_worker.error.connect(self.on_pdf_error)
            self.pdf_worker.start()

    def on_pdf_export_finished(self, pdf_path):
        self.statusBar().clearMessage()
        QMessageBox.information(self, "Export Success", f"Invoice PDF exported successfully to:\n{pdf_path}")

    def print_invoice_pdf(self):
        if hasattr(self, "pdf_worker") and self.pdf_worker.isRunning():
            QMessageBox.warning(self, "Task In Progress", "A PDF generation task is currently running. Please wait...")
            return
            
        if not self.save_invoice_data():
            return
            
        invoice_data = self.extract_invoice_data()
        if not invoice_data:
            return
            
        pdf_dir = os.path.join(get_base_dir(), "invoices_pdf")
        os.makedirs(pdf_dir, exist_ok=True)
        sanitized_no = self.txt_inv_number.text().strip().replace("/", "_").replace("\\", "_")
        pdf_path = os.path.join(pdf_dir, f"{sanitized_no}.pdf")
        
        self.statusBar().showMessage("Sending PDF to print queue...")
        self.pdf_worker = PDFWorker(invoice_data, pdf_path)
        self.pdf_worker.finished.connect(self.on_pdf_print_finished)
        self.pdf_worker.error.connect(self.on_pdf_error)
        self.pdf_worker.start()

    def on_pdf_print_finished(self, pdf_path):
        self.statusBar().clearMessage()
        try:
            os.startfile(pdf_path, "print")
        except Exception as e:
            confirm = QMessageBox.question(
                self, "Print Warning", 
                f"Could not directly print PDF:\n{e}\n\nWould you like to open it in your viewer?",
                QMessageBox.Yes | QMessageBox.No
            )
            if confirm == QMessageBox.Yes:
                try:
                    os.startfile(pdf_path)
                except Exception:
                    pass

    def export_accounting_data(self):
        items = ["Tally XML Format", "Standard Accounting CSV"]
        item, ok = QInputDialog.getItem(self, "Export for Accounting", "Choose format to export:", items, 0, False)
        
        if ok and item:
            invoices = get_all_invoices()
            if not invoices:
                QMessageBox.warning(self, "No Invoices", "There are no invoices in the database to export.")
                return
                
            if item == "Tally XML Format":
                default_name = self.get_safe_dialog_path("Tally_Sales_Vouchers.xml")
                file_path, _ = QFileDialog.getSaveFileName(self, "Export Tally XML Vouchers", default_name, "XML Files (*.xml)")
                if file_path:
                    from app.accounting_exporter import generate_tally_xml
                    try:
                        if generate_tally_xml(invoices, file_path):
                            base, _ = os.path.splitext(file_path)
                            QMessageBox.information(
                                self, "Export Success",
                                f"Tally XML files generated successfully!\n\n"
                                f"Please import them into Tally Prime in this exact order:\n"
                                f"1. Import Masters file first:\n   {base}_Masters.xml\n\n"
                                f"2. Import Vouchers file second:\n   {base}_Vouchers.xml"
                            )
                    except Exception as e:
                        QMessageBox.critical(self, "Export Error", f"Failed to generate Tally XML file: {e}")
            else:
                default_name = self.get_safe_dialog_path("Zoho_Sales_Register.csv")
                file_path, _ = QFileDialog.getSaveFileName(self, "Export Sales CSV", default_name, "CSV Files (*.csv)")
                if file_path:
                    from app.accounting_exporter import generate_accounting_csv
                    try:
                        if generate_accounting_csv(invoices, file_path):
                            QMessageBox.information(self, "Export Success", f"Sales register exported successfully for Zoho/QB!\nSaved at:\n{file_path}")
                    except Exception as e:
                        QMessageBox.critical(self, "Export Error", f"Failed to generate CSV: {e}")

    def import_bulk_invoices(self):
        default_dir = self.get_safe_dialog_path("")
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Bulk Invoices", default_dir, "Import Spreadsheet Files (*.xlsx *.csv)"
        )
        if file_path:
            self.statusBar().showMessage("Parsing Excel/CSV template and writing to SQLite DB...")
            self.import_worker = ImportWorker(file_path)
            self.import_worker.finished.connect(self.on_import_finished)
            self.import_worker.start()

    def on_import_finished(self, success_count, errors):
        self.statusBar().clearMessage()
        if success_count > 0:
            self.refresh_invoice_list()
            self.refresh_dashboard_data()
            self.new_invoice()
            msg = f"Successfully imported {success_count} invoices into the database!"
            if errors:
                msg += "\n\nWarnings/Errors encountered during import:\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    msg += f"\n...and {len(errors) - 10} more warnings."
            QMessageBox.information(self, "Bulk Import Success", msg)
        else:
            QMessageBox.critical(self, "Bulk Import Failed", error_msg)

    def init_timesheets_ui(self, parent_widget):
        layout = QVBoxLayout(parent_widget)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(12)
        
        splitter = QSplitter(Qt.Horizontal)
        
        # Left Panel (List of saved timesheets)
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        
        left_layout.addWidget(QLabel("<b>Saved Timesheets Summary</b>"))
        
        self.lst_timesheets = QListWidget()
        self.lst_timesheets.currentItemChanged.connect(self.on_timesheet_selected)
        left_layout.addWidget(self.lst_timesheets)
        
        btn_layout = QHBoxLayout()
        self.btn_ts_new = QPushButton("New Timesheet")
        self.btn_ts_new.setProperty("success", True)
        self.btn_ts_new.clicked.connect(self.new_timesheet)
        btn_layout.addWidget(self.btn_ts_new)
        
        self.btn_ts_delete = QPushButton("Delete")
        self.btn_ts_delete.setObjectName("btnDanger")
        self.btn_ts_delete.clicked.connect(self.delete_selected_timesheet)
        btn_layout.addWidget(self.btn_ts_delete)
        left_layout.addLayout(btn_layout)
        
        splitter.addWidget(left_widget)
        
        # Right Panel (Details & Entries Grid)
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(10, 0, 0, 0)
        
        details_group = QGroupBox("Timesheet Properties")
        details_form = QFormLayout(details_group)
        details_form.setSpacing(8)
        details_form.setContentsMargins(15, 15, 15, 15)
        
        self.txt_ts_id = QLineEdit()
        self.txt_ts_id.setReadOnly(True)
        self.txt_ts_id.setStyleSheet("background-color: #F1F5F9; color: #475569;")
        details_form.addRow("Timesheet ID:", self.txt_ts_id)
        
        client_row = QHBoxLayout()
        self.txt_ts_client_name = QLineEdit()
        self.txt_ts_client_name.setPlaceholderText("Enter Client / Customer Name")
        client_row.addWidget(self.txt_ts_client_name, 3)
        
        self.btn_ts_choose_client = QPushButton("CRM Directory")
        self.btn_ts_choose_client.setStyleSheet("font-size: 11px; padding: 4px 8px;")
        self.btn_ts_choose_client.clicked.connect(self.open_timesheet_client_directory)
        client_row.addWidget(self.btn_ts_choose_client, 1)
        details_form.addRow("Client Name:", client_row)
        
        dates_row = QHBoxLayout()
        self.date_ts_start = QDateEdit(QDate.currentDate().addDays(-30))
        self.date_ts_start.setCalendarPopup(True)
        self.date_ts_start.setDisplayFormat("yyyy-MM-dd")
        self.date_ts_end = QDateEdit(QDate.currentDate())
        self.date_ts_end.setCalendarPopup(True)
        self.date_ts_end.setDisplayFormat("yyyy-MM-dd")
        dates_row.addWidget(self.date_ts_start)
        dates_row.addWidget(QLabel("to"))
        dates_row.addWidget(self.date_ts_end)
        details_form.addRow("Filing Period:", dates_row)
        
        status_row = QHBoxLayout()
        self.lbl_ts_status = QLabel("<b>Status:</b> Draft")
        self.lbl_ts_invoice = QLabel("<b>Linked Invoice:</b> None")
        status_row.addWidget(self.lbl_ts_status)
        status_row.addSpacing(20)
        status_row.addWidget(self.lbl_ts_invoice)
        status_row.addStretch()
        details_form.addRow("Metadata:", status_row)
        
        right_layout.addWidget(details_group)
        
        # Entries Grid Section
        entries_group = QGroupBox("Activity Logs & Hours Tracking")
        entries_layout = QVBoxLayout(entries_group)
        entries_layout.setContentsMargins(15, 15, 15, 15)
        
        # Quick Entry Row
        quick_row = QHBoxLayout()
        self.date_ts_entry = QDateEdit(QDate.currentDate())
        self.date_ts_entry.setCalendarPopup(True)
        self.date_ts_entry.setDisplayFormat("yyyy-MM-dd")
        self.date_ts_entry.setMaximumWidth(110)
        
        self.txt_ts_activity = QLineEdit()
        self.txt_ts_activity.setPlaceholderText("Describe work/service done...")
        
        self.spin_ts_hours = QDoubleSpinBox()
        self.spin_ts_hours.setRange(0.1, 24.0)
        self.spin_ts_hours.setValue(1.0)
        self.spin_ts_hours.setSuffix(" hrs")
        self.spin_ts_hours.setMaximumWidth(80)
        
        self.spin_ts_rate = QDoubleSpinBox()
        self.spin_ts_rate.setRange(1.0, 99999.0)
        self.spin_ts_rate.setValue(1000.0)
        self.spin_ts_rate.setPrefix("₹ ")
        self.spin_ts_rate.setMaximumWidth(110)
        
        self.btn_ts_add_row = QPushButton("Add Entry")
        self.btn_ts_add_row.setProperty("success", True)
        self.btn_ts_add_row.clicked.connect(self.add_timesheet_entry_row)
        
        quick_row.addWidget(QLabel("Date:"), 0)
        quick_row.addWidget(self.date_ts_entry, 1)
        quick_row.addWidget(QLabel("Description:"), 0)
        quick_row.addWidget(self.txt_ts_activity, 3)
        quick_row.addWidget(QLabel("Hours:"), 0)
        quick_row.addWidget(self.spin_ts_hours, 1)
        quick_row.addWidget(QLabel("Rate:"), 0)
        quick_row.addWidget(self.spin_ts_rate, 1)
        quick_row.addWidget(self.btn_ts_add_row, 1)
        entries_layout.addLayout(quick_row)
        
        # Table of entries
        self.tbl_ts_entries = QTableWidget(0, 5)
        self.tbl_ts_entries.setHorizontalHeaderLabels(["Date", "Activity Description", "Hours", "Rate (₹/hr)", "Total (₹)"])
        self.tbl_ts_entries.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl_ts_entries.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.tbl_ts_entries.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.tbl_ts_entries.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.tbl_ts_entries.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.tbl_ts_entries.itemChanged.connect(self.on_timesheet_table_item_changed)
        entries_layout.addWidget(self.tbl_ts_entries)
        
        # Summary totals line
        totals_row = QHBoxLayout()
        self.lbl_ts_total_hours = QLabel("<b>Total Hours:</b> 0.00 hrs")
        self.lbl_ts_total_amount = QLabel("<b>Total Amount:</b> ₹ 0.00")
        self.lbl_ts_total_hours.setStyleSheet("font-size: 13px; color: #1E293B;")
        self.lbl_ts_total_amount.setStyleSheet("font-size: 13px; color: #1E293B; font-weight: bold;")
        
        totals_row.addWidget(self.lbl_ts_total_hours)
        totals_row.addSpacing(40)
        totals_row.addWidget(self.lbl_ts_total_amount)
        totals_row.addStretch()
        
        # Actions Layout
        self.btn_ts_save = QPushButton("Save Timesheet")
        self.btn_ts_save.clicked.connect(self.save_active_timesheet)
        totals_row.addWidget(self.btn_ts_save)
        
        self.btn_ts_export = QPushButton("Export Summary")
        self.btn_ts_export.setObjectName("btnPrimary")
        self.btn_ts_export.setProperty("primary", True)
        
        export_menu = QMenu(self)
        export_pdf_action = export_menu.addAction("Export as PDF Report")
        export_xls_action = export_menu.addAction("Export as Excel Spreadsheet")
        export_word_action = export_menu.addAction("Export as MS Word Draft")
        
        export_pdf_action.triggered.connect(lambda: self.export_active_timesheet("PDF"))
        export_xls_action.triggered.connect(lambda: self.export_active_timesheet("Excel"))
        export_word_action.triggered.connect(lambda: self.export_active_timesheet("Word"))
        self.btn_ts_export.setMenu(export_menu)
        totals_row.addWidget(self.btn_ts_export)
        
        self.btn_ts_invoice = QPushButton("Generate Invoice")
        self.btn_ts_invoice.setStyleSheet("background-color: #10B981; color: white; font-weight: bold;")
        self.btn_ts_invoice.clicked.connect(self.generate_invoice_from_timesheet)
        totals_row.addWidget(self.btn_ts_invoice)
        
        entries_layout.addLayout(totals_row)
        right_layout.addWidget(entries_group)
        
        splitter.addWidget(right_widget)
        splitter.setSizes([200, 600])
        layout.addWidget(splitter)
        
        self.load_timesheets_list()
        self.new_timesheet()

    def open_timesheet_client_directory(self):
        from app.gui.customer_dialog import CustomerDialog
        dialog = CustomerDialog(self)
        if dialog.exec_() == CustomerDialog.Accepted:
            cust = dialog.selected_customer
            if cust:
                self.txt_ts_client_name.setText(cust.get("name", ""))

    def load_timesheets_list(self):
        from app.database import get_all_timesheets
        self.lst_timesheets.clear()
        self.active_timesheets_data = get_all_timesheets()
        for ts in self.active_timesheets_data:
            item = QListWidgetItem(f"{ts.get('timesheet_id')} - {ts.get('client_name')}")
            item.setData(Qt.UserRole, ts.get('timesheet_id'))
            self.lst_timesheets.addItem(item)

    def on_timesheet_selected(self, current_item, previous_item=None):
        if not current_item:
            return
        ts_id = current_item.data(Qt.UserRole)
        from app.database import get_timesheet
        ts = get_timesheet(ts_id)
        if not ts:
            return
            
        self.txt_ts_id.setText(ts.get("timesheet_id", ""))
        self.txt_ts_client_name.setText(ts.get("client_name", ""))
        self.date_ts_start.setDate(QDate.fromString(ts.get("start_date", ""), "yyyy-MM-dd"))
        self.date_ts_end.setDate(QDate.fromString(ts.get("end_date", ""), "yyyy-MM-dd"))
        self.lbl_ts_status.setText(f"<b>Status:</b> {ts.get('status', 'Draft')}")
        
        linked = ts.get("linked_invoice_number")
        self.lbl_ts_invoice.setText(f"<b>Linked Invoice:</b> {linked if linked else 'None'}")
        
        try:
            self.tbl_ts_entries.itemChanged.disconnect(self.on_timesheet_table_item_changed)
        except TypeError:
            pass
        self.tbl_ts_entries.setRowCount(0)
        
        for entry in ts.get("entries", []):
            row = self.tbl_ts_entries.rowCount()
            self.tbl_ts_entries.insertRow(row)
            
            d_cell = QTableWidgetItem(entry.get("date", ""))
            d_cell.setTextAlignment(Qt.AlignCenter)
            self.tbl_ts_entries.setItem(row, 0, d_cell)
            
            self.tbl_ts_entries.setItem(row, 1, QTableWidgetItem(entry.get("activity", "")))
            
            h_cell = QTableWidgetItem(f"{entry.get('hours', 0.0):.2f}")
            h_cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_ts_entries.setItem(row, 2, h_cell)
            
            r_cell = QTableWidgetItem(f"{entry.get('rate', 0.0):.2f}")
            r_cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_ts_entries.setItem(row, 3, r_cell)
            
            t_cell = QTableWidgetItem(f"{entry.get('line_total', 0.0):.2f}")
            t_cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            t_cell.setFlags(t_cell.flags() & ~Qt.ItemIsEditable)
            self.tbl_ts_entries.setItem(row, 4, t_cell)
            
        self.tbl_ts_entries.itemChanged.connect(self.on_timesheet_table_item_changed)
        self.calculate_timesheet_totals()

    def new_timesheet(self):
        import random
        from datetime import datetime
        ts_id = f"TS-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
        self.txt_ts_id.setText(ts_id)
        self.txt_ts_client_name.clear()
        self.date_ts_start.setDate(QDate.currentDate().addDays(-30))
        self.date_ts_end.setDate(QDate.currentDate())
        self.lbl_ts_status.setText("<b>Status:</b> Draft")
        self.lbl_ts_invoice.setText("<b>Linked Invoice:</b> None")
        
        try:
            self.tbl_ts_entries.itemChanged.disconnect(self.on_timesheet_table_item_changed)
        except TypeError:
            pass
            
        self.tbl_ts_entries.setRowCount(0)
        self.tbl_ts_entries.itemChanged.connect(self.on_timesheet_table_item_changed)
        
        self.calculate_timesheet_totals()
        self.lst_timesheets.clearSelection()

    def delete_selected_timesheet(self):
        curr_item = self.lst_timesheets.currentItem()
        if not curr_item:
            QMessageBox.warning(self, "Selection Error", "Please select a timesheet from the list to delete.")
            return
            
        ts_id = curr_item.data(Qt.UserRole)
        confirm = QMessageBox.question(
            self, "Delete Timesheet",
            f"Are you sure you want to delete Timesheet {ts_id}?\nThis action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No
        )
        if confirm == QMessageBox.Yes:
            from app.database import delete_timesheet
            if delete_timesheet(ts_id):
                self.load_timesheets_list()
                self.new_timesheet()
                QMessageBox.information(self, "Deleted", f"Timesheet {ts_id} was successfully deleted.")
            else:
                QMessageBox.critical(self, "Error", f"Failed to delete timesheet {ts_id}.")

    def add_timesheet_entry_row(self):
        date_str = self.date_ts_entry.date().toString("yyyy-MM-dd")
        desc = self.txt_ts_activity.text().strip()
        hours = self.spin_ts_hours.value()
        rate = self.spin_ts_rate.value()
        line_total = hours * rate
        
        if not desc:
            QMessageBox.warning(self, "Validation Error", "Activity description cannot be empty.")
            self.txt_ts_activity.setFocus()
            return
            
        try:
            self.tbl_ts_entries.itemChanged.disconnect(self.on_timesheet_table_item_changed)
        except TypeError:
            pass
        
        row = self.tbl_ts_entries.rowCount()
        self.tbl_ts_entries.insertRow(row)
        
        d_cell = QTableWidgetItem(date_str)
        d_cell.setTextAlignment(Qt.AlignCenter)
        self.tbl_ts_entries.setItem(row, 0, d_cell)
        
        self.tbl_ts_entries.setItem(row, 1, QTableWidgetItem(desc))
        
        h_cell = QTableWidgetItem(f"{hours:.2f}")
        h_cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.tbl_ts_entries.setItem(row, 2, h_cell)
        
        r_cell = QTableWidgetItem(f"{rate:.2f}")
        r_cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.tbl_ts_entries.setItem(row, 3, r_cell)
        
        t_cell = QTableWidgetItem(f"{line_total:.2f}")
        t_cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        t_cell.setFlags(t_cell.flags() & ~Qt.ItemIsEditable)
        self.tbl_ts_entries.setItem(row, 4, t_cell)
        
        self.tbl_ts_entries.itemChanged.connect(self.on_timesheet_table_item_changed)
        self.calculate_timesheet_totals()
        
        self.txt_ts_activity.clear()
        self.txt_ts_activity.setFocus()

    def on_timesheet_table_item_changed(self, item):
        row = item.row()
        col = item.column()
        if col in [2, 3]:
            try:
                self.tbl_ts_entries.itemChanged.disconnect(self.on_timesheet_table_item_changed)
                
                h_item = self.tbl_ts_entries.item(row, 2)
                r_item = self.tbl_ts_entries.item(row, 3)
                
                hours = float(h_item.text()) if h_item else 0.0
                rate = float(r_item.text()) if r_item else 0.0
                line_total = hours * rate
                
                t_item = self.tbl_ts_entries.item(row, 4)
                if not t_item:
                    t_item = QTableWidgetItem()
                    t_item.setFlags(t_item.flags() & ~Qt.ItemIsEditable)
                    t_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.tbl_ts_entries.setItem(row, 4, t_item)
                t_item.setText(f"{line_total:.2f}")
                
            except ValueError:
                pass
            finally:
                self.tbl_ts_entries.itemChanged.connect(self.on_timesheet_table_item_changed)
        
        self.calculate_timesheet_totals()

    def calculate_timesheet_totals(self):
        tot_hours = 0.0
        tot_amount = 0.0
        
        for r in range(self.tbl_ts_entries.rowCount()):
            h_item = self.tbl_ts_entries.item(r, 2)
            t_item = self.tbl_ts_entries.item(r, 4)
            try:
                tot_hours += float(h_item.text()) if h_item else 0.0
                tot_amount += float(t_item.text()) if t_item else 0.0
            except ValueError:
                pass
                
        self.lbl_ts_total_hours.setText(f"<b>Total Hours:</b> {tot_hours:.2f} hrs")
        self.lbl_ts_total_amount.setText(f"<b>Total Amount:</b> ₹ {tot_amount:,.2f}")
        self.ts_total_hours_val = tot_hours
        self.ts_total_amount_val = tot_amount

    def save_active_timesheet(self):
        ts_id = self.txt_ts_id.text().strip()
        client = self.txt_ts_client_name.text().strip()
        start_d = self.date_ts_start.date().toString("yyyy-MM-dd")
        end_d = self.date_ts_end.date().toString("yyyy-MM-dd")
        
        if not client:
            QMessageBox.warning(self, "Validation Error", "Client Name is required before saving.")
            self.txt_ts_client_name.setFocus()
            return False
            
        entries = []
        for r in range(self.tbl_ts_entries.rowCount()):
            d_item = self.tbl_ts_entries.item(r, 0)
            a_item = self.tbl_ts_entries.item(r, 1)
            h_item = self.tbl_ts_entries.item(r, 2)
            r_item = self.tbl_ts_entries.item(r, 3)
            t_item = self.tbl_ts_entries.item(r, 4)
            
            entries.append({
                "date": d_item.text() if d_item else start_d,
                "activity": a_item.text() if a_item else "",
                "hours": float(h_item.text()) if h_item else 0.0,
                "rate": float(r_item.text()) if r_item else 0.0,
                "line_total": float(t_item.text()) if t_item else 0.0
            })
            
        status = self.lbl_ts_status.text().split(":")[-1].strip()
        linked = self.lbl_ts_invoice.text().split(":")[-1].strip()
        if linked == "None":
            linked = None
            
        ts_data = {
            "timesheet_id": ts_id,
            "client_name": client,
            "start_date": start_d,
            "end_date": end_d,
            "total_hours": self.ts_total_hours_val,
            "total_amount": self.ts_total_amount_val,
            "linked_invoice_number": linked,
            "status": status,
            "entries": entries
        }
        
        from app.database import save_timesheet
        if save_timesheet(ts_data):
            self.load_timesheets_list()
            for i in range(self.lst_timesheets.count()):
                item = self.lst_timesheets.item(i)
                if item.data(Qt.UserRole) == ts_id:
                    self.lst_timesheets.setCurrentItem(item)
                    break
            return True
        else:
            QMessageBox.critical(self, "Error", f"Failed to save timesheet {ts_id}.")
            return False

    def export_active_timesheet(self, format_type):
        ts_id = self.txt_ts_id.text().strip()
        client = self.txt_ts_client_name.text().strip()
        start_d = self.date_ts_start.date().toString("yyyy-MM-dd")
        end_d = self.date_ts_end.date().toString("yyyy-MM-dd")
        
        if not client:
            QMessageBox.warning(self, "Validation Error", "Please enter a Client Name before exporting.")
            return
            
        ext = ".pdf" if format_type == "PDF" else (".xlsx" if format_type == "Excel" else ".docx")
        filter_str = "PDF Files (*.pdf)" if format_type == "PDF" else ("Excel Files (*.xlsx)" if format_type == "Excel" else "Word Files (*.docx)")
        default_file = self.get_safe_dialog_path(f"Timesheet_{ts_id}{ext}")
        
        file_path, _ = QFileDialog.getSaveFileName(self, f"Export Timesheet as {format_type}", default_file, filter_str)
        if not file_path:
            return
            
        entries = []
        for r in range(self.tbl_ts_entries.rowCount()):
            d_item = self.tbl_ts_entries.item(r, 0)
            a_item = self.tbl_ts_entries.item(r, 1)
            h_item = self.tbl_ts_entries.item(r, 2)
            r_item = self.tbl_ts_entries.item(r, 3)
            t_item = self.tbl_ts_entries.item(r, 4)
            
            entries.append({
                "date": d_item.text() if d_item else start_d,
                "activity": a_item.text() if a_item else "",
                "hours": float(h_item.text()) if h_item else 0.0,
                "rate": float(r_item.text()) if r_item else 0.0,
                "line_total": float(t_item.text()) if t_item else 0.0
            })
            
        ts_data = {
            "timesheet_id": ts_id,
            "client_name": client,
            "start_date": start_d,
            "end_date": end_d,
            "total_hours": self.ts_total_hours_val,
            "total_amount": self.ts_total_amount_val,
            "status": self.lbl_ts_status.text().split(":")[-1].strip(),
            "entries": entries
        }
        
        try:
            if format_type == "PDF":
                from app.timesheet_exporter import export_timesheet_pdf
                export_timesheet_pdf(ts_data, file_path)
            elif format_type == "Excel":
                from app.timesheet_exporter import export_timesheet_excel
                export_timesheet_excel(ts_data, file_path)
            else:
                from app.timesheet_exporter import export_timesheet_word
                export_timesheet_word(ts_data, file_path)
                
            QMessageBox.information(self, "Success", f"Timesheet exported successfully to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export timesheet: {e}")

    def generate_invoice_from_timesheet(self):
        ts_id = self.txt_ts_id.text().strip()
        client = self.txt_ts_client_name.text().strip()
        
        if not client:
            QMessageBox.warning(self, "Validation Error", "Please specify a Client Name first.")
            return
            
        if self.tbl_ts_entries.rowCount() == 0:
            QMessageBox.warning(self, "Validation Error", "Cannot bill an empty timesheet.")
            return
            
        if not self.save_active_timesheet():
            return
            
        self.tab_widget.setCurrentIndex(0)
        self.new_invoice()
        
        self.rad_services.setChecked(True)
        self.on_invoice_type_changed()
        
        from app.database import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM customers WHERE name = ?", (client,))
        cust_row = cursor.fetchone()
        conn.close()
        
        if cust_row:
            cust = dict(cust_row)
            self.txt_cust_name.setText(cust.get("name", ""))
            self.txt_cust_mobile.setText(cust.get("mobile", ""))
            self.txt_cust_address.setText(cust.get("address", ""))
            self.txt_cust_gst.setText(cust.get("gstin", ""))
            self.txt_cust_pan.setText(cust.get("pan", ""))
            self.txt_cust_pin.setText(cust.get("pin", ""))
            self.txt_place_of_supply.setText(cust.get("place_of_supply", ""))
        else:
            self.txt_cust_name.setText(client)
            
        try:
            self.tbl_items.itemChanged.disconnect(self.on_table_item_changed)
        except TypeError:
            pass
            
        self.tbl_items.setRowCount(0)
        self.tbl_items.insertRow(0)
        desc = f"Professional Consultancy/Services rendered for Timesheet Ref: {ts_id} (Hours: {self.ts_total_hours_val:.2f})"
        self.tbl_items.setItem(0, 0, QTableWidgetItem(desc))
        self.tbl_items.setItem(0, 1, QTableWidgetItem("9982"))
        
        qty_cell = QTableWidgetItem("1")
        qty_cell.setTextAlignment(Qt.AlignCenter)
        self.tbl_items.setItem(0, 2, qty_cell)
        
        rate_cell = QTableWidgetItem(f"{self.ts_total_amount_val:.2f}")
        rate_cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.tbl_items.setItem(0, 3, rate_cell)
        
        gst_cell = QTableWidgetItem("18")
        gst_cell.setTextAlignment(Qt.AlignCenter)
        self.tbl_items.setItem(0, 4, gst_cell)
        
        cess_cell = QTableWidgetItem("0")
        cess_cell.setTextAlignment(Qt.AlignCenter)
        self.tbl_items.setItem(0, 5, cess_cell)
        
        val_cell = QTableWidgetItem(f"{self.ts_total_amount_val:.2f}")
        val_cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.tbl_items.setItem(0, 6, val_cell)
        
        # Initialize calculations columns 7 to 11
        for col in range(7, 12):
            cell = QTableWidgetItem("0.00")
            cell.setFlags(cell.flags() & ~Qt.ItemIsEditable)
            cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_items.setItem(0, col, cell)
            
        # Initialize Delete Button (col 12)
        btn_del = QPushButton("Delete")
        btn_del.setObjectName("btnDanger")
        btn_del.setProperty("danger", True)
        btn_del.setStyleSheet("padding: 2px 5px; font-size: 11px;")
        btn_del.clicked.connect(lambda _, r=0: self.remove_table_row(r))
        self.tbl_items.setCellWidget(0, 12, btn_del)
        
        # Calculate taxes and row totals immediately
        self.update_table_row_calculations(0)
        
        self.tbl_items.itemChanged.connect(self.on_table_item_changed)
        self.calculate_invoice_totals()
        
        from app.database import update_timesheet_status
        update_timesheet_status(ts_id, "Invoiced")
        
        self.load_timesheets_list()
        
        QMessageBox.information(
            self, "Invoice Generated",
            f"Timesheet {ts_id} converted to Invoice draft successfully!\n\n"
            f"Please verify client GST details and save the invoice to generate invoice number."
        )

    # --- USER AUTHENTICATION & PAYMENTS GUI INTERFACES ---

    def open_change_password(self):
        dialog = ChangePassDialog(self.current_user, self)
        dialog.exec_()

    def open_user_settings(self):
        dialog = UserSettingsDialog(self.current_user, self)
        dialog.exec_()

    def init_payments_ui(self, parent_widget):
        layout = QVBoxLayout(parent_widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        self.payments_sub_tabs = QTabWidget()
        layout.addWidget(self.payments_sub_tabs)
        
        # SUB-TAB 1: Outstanding & Aging
        sub1_widget = QWidget()
        sub1_layout = QVBoxLayout(sub1_widget)
        sub1_layout.setContentsMargins(10, 10, 10, 10)
        sub1_layout.setSpacing(10)
        
        # Top Settings Bar (within Sub-tab 1)
        top_bar = QHBoxLayout()
        top_bar.addWidget(QLabel("Aging Brackets (Days, comma-separated):"))
        
        self.txt_aging_brackets = QLineEdit("30, 60, 90")
        self.txt_aging_brackets.setStyleSheet("padding: 4px; max-width: 150px;")
        top_bar.addWidget(self.txt_aging_brackets)
        
        self.btn_refresh_payments = QPushButton("Refresh & Recalculate")
        self.btn_refresh_payments.setObjectName("btnPrimary")
        self.btn_refresh_payments.setProperty("primary", True)
        self.btn_refresh_payments.setStyleSheet("font-weight: bold; padding: 6px 12px;")
        self.btn_refresh_payments.clicked.connect(self.refresh_payments_tab)
        top_bar.addWidget(self.btn_refresh_payments)
        
        self.btn_export_outstanding = QPushButton("Export Outstanding (Excel)")
        self.btn_export_outstanding.setObjectName("btnSuccess")
        self.btn_export_outstanding.setProperty("success", True)
        self.btn_export_outstanding.setStyleSheet("font-weight: bold; padding: 6px 12px; margin-left: 5px;")
        self.btn_export_outstanding.clicked.connect(self.export_outstanding_list)
        top_bar.addWidget(self.btn_export_outstanding)

        self.btn_export_aging = QPushButton("Export Aging (Excel)")
        self.btn_export_aging.setObjectName("btnSuccess")
        self.btn_export_aging.setProperty("success", True)
        self.btn_export_aging.setStyleSheet("font-weight: bold; padding: 6px 12px; margin-left: 5px;")
        self.btn_export_aging.clicked.connect(self.export_aging_report_excel)
        top_bar.addWidget(self.btn_export_aging)
        
        top_bar.addStretch()
        sub1_layout.addLayout(top_bar)
        
        # Main Splitter
        main_splitter = QSplitter(Qt.Horizontal)
        sub1_layout.addWidget(main_splitter)
        
        # Left Panel (Pending Invoices & Payments Logging)
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        
        left_layout.addWidget(QLabel("<b>Invoices with Outstanding Balance</b>"))
        
        self.tbl_unpaid_invoices = QTableWidget()
        self.tbl_unpaid_invoices.setColumnCount(7)
        self.tbl_unpaid_invoices.setHorizontalHeaderLabels([
            "Invoice #", "Date", "Customer", "Grand Total", "Paid", "Outstanding", "Overdue Days"
        ])
        self.tbl_unpaid_invoices.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tbl_unpaid_invoices.horizontalHeader().setStretchLastSection(True)
        self.tbl_unpaid_invoices.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_unpaid_invoices.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_unpaid_invoices.itemSelectionChanged.connect(self.unpaid_invoice_selected)
        left_layout.addWidget(self.tbl_unpaid_invoices)
        
        # Log & Manage Actions for selected invoice
        act_box = QGroupBox("Payment Operations")
        act_layout = QVBoxLayout(act_box)
        
        btn_row = QHBoxLayout()
        self.btn_record_payment = QPushButton("Log Payment Receipt")
        self.btn_record_payment.setStyleSheet("font-weight: bold; padding: 8px;")
        self.btn_record_payment.clicked.connect(self.open_record_payment_dialog)
        btn_row.addWidget(self.btn_record_payment)
        
        self.btn_delete_payment = QPushButton("Delete Selected Payment")
        self.btn_delete_payment.setObjectName("btnDanger")
        self.btn_delete_payment.setProperty("danger", True)
        self.btn_delete_payment.setStyleSheet("font-weight: bold; padding: 8px;")
        self.btn_delete_payment.clicked.connect(self.delete_selected_payment)
        self.btn_delete_payment.setVisible(self.current_role == "admin")
        btn_row.addWidget(self.btn_delete_payment)
        act_layout.addLayout(btn_row)
        
        # Payment History Table for Selected Invoice
        act_layout.addWidget(QLabel("<b>Receipts History for Selected Invoice:</b>"))
        self.tbl_invoice_payments = QTableWidget()
        self.tbl_invoice_payments.setColumnCount(5)
        self.tbl_invoice_payments.setHorizontalHeaderLabels([
            "Payment ID", "Date", "Amount", "Mode", "Reference"
        ])
        self.tbl_invoice_payments.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tbl_invoice_payments.horizontalHeader().setStretchLastSection(True)
        self.tbl_invoice_payments.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_invoice_payments.setSelectionBehavior(QAbstractItemView.SelectRows)
        act_layout.addWidget(self.tbl_invoice_payments)
        
        left_layout.addWidget(act_box)
        main_splitter.addWidget(left_widget)
        
        # Right Panel (Customer Aging Summary Table)
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        
        right_layout.addWidget(QLabel("<b>Customer Outstanding Aging Summary</b>"))
        
        self.tbl_aging_report = QTableWidget()
        right_layout.addWidget(self.tbl_aging_report)
        
        main_splitter.addWidget(right_widget)
        main_splitter.setSizes([600, 500])
        
        self.payments_sub_tabs.addTab(sub1_widget, "Outstanding & Aging Analysis")
        
        # SUB-TAB 2: Received Payments Ledger
        sub2_widget = QWidget()
        sub2_layout = QVBoxLayout(sub2_widget)
        sub2_layout.setContentsMargins(10, 10, 10, 10)
        sub2_layout.setSpacing(10)
        
        # Filters Group
        filt_group = QGroupBox("Filter Received Receipts")
        filt_layout = QHBoxLayout(filt_group)
        filt_layout.setSpacing(10)
        
        filt_layout.addWidget(QLabel("Period:"))
        self.cmb_receipts_period = QComboBox()
        self.cmb_receipts_period.addItems([
            "All Receipts", 
            "This Month", 
            "Last Month", 
            "This Financial Year (Apr-Mar)", 
            "Custom Period"
        ])
        self.cmb_receipts_period.currentTextChanged.connect(self.on_receipts_period_changed)
        filt_layout.addWidget(self.cmb_receipts_period)
        
        self.lbl_rec_start = QLabel("Start Date:")
        filt_layout.addWidget(self.lbl_rec_start)
        self.dt_receipts_start = QDateEdit()
        self.dt_receipts_start.setCalendarPopup(True)
        self.dt_receipts_start.setDate(QDate.currentDate().addMonths(-1))
        self.dt_receipts_start.setEnabled(False)
        filt_layout.addWidget(self.dt_receipts_start)
        
        self.lbl_rec_end = QLabel("End Date:")
        filt_layout.addWidget(self.lbl_rec_end)
        self.dt_receipts_end = QDateEdit()
        self.dt_receipts_end.setCalendarPopup(True)
        self.dt_receipts_end.setDate(QDate.currentDate())
        self.dt_receipts_end.setEnabled(False)
        filt_layout.addWidget(self.dt_receipts_end)
        
        self.btn_refresh_receipts = QPushButton("Fetch Receipts")
        self.btn_refresh_receipts.setObjectName("btnPrimary")
        self.btn_refresh_receipts.setProperty("primary", True)
        self.btn_refresh_receipts.setStyleSheet("font-weight: bold; padding: 5px 12px;")
        self.btn_refresh_receipts.clicked.connect(self.refresh_receipts_ledger)
        filt_layout.addWidget(self.btn_refresh_receipts)
        
        self.btn_export_receipts = QPushButton("Export Receipts (Excel)")
        self.btn_export_receipts.setObjectName("btnSuccess")
        self.btn_export_receipts.setProperty("success", True)
        self.btn_export_receipts.setStyleSheet("font-weight: bold; padding: 5px 12px;")
        self.btn_export_receipts.clicked.connect(self.export_receipts_ledger_excel)
        filt_layout.addWidget(self.btn_export_receipts)
        
        filt_layout.addStretch()
        sub2_layout.addWidget(filt_group)
        
        # Receipts Ledger Table
        self.tbl_receipts_ledger = QTableWidget()
        self.tbl_receipts_ledger.setColumnCount(8)
        self.tbl_receipts_ledger.setHorizontalHeaderLabels([
            "Receipt ID", "Payment Date", "Invoice #", "Customer Name", "Amount Received (₹)", "Payment Mode", "Reference / UTR", "Internal Notes"
        ])
        self.tbl_receipts_ledger.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tbl_receipts_ledger.horizontalHeader().setStretchLastSection(True)
        self.tbl_receipts_ledger.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_receipts_ledger.setSelectionBehavior(QAbstractItemView.SelectRows)
        sub2_layout.addWidget(self.tbl_receipts_ledger)
        
        self.payments_sub_tabs.addTab(sub2_widget, "Received Payments Ledger")

    def refresh_payments_tab(self):
        # 1. Load unpaid invoices
        self.tbl_unpaid_invoices.setRowCount(0)
        self.tbl_invoice_payments.setRowCount(0)
        
        from app.database import get_all_invoices, get_invoice_payment_summary
        invoices = get_all_invoices()
        unpaid_count = 0
        
        from datetime import datetime
        today = datetime.today()
        
        for inv in invoices:
            inv_no = inv["invoice_number"]
            summary = get_invoice_payment_summary(inv_no)
            if not summary:
                continue
            outstanding = summary["outstanding"]
            if outstanding > 0.01:
                row = unpaid_count
                self.tbl_unpaid_invoices.insertRow(row)
                self.tbl_unpaid_invoices.setItem(row, 0, QTableWidgetItem(inv_no))
                self.tbl_unpaid_invoices.setItem(row, 1, QTableWidgetItem(inv.get("date", "")))
                self.tbl_unpaid_invoices.setItem(row, 2, QTableWidgetItem(inv.get("customer_name") or "Walk-in"))
                self.tbl_unpaid_invoices.setItem(row, 3, QTableWidgetItem(f"{summary['grand_total']:.2f}"))
                self.tbl_unpaid_invoices.setItem(row, 4, QTableWidgetItem(f"{summary['total_paid']:.2f}"))
                self.tbl_unpaid_invoices.setItem(row, 5, QTableWidgetItem(f"{outstanding:.2f}"))
                
                from app.database import parse_invoice_date
                inv_date = parse_invoice_date(inv.get("date", ""))
                overdue_days = (today - inv_date).days
                self.tbl_unpaid_invoices.setItem(row, 6, QTableWidgetItem(str(max(0, overdue_days))))
                unpaid_count += 1
                
        # 2. Populate Aging Summary Table
        brackets_str = self.txt_aging_brackets.text().strip()
        try:
            brackets = [int(b.strip()) for b in brackets_str.split(",") if b.strip()]
        except Exception:
            brackets = [30, 60, 90]
        brackets.sort()
        
        headers = ["Customer Name", "Total Invoiced", "Total Paid", "Outstanding"]
        prev_limit = 0
        for limit in brackets:
            headers.append(f"{prev_limit}-{limit} Days")
            prev_limit = limit + 1
        headers.append(f"{brackets[-1]}+ Days")
        
        self.tbl_aging_report.setColumnCount(len(headers))
        self.tbl_aging_report.setHorizontalHeaderLabels(headers)
        self.tbl_aging_report.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tbl_aging_report.horizontalHeader().setStretchLastSection(True)
        self.tbl_aging_report.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_aging_report.setRowCount(0)
        
        from app.database import get_aging_report
        report_data = get_aging_report(brackets_str)
        
        for row, r in enumerate(report_data):
            self.tbl_aging_report.insertRow(row)
            self.tbl_aging_report.setItem(row, 0, QTableWidgetItem(r["customer_name"]))
            self.tbl_aging_report.setItem(row, 1, QTableWidgetItem(f"{r['total_invoiced']:.2f}"))
            self.tbl_aging_report.setItem(row, 2, QTableWidgetItem(f"{r['total_paid']:.2f}"))
            self.tbl_aging_report.setItem(row, 3, QTableWidgetItem(f"{r['outstanding_balance']:.2f}"))
            
            for b_idx, val in enumerate(r["brackets"]):
                self.tbl_aging_report.setItem(row, 4 + b_idx, QTableWidgetItem(f"{val:.2f}"))
        
        # 3. Reload received receipts ledger
        self.refresh_receipts_ledger()

    def unpaid_invoice_selected(self):
        self.tbl_invoice_payments.setRowCount(0)
        selected_rows = self.tbl_unpaid_invoices.selectedItems()
        if not selected_rows:
            return
        inv_no = self.tbl_unpaid_invoices.item(selected_rows[0].row(), 0).text()
        
        from app.database import get_payments_for_invoice
        payments = get_payments_for_invoice(inv_no)
        self.tbl_invoice_payments.setRowCount(len(payments))
        for row, p in enumerate(payments):
            self.tbl_invoice_payments.setItem(row, 0, QTableWidgetItem(str(p["id"])))
            self.tbl_invoice_payments.setItem(row, 1, QTableWidgetItem(p["payment_date"]))
            self.tbl_invoice_payments.setItem(row, 2, QTableWidgetItem(f"{p['amount']:.2f}"))
            self.tbl_invoice_payments.setItem(row, 3, QTableWidgetItem(p["payment_mode"]))
            self.tbl_invoice_payments.setItem(row, 4, QTableWidgetItem(p["reference_number"] or ""))

    def open_record_payment_dialog(self):
        selected_rows = self.tbl_unpaid_invoices.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, "No Invoice Selected", "Please select an outstanding invoice from the left grid first.")
            return
        row_idx = selected_rows[0].row()
        inv_no = self.tbl_unpaid_invoices.item(row_idx, 0).text()
        outstanding = float(self.tbl_unpaid_invoices.item(row_idx, 5).text())
        
        dialog = RecordPaymentDialog(inv_no, outstanding, self)
        if dialog.exec_() == QDialog.Accepted:
            self.refresh_payments_tab()
            self.refresh_dashboard_data()

    def delete_selected_payment(self):
        if self.current_role != "admin":
            QMessageBox.critical(self, "Access Denied", "Only administrators can delete payment transactions.")
            return
            
        selected_rows = self.tbl_invoice_payments.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, "No Payment Selected", "Please select a payment receipt from the history table to delete.")
            return
        row_idx = selected_rows[0].row()
        payment_id = int(self.tbl_invoice_payments.item(row_idx, 0).text())
        amount = self.tbl_invoice_payments.item(row_idx, 2).text()
        
        reply = QMessageBox.question(self, "Confirm Delete", f"Are you sure you want to permanently delete payment receipt ID #{payment_id} (Amount: ₹ {amount})?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            from app.database import delete_payment
            if delete_payment(payment_id):
                QMessageBox.information(self, "Deleted", "Payment receipt deleted successfully.")
                self.refresh_payments_tab()
                self.refresh_dashboard_data()
            else:
                QMessageBox.critical(self, "Error", "Failed to delete payment transaction from database.")

    def handle_logout_action(self):
        confirm = QMessageBox.question(self, "Logout", "Are you sure you want to log out and switch user?",
                                     QMessageBox.Yes | QMessageBox.No)
        if confirm == QMessageBox.Yes:
            self.logout_requested = True
            self.close()

    def export_outstanding_list(self):
        default_path = self.get_safe_dialog_path("Outstanding_Invoices_Report.xlsx")
        path, _ = QFileDialog.getSaveFileName(self, "Save Outstanding Invoices Report", default_path, "Excel Files (*.xlsx)")
        if not path:
            return
            
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Outstanding Invoices"
        ws.views.sheetView[0].showGridLines = True
        
        title_font = Font(name="Segoe UI", size=14, bold=True, color="1e3a8a")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        cell_font = Font(name="Segoe UI", size=10)
        bold_cell_font = Font(name="Segoe UI", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_side = Side(border_style="thin", color="cbd5e1")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws["A1"] = "PuruNiti Smart Billing - Outstanding Invoices Report"
        ws["A1"].font = title_font
        ws.row_dimensions[1].height = 25
        
        ws["A2"] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(name="Segoe UI", size=9, italic=True)
        ws.row_dimensions[2].height = 18
        
        headers = ["Invoice #", "Invoice Date", "Customer Name", "Grand Total (₹)", "Total Paid (₹)", "Outstanding Balance (₹)", "Overdue Days"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[4].height = 24
        
        row_count = self.tbl_unpaid_invoices.rowCount()
        total_grand = 0.0
        total_paid = 0.0
        total_outstanding = 0.0
        
        for r_idx in range(row_count):
            row_excel = 5 + r_idx
            ws.row_dimensions[row_excel].height = 19
            
            c0 = ws.cell(row=row_excel, column=1, value=self.tbl_unpaid_invoices.item(r_idx, 0).text())
            c0.alignment = center_align
            
            c1 = ws.cell(row=row_excel, column=2, value=self.tbl_unpaid_invoices.item(r_idx, 1).text())
            c1.alignment = center_align
            
            c2 = ws.cell(row=row_excel, column=3, value=self.tbl_unpaid_invoices.item(r_idx, 2).text())
            c2.alignment = left_align
            
            g_total = float(self.tbl_unpaid_invoices.item(r_idx, 3).text())
            paid_val = float(self.tbl_unpaid_invoices.item(r_idx, 4).text())
            out_val = float(self.tbl_unpaid_invoices.item(r_idx, 5).text())
            
            total_grand += g_total
            total_paid += paid_val
            total_outstanding += out_val
            
            c3 = ws.cell(row=row_excel, column=4, value=g_total)
            c3.alignment = right_align
            c3.number_format = "#,##0.00"
            
            c4 = ws.cell(row=row_excel, column=5, value=paid_val)
            c4.alignment = right_align
            c4.number_format = "#,##0.00"
            
            c5 = ws.cell(row=row_excel, column=6, value=out_val)
            c5.alignment = right_align
            c5.number_format = "#,##0.00"
            
            c6 = ws.cell(row=row_excel, column=7, value=int(self.tbl_unpaid_invoices.item(r_idx, 6).text()))
            c6.alignment = center_align
            
            for c in [c0, c1, c2, c3, c4, c5, c6]:
                c.font = cell_font
                c.border = thin_border
                
        tot_row = 5 + row_count
        ws.row_dimensions[tot_row].height = 22
        
        c_lbl = ws.cell(row=tot_row, column=1, value="TOTALS")
        c_lbl.font = bold_cell_font
        c_lbl.alignment = center_align
        c_lbl.border = thin_border
        
        for col_idx in [2, 3, 7]:
            ws.cell(row=tot_row, column=col_idx, value="").border = thin_border
            
        c_tg = ws.cell(row=tot_row, column=4, value=total_grand)
        c_tg.number_format = "#,##0.00"
        
        c_tp = ws.cell(row=tot_row, column=5, value=total_paid)
        c_tp.number_format = "#,##0.00"
        
        c_to = ws.cell(row=tot_row, column=6, value=total_outstanding)
        c_to.number_format = "#,##0.00"
        
        for c in [c_tg, c_tp, c_to]:
            c.font = bold_cell_font
            c.alignment = right_align
            c.border = thin_border
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        try:
            wb.save(path)
            QMessageBox.information(self, "Export Successful", f"Outstanding Invoices Report saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Failed to save Excel file:\n{e}")

    def export_aging_report_excel(self):
        default_path = self.get_safe_dialog_path("Customer_Aging_Report.xlsx")
        path, _ = QFileDialog.getSaveFileName(self, "Save Customer Aging Report", default_path, "Excel Files (*.xlsx)")
        if not path:
            return
            
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aging Report"
        ws.views.sheetView[0].showGridLines = True
        
        title_font = Font(name="Segoe UI", size=14, bold=True, color="1e3a8a")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        cell_font = Font(name="Segoe UI", size=10)
        bold_cell_font = Font(name="Segoe UI", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_side = Side(border_style="thin", color="cbd5e1")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws["A1"] = "PuruNiti Smart Billing - Customer Aging Summary"
        ws["A1"].font = title_font
        ws.row_dimensions[1].height = 25
        
        ws["A2"] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} (Brackets: {self.txt_aging_brackets.text()})"
        ws["A2"].font = Font(name="Segoe UI", size=9, italic=True)
        ws.row_dimensions[2].height = 18
        
        col_count = self.tbl_aging_report.columnCount()
        headers = []
        for col in range(col_count):
            headers.append(self.tbl_aging_report.horizontalHeaderItem(col).text())
            
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[4].height = 24
        
        row_count = self.tbl_aging_report.rowCount()
        column_totals = [0.0] * (col_count - 1)
        
        for r_idx in range(row_count):
            row_excel = 5 + r_idx
            ws.row_dimensions[row_excel].height = 19
            
            cust_name = self.tbl_aging_report.item(r_idx, 0).text()
            c_name = ws.cell(row=row_excel, column=1, value=cust_name)
            c_name.font = cell_font
            c_name.alignment = left_align
            c_name.border = thin_border
            
            for c_idx in range(1, col_count):
                val_str = self.tbl_aging_report.item(r_idx, c_idx).text()
                try:
                    val = float(val_str)
                except ValueError:
                    val = 0.0
                column_totals[c_idx - 1] += val
                
                c_val = ws.cell(row=row_excel, column=c_idx + 1, value=val)
                c_val.font = cell_font
                c_val.alignment = right_align
                c_val.number_format = "#,##0.00"
                c_val.border = thin_border
                
        tot_row = 5 + row_count
        ws.row_dimensions[tot_row].height = 22
        
        c_lbl = ws.cell(row=tot_row, column=1, value="TOTALS")
        c_lbl.font = bold_cell_font
        c_lbl.alignment = left_align
        c_lbl.border = thin_border
        
        for col_idx in range(1, col_count):
            c_tot = ws.cell(row=tot_row, column=col_idx + 1, value=column_totals[col_idx - 1])
            c_tot.font = bold_cell_font
            c_tot.alignment = right_align
            c_tot.number_format = "#,##0.00"
            c_tot.border = thin_border
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 13)
            
        try:
            wb.save(path)
            QMessageBox.information(self, "Export Successful", f"Customer Aging Summary Report saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Failed to save Excel file:\n{e}")

    def on_receipts_period_changed(self, text):
        is_custom = (text == "Custom Period")
        self.dt_receipts_start.setEnabled(is_custom)
        self.dt_receipts_end.setEnabled(is_custom)
        
        from datetime import date, timedelta
        today = date.today()
        
        if text == "This Month":
            start_date = date(today.year, today.month, 1)
            import calendar
            last_day = calendar.monthrange(today.year, today.month)[1]
            end_date = date(today.year, today.month, last_day)
            self.dt_receipts_start.setDate(QDate(start_date))
            self.dt_receipts_end.setDate(QDate(end_date))
        elif text == "Last Month":
            first_day_this_month = date(today.year, today.month, 1)
            last_day_last_month = first_day_this_month - timedelta(days=1)
            first_day_last_month = date(last_day_last_month.year, last_day_last_month.month, 1)
            self.dt_receipts_start.setDate(QDate(first_day_last_month))
            self.dt_receipts_end.setDate(QDate(last_day_last_month))
        elif text == "This Financial Year (Apr-Mar)":
            if today.month >= 4:
                start_year = today.year
                end_year = today.year + 1
            else:
                start_year = today.year - 1
                end_year = today.year
            self.dt_receipts_start.setDate(QDate(start_year, 4, 1))
            self.dt_receipts_end.setDate(QDate(end_year, 3, 31))

    def refresh_receipts_ledger(self):
        self.tbl_receipts_ledger.setRowCount(0)
        period = self.cmb_receipts_period.currentText()
        
        start_date = None
        end_date = None
        
        if period != "All Receipts":
            start_date = self.dt_receipts_start.date().toString("yyyy-MM-dd")
            end_date = self.dt_receipts_end.date().toString("yyyy-MM-dd")
            
        payments = get_payments_by_period(start_date, end_date)
        self.tbl_receipts_ledger.setRowCount(len(payments))
        
        for row, p in enumerate(payments):
            self.tbl_receipts_ledger.setItem(row, 0, QTableWidgetItem(str(p["id"])))
            self.tbl_receipts_ledger.setItem(row, 1, QTableWidgetItem(p["payment_date"]))
            self.tbl_receipts_ledger.setItem(row, 2, QTableWidgetItem(p["invoice_number"]))
            self.tbl_receipts_ledger.setItem(row, 3, QTableWidgetItem(p.get("customer_name") or "Walk-in"))
            
            c_amt = QTableWidgetItem(f"{p['amount']:.2f}")
            c_amt.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_receipts_ledger.setItem(row, 4, c_amt)
            
            self.tbl_receipts_ledger.setItem(row, 5, QTableWidgetItem(p["payment_mode"]))
            self.tbl_receipts_ledger.setItem(row, 6, QTableWidgetItem(p["reference_number"] or ""))
            self.tbl_receipts_ledger.setItem(row, 7, QTableWidgetItem(p["notes"] or ""))

    def export_receipts_ledger_excel(self):
        default_path = self.get_safe_dialog_path("Received_Payments_Ledger.xlsx")
        path, _ = QFileDialog.getSaveFileName(self, "Save Received Payments Ledger", default_path, "Excel Files (*.xlsx)")
        if not path:
            return
            
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payments Received"
        ws.views.sheetView[0].showGridLines = True
        
        title_font = Font(name="Segoe UI", size=14, bold=True, color="1e3a8a")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        cell_font = Font(name="Segoe UI", size=10)
        bold_cell_font = Font(name="Segoe UI", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_side = Side(border_style="thin", color="cbd5e1")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        ws["A1"] = "PuruNiti Smart Billing - Received Payments Ledger"
        ws["A1"].font = title_font
        ws.row_dimensions[1].height = 25
        
        ws["A2"] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} (Filter: {self.cmb_receipts_period.currentText()})"
        if self.cmb_receipts_period.currentText() != "All Receipts":
            ws["A2"].value += f" ({self.dt_receipts_start.date().toString('yyyy-MM-dd')} to {self.dt_receipts_end.date().toString('yyyy-MM-dd')})"
        ws["A2"].font = Font(name="Segoe UI", size=9, italic=True)
        ws.row_dimensions[2].height = 18
        
        headers = ["Receipt ID", "Payment Date", "Invoice #", "Customer Name", "Amount Received (₹)", "Payment Mode", "Reference / UTR", "Internal Notes"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[4].height = 24
        
        row_count = self.tbl_receipts_ledger.rowCount()
        total_received = 0.0
        
        for r_idx in range(row_count):
            row_excel = 5 + r_idx
            ws.row_dimensions[row_excel].height = 19
            
            c0 = ws.cell(row=row_excel, column=1, value=self.tbl_receipts_ledger.item(r_idx, 0).text())
            c0.alignment = center_align
            
            c1 = ws.cell(row=row_excel, column=2, value=self.tbl_receipts_ledger.item(r_idx, 1).text())
            c1.alignment = center_align
            
            c2 = ws.cell(row=row_excel, column=3, value=self.tbl_receipts_ledger.item(r_idx, 2).text())
            c2.alignment = center_align
            
            c3 = ws.cell(row=row_excel, column=4, value=self.tbl_receipts_ledger.item(r_idx, 3).text())
            c3.alignment = left_align
            
            amt_str = self.tbl_receipts_ledger.item(r_idx, 4).text()
            try:
                amt = float(amt_str)
            except ValueError:
                amt = 0.0
            total_received += amt
            
            c4 = ws.cell(row=row_excel, column=5, value=amt)
            c4.alignment = right_align
            c4.number_format = "#,##0.00"
            
            c5 = ws.cell(row=row_excel, column=6, value=self.tbl_receipts_ledger.item(r_idx, 5).text())
            c5.alignment = center_align
            
            c6 = ws.cell(row=row_excel, column=7, value=self.tbl_receipts_ledger.item(r_idx, 6).text())
            c6.alignment = left_align
            
            c7 = ws.cell(row=row_excel, column=8, value=self.tbl_receipts_ledger.item(r_idx, 7).text())
            c7.alignment = left_align
            
            for c in [c0, c1, c2, c3, c4, c5, c6, c7]:
                c.font = cell_font
                c.border = thin_border
                
        tot_row = 5 + row_count
        ws.row_dimensions[tot_row].height = 22
        
        c_lbl = ws.cell(row=tot_row, column=1, value="TOTALS")
        c_lbl.font = bold_cell_font
        c_lbl.alignment = center_align
        c_lbl.border = thin_border
        
        for col_idx in [2, 3, 4, 6, 7, 8]:
            ws.cell(row=tot_row, column=col_idx, value="").border = thin_border
            
        c_tr = ws.cell(row=tot_row, column=5, value=total_received)
        c_tr.number_format = "#,##0.00"
        c_tr.font = bold_cell_font
        c_tr.alignment = right_align
        c_tr.border = thin_border
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        try:
            wb.save(path)
            QMessageBox.information(self, "Export Successful", f"Received Payments Ledger saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Failed to save Excel file:\n{e}")

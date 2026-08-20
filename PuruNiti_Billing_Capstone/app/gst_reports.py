import os
import sqlite3
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.database import get_db_connection

def query_report_data(start_date, end_date):
    """Fetches invoices and item data grouped for reports within date range."""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. Fetch Invoices list
    cursor.execute("""
        SELECT * FROM invoices 
        WHERE date BETWEEN ? AND ?
        ORDER BY date ASC, invoice_number ASC
    """, (start_date, end_date))
    invoices = [dict(row) for row in cursor.fetchall()]
    
    # 2. Fetch associated items
    cursor.execute("""
        SELECT item.*, inv.date, inv.customer_gstin 
        FROM invoice_items item
        JOIN invoices inv ON item.invoice_number = inv.invoice_number
        WHERE inv.date BETWEEN ? AND ?
    """, (start_date, end_date))
    items = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    return invoices, items

def generate_gstr1_excel(start_date, end_date, output_path):
    """Generates a clean GSTR-1 Excel Returns Report with sheets for B2B, B2CS, HSN, and Docs."""
    invoices, items = query_report_data(start_date, end_date)
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Border & Fills styles
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    align_center = Alignment(horizontal="center", vertical="center")
    
    # ------------------ SHEET 1: B2B ------------------
    ws_b2b = wb.create_sheet(title="B2B")
    b2b_headers = [
        "Recipient GSTIN", "Customer Name", "Invoice Number", "Invoice Date", 
        "Invoice Value", "Place of Supply", "Reverse Charge", "Tax Rate %", 
        "Taxable Value", "CGST Amount", "SGST Amount", "IGST Amount", "Cess Amount"
    ]
    ws_b2b.append(b2b_headers)
    ws_b2b.row_dimensions[1].height = 24
    
    b2b_count = 0
    for inv in invoices:
        cust_gst = inv.get("customer_gstin", "").strip()
        if not cust_gst:
            continue
        # Get items for this invoice to show per-tax-rate rows
        inv_items = [x for x in items if x["invoice_number"] == inv["invoice_number"]]
        
        # Group by GST rate
        rates_map = {}
        for it in inv_items:
            r = it["gst_rate"]
            if r not in rates_map:
                rates_map[r] = {
                    "taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "cess": 0.0
                }
            rates_map[r]["taxable"] += it["taxable_amount"]
            rates_map[r]["cgst"] += it["cgst_amount"]
            rates_map[r]["sgst"] += it["sgst_amount"]
            rates_map[r]["igst"] += it["igst_amount"]
            rates_map[r]["cess"] += it["cess_amount"]
            
        for rate, vals in rates_map.items():
            ws_b2b.append([
                cust_gst,
                inv.get("customer_name"),
                inv.get("invoice_number"),
                inv.get("date"),
                inv.get("grand_total"),
                inv.get("place_of_supply"),
                "Yes" if inv.get("rcm") == 1 else "No",
                rate,
                vals["taxable"],
                vals["cgst"],
                vals["sgst"],
                vals["igst"],
                vals["cess"]
            ])
            b2b_count += 1
            
    # ------------------ SHEET 2: B2CS ------------------
    ws_b2cs = wb.create_sheet(title="B2CS")
    b2cs_headers = [
        "Type", "Place of Supply", "Tax Rate %", "Taxable Value", 
        "CGST Amount", "SGST Amount", "IGST Amount", "Cess Amount"
    ]
    ws_b2cs.append(b2cs_headers)
    ws_b2cs.row_dimensions[1].height = 24
    
    # Aggregate consumer sales by State (Place of Supply) & GST rate
    b2cs_map = {}
    for inv in invoices:
        cust_gst = inv.get("customer_gstin", "").strip()
        if cust_gst:
            # Skip B2B
            continue
        inv_items = [x for x in items if x["invoice_number"] == inv["invoice_number"]]
        pos = inv.get("place_of_supply", "Other")
        
        for it in inv_items:
            r = it["gst_rate"]
            key = (pos, r)
            if key not in b2cs_map:
                b2cs_map[key] = {
                    "taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "cess": 0.0
                }
            b2cs_map[key]["taxable"] += it["taxable_amount"]
            b2cs_map[key]["cgst"] += it["cgst_amount"]
            b2cs_map[key]["sgst"] += it["sgst_amount"]
            b2cs_map[key]["igst"] += it["igst_amount"]
            b2cs_map[key]["cess"] += it["cess_amount"]
            
    for (pos, rate), vals in b2cs_map.items():
        ws_b2cs.append([
            "OE" if vals["igst"] > 0 else "PE",
            pos,
            rate,
            vals["taxable"],
            vals["cgst"],
            vals["sgst"],
            vals["igst"],
            vals["cess"]
        ])
        
    # ------------------ SHEET 3: HSN ------------------
    ws_hsn = wb.create_sheet(title="HSN")
    hsn_headers = [
        "HSN/SAC", "Description", "UQC", "Total Quantity", 
        "Total Value", "Taxable Value", "CGST Amount", "SGST Amount", "IGST Amount", "Cess Amount"
    ]
    ws_hsn.append(hsn_headers)
    ws_hsn.row_dimensions[1].height = 24
    
    # Group items by HSN code
    hsn_map = {}
    for it in items:
        hsn = it.get("hsn", "").strip() or "N/A"
        name = it.get("name", "Item/Service")
        if hsn not in hsn_map:
            hsn_map[hsn] = {
                "desc": name,
                "qty": 0.0,
                "total_val": 0.0,
                "taxable": 0.0,
                "cgst": 0.0,
                "sgst": 0.0,
                "igst": 0.0,
                "cess": 0.0
            }
        hsn_map[hsn]["qty"] += it["qty"]
        hsn_map[hsn]["total_val"] += it["total_amount"]
        hsn_map[hsn]["taxable"] += it["taxable_amount"]
        hsn_map[hsn]["cgst"] += it["cgst_amount"]
        hsn_map[hsn]["sgst"] += it["sgst_amount"]
        hsn_map[hsn]["igst"] += it["igst_amount"]
        hsn_map[hsn]["cess"] += it["cess_amount"]
        
    for hsn, vals in hsn_map.items():
        ws_hsn.append([
            hsn,
            vals["desc"],
            "NOS" if vals["qty"] % 1 == 0 else "OTH",
            vals["qty"],
            vals["total_val"],
            vals["taxable"],
            vals["cgst"],
            vals["sgst"],
            vals["igst"],
            vals["cess"]
        ])
        
    # ------------------ SHEET 4: Doc Summary ------------------
    ws_docs = wb.create_sheet(title="Doc_Summary")
    docs_headers = ["Nature of Document", "Sr. No. From", "Sr. No. To", "Total Number", "Cancelled"]
    ws_docs.append(docs_headers)
    ws_docs.row_dimensions[1].height = 24
    
    if invoices:
        from_inv = invoices[0]["invoice_number"]
        to_inv = invoices[-1]["invoice_number"]
        ws_docs.append([
            "Invoices for outward supply",
            from_inv,
            to_inv,
            len(invoices),
            0
        ])
    else:
        ws_docs.append(["Invoices for outward supply", "N/A", "N/A", 0, 0])
        
    # Apply global designs & formatting across all sheets
    for ws in wb.worksheets:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            
        for r_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 18
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=col_idx)
                cell.font = Font(name="Segoe UI", size=9.5)
                cell.border = thin_border
                
        # Fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(output_path)
    return True

def generate_gstr1_pdf(start_date, end_date, output_path):
    """Generates GSTR-1 Outward Return summary PDF with formatted sections."""
    invoices, items = query_report_data(start_date, end_date)
    
    # Standard landscape setup for wide columns
    doc = SimpleDocTemplate(
        output_path, 
        pagesize=landscape(A4),
        rightMargin=20, leftMargin=20, topMargin=25, bottomMargin=25
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'RepTitle', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=colors.HexColor('#1A365D'),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'RepSub', parent=styles['Normal'],
        fontName='Helvetica', fontSize=10, leading=12, textColor=colors.HexColor('#475569'),
        spaceAfter=15
    )
    section_style = ParagraphStyle(
        'RepSec', parent=styles['Heading2'],
        fontName='Helvetica-Bold', fontSize=12, leading=15, textColor=colors.HexColor('#1E3A8A'),
        spaceBefore=12, spaceAfter=6
    )
    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor('#1E293B')
    )
    cell_bold = ParagraphStyle(
        'CellBold', parent=cell_style,
        fontName='Helvetica-Bold'
    )
    header_style = ParagraphStyle(
        'Hdr', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white
    )
    
    story = []
    
    # Document Title
    story.append(Paragraph("PuruNiti Smart Billing system - GSTR-1 Outward Sales Return", title_style))
    story.append(Paragraph(f"Filing Period: {start_date} to {end_date} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    
    # ------------------ SECTION 1: B2B INVOICES ------------------
    story.append(Paragraph("Section 4A, 4B: B2B Outward Supplies (Registered Customers)", section_style))
    
    b2b_data = [[
        Paragraph("Recipient GSTIN", header_style), Paragraph("Name", header_style),
        Paragraph("Invoice No.", header_style), Paragraph("Date", header_style),
        Paragraph("Value", header_style), Paragraph("Place of Supply", header_style),
        Paragraph("Tax Rate %", header_style), Paragraph("Taxable Value", header_style),
        Paragraph("CGST", header_style), Paragraph("SGST", header_style),
        Paragraph("IGST", header_style), Paragraph("Cess", header_style)
    ]]
    
    b2b_total_taxable = 0.0
    b2b_total_cgst = 0.0
    b2b_total_sgst = 0.0
    b2b_total_igst = 0.0
    b2b_total_cess = 0.0
    
    for inv in invoices:
        cust_gst = inv.get("customer_gstin", "").strip()
        if not cust_gst:
            continue
        inv_items = [x for x in items if x["invoice_number"] == inv["invoice_number"]]
        
        rates_map = {}
        for it in inv_items:
            r = it["gst_rate"]
            if r not in rates_map:
                rates_map[r] = {"taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "cess": 0.0}
            rates_map[r]["taxable"] += it["taxable_amount"]
            rates_map[r]["cgst"] += it["cgst_amount"]
            rates_map[r]["sgst"] += it["sgst_amount"]
            rates_map[r]["igst"] += it["igst_amount"]
            rates_map[r]["cess"] += it["cess_amount"]
            
        for r, vals in rates_map.items():
            b2b_data.append([
                Paragraph(cust_gst, cell_style),
                Paragraph(inv.get("customer_name", "")[:18], cell_style),
                Paragraph(inv.get("invoice_number", ""), cell_style),
                Paragraph(inv.get("date", ""), cell_style),
                Paragraph(f"{inv.get('grand_total'):,.2f}", cell_style),
                Paragraph(inv.get("place_of_supply", ""), cell_style),
                Paragraph(f"{r}%", cell_style),
                Paragraph(f"{vals['taxable']:.2f}", cell_style),
                Paragraph(f"{vals['cgst']:.2f}", cell_style),
                Paragraph(f"{vals['sgst']:.2f}", cell_style),
                Paragraph(f"{vals['igst']:.2f}", cell_style),
                Paragraph(f"{vals['cess']:.2f}", cell_style)
            ])
            b2b_total_taxable += vals["taxable"]
            b2b_total_cgst += vals["cgst"]
            b2b_total_sgst += vals["sgst"]
            b2b_total_igst += vals["igst"]
            b2b_total_cess += vals["cess"]
            
    # Add B2B Total Row
    b2b_data.append([
        Paragraph("Total B2B Outward", cell_bold), Paragraph("", cell_style), Paragraph("", cell_style), Paragraph("", cell_style), Paragraph("", cell_style), Paragraph("", cell_style), Paragraph("", cell_style),
        Paragraph(f"{b2b_total_taxable:,.2f}", cell_bold),
        Paragraph(f"{b2b_total_cgst:,.2f}", cell_bold),
        Paragraph(f"{b2b_total_sgst:,.2f}", cell_bold),
        Paragraph(f"{b2b_total_igst:,.2f}", cell_bold),
        Paragraph(f"{b2b_total_cess:,.2f}", cell_bold)
    ])
    
    t_b2b = Table(b2b_data, colWidths=[80, 85, 55, 50, 50, 80, 40, 60, 50, 50, 50, 45])
    t_b2b.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
    ]))
    story.append(t_b2b)
    story.append(Spacer(1, 15))
    
    # ------------------ SECTION 2: B2C SMALL INVOICES ------------------
    story.append(Paragraph("Section 7: B2CS Outward Supplies (Unregistered Customers)", section_style))
    
    b2cs_data = [[
        Paragraph("Type", header_style), Paragraph("Place of Supply (State)", header_style),
        Paragraph("GST Rate %", header_style), Paragraph("Taxable Value", header_style),
        Paragraph("CGST Amount", header_style), Paragraph("SGST Amount", header_style),
        Paragraph("IGST Amount", header_style), Paragraph("Cess Amount", header_style)
    ]]
    
    b2cs_map = {}
    for inv in invoices:
        cust_gst = inv.get("customer_gstin", "").strip()
        if cust_gst:
            continue
        inv_items = [x for x in items if x["invoice_number"] == inv["invoice_number"]]
        pos = inv.get("place_of_supply", "Other")
        
        for it in inv_items:
            r = it["gst_rate"]
            key = (pos, r)
            if key not in b2cs_map:
                b2cs_map[key] = {"taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "cess": 0.0}
            b2cs_map[key]["taxable"] += it["taxable_amount"]
            b2cs_map[key]["cgst"] += it["cgst_amount"]
            b2cs_map[key]["sgst"] += it["sgst_amount"]
            b2cs_map[key]["igst"] += it["igst_amount"]
            b2cs_map[key]["cess"] += it["cess_amount"]
            
    b2cs_total_taxable = 0.0
    b2cs_total_cgst = 0.0
    b2cs_total_sgst = 0.0
    b2cs_total_igst = 0.0
    b2cs_total_cess = 0.0
    
    for (pos, rate), vals in b2cs_map.items():
        b2cs_data.append([
            Paragraph("OE" if vals["igst"] > 0 else "PE", cell_style),
            Paragraph(pos, cell_style),
            Paragraph(f"{rate}%", cell_style),
            Paragraph(f"{vals['taxable']:.2f}", cell_style),
            Paragraph(f"{vals['cgst']:.2f}", cell_style),
            Paragraph(f"{vals['sgst']:.2f}", cell_style),
            Paragraph(f"{vals['igst']:.2f}", cell_style),
            Paragraph(f"{vals['cess']:.2f}", cell_style)
        ])
        b2cs_total_taxable += vals["taxable"]
        b2cs_total_cgst += vals["cgst"]
        b2cs_total_sgst += vals["sgst"]
        b2cs_total_igst += vals["igst"]
        b2cs_total_cess += vals["cess"]
        
    b2cs_data.append([
        Paragraph("Total B2CS Outward", cell_bold), Paragraph("", cell_style), Paragraph("", cell_style),
        Paragraph(f"{b2cs_total_taxable:,.2f}", cell_bold),
        Paragraph(f"{b2cs_total_cgst:,.2f}", cell_bold),
        Paragraph(f"{b2cs_total_sgst:,.2f}", cell_bold),
        Paragraph(f"{b2cs_total_igst:,.2f}", cell_bold),
        Paragraph(f"{b2cs_total_cess:,.2f}", cell_bold)
    ])
    
    t_b2cs = Table(b2cs_data, colWidths=[60, 160, 60, 90, 80, 80, 80, 70])
    t_b2cs.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
    ]))
    story.append(t_b2cs)
    story.append(Spacer(1, 15))
    
    # ------------------ SECTION 3: HSN/SAC SUMMARY ------------------
    story.append(Paragraph("Section 12: HSN/SAC Summary of Outward Supplies", section_style))
    
    hsn_data = [[
        Paragraph("HSN/SAC", header_style), Paragraph("Description", header_style),
        Paragraph("UQC", header_style), Paragraph("Total Qty", header_style),
        Paragraph("Total Value", header_style), Paragraph("Taxable Value", header_style),
        Paragraph("CGST Amount", header_style), Paragraph("SGST Amount", header_style),
        Paragraph("IGST Amount", header_style), Paragraph("Cess Amount", header_style)
    ]]
    
    hsn_map = {}
    for it in items:
        hsn = it.get("hsn", "").strip() or "N/A"
        name = it.get("name", "Item/Service")
        if hsn not in hsn_map:
            hsn_map[hsn] = {
                "desc": name, "qty": 0.0, "total_val": 0.0, "taxable": 0.0,
                "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "cess": 0.0
            }
        hsn_map[hsn]["qty"] += it["qty"]
        hsn_map[hsn]["total_val"] += it["total_amount"]
        hsn_map[hsn]["taxable"] += it["taxable_amount"]
        hsn_map[hsn]["cgst"] += it["cgst_amount"]
        hsn_map[hsn]["sgst"] += it["sgst_amount"]
        hsn_map[hsn]["igst"] += it["igst_amount"]
        hsn_map[hsn]["cess"] += it["cess_amount"]
        
    for hsn, vals in hsn_map.items():
        hsn_data.append([
            Paragraph(hsn, cell_style),
            Paragraph(vals["desc"][:20], cell_style),
            Paragraph("NOS" if vals["qty"] % 1 == 0 else "OTH", cell_style),
            Paragraph(f"{vals['qty']:g}", cell_style),
            Paragraph(f"{vals['total_val']:.2f}", cell_style),
            Paragraph(f"{vals['taxable']:.2f}", cell_style),
            Paragraph(f"{vals['cgst']:.2f}", cell_style),
            Paragraph(f"{vals['sgst']:.2f}", cell_style),
            Paragraph(f"{vals['igst']:.2f}", cell_style),
            Paragraph(f"{vals['cess']:.2f}", cell_style)
        ])
        
    t_hsn = Table(hsn_data, colWidths=[65, 120, 45, 55, 75, 75, 65, 65, 65, 55])
    t_hsn.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
    ]))
    story.append(t_hsn)
    story.append(Spacer(1, 15))
    
    # ------------------ SECTION 4: DOCUMENT SEQUENCE SUMMARY ------------------
    story.append(Paragraph("Section 13: Documents Issued During Filing Period", section_style))
    
    from_inv = invoices[0]["invoice_number"] if invoices else "N/A"
    to_inv = invoices[-1]["invoice_number"] if invoices else "N/A"
    
    doc_summary_data = [
        [
            Paragraph("Document Type", header_style), Paragraph("Sr. From", header_style),
            Paragraph("Sr. To", header_style), Paragraph("Total Count", header_style),
            Paragraph("Cancelled", header_style)
        ],
        [
            Paragraph("Invoices for outward supply", cell_style),
            Paragraph(from_inv, cell_style),
            Paragraph(to_inv, cell_style),
            Paragraph(str(len(invoices)), cell_style),
            Paragraph("0", cell_style)
        ]
    ]
    t_docs = Table(doc_summary_data, colWidths=[180, 100, 100, 80, 80])
    t_docs.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
    ]))
    story.append(t_docs)
    
    doc.build(story)
    return True

def generate_dashboard_excel(start_date, end_date, output_path):
    """Generates a clean Dashboard Business Summary Report in Excel format."""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # KPI Totals
    cursor.execute("""
        SELECT 
            SUM(subtotal) as sub, SUM(cgst_total) as cgst, SUM(sgst_total) as sgst,
            SUM(igst_total) as igst, SUM(cess_total) as cess, SUM(grand_total) as grand
        FROM invoices 
        WHERE date BETWEEN ? AND ?
    """, (start_date, end_date))
    kpi = dict(cursor.fetchone())
    
    # Client-wise summary
    cursor.execute("""
        SELECT 
            customer_name, customer_mobile, COUNT(invoice_number) as count,
            SUM(subtotal) as sub, SUM(cgst_total+sgst_total+igst_total+cess_total) as tax,
            SUM(grand_total) as grand
        FROM invoices
        WHERE date BETWEEN ? AND ?
        GROUP BY customer_name, customer_mobile
        ORDER BY grand DESC
    """, (start_date, end_date))
    clients = [dict(row) for row in cursor.fetchall()]
    
    # Product ranking
    cursor.execute("""
        SELECT 
            name, hsn, SUM(qty) as qty, SUM(taxable_amount) as sub, SUM(total_amount) as grand
        FROM invoice_items
        WHERE invoice_number IN (SELECT invoice_number FROM invoices WHERE date BETWEEN ? AND ?)
        GROUP BY name, hsn
        ORDER BY grand DESC
    """, (start_date, end_date))
    products = [dict(row) for row in cursor.fetchall()]
    
    # HSN breakdown
    cursor.execute("""
        SELECT 
            hsn, SUM(taxable_amount) as sub, SUM(cgst_amount) as cgst,
            SUM(sgst_amount) as sgst, SUM(igst_amount) as igst,
            SUM(cess_amount) as cess, SUM(total_amount) as grand
        FROM invoice_items
        WHERE invoice_number IN (SELECT invoice_number FROM invoices WHERE date BETWEEN ? AND ?)
        GROUP BY hsn
        ORDER BY grand DESC
    """, (start_date, end_date))
    hsn_list = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Formatting
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    align_center = Alignment(horizontal="center", vertical="center")
    
    # ------------------ SHEET 1: KPIs ------------------
    ws_kpi = wb.create_sheet(title="KPI_Summary")
    ws_kpi.append(["Metric", "Total Value (INR)"])
    ws_kpi.append(["Taxable Value (Subtotal)", kpi.get("sub") or 0.0])
    ws_kpi.append(["CGST Total", kpi.get("cgst") or 0.0])
    ws_kpi.append(["SGST Total", kpi.get("sgst") or 0.0])
    ws_kpi.append(["IGST Total", kpi.get("igst") or 0.0])
    ws_kpi.append(["Cess Total", kpi.get("cess") or 0.0])
    ws_kpi.append(["Gross Grand Total", kpi.get("grand") or 0.0])
    
    # ------------------ SHEET 2: Client Rankings ------------------
    ws_clients = wb.create_sheet(title="Client_Billing")
    ws_clients.append(["Customer Name", "Phone", "Invoices Count", "Taxable Value", "Total Taxes", "Grand Total"])
    for c in clients:
        ws_clients.append([
            c.get("customer_name"), c.get("customer_mobile"), c.get("count"),
            c.get("sub"), c.get("tax"), c.get("grand")
        ])
        
    # ------------------ SHEET 3: Product Sales ------------------
    ws_products = wb.create_sheet(title="Product_Billing")
    ws_products.append(["Product / Service Name", "HSN/SAC", "Quantity Sold", "Taxable Revenue", "Grand Total Sales"])
    for p in products:
        ws_products.append([
            p.get("name"), p.get("hsn"), p.get("qty"), p.get("sub"), p.get("grand")
        ])
        
    # ------------------ SHEET 4: HSN Summary ------------------
    ws_hsn = wb.create_sheet(title="HSN_Billing")
    ws_hsn.append(["HSN/SAC", "Taxable Value", "CGST", "SGST", "IGST", "Cess", "Grand Total"])
    for h in hsn_list:
        ws_hsn.append([
            h.get("hsn"), h.get("sub"), h.get("cgst"), h.get("sgst"), h.get("igst"), h.get("cess"), h.get("grand")
        ])
        
    # Formatting
    for ws in wb.worksheets:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            
        for r_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 18
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=col_idx)
                cell.font = Font(name="Segoe UI", size=9.5)
                cell.border = thin_border
                
        # Fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(output_path)
    return True

def generate_dashboard_pdf(start_date, end_date, output_path):
    """Generates Dashboard business metrics summary PDF."""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # KPIs
    cursor.execute("""
        SELECT 
            SUM(subtotal) as sub, SUM(cgst_total) as cgst, SUM(sgst_total) as sgst,
            SUM(igst_total) as igst, SUM(cess_total) as cess, SUM(grand_total) as grand
        FROM invoices 
        WHERE date BETWEEN ? AND ?
    """, (start_date, end_date))
    kpi = dict(cursor.fetchone())
    
    # Client Summary
    cursor.execute("""
        SELECT 
            customer_name, customer_mobile, COUNT(invoice_number) as count,
            SUM(subtotal) as sub, SUM(cgst_total+sgst_total+igst_total+cess_total) as tax,
            SUM(grand_total) as grand
        FROM invoices
        WHERE date BETWEEN ? AND ?
        GROUP BY customer_name, customer_mobile
        ORDER BY grand DESC
    """, (start_date, end_date))
    clients = [dict(row) for row in cursor.fetchall()]
    
    # Products
    cursor.execute("""
        SELECT 
            name, hsn, SUM(qty) as qty, SUM(taxable_amount) as sub, SUM(total_amount) as grand
        FROM invoice_items
        WHERE invoice_number IN (SELECT invoice_number FROM invoices WHERE date BETWEEN ? AND ?)
        GROUP BY name, hsn
        ORDER BY grand DESC
    """, (start_date, end_date))
    products = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    doc = SimpleDocTemplate(
        output_path, 
        pagesize=A4,
        rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'RepTitle', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=colors.HexColor('#1A365D'),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'RepSub', parent=styles['Normal'],
        fontName='Helvetica', fontSize=10, leading=12, textColor=colors.HexColor('#475569'),
        spaceAfter=15
    )
    section_style = ParagraphStyle(
        'RepSec', parent=styles['Heading2'],
        fontName='Helvetica-Bold', fontSize=12, leading=15, textColor=colors.HexColor('#1E3A8A'),
        spaceBefore=15, spaceAfter=8
    )
    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=9, leading=11, textColor=colors.HexColor('#1E293B')
    )
    cell_bold = ParagraphStyle(
        'CellBold', parent=cell_style,
        fontName='Helvetica-Bold'
    )
    header_style = ParagraphStyle(
        'Hdr', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.white
    )
    
    story = []
    story.append(Paragraph("PuruNiti Smart Billing system - Business Performance Summary", title_style))
    story.append(Paragraph(f"Reporting Period: {start_date} to {end_date} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    
    # 1. KPI Cards Table
    story.append(Paragraph("Key Financial Indicators (INR)", section_style))
    kpi_data = [
        [Paragraph("Financial Metric", header_style), Paragraph("Total Accumulated Value", header_style)],
        [Paragraph("Taxable Subtotal (Sales value)", cell_style), Paragraph(f"INR {kpi.get('sub') or 0.0:,.2f}", cell_style)],
        [Paragraph("CGST Collected", cell_style), Paragraph(f"INR {kpi.get('cgst') or 0.0:,.2f}", cell_style)],
        [Paragraph("SGST Collected", cell_style), Paragraph(f"INR {kpi.get('sgst') or 0.0:,.2f}", cell_style)],
        [Paragraph("IGST Collected", cell_style), Paragraph(f"INR {kpi.get('igst') or 0.0:,.2f}", cell_style)],
        [Paragraph("Cess Collected", cell_style), Paragraph(f"INR {kpi.get('cess') or 0.0:,.2f}", cell_style)],
        [Paragraph("Gross Revenue (Grand Total)", cell_bold), Paragraph(f"INR {kpi.get('grand') or 0.0:,.2f}", cell_bold)]
    ]
    t_kpis = Table(kpi_data, colWidths=[250, 250])
    t_kpis.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E2E8F0')),
    ]))
    story.append(t_kpis)
    story.append(Spacer(1, 15))
    
    # 2. Top Clients
    story.append(Paragraph("Top Clients by Revenue", section_style))
    client_data = [[
        Paragraph("Client Name", header_style), Paragraph("Phone", header_style),
        Paragraph("Invoices", header_style), Paragraph("Taxable Value", header_style),
        Paragraph("Grand Total", header_style)
    ]]
    for c in clients[:10]: # Top 10 clients
        client_data.append([
            Paragraph(c.get("customer_name", "")[:28], cell_style),
            Paragraph(c.get("customer_mobile", "") or "N/A", cell_style),
            Paragraph(str(c.get("count")), cell_style),
            Paragraph(f"{c.get('sub'):,.2f}", cell_style),
            Paragraph(f"{c.get('grand'):,.2f}", cell_style)
        ])
    t_cl = Table(client_data, colWidths=[160, 90, 60, 95, 95])
    t_cl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
    ]))
    story.append(t_cl)
    story.append(Spacer(1, 15))
    
    # 3. Product Sales
    story.append(Paragraph("Top Selling Products / Services", section_style))
    prod_data = [[
        Paragraph("Product/Service Description", header_style), Paragraph("HSN/SAC", header_style),
        Paragraph("Qty Sold", header_style), Paragraph("Taxable Revenue", header_style),
        Paragraph("Total Billing", header_style)
    ]]
    for p in products[:10]: # Top 10 products
        prod_data.append([
            Paragraph(p.get("name", "")[:32], cell_style),
            Paragraph(p.get("hsn", "") or "N/A", cell_style),
            Paragraph(f"{p.get('qty'):g}", cell_style),
            Paragraph(f"{p.get('sub'):,.2f}", cell_style),
            Paragraph(f"{p.get('grand'):,.2f}", cell_style)
        ])
    t_pr = Table(prod_data, colWidths=[180, 80, 60, 90, 90])
    t_pr.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
    ]))
    story.append(t_pr)
    
    doc.build(story)
    return True

def generate_dynamic_dashboard_excel(dimension_name, measure_name, start_date, end_date, data_list, output_path):
    """Generates an Excel spreadsheet for the dynamic query results."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dynamic Analysis"
    
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    align_center = Alignment(horizontal="center", vertical="center")
    
    # Title Rows
    ws.append([f"PuruNiti Smart Billing system - Dynamic Query Analysis Report"])
    ws.append([f"Analysis Period: {start_date} to {end_date}"])
    ws.append([f"Group By (Dimension): {dimension_name} | Metric (Measure): {measure_name}"])
    ws.append([]) # empty separator row
    
    # Table headers
    headers = [dimension_name, f"Total {measure_name} (INR)", "Share %"]
    ws.append(headers)
    
    header_row_idx = 5
    ws.row_dimensions[header_row_idx].height = 24
    
    total_val = sum(val for label, val in data_list)
    if total_val == 0:
        total_val = 1.0
        
    for label, val in data_list:
        share = (val / total_val) * 100.0
        ws.append([label, val, f"{share:.2f}%"])
        
    # Append Total Row
    ws.append(["Total", total_val, "100.00%"])
    
    # Format headers
    for col_idx in range(1, 4):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    # Format data cells
    for r_idx in range(6, ws.max_row + 1):
        ws.row_dimensions[r_idx].height = 18
        for col_idx in range(1, 4):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Segoe UI", size=9.5)
            if r_idx == ws.max_row:
                cell.font = Font(name="Segoe UI", size=9.5, bold=True)
                
    # Fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    wb.save(output_path)
    return True

def generate_dynamic_dashboard_pdf(dimension_name, measure_name, start_date, end_date, data_list, output_path):
    """Generates a PDF report for the dynamic query results."""
    doc = SimpleDocTemplate(
        output_path, 
        pagesize=A4,
        rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'RepTitle', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=colors.HexColor('#1A365D'),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'RepSub', parent=styles['Normal'],
        fontName='Helvetica', fontSize=10, leading=12, textColor=colors.HexColor('#475569'),
        spaceAfter=15
    )
    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=9, leading=11, textColor=colors.HexColor('#1E293B')
    )
    cell_bold = ParagraphStyle(
        'CellBold', parent=cell_style,
        fontName='Helvetica-Bold'
    )
    header_style = ParagraphStyle(
        'Hdr', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.white
    )
    
    from datetime import datetime
    story = []
    story.append(Paragraph("PuruNiti Smart Billing system - Dynamic Query Report", title_style))
    story.append(Paragraph(f"Analysis Period: {start_date} to {end_date} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    story.append(Paragraph(f"<b>Grouped By:</b> {dimension_name} &nbsp;&nbsp;&nbsp;&nbsp; <b>Metric Measured:</b> {measure_name}", subtitle_style))
    story.append(Spacer(1, 10))
    
    # Table data
    table_data = [[
        Paragraph(dimension_name, header_style),
        Paragraph(f"Total {measure_name} (INR)", header_style),
        Paragraph("Percentage Share", header_style)
    ]]
    
    total_val = sum(val for label, val in data_list)
    if total_val == 0:
        total_val = 1.0
        
    for label, val in data_list:
        share = (val / total_val) * 100.0
        table_data.append([
            Paragraph(label, cell_style),
            Paragraph(f"INR {val:,.2f}", cell_style),
            Paragraph(f"{share:.2f}%", cell_style)
        ])
        
    # Total row
    table_data.append([
        Paragraph("Total Sum", cell_bold),
        Paragraph(f"INR {total_val:,.2f}", cell_bold),
        Paragraph("100.00%", cell_bold)
    ])
    
    t_results = Table(table_data, colWidths=[240, 160, 100])
    t_results.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E2E8F0')),
    ]))
    story.append(t_results)
    
    doc.build(story)
    return True

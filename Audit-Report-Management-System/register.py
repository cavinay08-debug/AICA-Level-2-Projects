from __future__ import annotations
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from threading import Lock
from config import REGISTER_HEADERS, REGISTER_PATH
_lock=Lock()

def ensure_register():
    REGISTER_PATH.parent.mkdir(parents=True,exist_ok=True)
    if not REGISTER_PATH.exists():
        wb=Workbook(); ws=wb.active; ws.title='REPORT_REGISTER'; ws.append(REGISTER_HEADERS)
        for c in ws[1]: c.font=Font(bold=True); c.alignment=Alignment(horizontal='center')
        ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions; wb.save(REGISTER_PATH); wb.close(); return
    with _lock:
        wb=load_workbook(REGISTER_PATH); ws=wb['REPORT_REGISTER']; existing=[c.value for c in ws[1]]
        changed=False
        for h in REGISTER_HEADERS:
            if h not in existing:
                ws.cell(1,ws.max_column+1).value=h; changed=True
        if changed:
            for c in ws[1]: c.font=Font(bold=True); c.alignment=Alignment(horizontal='center')
            ws.auto_filter.ref=ws.dimensions; wb.save(REGISTER_PATH)
        wb.close()

def next_report_id():
    ensure_register()
    with _lock:
        wb=load_workbook(REGISTER_PATH); ws=wb['REPORT_REGISTER'];
        nums=[]
        for row in range(2,ws.max_row+1):
            v=str(ws.cell(row,1).value or '')
            if v.startswith('AR-'):
                try: nums.append(int(v.split('-')[1]))
                except: pass
        n=max(nums,default=0)+1; wb.close(); return f'AR-{n:05d}'

def add_record(record):
    ensure_register()
    with _lock:
        wb=load_workbook(REGISTER_PATH); ws=wb['REPORT_REGISTER']; headers=[c.value for c in ws[1]]
        ws.append([record.get(h,'') for h in headers]); ws.auto_filter.ref=ws.dimensions; wb.save(REGISTER_PATH); wb.close()

def list_records(limit=500):
    ensure_register(); wb=load_workbook(REGISTER_PATH,read_only=True,data_only=True); ws=wb['REPORT_REGISTER']; headers=[c.value for c in ws[1]]; rows=[dict(zip(headers,row)) for row in ws.iter_rows(min_row=2,values_only=True)]; wb.close(); rows.reverse(); return rows[:limit]

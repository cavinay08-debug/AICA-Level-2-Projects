from __future__ import annotations
from datetime import datetime
from threading import Lock
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from config import *

_lock = Lock()

def _init_sheet(ws, headers):
    if ws.max_row == 1 and all(c.value is None for c in ws[1]):
        ws.append(headers)
    elif ws.max_row == 1 and [c.value for c in ws[1]] != headers:
        # Do not overwrite an existing sheet; migration is handled separately.
        pass
    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

def ensure_master():
    MASTER_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not MASTER_PATH.exists():
        wb = Workbook()
        specs = [
            ('FIRM_MASTER', FIRM_HEADERS), ('PARTNER_MASTER', PARTNER_HEADERS),
            ('CLIENT_MASTER', CLIENT_HEADERS), ('ENGAGEMENT_MASTER', ENGAGEMENT_HEADERS),
            ('TEMPLATE_MASTER', TEMPLATE_HEADERS), ('VARIABLE_LIBRARY', VARIABLE_HEADERS),
            ('WORDING_LIBRARY', WORDING_HEADERS), ('CLASSIFICATION_RESULTS', CLASSIFICATION_HEADERS), ('APPLICABILITY_RESULTS', APPLICABILITY_HEADERS)]
        for i, (name, headers) in enumerate(specs):
            ws = wb.active if i == 0 else wb.create_sheet()
            ws.title = name; ws.append(headers); _init_sheet(ws, headers)
        wb['FIRM_MASTER'].append(['FIRM-001','Your Firm Name','000000W/W000000','Office Address','Nashik','Yes'])
        wb['PARTNER_MASTER'].append(['P-001','FIRM-001','Partner 1','000000','Nashik','Yes'])
        wb['PARTNER_MASTER'].append(['P-002','FIRM-001','Partner 2','000001','Nashik','Yes'])
        _seed_reference_sheets(wb)
        wb.save(MASTER_PATH); wb.close()
        return

    # Non-destructive migration of a Stage 3 database.
    with _lock:
        wb = load_workbook(MASTER_PATH)
        for name, headers in [('CLIENT_MASTER', CLIENT_HEADERS), ('ENGAGEMENT_MASTER', ENGAGEMENT_HEADERS)]:
            if name not in wb.sheetnames:
                ws = wb.create_sheet(name); ws.append(headers); _init_sheet(ws, headers)
        # Ensure reference sheets exist too, without destroying existing data.
        for name, headers in [('FIRM_MASTER',FIRM_HEADERS),('PARTNER_MASTER',PARTNER_HEADERS),('TEMPLATE_MASTER',TEMPLATE_HEADERS),('VARIABLE_LIBRARY',VARIABLE_HEADERS),('WORDING_LIBRARY',WORDING_HEADERS),('CLASSIFICATION_RESULTS',CLASSIFICATION_HEADERS),('APPLICABILITY_RESULTS',APPLICABILITY_HEADERS)]:
            if name not in wb.sheetnames:
                ws=wb.create_sheet(name); ws.append(headers); _init_sheet(ws, headers)
        if 'TEMPLATE_MASTER' in wb.sheetnames and wb['TEMPLATE_MASTER'].max_row == 1:
            _seed_reference_sheets(wb)
        wb.save(MASTER_PATH); wb.close()

def _seed_reference_sheets(wb):
    ws = wb['TEMPLATE_MASTER']
    for i, ((entity, small, opinion), fn) in enumerate(TEMPLATE_RULES.items(), 1):
        ws.append([f'T-{i:03d}',fn,entity,'Yes' if small else 'No',opinion,'1.0',datetime.now().strftime('%Y-%m-%d'),'Active','Stage 4 master template'])
    ws = wb['VARIABLE_LIBRARY']
    for code,desc,req in CANONICAL_VARIABLES:
        aliases=PLACEHOLDER_ALIASES.get(code,[])+FIXED_SIGNATURE_ALIASES.get(code,[])
        source='User Input' if code in {'CLIENT_NAME','CIN','CITY','YEAR_ENDING','REPORT_DATE','UDIN'} else ('Firm Master' if code in {'FIRM_NAME','FRN','PLACE'} else 'Partner Master')
        ws.append([code,desc,req,' | '.join(aliases),source,'Yes'])
    wb['WORDING_LIBRARY'].append(['W-001','QUALIFICATION_BASIS','Qualified opinion – current template baseline','Private Ltd / Qualified','Use the wording contained in the approved master Word template. Edit only after professional review.','Active','1.0','Reference only; not an automated legal conclusion.'])

def _rows(sheet):
    ensure_master(); wb=load_workbook(MASTER_PATH,read_only=True,data_only=True); ws=wb[sheet]
    headers=[c.value for c in ws[1]]; rows=[dict(zip(headers,row)) for row in ws.iter_rows(min_row=2,values_only=True)]; wb.close(); return rows

def _active(rows, key='Active', active_only=True):
    return [r for r in rows if not active_only or str(r.get(key,'')).lower()=='yes']

def list_firms(active_only=True): return _active(_rows('FIRM_MASTER'),active_only=active_only)
def list_partners(firm_id=None,active_only=True):
    rows=_rows('PARTNER_MASTER'); return [r for r in rows if (not firm_id or r.get('FirmID')==firm_id) and (not active_only or str(r.get('Active','')).lower()=='yes')]
def get_firm(fid): return next((r for r in list_firms(False) if r.get('FirmID')==fid),None)
def get_partner(pid): return next((r for r in list_partners(None,False) if r.get('PartnerID')==pid),None)

def list_clients(active_only=True): return _active(_rows('CLIENT_MASTER'),active_only=active_only)
def get_client(cid): return next((r for r in list_clients(False) if r.get('ClientID')==cid),None)
def list_engagements(client_id=None, financial_year=None, active_only=False):
    rows=_rows('ENGAGEMENT_MASTER')
    return [r for r in rows if (not client_id or r.get('ClientID')==client_id) and (not financial_year or r.get('FinancialYear')==financial_year) and (not active_only or str(r.get('Status','')).lower()=='active')]
def get_engagement(eid): return next((r for r in list_engagements() if r.get('EngagementID')==eid),None)

def _append(sheet, values):
    ensure_master()
    with _lock:
        wb=load_workbook(MASTER_PATH); wb[sheet].append(values); wb[sheet].auto_filter.ref=wb[sheet].dimensions; wb.save(MASTER_PATH); wb.close()

def save_firm(d): _append('FIRM_MASTER',[d.get('FirmID') or 'FIRM-'+datetime.now().strftime('%Y%m%d%H%M%S'),d['FirmName'],d['FRN'],d.get('Address',''),d.get('DefaultPlace',''),d.get('Active','Yes')])
def save_partner(d): _append('PARTNER_MASTER',[d.get('PartnerID') or 'P-'+datetime.now().strftime('%Y%m%d%H%M%S'),d['FirmID'],d['PartnerName'],d['MembershipNo'],d.get('DefaultPlace',''),d.get('Active','Yes')])
def save_client(d): _append('CLIENT_MASTER',[d.get('ClientID') or 'CL-'+datetime.now().strftime('%Y%m%d%H%M%S'),d['ClientName'],d.get('CIN',''),d.get('PAN',''),d.get('City',''),d.get('Address',''),d.get('EntityType','private_ltd'),d.get('SmallCompany','No'),d.get('Active','Yes')])
def save_engagement(d): _append('ENGAGEMENT_MASTER',[d.get('EngagementID') or 'ENG-'+datetime.now().strftime('%Y%m%d%H%M%S'),d['ClientID'],d['FinancialYear'],d.get('YearEnding',''),d.get('EntityType','private_ltd'),d.get('SmallCompany','No'),d.get('Opinion','unmodified'),d.get('FirmID',''),d.get('PartnerID',''),d.get('Place',''),d.get('Status','Active'),d.get('PreviousEngagementID',''),d.get('Notes','')])

def _update(sheet,key,keyval,data,fields):
    ensure_master()
    with _lock:
        wb=load_workbook(MASTER_PATH); ws=wb[sheet]; headers=[c.value for c in ws[1]]; idx={h:i+1 for i,h in enumerate(headers)}
        for row in range(2,ws.max_row+1):
            if str(ws.cell(row,idx[key]).value)==str(keyval):
                for field in fields:
                    if field in data and field in idx: ws.cell(row,idx[field]).value=data[field]
                wb.save(MASTER_PATH); wb.close(); return True
        wb.close(); return False

def update_firm(fid,d): return _update('FIRM_MASTER','FirmID',fid,d,['FirmName','FRN','Address','DefaultPlace','Active'])
def update_partner(pid,d): return _update('PARTNER_MASTER','PartnerID',pid,d,['FirmID','PartnerName','MembershipNo','DefaultPlace','Active'])
def update_client(cid,d): return _update('CLIENT_MASTER','ClientID',cid,d,['ClientName','CIN','PAN','City','Address','EntityType','SmallCompany','Active'])
def update_engagement(eid,d): return _update('ENGAGEMENT_MASTER','EngagementID',eid,d,['FinancialYear','YearEnding','EntityType','SmallCompany','Opinion','FirmID','PartnerID','Place','Status','PreviousEngagementID','Notes'])

def list_templates(): return _rows('TEMPLATE_MASTER')
def list_variables(): return _rows('VARIABLE_LIBRARY')
def list_wordings(): return _rows('WORDING_LIBRARY')


def list_classification(engagement_id=None):
    rows=_rows('CLASSIFICATION_RESULTS')
    return [r for r in rows if not engagement_id or r.get('EngagementID')==engagement_id]

def get_classification(engagement_id):
    rows=list_classification(engagement_id)
    return rows[-1] if rows else None

def save_classification(d):
    values=[
        d.get('EngagementID',''),d.get('AssessmentBasis','Current ICAI / Companies Act'),
        d.get('AssessmentDate',''),d.get('EntityForm',''),
        d.get('Listed','No'),d.get('InProcessListing','No'),d.get('Bank','No'),
        d.get('FinancialInstitution','No'),d.get('Insurance','No'),d.get('Section8','No'),
        d.get('OPC','No'),d.get('SmallCompany','No'),d.get('PrivateCompany','No'),
        d.get('PublicCompany','No'),d.get('HoldingCompany','No'),d.get('SubsidiaryCompany','No'),d.get('HoldingSubsidiaryOfNonMSME','No'),
        d.get('SpecialActEntity','No'),d.get('PaidUpCapital',''),d.get('ReservesSurplus',''),
        d.get('TurnoverPriorYear',''),d.get('RevenueCurrentYear',''),d.get('BorrowingsMax',''),
        d.get('NetWorth',''),d.get('IndASRequired','No'),d.get('AccountingFramework',''),
        d.get('NonCompanyASCategory',''),d.get('LegacyLevel',''),d.get('CorporateSMCResult',''),
        d.get('CAROSystemResult',''),d.get('IFCSystemResult',''),
        d.get('CashFlowSystemResult',''),d.get('ScheduleIIISystemResult',''),
        d.get('ProfessionalNotes',''),datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ]
    _append('CLASSIFICATION_RESULTS',values)

def save_applicability_results(engagement_id, results):
    ensure_master()
    with _lock:
        wb=load_workbook(MASTER_PATH)
        ws=wb['APPLICABILITY_RESULTS']
        # Replace the current result set for this engagement so the sheet
        # remains a current-state register rather than an uncontrolled duplicate.
        headers=[c.value for c in ws[1]]
        idx={h:i+1 for i,h in enumerate(headers)}
        keep=[]
        for row in range(2,ws.max_row+1):
            if str(ws.cell(row,idx['EngagementID']).value)!=str(engagement_id):
                keep.append([ws.cell(row,col).value for col in range(1,len(headers)+1)])
        ws.delete_rows(2, max(ws.max_row-1,0))
        for vals in keep: ws.append(vals)
        now=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for r in results:
            ws.append([engagement_id,r.get('RuleCode',''),r.get('Requirement',''),
                       r.get('SystemResult',''),r.get('FinalResult',r.get('SystemResult','')),
                       r.get('Override','No'),r.get('Rationale',''),r.get('Source',''),
                       r.get('RuleVersion',''),now])
        ws.auto_filter.ref=ws.dimensions
        wb.save(MASTER_PATH); wb.close()

def list_applicability(engagement_id=None):
    rows=_rows('APPLICABILITY_RESULTS')
    return [r for r in rows if not engagement_id or r.get('EngagementID')==engagement_id]

import unittest
import os
import sys
import openpyxl
import io

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.db import init_db, get_connection
from src.tally_converter import convert_tally_excel
from src.import_engine import save_imported_ledgers
from src.reconciliation_engine import run_reconciliation_check
from src.mapping_engine import execute_auto_mapping

class TestTallyHierarchyParser(unittest.TestCase):

    def setUp(self):
        init_db()

    def _create_mock_tally_excel(self, rows_data):
        """Creates an in-memory openpyxl workbook formatted like Tally export."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trial Balance"

        # Headers
        ws.append(["Particulars", "Opening Balance", "Debit", "Credit", "Closing Balance"])

        for item in rows_data:
            ws.append([
                item.get('name', ''),
                item.get('opening', 0.0),
                item.get('debit', 0.0),
                item.get('credit', 0.0),
                item.get('closing_str', '')
            ])
            cur_row = ws.max_row
            if item.get('bold', False):
                ws.cell(row=cur_row, column=1).font = openpyxl.styles.Font(bold=True)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    def test_summary_trial_balance_parsing(self):
        """Tests parsing of a summary trial balance with top-level groups and standalone ledgers."""
        rows = [
            {'name': 'Capital Account', 'bold': True, 'credit': 1000000.0, 'closing_str': '1000000.00 Cr'},
            {'name': 'Share Capital', 'bold': False, 'credit': 1000000.0, 'closing_str': '1000000.00 Cr'},
            {'name': 'Current Liabilities', 'bold': True, 'debit': 25520.0, 'credit': 716000.0, 'closing_str': '690480.00 Cr'},
            {'name': 'Duties & Taxes', 'bold': False, 'debit': 25520.0, 'credit': 207500.0, 'closing_str': '181980.00 Cr'},
            {'name': 'CGST', 'bold': False, 'debit': 10260.0, 'credit': 49500.0, 'closing_str': '39240.00 Cr'},
            {'name': 'IGST', 'bold': False, 'credit': 103500.0, 'closing_str': '103500.00 Cr'},
            {'name': 'SGST', 'bold': False, 'debit': 10260.0, 'credit': 49500.0, 'closing_str': '39240.00 Cr'},
            {'name': 'TDS Payable (Professionals)', 'bold': False, 'debit': 5000.0, 'credit': 5000.0, 'closing_str': '0.00'},
            {'name': 'Sundry Creditors', 'bold': False, 'credit': 278500.0, 'closing_str': '278500.00 Cr'},
            {'name': 'Advance from Customers', 'bold': False, 'credit': 50000.0, 'closing_str': '50000.00 Cr'},
            {'name': 'Salary Payable', 'bold': False, 'credit': 180000.0, 'closing_str': '180000.00 Cr'},
            {'name': 'Fixed Assets', 'bold': True, 'debit': 195000.0, 'credit': 30000.0, 'closing_str': '165000.00 Dr'},
            {'name': 'Computers & Software', 'bold': False, 'debit': 150000.0, 'credit': 30000.0, 'closing_str': '120000.00 Dr'},
            {'name': 'Furniture & Fittings', 'bold': False, 'debit': 45000.0, 'closing_str': '45000.00 Dr'},
            {'name': 'Current Assets', 'bold': True, 'debit': 3040500.0, 'credit': 753620.0, 'closing_str': '2286880.00 Dr'},
            {'name': 'Sundry Debtors', 'bold': False, 'debit': 1327500.0, 'credit': 569000.0, 'closing_str': '758500.00 Dr'},
            {'name': 'Cash-in-Hand', 'bold': False, 'debit': 50000.0, 'credit': 34600.0, 'closing_str': '15400.00 Dr'},
            {'name': 'Petty Cash', 'bold': False, 'debit': 50000.0, 'credit': 34600.0, 'closing_str': '15400.00 Dr'},
            {'name': 'Bank Accounts', 'bold': False, 'debit': 1604000.0, 'credit': 150020.0, 'closing_str': '1453980.00 Dr'},
            {'name': 'Axis Bank', 'bold': False, 'debit': 1604000.0, 'credit': 150020.0, 'closing_str': '1453980.00 Dr'},
            {'name': 'Accrued Income', 'bold': False, 'debit': 35000.0, 'closing_str': '35000.00 Dr'},
            {'name': 'Prepaid Expenses', 'bold': False, 'debit': 24000.0, 'closing_str': '24000.00 Dr'},
            {'name': 'Sales Accounts', 'bold': True, 'credit': 1160000.0, 'closing_str': '1160000.00 Cr'},
            {'name': 'Consulting Revenue', 'bold': False, 'credit': 585000.0, 'closing_str': '585000.00 Cr'},
            {'name': 'Software Dev Revenue', 'bold': False, 'credit': 575000.0, 'closing_str': '575000.00 Cr'},
            {'name': 'Indirect Expenses', 'bold': True, 'debit': 398600.0, 'closing_str': '398600.00 Dr'},
            {'name': 'Salary Expense', 'bold': False, 'debit': 180000.0, 'closing_str': '180000.00 Dr'},
            {'name': 'Rent Expense', 'bold': False, 'debit': 20000.0, 'closing_str': '20000.00 Dr'},
            {'name': 'Professional Fees', 'bold': False, 'debit': 50000.0, 'closing_str': '50000.00 Dr'},
            {'name': 'Bad Debts', 'bold': False, 'debit': 15000.0, 'closing_str': '15000.00 Dr'},
            {'name': 'Depreciation Expense', 'bold': False, 'debit': 30000.0, 'closing_str': '30000.00 Dr'},
            {'name': 'Marketing & Advertising', 'bold': False, 'debit': 25000.0, 'closing_str': '25000.00 Dr'},
            {'name': 'Office Maintenance', 'bold': False, 'debit': 25000.0, 'closing_str': '25000.00 Dr'},
            {'name': 'Internet Expense', 'bold': False, 'debit': 18000.0, 'closing_str': '18000.00 Dr'},
            {'name': 'Travelling & Conveyance', 'bold': False, 'debit': 25600.0, 'closing_str': '25600.00 Dr'},
            {'name': 'Staff Welfare', 'bold': False, 'debit': 9000.0, 'closing_str': '9000.00 Dr'},
            {'name': 'Bank Charges', 'bold': False, 'debit': 1000.0, 'closing_str': '1000.00 Dr'},
        ]

        excel_bytes = self._create_mock_tally_excel(rows)
        ledgers, stats, errors = convert_tally_excel(excel_bytes)

        self.assertEqual(len(errors), 0, f"Conversion should have 0 errors, got {errors}")
        self.assertEqual(len(ledgers), 28, f"Expected 28 leaf ledgers, got {len(ledgers)}")

        tot_dr = sum(l['closing_dr'] for l in ledgers)
        tot_cr = sum(l['closing_cr'] for l in ledgers)
        self.assertAlmostEqual(tot_dr, 2850480.0, delta=1.0)
        self.assertAlmostEqual(tot_cr, 2850480.0, delta=1.0)
        self.assertAlmostEqual(tot_dr, tot_cr, delta=1.0, msg="Total Dr must equal Total Cr")

    def test_party_level_trial_balance_parsing(self):
        """Tests parsing of a detailed trial balance with individual customer and vendor party ledgers."""
        rows = [
            {'name': 'Capital Account', 'bold': True, 'credit': 1000000.0, 'closing_str': '1000000.00 Cr'},
            {'name': 'Share Capital', 'bold': False, 'credit': 1000000.0, 'closing_str': '1000000.00 Cr'},
            {'name': 'Current Liabilities', 'bold': True, 'debit': 25520.0, 'credit': 747500.0, 'closing_str': '721980.00 Cr'},
            {'name': 'Duties & Taxes', 'bold': False, 'debit': 25520.0, 'credit': 228200.0, 'closing_str': '202680.00 Cr'},
            {'name': 'CGST', 'bold': False, 'debit': 10260.0, 'credit': 54900.0, 'closing_str': '44640.00 Cr'},
            {'name': 'IGST', 'bold': False, 'credit': 113400.0, 'closing_str': '113400.00 Cr'},
            {'name': 'SGST', 'bold': False, 'debit': 10260.0, 'credit': 54900.0, 'closing_str': '44640.00 Cr'},
            {'name': 'TDS Payable (Professionals)', 'bold': False, 'debit': 5000.0, 'credit': 5000.0, 'closing_str': '0.00'},
            {'name': 'Sundry Creditors', 'bold': False, 'credit': 278500.0, 'closing_str': '278500.00 Cr'},
            {'name': 'Hardware Hub', 'bold': False, 'credit': 150000.0, 'closing_str': '150000.00 Cr'},
            {'name': 'LegalMind Associates', 'bold': False, 'credit': 54000.0, 'closing_str': '54000.00 Cr'},
            {'name': 'Office Needs Supplier', 'bold': False, 'credit': 74500.0, 'closing_str': '74500.00 Cr'},
            {'name': 'Advance from Customers', 'bold': False, 'credit': 50000.0, 'closing_str': '50000.00 Cr'},
            {'name': 'Salary Payable', 'bold': False, 'credit': 180000.0, 'closing_str': '180000.00 Cr'},
            {'name': 'Fixed Assets', 'bold': True, 'debit': 195000.0, 'credit': 30000.0, 'closing_str': '165000.00 Dr'},
            {'name': 'Computers & Software', 'bold': False, 'debit': 150000.0, 'credit': 30000.0, 'closing_str': '120000.00 Dr'},
            {'name': 'Furniture & Fittings', 'bold': False, 'debit': 45000.0, 'closing_str': '45000.00 Dr'},
            {'name': 'Current Assets', 'bold': True, 'debit': 3186200.0, 'credit': 765420.0, 'closing_str': '2420780.00 Dr'},
            {'name': 'Sundry Debtors', 'bold': False, 'debit': 1463200.0, 'credit': 569000.0, 'closing_str': '894200.00 Dr'},
            {'name': 'Global Tech Solutions', 'bold': False, 'debit': 566400.0, 'credit': 177000.0, 'closing_str': '389400.00 Dr'},
            {'name': 'Stark Industries (Out of State)', 'bold': False, 'debit': 177000.0, 'closing_str': '177000.00 Dr'},
            {'name': 'TechNova Solutions', 'bold': False, 'debit': 719800.0, 'credit': 392000.0, 'closing_str': '327800.00 Dr'},
            {'name': 'Cash-in-Hand', 'bold': False, 'debit': 60000.0, 'credit': 36400.0, 'closing_str': '23600.00 Dr'},
            {'name': 'Petty Cash', 'bold': False, 'debit': 60000.0, 'credit': 36400.0, 'closing_str': '23600.00 Dr'},
            {'name': 'Bank Accounts', 'bold': False, 'debit': 1604000.0, 'credit': 160020.0, 'closing_str': '1443980.00 Dr'},
            {'name': 'Axis Bank', 'bold': False, 'debit': 1604000.0, 'credit': 160020.0, 'closing_str': '1443980.00 Dr'},
            {'name': 'Accrued Income', 'bold': False, 'debit': 35000.0, 'closing_str': '35000.00 Dr'},
            {'name': 'Prepaid Expenses', 'bold': False, 'debit': 24000.0, 'closing_str': '24000.00 Dr'},
            {'name': 'Sales Accounts', 'bold': True, 'credit': 1275000.0, 'closing_str': '1275000.00 Cr'},
            {'name': 'Consulting Revenue', 'bold': False, 'credit': 645000.0, 'closing_str': '645000.00 Cr'},
            {'name': 'Software Dev Revenue', 'bold': False, 'credit': 630000.0, 'closing_str': '630000.00 Cr'},
            {'name': 'Indirect Expenses', 'bold': True, 'debit': 400400.0, 'closing_str': '400400.00 Dr'},
            {'name': 'Salary Expense', 'bold': False, 'debit': 180000.0, 'closing_str': '180000.00 Dr'},
            {'name': 'Rent Expense', 'bold': False, 'debit': 20000.0, 'closing_str': '20000.00 Dr'},
            {'name': 'Professional Fees', 'bold': False, 'debit': 50000.0, 'closing_str': '50000.00 Dr'},
            {'name': 'Bad Debts', 'bold': False, 'debit': 15000.0, 'closing_str': '15000.00 Dr'},
            {'name': 'Depreciation Expense', 'bold': False, 'debit': 30000.0, 'closing_str': '30000.00 Dr'},
            {'name': 'Marketing & Advertising', 'bold': False, 'debit': 25000.0, 'closing_str': '25000.00 Dr'},
            {'name': 'Office Maintenance', 'bold': False, 'debit': 25000.0, 'closing_str': '25000.00 Dr'},
            {'name': 'Internet Expense', 'bold': False, 'debit': 18000.0, 'closing_str': '18000.00 Dr'},
            {'name': 'Travelling & Conveyance', 'bold': False, 'debit': 27400.0, 'closing_str': '27400.00 Dr'},
            {'name': 'Staff Welfare', 'bold': False, 'debit': 9000.0, 'closing_str': '9000.00 Dr'},
            {'name': 'Bank Charges', 'bold': False, 'debit': 1000.0, 'closing_str': '1000.00 Dr'},
        ]

        excel_bytes = self._create_mock_tally_excel(rows)
        ledgers, stats, errors = convert_tally_excel(excel_bytes)

        self.assertEqual(len(errors), 0, f"Conversion should have 0 errors, got {errors}")
        self.assertEqual(len(ledgers), 32, f"Expected 32 leaf ledgers, got {len(ledgers)}")

        tot_dr = sum(l['closing_dr'] for l in ledgers)
        tot_cr = sum(l['closing_cr'] for l in ledgers)
        self.assertAlmostEqual(tot_dr, 2986180.0, delta=1.0)
        self.assertAlmostEqual(tot_cr, 2986180.0, delta=1.0)
        self.assertAlmostEqual(tot_dr, tot_cr, delta=1.0, msg="Total Dr must equal Total Cr")

        # Verify DB save and reconciliation
        save_imported_ledgers(ledgers, source_filename="test_party_tb.xlsx")
        execute_auto_mapping()
        recon = run_reconciliation_check()
        self.assertTrue(recon['tb_stats']['is_balanced'], "Reconciliation engine should report balanced TB")

if __name__ == '__main__':
    unittest.main()

"""
AIS Broker Reconciliation - upgraded launcher

Keep this file in the same folder as ais_broker_reco_final.py and run:

    python ais_broker_reco_modified.py

The original file is imported and upgraded at runtime. It is not overwritten.
"""

import os
import re
import datetime
import traceback
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import ais_broker_reco_final as base


# ---------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------
base.QUANTITY_TOLERANCE = 0.0001
base.QUANTITY_RELATIVE_TOLERANCE = 0.000001
base.TRANSACTION_COST_PERCENT_LIMIT = 0.0075
base.TRANSACTION_COST_MAX_LIMIT = 2500.0


def quantity_matches(a_qty, b_qty):
    a_qty = base.clean_number(a_qty)
    b_qty = base.clean_number(b_qty)
    scale = max(abs(a_qty), abs(b_qty), 1.0)
    tolerance = max(
        base.QUANTITY_TOLERANCE,
        scale * base.QUANTITY_RELATIVE_TOLERANCE
    )
    return abs(a_qty - b_qty) <= tolerance


def sale_cost_tolerance(ais_value, broker_value):
    """
    Dynamic limit for brokerage, exchange charges, GST, stamp duty and
    similar presentation differences.
    """
    value = max(
        abs(base.clean_number(ais_value)),
        abs(base.clean_number(broker_value)),
        1.0
    )
    return min(
        max(
            base.SALE_CONSIDERATION_IGNORE_LIMIT,
            value * base.TRANSACTION_COST_PERCENT_LIMIT
        ),
        base.TRANSACTION_COST_MAX_LIMIT
    )


def nominal_sale_variance(ais_txn, broker_txn):
    return (
        quantity_matches(ais_txn.quantity, broker_txn.quantity)
        and abs(
            ais_txn.sale_consideration - broker_txn.sale_consideration
        ) <= sale_cost_tolerance(
            ais_txn.sale_consideration,
            broker_txn.sale_consideration
        )
    )


def security_matches(ais_txn, broker_txn, soft=False):
    if (
        ais_txn.isin and broker_txn.isin
        and ais_txn.isin == broker_txn.isin
    ):
        return True

    ais_name = re.sub(
        r"[^A-Z0-9]", "",
        base.clean_security_name(ais_txn.security_name)
    )
    broker_name = re.sub(
        r"[^A-Z0-9]", "",
        base.clean_security_name(broker_txn.security_name)
    )
    if not ais_name or not broker_name:
        return False
    if ais_name == broker_name:
        return True
    return soft and (
        (len(ais_name) >= 6 and ais_name in broker_name)
        or (len(broker_name) >= 6 and broker_name in ais_name)
    )


# ---------------------------------------------------------------------
# TAX RULE ENGINE
# ---------------------------------------------------------------------
class TaxRuleEngine:
    """
    Holding-period classification for securities and mutual funds.

    STT is supporting evidence of a listed/exchange-traded equity
    transaction. STT does not by itself alter the statutory holding period.
    """

    LAW_CHANGE_DATE = datetime.date(2024, 7, 23)
    SECTION_50AA_DATE = datetime.date(2023, 4, 1)

    def __init__(self, financial_year):
        self.financial_year = financial_year

    @staticmethod
    def contains(text, words):
        return any(word in text for word in words)

    @staticmethod
    def add_months(value, months):
        month_index = value.month - 1 + months
        year = value.year + month_index // 12
        month = month_index % 12 + 1
        leap = year % 400 == 0 or (
            year % 4 == 0 and year % 100 != 0
        )
        month_days = [
            31, 29 if leap else 28, 31, 30, 31, 30,
            31, 31, 30, 31, 30, 31
        ]
        return datetime.date(
            year, month, min(value.day, month_days[month - 1])
        )

    def held_more_than_months(self, purchase_date, sale_date, months):
        return sale_date > self.add_months(purchase_date, months)

    def classify_security_type(
        self, isin, security_name, declared_asset_type=""
    ):
        isin = base.clean_isin_code(isin)
        name = base.clean_security_name(security_name)
        declared = base.clean_security_name(declared_asset_type)
        text = f"{declared} {name}"

        if self.contains(text, [
            "MARKET LINKED DEBENTURE",
            "MARKET-LINKED DEBENTURE",
            " MLD"
        ]):
            return "MARKET_LINKED_DEBENTURE"

        if self.contains(text, ["SOVEREIGN GOLD BOND", " SGB"]):
            return "SGB"

        if self.contains(text, [
            "PREFERENCE SHARE", "PREFERENCE SHARES", "PREF SHARE"
        ]):
            return "PREFERENCE_SHARE"

        if self.contains(text, [
            "BOND", "NCD", "DEBENTURE", "GOVERNMENT SECURITY",
            "G-SEC", "TREASURY BILL", "T-BILL"
        ]):
            if self.contains(text, ["UNLISTED", "PRIVATE PLACEMENT"]):
                return "UNLISTED_BOND"
            if self.contains(text, ["LISTED", "EXCHANGE TRADED"]):
                return "LISTED_BOND"
            return (
                "LISTED_BOND"
                if isin.startswith(("INE", "IN9"))
                else "UNLISTED_BOND"
            )

        is_mf = (
            isin.startswith("INF")
            or self.contains(text, [
                "MUTUAL FUND", " FUND", "SCHEME", "ETF", "ELSS"
            ])
        )
        if is_mf:
            if self.contains(text, [
                "DEBT", "LIQUID", "TREASURY", "GILT", "OVERNIGHT",
                "ULTRA SHORT", "LOW DURATION", "SHORT DURATION",
                "MEDIUM DURATION", "LONG DURATION", "DYNAMIC BOND",
                "CORPORATE BOND", "CREDIT RISK", "MONEY MARKET",
                "BANKING AND PSU", "FIXED MATURITY", "FMP",
                "INCOME FUND", "SAVINGS FUND", "FLOATER"
            ]):
                return "DEBT_MF"
            if self.contains(text, [
                "EQUITY", "ELSS", "ARBITRAGE", "INDEX FUND",
                "INDEX ETF", "EQUITY ETF", "LARGE CAP", "MID CAP",
                "SMALL CAP", "FLEXI CAP", "MULTI CAP", "FOCUSED",
                "VALUE FUND", "CONTRA", "SECTORAL", "THEMATIC"
            ]):
                return "EQUITY_MF"
            return "OTHER_MF"

        if self.contains(text, [
            "UNLISTED EQUITY", "UNLISTED SHARE", "PRIVATE LIMITED"
        ]):
            return "UNLISTED_EQUITY"

        return "EQUITY_SHARE"

    def probably_listed(
        self, asset_type, isin, security_name, stt_paid,
        declared_asset_type=""
    ):
        text = (
            base.clean_security_name(declared_asset_type) + " "
            + base.clean_security_name(security_name)
        )
        if self.contains(text, ["UNLISTED", "PRIVATE LIMITED"]):
            return False
        if self.contains(text, ["LISTED", "EXCHANGE TRADED", "ETF"]):
            return True
        if base.clean_number(stt_paid) > 0:
            return True
        if asset_type in ("EQUITY_MF", "LISTED_BOND"):
            return True
        if asset_type in ("UNLISTED_BOND", "UNLISTED_EQUITY"):
            return False
        return base.clean_isin_code(isin).startswith(
            ("INE", "IN9", "INF")
        )

    def classify_transaction(
        self, isin, security_name, purchase_date, sale_date,
        stt_paid=None, declared_asset_type=""
    ):
        purchase_date = base.clean_date(purchase_date)
        sale_date = base.clean_date(sale_date)
        if not purchase_date or not sale_date:
            return "Review Required", None, (
                "Missing or invalid purchase/sale date"
            )

        holding_days = (sale_date - purchase_date).days
        if holding_days < 0:
            return "Review Required", holding_days, (
                "Sale date is before purchase date"
            )

        asset_type = self.classify_security_type(
            isin, security_name, declared_asset_type
        )
        listed = self.probably_listed(
            asset_type, isin, security_name, stt_paid,
            declared_asset_type
        )
        evidence = (
            f"Asset={asset_type}; "
            f"probably listed={'Yes' if listed else 'No'}; "
            f"STT paid={'Yes' if base.clean_number(stt_paid) > 0 else 'No'}"
        )

        if asset_type == "MARKET_LINKED_DEBENTURE":
            return "STCG", holding_days, (
                f"{evidence}; deemed STCG under Section 50AA"
            )

        if (
            asset_type == "UNLISTED_BOND"
            and sale_date >= self.LAW_CHANGE_DATE
        ):
            return "STCG", holding_days, (
                f"{evidence}; unlisted bond/debenture transferred on or "
                f"after 23-07-2024 is deemed STCG under Section 50AA"
            )

        if (
            asset_type == "DEBT_MF"
            and purchase_date >= self.SECTION_50AA_DATE
        ):
            return "STCG", holding_days, (
                f"{evidence}; specified/debt mutual fund acquired on or "
                f"after 01-04-2023 is treated as STCG under Section 50AA"
            )

        if asset_type == "EQUITY_SHARE":
            months = 12 if listed else 24
        elif asset_type == "UNLISTED_EQUITY":
            months = 24
        elif asset_type == "EQUITY_MF":
            months = 12
        elif asset_type == "PREFERENCE_SHARE":
            months = 12 if listed else 24
        elif asset_type == "LISTED_BOND":
            months = 12
        elif asset_type == "UNLISTED_BOND":
            months = 36
        elif asset_type == "SGB":
            months = 12 if listed else (
                36 if sale_date < self.LAW_CHANGE_DATE else 24
            )
        elif asset_type in ("DEBT_MF", "OTHER_MF"):
            if sale_date < self.LAW_CHANGE_DATE:
                months = 36
            else:
                months = 12 if listed else 24
        else:
            if sale_date < self.LAW_CHANGE_DATE:
                months = 12 if listed else 36
            else:
                months = 12 if listed else 24

        is_long_term = self.held_more_than_months(
            purchase_date, sale_date, months
        )
        classification = "LTCG" if is_long_term else "STCG"
        relation = "more than" if is_long_term else "not more than"
        return classification, holding_days, (
            f"{evidence}; held {relation} {months} months "
            f"({holding_days} days)"
        )


base.TaxRuleEngine = TaxRuleEngine


# ---------------------------------------------------------------------
# NORMALISATION WRAPPER
# ---------------------------------------------------------------------
_original_loader = base.load_and_normalize_workbook


def _find_original_value(txn, possible_names):
    normalized = {
        base.normalize_header_text(key): value
        for key, value in txn.original_row_data.items()
    }
    for key, value in normalized.items():
        if any(name in key for name in possible_names):
            return value
    return ""


def upgraded_loader(file_path, file_label, log_callback):
    transactions = _original_loader(
        file_path, file_label, log_callback
    )
    for txn in transactions:
        txn.declared_asset_type = base.clean_security_name(
            _find_original_value(txn, [
                "type of asset", "asset type", "security type",
                "instrument type", "nature of asset", "asset class"
            ])
        )
        txn.calculated_asset_type = ""
        txn.is_ltcg = "Review Required"

        original_ltcg = str(_find_original_value(txn, [
            "is it ltcg", "is ltcg", "ltcg yes no"
        ])).strip().upper()
        if original_ltcg in ("YES", "Y", "TRUE", "1"):
            txn.original_classification = "LTCG"
        elif original_ltcg in ("NO", "N", "FALSE", "0"):
            txn.original_classification = "STCG"
    return transactions


base.load_and_normalize_workbook = upgraded_loader


# ---------------------------------------------------------------------
# MATCHING ENGINE
# ---------------------------------------------------------------------
def match_and_reconcile(
    ais_txns, broker_txns, tax_engine, log_callback
):
    log_callback("Starting quantity-led matching engine...")
    remaining_broker = list(broker_txns)
    unmatched_ais = [txn for txn in ais_txns if not txn.is_duplicate]
    matched_pairs = []
    match_counter = 1

    passes = [
        ("Exact security/date/quantity", False, False, True),
        ("Exact security/date-window/quantity", False, True, True),
        ("Description/date/quantity", True, False, True),
        ("Description/date-window/quantity", True, True, True),
        # Keeps genuine value differences together for difference reporting.
        ("Quantity matched - value difference", True, True, False),
    ]

    for label, soft_name, date_window, require_nominal in passes:
        log_callback(f"Matching pass: {label}")
        current = unmatched_ais
        unmatched_ais = []

        for ais in current:
            candidates = []
            for index, broker in enumerate(remaining_broker):
                if not security_matches(ais, broker, soft_name):
                    continue
                if not quantity_matches(ais.quantity, broker.quantity):
                    continue
                if not ais.sale_date or not broker.sale_date:
                    continue

                days = abs((ais.sale_date - broker.sale_date).days)
                if date_window:
                    if days > base.DATE_WINDOW_DAYS:
                        continue
                elif days != 0:
                    continue

                if require_nominal and not nominal_sale_variance(
                    ais, broker
                ):
                    continue

                value_difference = abs(
                    ais.sale_consideration
                    - broker.sale_consideration
                )
                candidates.append(
                    (days, value_difference, index)
                )

            if not candidates:
                unmatched_ais.append(ais)
                continue

            _, value_difference, found_index = min(candidates)
            broker = remaining_broker.pop(found_index)
            ais.match_id = f"M-{match_counter:04d}"
            broker.match_id = ais.match_id

            if (
                require_nominal
                and value_difference > base.TOLERANCE_AMOUNT
            ):
                limit = sale_cost_tolerance(
                    ais.sale_consideration,
                    broker.sale_consideration
                )
                mode = (
                    f"{label}; nominal cost variance "
                    f"Rs {value_difference:.2f} <= Rs {limit:.2f}"
                )
            else:
                mode = label

            matched_pairs.append((ais, broker, mode))
            match_counter += 1

    log_callback(
        f"Matched={len(matched_pairs)}, "
        f"unmatched AIS={len(unmatched_ais)}, "
        f"unmatched Broker={len(remaining_broker)}"
    )
    return matched_pairs, unmatched_ais, remaining_broker


base.match_and_reconcile = match_and_reconcile


# ---------------------------------------------------------------------
# EXCEL NAVIGATION AND TAX SOFTWARE OUTPUT
# ---------------------------------------------------------------------
_original_style_header = base.RecoReportGenerator.style_header
_original_summary = base.RecoReportGenerator.write_summary_sheet


def style_header_with_home(self, ws, columns):
    _original_style_header(self, ws, columns)
    if ws.title != "Summary":
        cell = ws.cell(row=1, column=len(columns) + 2)
        cell.value = "Back to Summary"
        cell.hyperlink = "#'Summary'!A1"
        cell.font = Font(
            name="Segoe UI", size=10, bold=True,
            color="FFFFFF", underline="single"
        )
        cell.fill = PatternFill(
            start_color=self.COLOR_INDIGO_HEADER,
            end_color=self.COLOR_INDIGO_HEADER,
            fill_type="solid"
        )
        cell.alignment = self.ALIGN_CENTER


def summary_with_links(self, summary_data):
    _original_summary(self, summary_data)
    ws = self.wb["Summary"]
    start_row = ws.max_row + 3
    ws.cell(
        row=start_row, column=1,
        value="DIRECT SHEET NAVIGATION"
    ).font = Font(
        name="Segoe UI", size=12, bold=True, color="1E3A8A"
    )

    sheets = [
        "Tax Software Output",
        "Difference in Long Term",
        "Difference in Short Term",
        "Other Differences",
        "Ignored Sales Variances",
        "Duplicate Entries in AIS",
        "Missing Entries in AIS",
        "Missing Entries in Broker",
        "Matched Transactions",
        "Review Required",
    ]
    for offset, sheet_name in enumerate(sheets, 1):
        cell = ws.cell(
            row=start_row + offset,
            column=1,
            value=f"Open {sheet_name}"
        )
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.style = "Hyperlink"
    ws.column_dimensions["A"].width = 48


def write_tax_software_output(self, transactions):
    ws = self.wb.create_sheet(title="Tax Software Output")
    headers = [
        "Source Row", "Source Sheet", "Security", "ISIN",
        "Type of Asset", "Sale Date", "Purchase Date",
        "Holding Days", "Quantity", "Sale Consideration",
        "Purchase Cost", "STT", "STT Paid?",
        "Original Classification", "Calculated Classification",
        "Is it LTCG?", "Tax Rule Remarks"
    ]
    self.style_header(ws, headers)

    row = 2
    for txn in transactions:
        ws.append([
            txn.source_row, txn.source_sheet, txn.security_name,
            txn.isin, txn.calculated_asset_type, txn.sale_date,
            txn.purchase_date, txn.holding_days, txn.quantity,
            txn.sale_consideration, txn.purchase_cost, txn.stt,
            "Yes" if txn.stt > 0 else "No",
            txn.original_classification,
            txn.calculated_classification,
            txn.is_ltcg,
            txn.remarks
        ])
        ws.cell(row=row, column=6).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=7).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=9).number_format = "#,##0.0000"
        for column in (10, 11, 12):
            ws.cell(row=row, column=column).number_format = "#,##0.00"

        if txn.calculated_classification == "LTCG":
            fill_color = self.COLOR_GREEN_MATCH
        elif txn.calculated_classification == "Review Required":
            fill_color = self.COLOR_YELLOW_WARN
        else:
            fill_color = None

        if fill_color:
            fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type="solid"
            )
            ws.cell(row=row, column=15).fill = fill
            ws.cell(row=row, column=16).fill = fill
        row += 1

    if row > 2:
        self.format_rows(ws, 2, row - 1, len(headers))
    self.autofit_columns(ws)


base.RecoReportGenerator.style_header = style_header_with_home
base.RecoReportGenerator.write_summary_sheet = summary_with_links
base.RecoReportGenerator.write_tax_software_output = (
    write_tax_software_output
)


# ---------------------------------------------------------------------
# UPDATED WORKFLOW
# ---------------------------------------------------------------------
def apply_tax_rules(transactions, engine):
    for txn in transactions:
        txn.calculated_asset_type = engine.classify_security_type(
            txn.isin, txn.security_name, txn.declared_asset_type
        )
        result, holding_days, remarks = engine.classify_transaction(
            txn.isin,
            txn.security_name,
            txn.purchase_date,
            txn.sale_date,
            stt_paid=txn.stt,
            declared_asset_type=txn.declared_asset_type
        )
        txn.calculated_classification = result
        txn.holding_days = holding_days
        txn.is_ltcg = (
            "Yes" if result == "LTCG"
            else "No" if result == "STCG"
            else "Review Required"
        )
        txn.remarks = remarks


def run_reconciliation_workflow(
    ais_file, broker_file, output_path, fy,
    progress_callback, log_callback
):
    try:
        progress_callback(5, "Initialising reconciliation...")
        engine = TaxRuleEngine(fy)

        progress_callback(10, "Loading AIS workbook...")
        ais_rows = base.load_and_normalize_workbook(
            ais_file, "AIS Data File", log_callback
        )
        progress_callback(25, "Loading Broker workbook...")
        broker_rows = base.load_and_normalize_workbook(
            broker_file, "Broker Data File", log_callback
        )

        if not ais_rows or not broker_rows:
            raise ValueError(
                "Both files must contain valid transaction rows."
            )

        progress_callback(32, "Applying tax holding-period rules...")
        apply_tax_rules(ais_rows, engine)
        apply_tax_rules(broker_rows, engine)

        progress_callback(40, "Checking AIS duplicates...")
        duplicates = base.detect_ais_duplicates(
            ais_rows, broker_rows, log_callback
        )

        progress_callback(
            55, "Matching scrip, quantity and consideration..."
        )
        matched, unmatched_ais, unmatched_broker = (
            match_and_reconcile(
                ais_rows, broker_rows, engine, log_callback
            )
        )

        ltcg_differences = []
        stcg_differences = []
        other_differences = []
        ignored_variances = []
        review_rows = []

        for txn in unmatched_ais + unmatched_broker:
            if (
                txn.calculated_classification == "Review Required"
                or txn.quantity <= 0
                or txn.sale_consideration <= 0
            ):
                review_rows.append(txn)

        progress_callback(70, "Analysing genuine differences...")
        for ais, broker, match_mode in matched:
            purchase_date = broker.purchase_date or ais.purchase_date
            declared_type = (
                broker.declared_asset_type
                or ais.declared_asset_type
            )
            result, holding_days, tax_note = (
                engine.classify_transaction(
                    ais.isin or broker.isin,
                    ais.security_name or broker.security_name,
                    purchase_date,
                    ais.sale_date or broker.sale_date,
                    stt_paid=max(ais.stt, broker.stt),
                    declared_asset_type=declared_type
                )
            )
            asset_type = engine.classify_security_type(
                ais.isin or broker.isin,
                ais.security_name or broker.security_name,
                declared_type
            )

            for txn in (ais, broker):
                txn.calculated_asset_type = asset_type
                txn.calculated_classification = result
                txn.holding_days = holding_days
                txn.is_ltcg = (
                    "Yes" if result == "LTCG"
                    else "No" if result == "STCG"
                    else "Review Required"
                )

            ais_class = (
                ais.original_classification
                if ais.original_classification in ("LTCG", "STCG")
                else None
            )
            broker_class = (
                broker.original_classification
                if broker.original_classification
                in ("LTCG", "STCG")
                else None
            )

            sale_diff = abs(
                ais.sale_consideration - broker.sale_consideration
            )
            cost_diff = abs(
                ais.purchase_cost - broker.purchase_cost
            )
            stt_diff = abs(ais.stt - broker.stt)
            allowed_sale_diff = sale_cost_tolerance(
                ais.sale_consideration,
                broker.sale_consideration
            )

            notes = [tax_note, f"Match mode: {match_mode}."]
            class_mismatch = False
            amount_mismatch = False

            if result == "Review Required":
                if ais not in review_rows:
                    review_rows.append(ais)
            else:
                if ais_class and ais_class != result:
                    class_mismatch = True
                    notes.append(
                        f"AIS class {ais_class} corrected to {result}."
                    )
                if broker_class and broker_class != result:
                    class_mismatch = True
                    notes.append(
                        f"Broker/tax-software class {broker_class} "
                        f"corrected to {result}."
                    )

            if not quantity_matches(
                ais.quantity, broker.quantity
            ):
                amount_mismatch = True
                notes.append(
                    f"Quantity mismatch: AIS={ais.quantity:.4f}, "
                    f"Broker={broker.quantity:.4f}."
                )

            if sale_diff > allowed_sale_diff:
                amount_mismatch = True
                notes.append(
                    f"Sale consideration mismatch: "
                    f"Diff={sale_diff:.2f}; "
                    f"nominal limit={allowed_sale_diff:.2f}."
                )
            elif sale_diff > base.TOLERANCE_AMOUNT:
                notes.append(
                    f"Ignored nominal brokerage/transaction-cost "
                    f"variance: Diff={sale_diff:.2f}; "
                    f"allowed={allowed_sale_diff:.2f}."
                )

            if cost_diff > base.TOLERANCE_AMOUNT:
                amount_mismatch = True
                notes.append(
                    f"Acquisition cost mismatch: Diff={cost_diff:.2f}."
                )

            if stt_diff > base.TOLERANCE_AMOUNT:
                notes.append(
                    f"STT/charge presentation variance ignored: "
                    f"Diff={stt_diff:.2f}."
                )

            remarks = " ".join(notes)
            ais.remarks = remarks
            broker.remarks = remarks

            if class_mismatch:
                target = (
                    ltcg_differences
                    if result == "LTCG"
                    else stcg_differences
                )
                target.append((
                    ais, broker,
                    "Classification Mismatch", remarks
                ))
            elif amount_mismatch:
                other_differences.append((
                    ais, broker,
                    "Quantity/Amount Mismatch", remarks
                ))
            elif sale_diff > base.TOLERANCE_AMOUNT:
                ignored_variances.append((
                    ais, broker,
                    "Ignored Transaction-Cost Variance", remarks
                ))

        summary = {
            "total_ais": len(ais_rows),
            "total_broker": len(broker_rows),
            "matched": len(matched),
            "duplicates": len(duplicates),
            "missing_in_ais": len(unmatched_broker),
            "missing_in_broker": len(unmatched_ais),
            "ltcg_diffs": len(ltcg_differences),
            "stcg_diffs": len(stcg_differences),
            "sales_mismatch": sum(
                1 for ais, broker, _, _ in other_differences
                if abs(
                    ais.sale_consideration
                    - broker.sale_consideration
                ) > sale_cost_tolerance(
                    ais.sale_consideration,
                    broker.sale_consideration
                )
            ),
            "ignored_sales_variances": len(ignored_variances),
            "other_diffs": len(other_differences),
            "review_required": len(review_rows),
        }

        progress_callback(85, "Writing Excel report...")
        report = base.RecoReportGenerator(
            output_path, fy, ais_file, broker_file
        )
        report.write_summary_sheet(summary)
        report.write_tax_software_output(broker_rows)
        report.write_difference_sheet(
            "Difference in Long Term", ltcg_differences
        )
        report.write_difference_sheet(
            "Difference in Short Term", stcg_differences
        )
        report.write_duplicates_sheet(duplicates)
        report.write_missing_sheet(
            "Missing Entries in AIS", unmatched_broker
        )
        report.write_missing_sheet(
            "Missing Entries in Broker", unmatched_ais
        )
        report.write_difference_sheet(
            "Other Differences", other_differences
        )
        report.write_difference_sheet(
            "Ignored Sales Variances", ignored_variances
        )
        report.write_matched_sheet(matched)
        report.write_review_sheet(review_rows)
        report.save()

        log_path = os.path.splitext(output_path)[0] + (
            "_TECHNICAL_LOG.txt"
        )
        with open(log_path, "w", encoding="utf-8") as log_file:
            log_file.write(
                "AIS BROKER RECO - UPGRADED TECHNICAL LOG\n"
            )
            log_file.write(
                f"Generated: {datetime.datetime.now()}\n"
            )
            log_file.write(f"Financial year: {fy}\n")
            for key, value in summary.items():
                log_file.write(f"{key}: {value}\n")

        progress_callback(100, "Completed successfully.")
        log_callback(f"Report created: {output_path}")
        return summary

    except Exception as error:
        log_callback(f"FATAL ERROR: {error}")
        log_callback(traceback.format_exc())
        progress_callback(100, f"Error: {error}")
        raise


base.run_reconciliation_workflow = run_reconciliation_workflow


if __name__ == "__main__":
    app = base.AISBrokerRecoApp()
    app.mainloop()

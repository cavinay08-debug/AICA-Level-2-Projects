"""
AIS Broker Reconciliation - ISIN Auto-Fill Upgrade v4

Adds, on top of v3 (report_v2 + modified + final):

    1. Auto-fill of blank ISIN codes in the BROKER data sheet only, using
       the NSE Equity ISIN master and the AMFI Mutual Fund ISIN master
       (both read from ISIN_Code_List.xlsx).
         - Equity: match by exact SYMBOL first, then by fuzzy company name.
         - Mutual Funds: match by fuzzy scheme name. If the plan type
           (Growth/Payout vs Dividend-Reinvestment/IDCW) cannot be
           determined from the broker row, the ISIN is NOT guessed -
           the row is flagged "Manual Review Required" instead.
         - Every fill/flag decision is written to a new "ISIN Auto-Fill
           Log" sheet appended to the output workbook, plus a note in the
           row's own Remarks column.
    2. AIS-vs-Broker date matching window reduced to 2 days (T+1/T+2
       settlement offset) instead of the previous 3-day default.
    3. Sale-value tolerance for brokerage/other charges: unchanged: v2/
       modified already apply a dynamic tolerance
       (max(Rs 100, 0.75% of value), capped at Rs 2,500) to absorb the
       AIS-net-value vs Broker-net-value gap caused by brokerage and
       other transaction charges. This module keeps that logic as-is.

Keep this file, and the ISIN_Code_List.xlsx master file, in the same
folder as:
    ais_broker_reco_final.py
    ais_broker_reco_modified.py
    ais_broker_reco_report_v2.py

Run:
    python ais_broker_reco_isin_v4.py

The ISIN master file is looked for, in order:
    1. Path in the ISIN_MASTER_FILE environment variable (if set)
    2. "ISIN_Code_List.xlsx" in this script's own folder

If it is not found, the app still runs normally - ISIN auto-fill is
simply skipped for that run (a warning is written to the log/output).
"""

import os
import re
import datetime
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import ais_broker_reco_report_v2 as v2

mod = v2.mod
base = v2.base

# v2's report-formatting functions reference Font/PatternFill from its own
# module globals but never import them there (this is the same gap that
# ais_broker_reco_report_v3.py patched). Inject them so those functions work.
v2.Font = Font
v2.PatternFill = PatternFill

# =====================================================================
# 1. CONFIGURATION
# =====================================================================

# Settlement-offset tolerance for AIS-vs-Broker date matching.
base.DATE_WINDOW_DAYS = 2

ISIN_MASTER_FILENAME = "ISIN_Code_List.xlsx"
_ISIN_RE = re.compile(r'^IN[A-Z0-9]{10}$')

_ISIN_MASTER_CACHE = {}
_ISIN_FILL_LOG = []  # reset at the start of every reconciliation run

# Words that indicate a mutual fund plan type. Kept separate from
# LEGAL_AND_COMMON_WORDS so they remain significant tokens for
# company_name_match - a broker row that says "...Growth" should match
# master rows that also say "...Growth" better than "...IDCW" rows.
_GROWTH_PLAN_WORDS = ("GROWTH", "PAYOUT")
_REINVEST_PLAN_WORDS = ("REINVESTMENT", "REINVEST", "IDCW", "DIVIDEND")

_MF_KEYWORDS = (
    "MUTUAL FUND", " FUND", "SCHEME", "ETF", "ELSS", "MF-", "MF "
)

_SYMBOL_FIELD_NAMES = (
    "symbol", "trading symbol", "nse symbol", "bse symbol",
    "scrip code", "script code", "scrip symbol", "security code",
    "trading code", "isin symbol"
)


def _default_isin_master_path():
    here = os.path.dirname(os.path.abspath(__file__))
    candidate = os.path.join(here, ISIN_MASTER_FILENAME)
    return candidate if os.path.exists(candidate) else ""


def _resolve_isin_master_path():
    env_path = os.environ.get("ISIN_MASTER_FILE", "").strip()
    if env_path and os.path.exists(env_path):
        return env_path
    return _default_isin_master_path()


# =====================================================================
# 2. LOAD & INDEX THE ISIN MASTER
# =====================================================================
def _looks_like_isin(value):
    return isinstance(value, str) and bool(_ISIN_RE.match(value.strip().upper()))


def _index_by_block(rows_iterable):
    """
    rows_iterable: iterable of (compact_name, ...) tuples.
    Returns dict: first-4-chars-of-compact -> list of row tuples.
    Used to cut down fuzzy-match comparisons for large master sheets.
    """
    index = {}
    for item in rows_iterable:
        compact = item[0]
        if not compact:
            continue
        block_key = compact[:4]
        index.setdefault(block_key, []).append(item)
    return index


def load_isin_master(path, log_callback):
    if not path or not os.path.exists(path):
        log_callback(
            f"ISIN master file not found (looked for '{ISIN_MASTER_FILENAME}' "
            f"next to the app, or ISIN_MASTER_FILE env var). "
            f"Skipping ISIN auto-fill for this run."
        )
        return None

    cache_key = (path, os.path.getmtime(path))
    if cache_key in _ISIN_MASTER_CACHE:
        return _ISIN_MASTER_CACHE[cache_key]

    log_callback(f"Loading ISIN master file: {os.path.basename(path)} ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    equity_by_symbol = {}
    equity_rows = []  # (compact_name, display_name, isin)

    if "EQUITY_ISIN" in wb.sheetnames:
        ws = wb["EQUITY_ISIN"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            symbol = row[0] if len(row) > 0 else None
            company_name = row[1] if len(row) > 1 else None
            isin = next((c for c in row if _looks_like_isin(c)), None)
            if not isin:
                continue
            isin = isin.strip().upper()
            if symbol:
                sym_clean = base.clean_isin_code(symbol)
                if sym_clean:
                    equity_by_symbol[sym_clean] = isin
            if company_name:
                compact = _own_compact(company_name)
                if compact:
                    equity_rows.append(
                        (compact, base.clean_security_name(company_name), isin)
                    )

    mf_growth_rows = []     # (compact_name, display_name, isin)
    mf_reinvest_rows = []   # (compact_name, display_name, isin)

    if "Mutual Fund_ISIN" in wb.sheetnames:
        ws = wb["Mutual Fund_ISIN"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            isin_growth = row[1]
            isin_reinvest = row[2]
            scheme_name = row[3]
            if not scheme_name or not str(scheme_name).strip():
                continue
            display_name = base.clean_security_name(scheme_name)
            compact = _own_compact(scheme_name)
            if not compact:
                continue
            if _looks_like_isin(isin_growth):
                mf_growth_rows.append((compact, display_name, isin_growth.strip().upper()))
            if _looks_like_isin(isin_reinvest):
                mf_reinvest_rows.append((compact, display_name, isin_reinvest.strip().upper()))

    master = {
        "equity_by_symbol": equity_by_symbol,
        "equity_block_index": _index_by_block(equity_rows),
        "mf_growth_block_index": _index_by_block(mf_growth_rows),
        "mf_reinvest_block_index": _index_by_block(mf_reinvest_rows),
    }
    log_callback(
        f"ISIN master indexed: {len(equity_by_symbol)} equity symbols, "
        f"{len(equity_rows)} equity names, "
        f"{len(mf_growth_rows)} MF growth/payout schemes, "
        f"{len(mf_reinvest_rows)} MF dividend-reinvestment schemes."
    )

    _ISIN_MASTER_CACHE[cache_key] = master
    return master


# =====================================================================
# 3. MATCHING HELPERS
# =====================================================================
def _find_broker_symbol_value(txn):
    normalized = {
        base.normalize_header_text(key): value
        for key, value in txn.original_row_data.items()
    }
    for key, value in normalized.items():
        if any(name in key for name in _SYMBOL_FIELD_NAMES):
            if value not in (None, ""):
                return str(value)
    return ""


def _is_mf_name(name):
    upper = f" {name.upper()} "
    return any(kw in upper for kw in _MF_KEYWORDS)


_MINIMAL_LEGAL_WORDS = {"LIMITED", "LTD", "PRIVATE", "PVT", "CO", "COMPANY"}


def _own_tokens(value):
    return re.findall(r'[A-Z0-9]+', base.clean_security_name(value))


def _own_compact(value):
    tokens = [t for t in _own_tokens(value) if t not in _MINIMAL_LEGAL_WORDS]
    return "".join(tokens)


def _name_match_strict(query, candidate):
    """
    Deliberately stricter than v2.company_name_match, which is tuned for
    matching two representations of the SAME trade (AIS row vs Broker
    row) and, e.g., strips "INDUSTRIES"/"OF"/"INDIA" as noise words and
    treats a shared first token as a match. Both of those are too loose
    when a single query is compared against a master of thousands of
    unrelated company/scheme names - that logic would equate "Reliance
    Industries" with "Reliance Power" or "Reliance Infrastructure".
    Only LIMITED/LTD/PRIVATE/PVT/CO/COMPANY are treated as noise here.
    """
    q = _own_compact(query)
    c = _own_compact(candidate)
    if not q or not c:
        return False
    if q == c:
        return True

    shorter, longer = (q, c) if len(q) <= len(c) else (c, q)
    if len(shorter) >= 6 and shorter in longer:
        return True

    if len(shorter) >= 4:
        ratio = SequenceMatcher(None, q, c).ratio()
        if ratio >= 0.90:
            return True

    return False


def _fuzzy_lookup(name, block_index):
    """
    Looks up `name` in a block-indexed master list.
    Returns (isin, display_name) if exactly one distinct ISIN is found
    among matches, ("AMBIGUOUS", None) if more than one distinct ISIN
    matches, or (None, None) if nothing matches.
    """
    compact_query = _own_compact(name)
    if not compact_query:
        return None, None

    candidates = block_index.get(compact_query[:4], [])
    matches = []
    for compact_master, display_name, isin in candidates:
        if _name_match_strict(name, display_name):
            matches.append((isin, display_name))

    if not matches:
        return None, None
    distinct_isins = {isin for isin, _ in matches}
    if len(distinct_isins) > 1:
        return "AMBIGUOUS", None
    return matches[0]


def _match_equity_isin(txn, master):
    symbol_value = _find_broker_symbol_value(txn)
    if symbol_value:
        sym_clean = base.clean_isin_code(symbol_value)
        isin = master["equity_by_symbol"].get(sym_clean)
        if isin:
            return isin, "Equity Symbol Match", False

    isin, _ = _fuzzy_lookup(txn.security_name, master["equity_block_index"])
    if isin == "AMBIGUOUS":
        return None, "Equity Company Name Match", True
    if isin:
        return isin, "Equity Company Name Match", False
    return None, "Equity - No Match", False


def _match_mf_isin(txn, master):
    name_upper = txn.security_name.upper()
    wants_growth = any(w in name_upper for w in _GROWTH_PLAN_WORDS)
    wants_reinvest = any(w in name_upper for w in _REINVEST_PLAN_WORDS)

    growth_isin, _ = _fuzzy_lookup(txn.security_name, master["mf_growth_block_index"])
    reinvest_isin, _ = _fuzzy_lookup(txn.security_name, master["mf_reinvest_block_index"])

    growth_found = growth_isin not in (None, "AMBIGUOUS")
    reinvest_found = reinvest_isin not in (None, "AMBIGUOUS")

    if growth_isin == "AMBIGUOUS" or reinvest_isin == "AMBIGUOUS":
        return None, "MF Scheme Match", True  # ambiguous -> manual review

    if growth_found and not reinvest_found:
        return growth_isin, "MF Growth/Payout Match", False
    if reinvest_found and not growth_found:
        return reinvest_isin, "MF Dividend-Reinvestment Match", False

    if growth_found and reinvest_found:
        # Same scheme has both plan types on record - only safe to pick
        # one if the broker row itself states the plan type.
        if wants_growth and not wants_reinvest:
            return growth_isin, "MF Growth/Payout Match (plan stated)", False
        if wants_reinvest and not wants_growth:
            return reinvest_isin, "MF Dividend-Reinvestment Match (plan stated)", False
        # Plan type not stated on the broker row and both plan ISINs
        # exist for this scheme -> do not guess.
        return None, "MF Scheme Match - Plan Type Not Stated", True

    return None, "MF - No Match", False


# =====================================================================
# 4. FILL BROKER ISINS
# =====================================================================
def fill_missing_broker_isins(transactions, master, log_callback):
    if master is None:
        return

    filled = 0
    ambiguous = 0
    unmatched = 0

    for txn in transactions:
        if txn.isin:
            continue
        if not txn.security_name:
            continue

        if _is_mf_name(txn.security_name):
            isin, method, is_ambiguous = _match_mf_isin(txn, master)
        else:
            isin, method, is_ambiguous = _match_equity_isin(txn, master)

        if isin:
            txn.isin = isin
            note = f"ISIN auto-filled from ISIN master ({method})."
            txn.remarks = f"{txn.remarks}; {note}".strip("; ")
            _ISIN_FILL_LOG.append((
                txn.source_sheet, txn.source_row, txn.security_name,
                isin, method, "Filled"
            ))
            filled += 1
        elif is_ambiguous:
            note = (
                f"ISIN left blank - {method}: multiple/ambiguous candidates "
                f"or plan type not stated in Broker data. Manual review required."
            )
            txn.remarks = f"{txn.remarks}; {note}".strip("; ")
            _ISIN_FILL_LOG.append((
                txn.source_sheet, txn.source_row, txn.security_name,
                "", method, "Ambiguous - Manual Review"
            ))
            ambiguous += 1
        else:
            note = (
                "ISIN left blank - no match found in ISIN master for this "
                "security/scheme name. Manual review required."
            )
            txn.remarks = f"{txn.remarks}; {note}".strip("; ")
            _ISIN_FILL_LOG.append((
                txn.source_sheet, txn.source_row, txn.security_name,
                "", method, "Not Matched - Manual Review"
            ))
            unmatched += 1

    log_callback(
        f"ISIN auto-fill (Broker data): {filled} filled, "
        f"{ambiguous} ambiguous/plan-unclear, {unmatched} unmatched "
        f"(left blank for manual review)."
    )


# =====================================================================
# 5. HOOK INTO THE LOADER
# =====================================================================
_previous_loader = base.load_and_normalize_workbook


def load_with_isin_autofill(file_path, file_label, log_callback):
    global _ISIN_FILL_LOG

    if file_label == "AIS Data File":
        _ISIN_FILL_LOG = []  # reset once per reconciliation run

    transactions = _previous_loader(file_path, file_label, log_callback)

    if file_label == "Broker Data File":
        master_path = _resolve_isin_master_path()
        master = load_isin_master(master_path, log_callback)
        fill_missing_broker_isins(transactions, master, log_callback)

    return transactions


base.load_and_normalize_workbook = load_with_isin_autofill
mod.base.load_and_normalize_workbook = load_with_isin_autofill


# =====================================================================
# 6. APPEND "ISIN AUTO-FILL LOG" SHEET TO THE OUTPUT WORKBOOK
# =====================================================================
_previous_run_workflow = base.run_reconciliation_workflow

_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_STATUS_COLORS = {
    "Filled": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "Ambiguous - Manual Review": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "Not Matched - Manual Review": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
}


def _append_isin_fill_log_sheet(output_path, log_callback):
    if not _ISIN_FILL_LOG:
        return
    if not os.path.exists(output_path):
        return

    wb = openpyxl.load_workbook(output_path)
    ws = wb.create_sheet(title="ISIN Auto-Fill Log")

    headers = [
        "Source Sheet", "Source Row", "Broker Security/Scheme Name",
        "Filled ISIN", "Match Method", "Status"
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for row_idx, record in enumerate(_ISIN_FILL_LOG, start=2):
        ws.append(list(record))
        status = record[-1]
        fill = _STATUS_COLORS.get(status)
        if fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    widths = [30, 12, 45, 16, 34, 28]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    log_callback(
        f"ISIN Auto-Fill Log sheet appended to output "
        f"({len(_ISIN_FILL_LOG)} rows)."
    )


def run_reconciliation_workflow_with_isin_log(
    ais_file, broker_file, output_path, fy,
    progress_callback, log_callback
):
    summary = _previous_run_workflow(
        ais_file, broker_file, output_path, fy,
        progress_callback, log_callback
    )
    try:
        _append_isin_fill_log_sheet(output_path, log_callback)
    except Exception as exc:
        log_callback(f"WARNING: Could not append ISIN Auto-Fill Log sheet: {exc}")

    summary["isin_filled"] = sum(1 for r in _ISIN_FILL_LOG if r[-1] == "Filled")
    summary["isin_manual_review"] = sum(1 for r in _ISIN_FILL_LOG if r[-1] != "Filled")
    return summary


base.run_reconciliation_workflow = run_reconciliation_workflow_with_isin_log


if __name__ == "__main__":
    app = base.AISBrokerRecoApp()
    app.mainloop()

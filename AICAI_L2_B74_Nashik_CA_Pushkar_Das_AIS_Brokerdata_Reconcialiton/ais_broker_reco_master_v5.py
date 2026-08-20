"""
AIS BROKER RECO - MASTER CONSOLIDATED BUILD (v5)
=================================================
Combines, in one file, everything previously spread across:
    ais_broker_reco_final.py        (core engine + GUI)
    ais_broker_reco_modified.py     (tax software output, nominal-cost tolerance)
    ais_broker_reco_report_v2.py    (STT-sheet inference, value dashboard)
    ais_broker_reco_report_v3.py    (Font/PatternFill fix - not needed, imported directly here)
    ais_broker_reco_isin_v4.py      (ISIN auto-fill for blank broker ISINs)

New in this build (per latest requirements):
    1. Summary sheet value-dashboard (Value as per AIS / Broker / Actual
       Difference) sits alongside the record-count metrics.
    2. Every metric/dashboard row in Summary has a working hyperlink to
       jump straight to the relevant sheet, and every detail sheet has a
       "Back to Summary" link at the top-left cell.
    3. Sheet naming and the Summary link label for "Other Financial
       Differences" are identical (no more mismatch).
    4. "Is it LTCG?" column (Yes / No / Review Required) is written in the
       Tax Software Output sheet, derived from Period of Holding + Type
       of Asset (+ STT-paid inference used only as a supporting remark,
       not as the classification test itself).
    5. STT Paid amount and Holding Period columns removed from all output
       sheets (these are not reliably present in either source file); the
       information is retained internally only for classification.
    6. Nominal Sale Value differences caused purely by brokerage/other
       transaction charges are ignored automatically using a dynamic
       tolerance (greater of Rs 100 or 0.75% of value, capped at
       Rs 2,500) whenever the underlying Qty sold/redeemed matches.
    7. Name-matching recognises broker-side trading symbols/short codes
       against the AIS full security name, using the NSE Equity master
       (SYMBOL <-> NAME OF COMPANY) inside ISIN_Code_List.xlsx when
       available, plus a generated-acronym and fuzzy-ratio fallback.
    8. Blank Broker ISINs are auto-filled from ISIN_Code_List.xlsx
       (Equity + Mutual Fund masters), never guessed when ambiguous.

Run:
    python ais_broker_reco_master_v5.py
"""

import os
import re
import sys
import datetime
import threading
import traceback
from difflib import SequenceMatcher

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter import scrolledtext

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# 1. CONSTANTS & CONFIGURATION
# =====================================================================
TOLERANCE_AMOUNT = 10.0            # Strict monetary tolerance (Rs)
SALE_TOL_MIN = 100.0               # Minimum nominal-variance tolerance (Rs)
SALE_TOL_PCT = 0.0075              # 0.75% of transaction value
SALE_TOL_CAP = 2500.0              # Cap on the dynamic tolerance (Rs)
DATE_WINDOW_DAYS = 2                # T+1/T+2 settlement offset window

ISIN_MASTER_FILENAME = "ISIN_Code_List.xlsx"
_ISIN_RE = re.compile(r'^IN[A-Z0-9]{10}$')

COLUMN_SYNONYMS = {
    'isin': [
        'isin', 'isin code', 'isin_code', 'security code', 'isin number',
        'security description isin', 'instrument isin', 'isin id',
        'security isin', 'asset isin'
    ],
    'security_name': [
        'security name', 'security description', 'security', 'scrip name',
        'scrip', 'company name', 'description', 'stock name', 'scheme name',
        'scheme', 'particulars', 'scrip/security description', 'instrument name',
        'name of security', 'name of share', 'name of shares', 'name of unit',
        'asset description'
    ],
    'sale_date': [
        'sale date', 'transaction date', 'trade date', 'sell date',
        'date of sale', 'sale_date', 'date of transfer', 'transfer date',
        'date of transaction', 'redemption date', 'sell_date', 'disposal date',
        'date of sale/transfer', 'sale transfer date'
    ],
    'purchase_date': [
        'purchase date', 'buy date', 'date of purchase', 'purchase_date',
        'date of acquisition', 'acquisition date', 'buy_date', 'acquisition_date'
    ],
    'quantity': [
        'quantity', 'qty', 'units', 'volume', 'sale quantity', 'sell quantity',
        'number of shares', 'no of shares', 'no. of shares', 'shares qty',
        'sell qty', 'sale qty', 'qty sold', 'number of units', 'no of units',
        'no. of units', 'units sold'
    ],
    'sale_consideration': [
        'sale consideration', 'sales consideration', 'sale value', 'sale amount',
        'sell value', 'sales value', 'gross sales', 'sales amount', 'consideration',
        'redemption value', 'redemption amount', 'amount', 'value of transaction',
        'transaction value', 'sale_value', 'consideration amount', 'sell value (gross)',
        'full value of consideration', 'sales proceeds', 'sale proceeds',
        'gross consideration', 'net consideration'
    ],
    'purchase_cost': [
        'purchase cost', 'cost of acquisition', 'acquisition cost', 'purchase value',
        'purchase amount', 'buy value', 'buy cost', 'cost value', 'cost',
        'cost_of_acquisition', 'purchase price/cost', 'buy value (gross)',
        'purchase price', 'indexed cost of acquisition'
    ],
    'stt': [
        'stt', 'securities transaction tax', 'stt paid', 'stt charges',
        'stt_charges', 'stt amt', 'stt amount', 'securities transaction tax (stt)'
    ],
    'gain_class': [
        'class', 'type of gain', 'classification', 'gain type',
        'ltcg/stcg', 'capital gain type', 'nature of income', 'gain_class',
        'term', 'gain classification'
    ]
}

_NAME_STOPWORDS = {
    "LIMITED", "LTD", "LTD.", "PRIVATE", "PVT", "CO", "COMPANY", "CORP",
    "CORPORATION", "INDIA", "INDIAN", "IND", "THE", "OF", "AND", "GROUP",
    "HOLDINGS", "ENTERPRISES", "INDUSTRIES", "SERVICES", "PLAN", "OPTION",
    "GROWTH", "DIRECT", "REGULAR", "FUND", "SCHEME"
}


def sale_cost_tolerance(v1, v2):
    """Dynamic tolerance for brokerage/other-charge driven nominal variance."""
    base = max(abs(v1), abs(v2), 1.0)
    return min(max(SALE_TOL_MIN, SALE_TOL_PCT * base), SALE_TOL_CAP)


# =====================================================================
# 2. NAME MATCHING & OPTIONAL ISIN MASTER (Equity symbols + AMFI MF)
# =====================================================================
class ISINMasterLookup:
    """
    Loads ISIN_Code_List.xlsx (if present) to help:
        a) recognise broker-side trading SYMBOLS / short codes against
           full AIS security names, and
        b) auto-fill blank ISIN codes in broker rows.
    Fails silently (empty lookup) if the file is not found or malformed.
    """
    def __init__(self, log_callback=print):
        self.log_callback = log_callback
        self.symbol_to_name = {}     # SYMBOL -> NAME OF COMPANY
        self.name_to_isin_eq = {}    # NAME OF COMPANY (norm) -> ISIN
        self.symbol_to_isin = {}     # SYMBOL -> ISIN
        self.mf_growth = {}          # normalized scheme core name -> ISIN
        self.mf_div = {}             # normalized scheme core name -> ISIN
        self.loaded = False
        self._load()

    def _find_master_path(self):
        env_path = os.environ.get("ISIN_MASTER_FILE")
        if env_path and os.path.exists(env_path):
            return env_path
        local_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), ISIN_MASTER_FILENAME
        )
        if os.path.exists(local_path):
            return local_path
        if os.path.exists(ISIN_MASTER_FILENAME):
            return ISIN_MASTER_FILENAME
        return None

    def _load(self):
        path = self._find_master_path()
        if not path:
            self.log_callback("ISIN master file not found - ISIN auto-fill and symbol lookup will be skipped.")
            return
        try:
            xl = pd.ExcelFile(path, engine="openpyxl")
            for sheet_name in xl.sheet_names:
                sname = sheet_name.strip().lower()
                df = xl.parse(sheet_name=sheet_name, header=None)
                if df.empty:
                    continue
                if "mutual" in sname or "mf" in sname or "amfi" in sname:
                    self._load_mf_sheet(df)
                else:
                    self._load_equity_sheet(df)
            self.loaded = bool(self.symbol_to_name or self.mf_growth or self.mf_div)
            self.log_callback(
                f"ISIN master loaded: {len(self.symbol_to_name)} equity symbols, "
                f"{len(self.mf_growth)} MF growth schemes, {len(self.mf_div)} MF dividend schemes."
            )
        except Exception as exc:
            self.log_callback(f"WARNING: Could not read ISIN master file: {exc}")

    def _load_equity_sheet(self, df):
        # Detect a header row (SYMBOL / NAME OF COMPANY / ISIN)
        header_row_idx = None
        for i in range(min(5, len(df))):
            row_text = " ".join(str(x).upper() for x in df.iloc[i].values if pd.notna(x))
            if "SYMBOL" in row_text and "ISIN" in row_text:
                header_row_idx = i
                break

        if header_row_idx is not None:
            headers = [str(x).strip().upper() for x in df.iloc[header_row_idx].values]
            data = df.iloc[header_row_idx + 1:]
            col_symbol = next((i for i, h in enumerate(headers) if "SYMBOL" in h), 0)
            col_name = next((i for i, h in enumerate(headers) if "NAME" in h), 1)
            col_isin = next((i for i, h in enumerate(headers) if "ISIN" in h), 6)
        else:
            # Positional fallback: standard NSE EQUITY_L layout
            # SYMBOL, NAME OF COMPANY, SERIES, DATE OF LISTING, PAID UP VALUE,
            # MARKET LOT, ISIN NUMBER, FACE VALUE
            data = df
            col_symbol, col_name, col_isin = 0, 1, 6

        for _, row in data.iterrows():
            try:
                symbol = str(row.iloc[col_symbol]).strip().upper()
                name = str(row.iloc[col_name]).strip().upper()
                isin = str(row.iloc[col_isin]).strip().upper()
            except Exception:
                continue
            if not symbol or symbol == "NAN" or not _ISIN_RE.match(isin):
                continue
            self.symbol_to_name[symbol] = name
            self.symbol_to_isin[symbol] = isin
            self.name_to_isin_eq[normalize_company_core(name)] = isin

    def _load_mf_sheet(self, df):
        # Detect header row for AMFI-style layout
        header_row_idx = None
        for i in range(min(5, len(df))):
            row_text = " ".join(str(x).upper() for x in df.iloc[i].values if pd.notna(x))
            if "SCHEME" in row_text and "ISIN" in row_text:
                header_row_idx = i
                break

        if header_row_idx is not None:
            headers = [str(x).strip().upper() for x in df.iloc[header_row_idx].values]
            data = df.iloc[header_row_idx + 1:]
            col_name = next((i for i, h in enumerate(headers) if "SCHEME NAME" in h), None)
            col_growth = next((i for i, h in enumerate(headers) if "GROWTH" in h or "PAYOUT" in h), None)
            col_div = next((i for i, h in enumerate(headers) if "REINVEST" in h or "DIVIDEND" in h or "IDCW" in h), None)
        else:
            data = df
            col_name, col_growth, col_div = 3, 1, 2

        if col_name is None:
            return

        for _, row in data.iterrows():
            try:
                name = str(row.iloc[col_name]).strip().upper()
            except Exception:
                continue
            if not name or name == "NAN":
                continue
            core = normalize_company_core(name)
            if col_growth is not None:
                isin_g = str(row.iloc[col_growth]).strip().upper() if col_growth < len(row) else ""
                if _ISIN_RE.match(isin_g):
                    self.mf_growth[core] = isin_g
            if col_div is not None:
                isin_d = str(row.iloc[col_div]).strip().upper() if col_div < len(row) else ""
                if _ISIN_RE.match(isin_d):
                    self.mf_div[core] = isin_d

    def resolve_symbol_to_name(self, token):
        """If token is a known NSE trading symbol, return the company name."""
        return self.symbol_to_name.get(token.strip().upper())

    def best_fuzzy_equity_name(self, query_core, threshold=0.72):
        best, best_score = None, 0.0
        for core_name in self.name_to_isin_eq.keys():
            score = SequenceMatcher(None, query_core, core_name).ratio()
            if score > best_score:
                best, best_score = core_name, score
        if best_score >= threshold:
            return best, self.name_to_isin_eq[best], best_score
        return None, None, 0.0

    def best_fuzzy_mf(self, query_core, plan_hint, threshold=0.72):
        """plan_hint: 'growth', 'div', or None (unknown)."""
        candidates = []
        if plan_hint in (None, "growth"):
            for core_name, isin in self.mf_growth.items():
                score = SequenceMatcher(None, query_core, core_name).ratio()
                if score >= threshold:
                    candidates.append(("growth", core_name, isin, score))
        if plan_hint in (None, "div"):
            for core_name, isin in self.mf_div.items():
                score = SequenceMatcher(None, query_core, core_name).ratio()
                if score >= threshold:
                    candidates.append(("div", core_name, isin, score))
        if not candidates:
            return None
        candidates.sort(key=lambda c: c[3], reverse=True)
        if plan_hint is None and len(candidates) > 1 and candidates[0][0] != candidates[1][0]:
            # Ambiguous across plan types when plan cannot be determined
            return "AMBIGUOUS"
        return candidates[0]


def normalize_company_core(name):
    """Uppercase, strip punctuation and common stopwords -> ordered token list joined."""
    text = re.sub(r'[^A-Z0-9& ]', ' ', str(name).upper())
    tokens = [t for t in text.split() if t and t not in _NAME_STOPWORDS and t != '&']
    return " ".join(tokens)


def generate_acronym(core_text):
    return "".join(t[0] for t in core_text.split() if t)


def names_match(name_a, name_b, isin_a="", isin_b="", isin_master=None):
    """
    Enhanced name matcher. Recognises:
        - exact / substring match on cleaned names (legacy behaviour)
        - broker-side trading SYMBOL vs AIS full company name (via master)
        - generated-acronym match (e.g. RIL vs Reliance Industries Ltd)
        - fuzzy ratio fallback
    """
    if isin_a and isin_b:
        return isin_a == isin_b

    a_clean = clean_security_name(name_a).replace(" ", "")
    b_clean = clean_security_name(name_b).replace(" ", "")
    if a_clean and b_clean and (a_clean == b_clean or a_clean in b_clean or b_clean in a_clean):
        return True

    a_core = normalize_company_core(name_a)
    b_core = normalize_company_core(name_b)
    if a_core and b_core and a_core == b_core:
        return True

    # Symbol-vs-name recognition using the NSE equity master
    if isin_master is not None and isin_master.loaded:
        b_token = re.sub(r'[^A-Z0-9]', '', name_b.upper())
        a_token = re.sub(r'[^A-Z0-9]', '', name_a.upper())
        resolved_b = isin_master.resolve_symbol_to_name(b_token)
        if resolved_b and normalize_company_core(resolved_b) == a_core:
            return True
        resolved_a = isin_master.resolve_symbol_to_name(a_token)
        if resolved_a and normalize_company_core(resolved_a) == b_core:
            return True

    # Generated-acronym match (e.g. broker uses "RIL"/"TCS"-style short code)
    if a_core and b_core:
        a_acr = generate_acronym(a_core)
        b_acr = generate_acronym(b_core)
        b_compact = b_core.replace(" ", "")
        a_compact = a_core.replace(" ", "")
        if len(b_compact) <= 6 and b_compact == a_acr:
            return True
        if len(a_compact) <= 6 and a_compact == b_acr:
            return True

    # Fuzzy fallback on the stopword-stripped core names
    if a_core and b_core:
        ratio = SequenceMatcher(None, a_core, b_core).ratio()
        if ratio >= 0.72:
            return True

    return False


# =====================================================================
# 3. TAX RULE ENGINE
# =====================================================================
class TaxRuleEngine:
    """Indian Income Tax Capital Gains classification engine."""

    def __init__(self, financial_year):
        self.financial_year = financial_year

    def classify_security_type(self, isin, security_name):
        isin = str(isin).strip().upper()
        name = str(security_name).strip().upper()

        if isin.startswith("INF"):
            debt_keywords = [
                "DEBT", "BOND", "LIQUID", "TREASURY", "GILT", "OVERNIGHT", "ULTRA SHORT",
                "SHORT TERM", "MEDIUM TERM", "CASH", "CONSERVATIVE", "HYBRID DEBT",
                "FIXED MATURITY", "FMP", "INCOME FUND", "DYNAMIC BOND", "SAVINGS FUND"
            ]
            is_debt = any(kw in name and "ARBITRAGE" not in name and "EQUITY" not in name
                          for kw in debt_keywords)
            return "DEBT_MF" if is_debt else "EQUITY_MF"

        elif isin.startswith("INE") or isin.startswith("IN9"):
            bond_keywords = ["BOND", "NCD", "DEBENTURE", "DEB", "BND", "GOLD BOND", "SGB"]
            if any(kw in name for kw in bond_keywords):
                return "OTHER_SECURITY"
            return "EQUITY_SHARE"
        else:
            if "MUTUAL FUND" in name or "MF" in name or "FUND" in name:
                if any(kw in name for kw in ["DEBT", "LIQUID", "GILT", "BOND"]):
                    return "DEBT_MF"
                return "EQUITY_MF"
            elif any(kw in name for kw in ["BOND", "DEBENTURE", "NCD"]):
                return "OTHER_SECURITY"
            return "EQUITY_SHARE"

    def classify_transaction(self, isin, security_name, purchase_date, sale_date):
        """Returns (classification, holding_days, remarks). classification in
        {"LTCG", "STCG", "Review Required"}."""
        if not purchase_date or not sale_date:
            return "Review Required", None, "Missing purchase or sale date"

        if isinstance(purchase_date, datetime.datetime):
            purchase_date = purchase_date.date()
        if isinstance(sale_date, datetime.datetime):
            sale_date = sale_date.date()

        holding_days = (sale_date - purchase_date).days
        if holding_days < 0:
            return "Review Required", holding_days, f"Sale date ({sale_date}) is before purchase date ({purchase_date})"

        asset_type = self.classify_security_type(isin, security_name)

        if asset_type in ("EQUITY_SHARE", "EQUITY_MF"):
            if holding_days > 365:
                return "LTCG", holding_days, f"Equity-oriented asset held > 12 months ({holding_days} days)"
            return "STCG", holding_days, f"Equity-oriented asset held <= 12 months ({holding_days} days)"

        elif asset_type == "DEBT_MF":
            cutoff_date = datetime.date(2023, 4, 1)
            if purchase_date >= cutoff_date:
                return "STCG", holding_days, f"Debt MF acquired on/after 01-Apr-2023 (Sec 50AA) - always STCG ({holding_days} days)"
            if holding_days > 1095:
                return "LTCG", holding_days, f"Debt MF acquired before 01-Apr-2023, held > 36 months ({holding_days} days)"
            return "STCG", holding_days, f"Debt MF acquired before 01-Apr-2023, held <= 36 months ({holding_days} days)"

        elif asset_type == "OTHER_SECURITY":
            if any(kw in security_name.upper() for kw in ["BOND", "NCD", "DEBENTURE"]):
                limit, label = 1095, "Debt security"
            else:
                limit, label = 730, "Other unlisted security"
            if holding_days > limit:
                return "LTCG", holding_days, f"{label} held > limit ({holding_days} days)"
            return "STCG", holding_days, f"{label} held <= limit ({holding_days} days)"

        return "Review Required", holding_days, f"Could not determine rules for asset type: {asset_type}"


def ltcg_yes_no(calc_cls):
    if calc_cls == "LTCG":
        return "Yes"
    if calc_cls == "STCG":
        return "No"
    return "Review Required"


def infer_stt_status_from_sheet(sheet_name):
    """
    Returns (stt_paid: bool|None, remark: str)
    'Gain arising of STT Paid' style sheet names => STT Paid: Yes.
    'Without STT' / 'STT Not Paid' / 'No STT' style names => STT Paid: No.
    """
    name = normalize_header_text(sheet_name)
    negative_phrases = ("without stt", "stt not paid", "stt unpaid", "no stt", "non stt")
    if any(p in name for p in negative_phrases):
        return False, "STT Paid: No (inferred from source-sheet name)"
    if "stt paid" in name or "with stt" in name:
        return True, "STT Paid: Yes (inferred from source-sheet name; security treated as listed)"
    return None, "STT Paid: Not determinable from source-sheet name"


# =====================================================================
# 4. DATA NORMALISATION & WORKBOOK INSPECTOR
# =====================================================================
class NormalizedTransaction:
    def __init__(self):
        self.match_id = None
        self.source_file = ""
        self.source_sheet = ""
        self.source_row = None
        self.isin = ""
        self.security_name = ""
        self.sale_date = None
        self.purchase_date = None
        self.quantity = 0.0
        self.sale_consideration = 0.0
        self.purchase_cost = 0.0
        self.stt = 0.0                      # internal use only, never printed
        self.original_classification = ""
        self.calculated_classification = ""
        self.holding_days = None            # internal use only, never printed
        self.stt_status_remark = ""
        self.remarks = ""
        self.is_duplicate = False
        self.original_row_data = {}


def clean_isin_code(val):
    if pd.isna(val) or not val:
        return ""
    return re.sub(r'[^A-Za-z0-9]', '', str(val)).strip().upper()


def clean_security_name(val):
    if pd.isna(val) or not val:
        return ""
    cleaned = str(val).strip().upper()
    return re.sub(r'\s+', ' ', cleaned)


def extract_isin_from_text(val):
    if pd.isna(val) or not val:
        return ""
    match = re.search(r'\b(IN[A-Z0-9]{10})\b', str(val).upper())
    return match.group(1) if match else ""


def clean_date(val):
    if pd.isna(val) or val is None or str(val).strip() == "":
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val

    date_str = str(val).strip()
    if " " in date_str:
        date_str = date_str.split(" ")[0]

    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%Y/%m/%d",
                "%d-%b-%Y", "%d-%B-%Y", "%d-%b-%y", "%d-%B-%y"):
        try:
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    try:
        parsed = pd.to_datetime(date_str, dayfirst=True, errors="coerce")
        if pd.notna(parsed):
            return parsed.date()
    except Exception:
        pass
    return None


def clean_number(val):
    if pd.isna(val) or val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip()
    if not val_str:
        return 0.0
    val_str = val_str.replace("₹", "").replace(",", "").replace("Rs", "").replace(" ", "")
    if val_str.startswith("(") and val_str.endswith(")"):
        val_str = "-" + val_str[1:-1]
    try:
        return float(val_str)
    except ValueError:
        return 0.0


def normalize_header_text(val):
    if pd.isna(val) or val is None:
        return ""
    text = str(val).strip().lower()
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r'[_/\\-]+', ' ', text)
    text = re.sub(r'[^a-z0-9(). ]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def header_match_score(cell_value, synonym):
    if not cell_value or not synonym:
        return 0
    if cell_value == synonym:
        return 5
    if cell_value.startswith(synonym) or cell_value.endswith(synonym):
        return 4
    cell_tokens = set(cell_value.split())
    synonym_tokens = set(synonym.split())
    if synonym_tokens and synonym_tokens.issubset(cell_tokens):
        return 3
    if synonym in cell_value:
        return 2
    return 0


def detect_header_row(df_raw, sheet_name=""):
    best_row_idx = None
    best_mapping = {}
    best_header_depth = 1
    best_score = None
    is_mf_units_sheet = "units of mf except equity fund" in str(sheet_name).strip().lower()

    scan_limit = min(75, len(df_raw))
    for row_idx in range(scan_limit):
        row_values = [normalize_header_text(x) for x in df_raw.iloc[row_idx].values]
        candidate_rows = [(row_values, 1)]

        if row_idx + 1 < scan_limit:
            next_row_values = [normalize_header_text(x) for x in df_raw.iloc[row_idx + 1].values]
            combined_values = []
            max_len = max(len(row_values), len(next_row_values))
            for col_idx in range(max_len):
                top = row_values[col_idx] if col_idx < len(row_values) else ""
                bottom = next_row_values[col_idx] if col_idx < len(next_row_values) else ""
                combined_values.append(f"{top} {bottom}".strip())
            candidate_rows.append((combined_values, 2))

        for candidate_values, header_depth in candidate_rows:
            mapping = {}
            used_columns = set()
            quality_score = 0
            non_empty_cells = sum(1 for cell in candidate_values if cell and cell != 'nan')

            for std_col, synonyms in COLUMN_SYNONYMS.items():
                normalized_synonyms = [normalize_header_text(syn) for syn in synonyms]
                best_for_field = None
                for col_idx, cell_value in enumerate(candidate_values):
                    if col_idx in used_columns or not cell_value or cell_value == 'nan':
                        continue
                    col_score = max((header_match_score(cell_value, syn) for syn in normalized_synonyms), default=0)
                    if col_score <= 0:
                        continue
                    if best_for_field is None or col_score > best_for_field[0]:
                        best_for_field = (col_score, col_idx)
                if best_for_field is not None:
                    col_score, col_idx = best_for_field
                    mapping[std_col] = col_idx
                    used_columns.add(col_idx)
                    quality_score += col_score

            distinct_count = len(mapping)
            has_dates = 'sale_date' in mapping or 'purchase_date' in mapping
            has_values = 'sale_consideration' in mapping or 'purchase_cost' in mapping
            has_quantity_or_sheet_exception = 'quantity' in mapping or is_mf_units_sheet
            has_useful_shape = distinct_count >= 4 and has_quantity_or_sheet_exception and has_dates and has_values
            if not has_useful_shape:
                continue

            score_tuple = (distinct_count, quality_score, non_empty_cells, row_idx, header_depth)
            if best_score is None or score_tuple > best_score:
                best_score = score_tuple
                best_row_idx = row_idx
                best_mapping = mapping
                best_header_depth = header_depth

    return best_row_idx, best_mapping, best_header_depth


def resolve_excel_engine(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xls':
        try:
            import xlrd  # noqa: F401
        except ImportError as exc:
            raise ImportError(
                "Legacy .xls files require the optional 'xlrd' package. "
                "Install it with: python -m pip install xlrd"
            ) from exc
        return 'xlrd'
    return 'openpyxl'


def load_and_normalize_workbook(file_path, file_label, log_callback):
    log_callback(f"Inspecting workbook: {os.path.basename(file_path)}...")
    transactions = []

    try:
        engine = resolve_excel_engine(file_path)
        xl = pd.ExcelFile(file_path, engine=engine)
        sheet_names = xl.sheet_names
        log_callback(f"Found sheets in {file_label}: {sheet_names}")

        for sheet_name in sheet_names:
            df_raw = xl.parse(sheet_name=sheet_name, header=None)
            if len(df_raw) == 0:
                log_callback(f"Skipping empty sheet: {sheet_name}")
                continue

            is_mf_units_sheet = "units of mf except equity fund" in sheet_name.strip().lower()
            header_row_idx, mapping, header_depth = detect_header_row(df_raw, sheet_name=sheet_name)
            if header_row_idx is None:
                log_callback(f"Skipping sheet '{sheet_name}': Could not identify required columns.")
                continue

            stt_status, stt_remark = infer_stt_status_from_sheet(sheet_name)

            log_callback(
                f"Sheet '{sheet_name}' - Header detected at row {header_row_idx + 1}"
                f"{' (2-row header)' if header_depth == 2 else ''}"
            )
            log_callback(f"Mapped columns: {list(mapping.keys())}")

            data_start_idx = header_row_idx + header_depth
            df_data = df_raw.iloc[data_start_idx:].copy()

            if header_depth == 2 and header_row_idx + 1 < len(df_raw):
                combined_headers = []
                top_headers = df_raw.iloc[header_row_idx].values
                bottom_headers = df_raw.iloc[header_row_idx + 1].values
                max_len = max(len(top_headers), len(bottom_headers))
                for col_idx in range(max_len):
                    top = normalize_header_text(top_headers[col_idx]) if col_idx < len(top_headers) else ""
                    bottom = normalize_header_text(bottom_headers[col_idx]) if col_idx < len(bottom_headers) else ""
                    combined_headers.append((f"{top} {bottom}".strip()) or f"column_{col_idx + 1}")
                df_data.columns = combined_headers[:len(df_data.columns)]
            else:
                raw_headers = df_raw.iloc[header_row_idx].values
                df_data.columns = [
                    normalize_header_text(val) or f"column_{idx + 1}"
                    for idx, val in enumerate(raw_headers[:len(df_data.columns)])
                ]

            sheet_valid_count = 0
            for i in range(len(df_data)):
                raw_row = df_data.iloc[i]
                row_str_val = " ".join([str(x).lower() for x in raw_row.values if not pd.isna(x)])
                if not row_str_val or "total" in row_str_val or "grand total" in row_str_val:
                    continue

                t = NormalizedTransaction()
                t.source_file = os.path.basename(file_path)
                t.source_sheet = sheet_name
                t.source_row = int(df_data.index[i]) + 2
                t.stt_status_remark = stt_remark

                if 'isin' in mapping:
                    t.isin = clean_isin_code(raw_row.iloc[mapping['isin']])
                if 'security_name' in mapping:
                    t.security_name = clean_security_name(raw_row.iloc[mapping['security_name']])
                    if not t.isin:
                        t.isin = extract_isin_from_text(t.security_name)
                if 'sale_date' in mapping:
                    t.sale_date = clean_date(raw_row.iloc[mapping['sale_date']])
                if 'purchase_date' in mapping:
                    t.purchase_date = clean_date(raw_row.iloc[mapping['purchase_date']])
                if 'quantity' in mapping:
                    t.quantity = clean_number(raw_row.iloc[mapping['quantity']])
                elif is_mf_units_sheet:
                    t.quantity = 1.0
                if 'sale_consideration' in mapping:
                    t.sale_consideration = clean_number(raw_row.iloc[mapping['sale_consideration']])
                if 'purchase_cost' in mapping:
                    t.purchase_cost = clean_number(raw_row.iloc[mapping['purchase_cost']])
                if 'stt' in mapping:
                    t.stt = clean_number(raw_row.iloc[mapping['stt']])
                if 'gain_class' in mapping:
                    raw_cls = str(raw_row.iloc[mapping['gain_class']]).strip().upper()
                    if "LONG" in raw_cls or "LTCG" in raw_cls:
                        t.original_classification = "LTCG"
                    elif "SHORT" in raw_cls or "STCG" in raw_cls:
                        t.original_classification = "STCG"
                    else:
                        t.original_classification = raw_cls

                t.original_row_data = {str(k): v for k, v in zip(df_data.columns, raw_row.values)}

                if t.quantity > 0 and t.sale_date is not None:
                    if t.isin or t.security_name:
                        transactions.append(t)
                        sheet_valid_count += 1

            log_callback(f"Sheet '{sheet_name}' contributed {sheet_valid_count} valid rows.")

        log_callback(f"Successfully loaded {len(transactions)} transaction records from {file_label}.")

    except Exception as e:
        log_callback(f"ERROR reading workbook {file_label}: {str(e)}")
        log_callback(traceback.format_exc())
        raise e

    return transactions


# =====================================================================
# 5. ISIN AUTO-FILL FOR BLANK BROKER ISINs
# =====================================================================
_ISIN_FILL_LOG = []  # (Source Sheet, Source Row, Name, Filled ISIN, Match Method, Status)


def autofill_blank_isins(broker_txns, isin_master, log_callback):
    if not isin_master.loaded:
        return
    _ISIN_FILL_LOG.clear()
    filled, flagged = 0, 0

    for t in broker_txns:
        if t.isin:
            continue
        name = t.security_name
        core = normalize_company_core(name)
        asset_hint_mf = any(k in name.upper() for k in ("FUND", "SCHEME", "MF", "MUTUAL"))

        if not asset_hint_mf:
            # Try equity: exact symbol first, then fuzzy company name
            token = re.sub(r'[^A-Z0-9]', '', name.upper())
            isin = isin_master.symbol_to_isin.get(token)
            if isin:
                t.isin = isin
                t.remarks = (t.remarks + "; " if t.remarks else "") + f"ISIN auto-filled via symbol match ({token})."
                _ISIN_FILL_LOG.append((t.source_sheet, t.source_row, name, isin, "Symbol Match", "Filled"))
                filled += 1
                continue

            best_name, best_isin, score = isin_master.best_fuzzy_equity_name(core)
            if best_isin:
                t.isin = best_isin
                t.remarks = (t.remarks + "; " if t.remarks else "") + f"ISIN auto-filled via fuzzy company-name match ({score:.2f})."
                _ISIN_FILL_LOG.append((t.source_sheet, t.source_row, name, best_isin, f"Fuzzy Name ({score:.2f})", "Filled"))
                filled += 1
                continue

            t.remarks = (t.remarks + "; " if t.remarks else "") + "ISIN not found - Manual Review Required."
            _ISIN_FILL_LOG.append((t.source_sheet, t.source_row, name, "", "None", "Not Matched - Manual Review"))
            flagged += 1
        else:
            name_upper = name.upper()
            if any(k in name_upper for k in ("IDCW", "DIVIDEND", "REINVEST", "PAYOUT")) and "GROWTH" not in name_upper:
                plan_hint = "div"
            elif "GROWTH" in name_upper:
                plan_hint = "growth"
            else:
                plan_hint = None

            result = isin_master.best_fuzzy_mf(core, plan_hint)
            if result == "AMBIGUOUS" or result is None:
                status = "Ambiguous - Manual Review" if result == "AMBIGUOUS" else "Not Matched - Manual Review"
                t.remarks = (t.remarks + "; " if t.remarks else "") + f"MF ISIN {status} - not guessed."
                _ISIN_FILL_LOG.append((t.source_sheet, t.source_row, name, "", "MF Fuzzy", status))
                flagged += 1
            else:
                plan_type, core_name, isin, score = result
                t.isin = isin
                t.remarks = (t.remarks + "; " if t.remarks else "") + f"ISIN auto-filled via MF fuzzy match ({plan_type}, {score:.2f})."
                _ISIN_FILL_LOG.append((t.source_sheet, t.source_row, name, isin, f"MF Fuzzy {plan_type} ({score:.2f})", "Filled"))
                filled += 1

    log_callback(f"ISIN auto-fill complete: {filled} filled, {flagged} flagged for manual review.")


def append_isin_fill_log_sheet(wb, styler):
    if not _ISIN_FILL_LOG:
        return
    ws = wb.create_sheet(title="ISIN Auto-Fill Log")
    styler.write_back_link(ws)
    headers = ["Source Sheet", "Source Row", "Broker Security/Scheme Name", "Filled ISIN", "Match Method", "Status"]
    header_row = styler.style_header(ws, headers)

    status_colors = {
        "Filled": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "Ambiguous - Manual Review": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Not Matched - Manual Review": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    }
    row_num = header_row + 1
    for record in _ISIN_FILL_LOG:
        ws.append(list(record))
        status = record[-1]
        fill = status_colors.get(status)
        if fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_idx).fill = fill
        row_num += 1

    if row_num > header_row + 1:
        styler.format_rows(ws, header_row + 1, row_num - 1, len(headers))
    styler.autofit_columns(ws)


# =====================================================================
# 6. DUPLICATE & TRANSACTION MATCHING ENGINE
# =====================================================================
def detect_ais_duplicates(ais_txns, broker_txns, log_callback):
    log_callback("Running AIS duplicate detection analysis...")
    broker_counts = {}
    for t in broker_txns:
        key = (t.isin, t.sale_date, round(t.quantity, 4), round(t.sale_consideration, 2))
        broker_counts[key] = broker_counts.get(key, 0) + 1

    ais_groups = {}
    for t in ais_txns:
        key = (t.isin, t.sale_date, round(t.quantity, 4), round(t.sale_consideration, 2))
        ais_groups.setdefault(key, []).append(t)

    duplicates_list = []
    dup_group_id = 1
    exact_duplicates_count = 0
    for key, group in ais_groups.items():
        ais_count = len(group)
        if ais_count > 1:
            broker_count = broker_counts.get(key, 0)
            if ais_count > broker_count:
                for idx in range(broker_count, ais_count):
                    dup_txn = group[idx]
                    dup_txn.is_duplicate = True
                    dup_txn.remarks = f"Duplicate Entry: Found {ais_count} occurrences in AIS but only {broker_count} in Broker data."
                    duplicates_list.append((dup_group_id, dup_txn, ais_count, broker_count))
                    exact_duplicates_count += 1
                dup_group_id += 1

    log_callback(f"Duplicate detection completed: Flagged {exact_duplicates_count} duplicates in AIS.")
    return duplicates_list


def match_and_reconcile(ais_txns, broker_txns, isin_master, log_callback):
    log_callback("Starting transaction matching engine...")
    matching_ais = [t for t in ais_txns if not t.is_duplicate]
    matching_broker = list(broker_txns)
    matched_pairs = []
    match_counter = 1

    def qty_ok(a, b):
        return abs(a.quantity - b.quantity) < 1e-4

    def value_ok(a, b):
        return abs(a.sale_consideration - b.sale_consideration) <= sale_cost_tolerance(
            a.sale_consideration, b.sale_consideration
        )

    def id_match(a, b):
        if a.isin and b.isin:
            return a.isin == b.isin
        return names_match(a.security_name, b.security_name, a.isin, b.isin, isin_master)

    passes = [
        ("Strict Match", lambda a, b: id_match(a, b) and a.sale_date == b.sale_date and qty_ok(a, b) and value_ok(a, b)),
        ("Date Window Match", lambda a, b: id_match(a, b) and a.sale_date and b.sale_date and
            abs((a.sale_date - b.sale_date).days) <= DATE_WINDOW_DAYS and qty_ok(a, b) and value_ok(a, b)),
        ("Description Match", lambda a, b: names_match(a.security_name, b.security_name, a.isin, b.isin, isin_master)
            and a.sale_date == b.sale_date and qty_ok(a, b) and value_ok(a, b)),
        ("Soft Match", lambda a, b: names_match(a.security_name, b.security_name, a.isin, b.isin, isin_master)
            and a.sale_date and b.sale_date and abs((a.sale_date - b.sale_date).days) <= DATE_WINDOW_DAYS
            and qty_ok(a, b) and value_ok(a, b)),
        ("Tolerance Value Match", lambda a, b: id_match(a, b) and a.sale_date and b.sale_date and
            abs((a.sale_date - b.sale_date).days) <= DATE_WINDOW_DAYS and qty_ok(a, b) and
            (abs(a.sale_consideration - b.sale_consideration) / max(a.sale_consideration, 1.0)) <= 0.01),
    ]

    unmatched_ais = matching_ais
    for pass_label, pass_fn in passes:
        log_callback(f"Matching Pass: {pass_label}...")
        next_unmatched = []
        for a in unmatched_ais:
            found_idx = None
            for i, b in enumerate(matching_broker):
                if pass_fn(a, b):
                    found_idx = i
                    break
            if found_idx is not None:
                b_match = matching_broker.pop(found_idx)
                a.match_id = f"M-{match_counter:04d}"
                b_match.match_id = a.match_id
                matched_pairs.append((a, b_match, pass_label))
                match_counter += 1
            else:
                next_unmatched.append(a)
        unmatched_ais = next_unmatched

    log_callback(f"Matching finished: Reconciled {len(matched_pairs)} transactions successfully.")
    log_callback(f"Unmatched AIS records remaining: {len(unmatched_ais)}")
    log_callback(f"Unmatched Broker records remaining: {len(matching_broker)}")

    return matched_pairs, unmatched_ais, matching_broker


def material_pairs(matched_pairs):
    """Matched pairs whose sale-value gap exceeds the dynamic nominal tolerance."""
    result = []
    for a, b, mode in matched_pairs:
        tol = sale_cost_tolerance(a.sale_consideration, b.sale_consideration)
        if abs(a.sale_consideration - b.sale_consideration) > tol:
            result.append((a, b, mode))
    return result


def nominal_variance_pairs(matched_pairs):
    """Matched pairs with a small non-zero sale-value gap absorbed as brokerage/charges."""
    result = []
    for a, b, mode in matched_pairs:
        diff = abs(a.sale_consideration - b.sale_consideration)
        tol = sale_cost_tolerance(a.sale_consideration, b.sale_consideration)
        if 0 < diff <= tol:
            result.append((a, b, mode))
    return result


# =====================================================================
# 7. REPORT GENERATOR (OPENPYXL WRAPPER)
# =====================================================================
class RecoReportGenerator:
    def __init__(self, output_path, fy, ais_file, broker_file):
        self.output_path = output_path
        self.fy = fy
        self.ais_file = os.path.basename(ais_file)
        self.broker_file = os.path.basename(broker_file)
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

        self.COLOR_INDIGO_HEADER = "1E3A8A"
        self.COLOR_LIGHT_ZEBRA = "F8FAFC"
        self.COLOR_GREEN_MATCH = "E8F5E9"
        self.COLOR_RED_DIFF = "FFEBEE"
        self.COLOR_YELLOW_WARN = "FFFDE7"

        self.FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        self.FONT_BODY = Font(name="Segoe UI", size=10)
        self.FONT_BOLD = Font(name="Segoe UI", size=10, bold=True)
        self.FONT_LINK = Font(name="Segoe UI", size=10, bold=True, color="1D4ED8", underline="single")

        self.ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
        self.ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
        self.ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

        self.BORDER_THIN = Border(
            left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
        )

        # sheet titles known ahead of time, for Summary hyperlinks
        self.SHEET_TAX_OUTPUT = "Tax Software Output"
        self.SHEET_LTCG_DIFF = "Difference in Long Term"
        self.SHEET_STCG_DIFF = "Difference in Short Term"
        self.SHEET_DUPLICATES = "Duplicate Entries in AIS"
        self.SHEET_MISSING_IN_AIS = "Missing Entries in AIS"
        self.SHEET_MISSING_IN_BROKER = "Missing Entries in Broker"
        self.SHEET_OTHER_DIFF = "Other Financial Differences"
        self.SHEET_IGNORED_VAR = "Ignored Sales Variances"
        self.SHEET_MATCHED = "Matched Transactions"
        self.SHEET_REVIEW = "Review Required"

    # ---------- shared helpers ----------
    def write_back_link(self, ws):
        cell = ws.cell(row=1, column=1, value="\u2b05 Back to Summary")
        cell.font = self.FONT_LINK
        cell.hyperlink = "#'Summary'!A1"
        ws.row_dimensions[1].height = 18

    def style_header(self, ws, columns):
        ws.append(columns)
        header_row = ws.max_row
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = self.FONT_HEADER
            cell.fill = PatternFill(start_color=self.COLOR_INDIGO_HEADER, end_color=self.COLOR_INDIGO_HEADER, fill_type="solid")
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        ws.row_dimensions[header_row].height = 28
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
        return header_row

    def format_rows(self, ws, start_row, end_row, num_cols):
        for row in range(start_row, end_row + 1):
            ws.row_dimensions[row].height = 20
            fill_color = self.COLOR_LIGHT_ZEBRA if row % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.FONT_BODY
                cell.border = self.BORDER_THIN
                if cell.fill.fill_type is None:
                    cell.fill = row_fill

    def autofit_columns(self, ws):
        ws.views.sheetView[0].showGridLines = True
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format and ('#,##0.00' in cell.number_format or '%' in cell.number_format):
                    val += "   "
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    def add_nav_link(self, ws, row, col, sheet_title, label=None):
        label = label or f"Open {sheet_title} \u279c"
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = self.FONT_LINK
        cell.hyperlink = f"#'{sheet_title}'!A1"

    # ---------- Summary ----------
    def write_summary_sheet(self, summary_data, dashboard_rows):
        ws = self.wb.create_sheet(title="Summary")
        ws.views.sheetView[0].showGridLines = True

        ws.cell(row=1, column=1, value="RECONCILIATION SUMMARY REPORT").font = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
        ws.row_dimensions[1].height = 30

        meta = [
            ("Application:", "AIS BROKER RECO (Master v5)"),
            ("Selected Financial Year:", self.fy),
            ("AIS Source Workbook:", self.ais_file),
            ("Broker Source Workbook:", self.broker_file),
            ("Reconciliation Date:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        row_idx = 3
        for k, v in meta:
            ws.cell(row=row_idx, column=1, value=k).font = self.FONT_BOLD
            ws.cell(row=row_idx, column=2, value=v).font = self.FONT_BODY
            row_idx += 1

        row_idx += 2
        ws.cell(row=row_idx, column=1, value="RECONCILIATION METRICS").font = Font(name="Segoe UI", size=12, bold=True, color="1E3A8A")
        row_idx += 1

        metrics_headers = ["Metric Description", "Record Count / Value", "Navigate"]
        for col, h in enumerate(metrics_headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=h)
            cell.font = self.FONT_HEADER
            cell.fill = PatternFill(start_color=self.COLOR_INDIGO_HEADER, end_color=self.COLOR_INDIGO_HEADER, fill_type="solid")
            cell.alignment = self.ALIGN_CENTER

        metrics = [
            ("Total AIS Records Loaded", summary_data['total_ais'], None),
            ("Total Broker Records Loaded", summary_data['total_broker'], None),
            ("Successfully Matched Records", summary_data['matched'], self.SHEET_MATCHED),
            ("Duplicate Records in AIS", summary_data['duplicates'], self.SHEET_DUPLICATES),
            ("Missing Records in AIS (Broker only)", summary_data['missing_in_ais'], self.SHEET_MISSING_IN_AIS),
            ("Missing Records in Broker (AIS only)", summary_data['missing_in_broker'], self.SHEET_MISSING_IN_BROKER),
            ("Long-Term Capital Gain Differences", summary_data['ltcg_diffs'], self.SHEET_LTCG_DIFF),
            ("Short-Term Capital Gain Differences", summary_data['stcg_diffs'], self.SHEET_STCG_DIFF),
            ("Other Financial Differences", summary_data['other_diffs'], self.SHEET_OTHER_DIFF),
            ("Ignored Minor Sales Variances", summary_data['ignored_sales_variances'], self.SHEET_IGNORED_VAR),
            ("Records Requiring Manual Review", summary_data['review_required'], self.SHEET_REVIEW),
            ("Tax Software Output Rows", summary_data['matched'], self.SHEET_TAX_OUTPUT),
        ]

        for k, v, sheet_link in metrics:
            row_idx += 1
            c1 = ws.cell(row=row_idx, column=1, value=k)
            c2 = ws.cell(row=row_idx, column=2, value=v)
            c1.font = self.FONT_BODY
            c1.border = self.BORDER_THIN
            c2.font = self.FONT_BOLD
            c2.border = self.BORDER_THIN
            c2.alignment = self.ALIGN_CENTER
            c2.number_format = '#,##0'

            if k == "Successfully Matched Records":
                c2.fill = PatternFill(start_color=self.COLOR_GREEN_MATCH, end_color=self.COLOR_GREEN_MATCH, fill_type="solid")
            elif k not in ("Total AIS Records Loaded", "Total Broker Records Loaded") and v > 0:
                c2.fill = PatternFill(start_color=self.COLOR_RED_DIFF, end_color=self.COLOR_RED_DIFF, fill_type="solid")

            c3 = ws.cell(row=row_idx, column=3)
            c3.border = self.BORDER_THIN
            if sheet_link and sheet_link in [ws2.title for ws2 in self.wb.worksheets] + [sheet_link]:
                self.add_nav_link(ws, row_idx, 3, sheet_link, label=f"Open {sheet_link} \u279c")

        # ---- Value Dashboard ----
        row_idx += 3
        ws.cell(row=row_idx, column=1, value="VALUE DASHBOARD (AIS vs BROKER)").font = Font(name="Segoe UI", size=12, bold=True, color="1E3A8A")
        row_idx += 1
        dash_headers = ["Reconciliation Category", "Value as per AIS", "Value as per Broker Report", "Actual Difference"]
        for col, h in enumerate(dash_headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=h)
            cell.font = self.FONT_HEADER
            cell.fill = PatternFill(start_color=self.COLOR_INDIGO_HEADER, end_color=self.COLOR_INDIGO_HEADER, fill_type="solid")
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN

        for label, ais_value, broker_value in dashboard_rows:
            row_idx += 1
            difference = ais_value - broker_value
            values = [label, ais_value, broker_value, difference]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.BORDER_THIN
                cell.font = self.FONT_BODY
                if col > 1:
                    cell.number_format = '#,##0.00;[Red]-#,##0.00'
                    cell.alignment = self.ALIGN_RIGHT

        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20

    # ---------- Tax Software Output ----------
    def write_tax_software_output(self, tax_rows):
        ws = self.wb.create_sheet(title=self.SHEET_TAX_OUTPUT)
        self.write_back_link(ws)
        headers = [
            "Match ID", "Security", "ISIN", "Transaction Date", "Purchase Date", "Quantity",
            "Sale Consideration", "Purchase Cost", "Capital Gain / (Loss)",
            "Is it LTCG ?", "Basis of Classification", "STT Status Remark"
        ]
        header_row = self.style_header(ws, headers)
        row_num = header_row + 1
        for r in tax_rows:
            gain = r["sale_consideration"] - r["purchase_cost"]
            ws.append([
                r["match_id"], r["security_name"], r["isin"], r["sale_date"], r["purchase_date"],
                r["quantity"], r["sale_consideration"], r["purchase_cost"], gain,
                r["is_ltcg"], r["basis"], r["stt_remark"]
            ])
            ws.cell(row=row_num, column=4).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_num, column=5).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            ws.cell(row=row_num, column=7).number_format = '#,##0.00'
            ws.cell(row=row_num, column=8).number_format = '#,##0.00'
            ws.cell(row=row_num, column=9).number_format = '#,##0.00'
            ws.cell(row=row_num, column=1).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=3).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=10).alignment = self.ALIGN_CENTER
            if r["is_ltcg"] == "Yes":
                fill = PatternFill(start_color=self.COLOR_GREEN_MATCH, end_color=self.COLOR_GREEN_MATCH, fill_type="solid")
                ws.cell(row=row_num, column=10).fill = fill
            elif r["is_ltcg"] == "Review Required":
                fill = PatternFill(start_color=self.COLOR_YELLOW_WARN, end_color=self.COLOR_YELLOW_WARN, fill_type="solid")
                ws.cell(row=row_num, column=10).fill = fill
            row_num += 1

        if row_num > header_row + 1:
            self.format_rows(ws, header_row + 1, row_num - 1, len(headers))
        self.autofit_columns(ws)

    # ---------- Difference sheets (no STT amount / holding-days columns) ----------
    def write_difference_sheet(self, title, matched_pairs_diffs):
        ws = self.wb.create_sheet(title=title)
        self.write_back_link(ws)
        headers = [
            "Match ID", "Security", "ISIN", "Transaction Date", "Purchase Date", "Quantity",
            "AIS Class", "Broker Class", "Calc Class", "AIS Sale Value", "Broker Sale Value",
            "Diff Sale Value", "AIS Cost", "Broker Cost", "Diff Cost",
            "Difference Type", "Remarks", "Source Sheet", "Source Row"
        ]
        header_row = self.style_header(ws, headers)
        row_num = header_row + 1
        for ais, broker, diff_type, remarks in matched_pairs_diffs:
            diff_sale = ais.sale_consideration - broker.sale_consideration
            diff_cost = ais.purchase_cost - broker.purchase_cost
            row_values = [
                ais.match_id, ais.security_name, ais.isin, ais.sale_date, broker.purchase_date, ais.quantity,
                ais.original_classification, broker.original_classification, ais.calculated_classification,
                ais.sale_consideration, broker.sale_consideration, diff_sale,
                ais.purchase_cost, broker.purchase_cost, diff_cost,
                diff_type, remarks, ais.source_sheet, ais.source_row
            ]
            ws.append(row_values)

            ws.cell(row=row_num, column=4).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_num, column=5).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            for c in (10, 11, 12, 13, 14, 15):
                ws.cell(row=row_num, column=c).number_format = '#,##0.00'

            ws.cell(row=row_num, column=1).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=3).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=4).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=5).alignment = self.ALIGN_CENTER

            red_fill = PatternFill(start_color=self.COLOR_RED_DIFF, end_color=self.COLOR_RED_DIFF, fill_type="solid")
            if ais.original_classification != broker.original_classification or ais.calculated_classification != broker.original_classification:
                for c in (7, 8, 9):
                    ws.cell(row=row_num, column=c).fill = red_fill

            tol = sale_cost_tolerance(ais.sale_consideration, broker.sale_consideration)
            if abs(diff_sale) > tol:
                for c in (10, 11, 12):
                    ws.cell(row=row_num, column=c).fill = red_fill
            if abs(diff_cost) > TOLERANCE_AMOUNT:
                for c in (13, 14, 15):
                    ws.cell(row=row_num, column=c).fill = red_fill

            row_num += 1

        if row_num > header_row + 1:
            self.format_rows(ws, header_row + 1, row_num - 1, len(headers))
        self.autofit_columns(ws)

    def write_duplicates_sheet(self, duplicate_groups):
        ws = self.wb.create_sheet(title=self.SHEET_DUPLICATES)
        self.write_back_link(ws)
        headers = ["Dup Group Ref", "Security", "ISIN", "Transaction Date", "Quantity",
                   "Sale Consideration", "Source Row", "Source Sheet", "Remarks"]
        header_row = self.style_header(ws, headers)
        row_num = header_row + 1
        for dup_id, t, ais_count, broker_count in duplicate_groups:
            ws.append([f"DUP-{dup_id:03d}", t.security_name, t.isin, t.sale_date, t.quantity,
                       t.sale_consideration, t.source_row, t.source_sheet, t.remarks])
            ws.cell(row=row_num, column=1).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=3).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=4).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=5).number_format = '#,##0.00'
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            yellow_fill = PatternFill(start_color=self.COLOR_YELLOW_WARN, end_color=self.COLOR_YELLOW_WARN, fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = yellow_fill
            row_num += 1
        if row_num > header_row + 1:
            self.format_rows(ws, header_row + 1, row_num - 1, len(headers))
        self.autofit_columns(ws)

    def write_missing_sheet(self, title, transactions_list):
        ws = self.wb.create_sheet(title=title)
        self.write_back_link(ws)
        headers = ["Source Row", "Source Sheet", "Security", "ISIN", "Transaction Date",
                   "Quantity", "Sale Consideration", "Purchase Date", "Purchase Cost", "Reported Classification"]
        header_row = self.style_header(ws, headers)
        row_num = header_row + 1
        for t in transactions_list:
            ws.append([t.source_row, t.source_sheet, t.security_name, t.isin, t.sale_date,
                       t.quantity, t.sale_consideration, t.purchase_date, t.purchase_cost, t.original_classification])
            ws.cell(row=row_num, column=1).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=4).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=5).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            ws.cell(row=row_num, column=7).number_format = '#,##0.00'
            ws.cell(row=row_num, column=8).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_num, column=9).number_format = '#,##0.00'
            red_fill = PatternFill(start_color=self.COLOR_RED_DIFF, end_color=self.COLOR_RED_DIFF, fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = red_fill
            row_num += 1
        if row_num > header_row + 1:
            self.format_rows(ws, header_row + 1, row_num - 1, len(headers))
        self.autofit_columns(ws)

    def write_matched_sheet(self, matched_pairs):
        ws = self.wb.create_sheet(title=self.SHEET_MATCHED)
        self.write_back_link(ws)
        headers = ["Match ID", "Security", "ISIN", "Transaction Date", "Quantity",
                   "AIS Sale Value", "Broker Sale Value", "AIS Cost", "Broker Cost",
                   "AIS Class", "Broker Class", "Match Mode"]
        header_row = self.style_header(ws, headers)
        row_num = header_row + 1
        for ais, broker, match_mode in matched_pairs:
            ws.append([ais.match_id, ais.security_name, ais.isin, ais.sale_date, ais.quantity,
                       ais.sale_consideration, broker.sale_consideration,
                       ais.purchase_cost, broker.purchase_cost,
                       ais.original_classification, broker.original_classification, match_mode])
            for c in (5, 6, 7, 8, 9):
                ws.cell(row=row_num, column=c).number_format = '#,##0.00'
            ws.cell(row=row_num, column=1).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=3).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=4).alignment = self.ALIGN_CENTER

            diff_sale = abs(ais.sale_consideration - broker.sale_consideration)
            diff_cost = abs(ais.purchase_cost - broker.purchase_cost)
            tol = sale_cost_tolerance(ais.sale_consideration, broker.sale_consideration)
            class_match = (ais.original_classification == broker.original_classification)
            if diff_sale <= tol and diff_cost <= TOLERANCE_AMOUNT and class_match:
                green_fill = PatternFill(start_color=self.COLOR_GREEN_MATCH, end_color=self.COLOR_GREEN_MATCH, fill_type="solid")
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = green_fill
            row_num += 1
        if row_num > header_row + 1:
            self.format_rows(ws, header_row + 1, row_num - 1, len(headers))
        self.autofit_columns(ws)

    def write_review_sheet(self, review_txns):
        ws = self.wb.create_sheet(title=self.SHEET_REVIEW)
        self.write_back_link(ws)
        headers = ["Source File", "Source Row", "Source Sheet", "Security", "ISIN",
                   "Transaction Date", "Quantity", "Sale Consideration", "Purchase Date", "Remarks"]
        header_row = self.style_header(ws, headers)
        row_num = header_row + 1
        for t in review_txns:
            ws.append([t.source_file, t.source_row, t.source_sheet, t.security_name, t.isin,
                       t.sale_date, t.quantity, t.sale_consideration, t.purchase_date, t.remarks])
            ws.cell(row=row_num, column=2).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=5).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=6).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=7).number_format = '#,##0.00'
            ws.cell(row=row_num, column=8).number_format = '#,##0.00'
            ws.cell(row=row_num, column=9).alignment = self.ALIGN_CENTER
            yellow_fill = PatternFill(start_color=self.COLOR_YELLOW_WARN, end_color=self.COLOR_YELLOW_WARN, fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = yellow_fill
            row_num += 1
        if row_num > header_row + 1:
            self.format_rows(ws, header_row + 1, row_num - 1, len(headers))
        self.autofit_columns(ws)

    def save(self):
        self.wb.save(self.output_path)


# =====================================================================
# 8. APP CONTROLLER / CORE RECONCILIATION PROCESSOR
# =====================================================================
def run_reconciliation_workflow(ais_file, broker_file, output_path, fy, progress_callback, log_callback):
    try:
        progress_callback(5, "Initialising reconciliation run...")
        log_callback("----------------------------------------------------------------------")
        log_callback(f"Starting AIS BROKER RECO Engine (Master v5) for Financial Year: {fy}")
        log_callback("----------------------------------------------------------------------")

        tax_engine = TaxRuleEngine(fy)
        isin_master = ISINMasterLookup(log_callback)

        progress_callback(10, "Loading AIS data workbook...")
        ais_raw_list = load_and_normalize_workbook(ais_file, "AIS Data File", log_callback)

        progress_callback(25, "Loading Broker data workbook...")
        broker_raw_list = load_and_normalize_workbook(broker_file, "Broker Data File", log_callback)

        if not ais_raw_list:
            raise ValueError("Zero valid transaction entries extracted from AIS file. Process stopped.")
        if not broker_raw_list:
            raise ValueError("Zero valid transaction entries extracted from Broker file. Process stopped.")

        progress_callback(35, "Auto-filling blank Broker ISIN codes (if master available)...")
        autofill_blank_isins(broker_raw_list, isin_master, log_callback)

        progress_callback(45, "Analysing AIS duplicate entries...")
        duplicate_groups = detect_ais_duplicates(ais_raw_list, broker_raw_list, log_callback)

        progress_callback(58, "Matching transactions between sheets...")
        matched_pairs, unmatched_ais, unmatched_broker = match_and_reconcile(
            ais_raw_list, broker_raw_list, isin_master, log_callback
        )

        progress_callback(70, "Evaluating Tax Classifications and differences...")

        ltcg_mismatches, stcg_mismatches, other_differences = [], [], []
        ignored_sales_variances, review_required_txns, tax_output_rows = [], [], []

        for t in unmatched_ais + unmatched_broker:
            if not t.isin and not t.security_name:
                t.remarks = "Empty security identifier and ISIN."
                review_required_txns.append(t)
            elif t.quantity <= 0 or t.sale_consideration <= 0:
                t.remarks = "Invalid numeric parameters (quantity/value <= 0)."
                review_required_txns.append(t)

        for ais, broker, match_mode in matched_pairs:
            purchase_date = broker.purchase_date if broker.purchase_date else ais.purchase_date
            calc_cls, h_days, tax_remarks = tax_engine.classify_transaction(
                ais.isin, ais.security_name, purchase_date, ais.sale_date
            )
            ais.calculated_classification = calc_cls
            broker.calculated_classification = calc_cls
            ais.holding_days = h_days
            broker.holding_days = h_days

            ais_cls_norm = ais.original_classification if ais.original_classification in ("LTCG", "STCG") else None
            broker_cls_norm = broker.original_classification if broker.original_classification in ("LTCG", "STCG") else None

            stt_remark = ais.stt_status_remark or broker.stt_status_remark or "STT Paid: Not determinable from source-sheet name"

            diff_remarks = [f"Tax basis: {tax_remarks}.", stt_remark]
            diff_sale = abs(ais.sale_consideration - broker.sale_consideration)
            diff_cost = abs(ais.purchase_cost - broker.purchase_cost)
            dyn_tol = sale_cost_tolerance(ais.sale_consideration, broker.sale_consideration)

            is_class_mismatch = False
            is_amount_mismatch = False

            if ais_cls_norm and broker_cls_norm and ais_cls_norm != broker_cls_norm:
                is_class_mismatch = True
                diff_remarks.append(f"Classification Mismatch: AIS={ais_cls_norm}, Broker={broker_cls_norm}.")

            if broker_cls_norm and calc_cls != "Review Required" and broker_cls_norm != calc_cls:
                is_class_mismatch = True
                diff_remarks.append(f"Holding Period Rule Discrepancy: Broker classified as {broker_cls_norm} but rule indicates {calc_cls}.")

            if calc_cls == "Review Required":
                diff_remarks.append(f"Tax calculation review: {tax_remarks}")
                review_required_txns.append(ais)

            # Qty-matched dynamic tolerance: ignore nominal Sale Value gap from brokerage/charges
            qty_matches = abs(ais.quantity - broker.quantity) < 1e-4
            if diff_sale > dyn_tol:
                is_amount_mismatch = True
                diff_remarks.append(
                    f"Sale Consideration Mismatch: Diff={diff_sale:.2f} "
                    f"(AIS={ais.sale_consideration:.2f}, Broker={broker.sale_consideration:.2f})."
                )
            elif diff_sale > 0 and qty_matches:
                diff_remarks.append(
                    f"Ignored nominal Sale Value variance (Qty matches; within Rs {dyn_tol:.2f} "
                    f"transaction-cost tolerance): Diff={diff_sale:.2f} "
                    f"(AIS gross={ais.sale_consideration:.2f}, Broker net={broker.sale_consideration:.2f})."
                )

            if diff_cost > TOLERANCE_AMOUNT:
                is_amount_mismatch = True
                diff_remarks.append(f"Acquisition Cost Mismatch: Diff={diff_cost:.2f} (AIS={ais.purchase_cost:.2f}, Broker={broker.purchase_cost:.2f}).")

            final_remarks = "; ".join(diff_remarks)
            ais.remarks = final_remarks
            broker.remarks = final_remarks

            if is_class_mismatch:
                target = ltcg_mismatches if (calc_cls == "LTCG" or broker_cls_norm == "LTCG") else stcg_mismatches
                target.append((ais, broker, "Classification Mismatch", final_remarks))
            elif is_amount_mismatch:
                other_differences.append((ais, broker, "Amount Mismatch", final_remarks))
            elif diff_sale > 0 and qty_matches:
                ignored_sales_variances.append((ais, broker, "Ignored Transaction-Cost Variance", final_remarks))

            tax_output_rows.append({
                "match_id": ais.match_id,
                "security_name": ais.security_name,
                "isin": ais.isin,
                "sale_date": ais.sale_date,
                "purchase_date": purchase_date,
                "quantity": ais.quantity,
                "sale_consideration": broker.sale_consideration if broker.sale_consideration else ais.sale_consideration,
                "purchase_cost": broker.purchase_cost if broker.purchase_cost else ais.purchase_cost,
                "is_ltcg": ltcg_yes_no(calc_cls),
                "basis": tax_remarks,
                "stt_remark": stt_remark,
            })

        progress_callback(85, "Generating Reconciliation Excel Workbook sheets...")

        # ---- Value dashboard computation ----
        total_ais_value = sum(t.sale_consideration for t in ais_raw_list)
        total_broker_value = sum(t.sale_consideration for t in broker_raw_list)
        mat_pairs = material_pairs(matched_pairs)
        nom_pairs = nominal_variance_pairs(matched_pairs)
        matched_ais_value = sum(a.sale_consideration for a, b, _ in matched_pairs)
        matched_broker_value = sum(b.sale_consideration for a, b, _ in matched_pairs)
        material_ais_value = sum(a.sale_consideration for a, b, _ in mat_pairs)
        material_broker_value = sum(b.sale_consideration for a, b, _ in mat_pairs)
        nominal_ais_value = sum(a.sale_consideration for a, b, _ in nom_pairs)
        nominal_broker_value = sum(b.sale_consideration for a, b, _ in nom_pairs)
        missing_in_broker_ais_value = sum(t.sale_consideration for t in unmatched_ais)
        missing_in_ais_broker_value = sum(t.sale_consideration for t in unmatched_broker)

        dashboard_rows = [
            ("Matched Transactions (Gross)", matched_ais_value, matched_broker_value),
            ("Material Differences (Above Tolerance)", material_ais_value, material_broker_value),
            ("Ignored Transaction-Cost Variances (Within Tolerance)", nominal_ais_value, nominal_broker_value),
            ("Missing in Broker (Present only in AIS)", missing_in_broker_ais_value, 0.0),
            ("Missing in AIS (Present only in Broker)", 0.0, missing_in_ais_broker_value),
            ("Grand Total (All Records)", total_ais_value, total_broker_value),
        ]

        summary_stats = {
            'total_ais': len(ais_raw_list),
            'total_broker': len(broker_raw_list),
            'matched': len(matched_pairs),
            'duplicates': len(duplicate_groups),
            'missing_in_ais': len(unmatched_broker),
            'missing_in_broker': len(unmatched_ais),
            'ltcg_diffs': len(ltcg_mismatches),
            'stcg_diffs': len(stcg_mismatches),
            'ignored_sales_variances': len(ignored_sales_variances),
            'other_diffs': len(other_differences),
            'review_required': len(review_required_txns)
        }

        rep = RecoReportGenerator(output_path, fy, ais_file, broker_file)
        rep.write_summary_sheet(summary_stats, dashboard_rows)
        rep.write_tax_software_output(tax_output_rows)
        rep.write_difference_sheet(rep.SHEET_LTCG_DIFF, ltcg_mismatches)
        rep.write_difference_sheet(rep.SHEET_STCG_DIFF, stcg_mismatches)
        rep.write_duplicates_sheet(duplicate_groups)
        rep.write_missing_sheet(rep.SHEET_MISSING_IN_AIS, unmatched_broker)
        rep.write_missing_sheet(rep.SHEET_MISSING_IN_BROKER, unmatched_ais)
        rep.write_difference_sheet(rep.SHEET_OTHER_DIFF, other_differences)
        rep.write_difference_sheet(rep.SHEET_IGNORED_VAR, ignored_sales_variances)
        rep.write_matched_sheet(matched_pairs)
        rep.write_review_sheet(review_required_txns)
        append_isin_fill_log_sheet(rep.wb, rep)

        rep.save()

        log_dir = os.path.dirname(output_path)
        log_filename = os.path.splitext(os.path.basename(output_path))[0] + "_TECHNICAL_LOG.txt"
        tech_log_path = os.path.join(log_dir, log_filename)

        with open(tech_log_path, 'w', encoding='utf-8') as f_log:
            f_log.write("=== AIS BROKER RECO TECHNICAL LOG (Master v5) ===\n")
            f_log.write(f"Date of Reco: {datetime.datetime.now()}\n")
            f_log.write(f"Financial Year: {fy}\n")
            f_log.write(f"AIS Input: {ais_file}\n")
            f_log.write(f"Broker Input: {broker_file}\n")
            f_log.write(f"Reconciliation Output: {output_path}\n\n")
            for key, value in summary_stats.items():
                f_log.write(f"{key}: {value}\n")

        progress_callback(100, "Processing completed! Excel file saved successfully.")
        log_callback("----------------------------------------------------------------------")
        log_callback("SUCCESS: Reconciliation workbook created at:")
        log_callback(f" > {output_path}")
        log_callback(f"Technical log saved next to it at:")
        log_callback(f" > {tech_log_path}")
        log_callback("----------------------------------------------------------------------")

        return summary_stats

    except Exception as e:
        log_callback(f"FATAL ERROR during reconciliation: {str(e)}")
        log_callback(traceback.format_exc())
        progress_callback(100, f"Error: {str(e)}")
        raise e


# =====================================================================
# 9. TKINTER USER INTERFACE (IDLE COMPATIBLE)
# =====================================================================
class CanvasProgressBar(tk.Canvas):
    def __init__(self, parent, bg="#1e293b", fg="#6366f1", height=18, **kwargs):
        super().__init__(parent, bg=bg, height=height, highlightthickness=1, highlightbackground="#475569", bd=0, **kwargs)
        self.fg = fg
        self.progress = 0.0
        self.bind("<Configure>", self.draw)

    def set(self, val):
        self.progress = min(max(val, 0.0), 1.0)
        self.draw()

    def draw(self, event=None):
        self.delete("all")
        w = self.winfo_width()
        h = self.winfo_height()
        self.create_rectangle(2, 2, w - 2, h - 2, fill=self["bg"], width=0)
        if self.progress > 0:
            fill_width = int((w - 4) * self.progress)
            if fill_width > 0:
                self.create_rectangle(2, 2, 2 + fill_width, h - 2, fill=self.fg, width=0)
                pct_text = f"{int(self.progress * 100)}%"
                text_color = "#ffffff" if self.progress > 0.55 else "#a5b4fc"
                self.create_text(w / 2, h / 2, text=pct_text, fill=text_color, font=("Segoe UI", 9, "bold"))


class AISBrokerRecoApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("AIS BROKER RECO - Tax Capital Gains Engine (Master v5)")
        self.geometry("820x680")
        self.configure(bg="#0f172a")
        self.minsize(750, 600)

        self.ais_file_path = ""
        self.broker_file_path = ""
        self.is_running = False

        self.setup_ui()
        self.check_dependencies()

    def setup_ui(self):
        header_frame = tk.Frame(self, bg="#1e1b4b", height=90, bd=0)
        header_frame.pack(fill="x", side="top")
        header_frame.pack_propagate(False)

        tk.Label(header_frame, text="AIS BROKER RECO", font=("Segoe UI", 22, "bold"),
                  fg="#f8fafc", bg="#1e1b4b").pack(pady=(12, 1))
        tk.Label(header_frame, text="Capital Gains Reconciliation & Audit Tool (Income Tax Act, 1961) - Master v5",
                  font=("Segoe UI", 9, "italic"), fg="#c7d2fe", bg="#1e1b4b").pack()

        main_container = tk.Frame(self, bg="#0f172a")
        main_container.pack(fill="both", expand=True, padx=25, pady=20)

        files_frame = tk.Frame(main_container, bg="#1e293b", bd=1, relief="flat", padx=15, pady=15)
        files_frame.pack(fill="x", pady=(0, 15))

        tk.Label(files_frame, text="AIS Data Sheet (Excel):", font=("Segoe UI", 10, "bold"),
                  fg="#f1f5f9", bg="#1e293b").grid(row=0, column=0, sticky="w", pady=(0, 5))
        self.ais_entry = tk.Entry(files_frame, font=("Segoe UI", 10), bg="#334155", fg="#ffffff",
                                    insertbackground="#ffffff", bd=1, relief="solid")
        self.ais_entry.grid(row=1, column=0, columnspan=2, sticky="ew", ipady=5, padx=(0, 10))
        ais_btn = tk.Button(files_frame, text="Browse AIS File", font=("Segoe UI", 10, "bold"),
                              bg="#6366f1", fg="#ffffff", activebackground="#4f46e5", activeforeground="#ffffff",
                              bd=0, cursor="hand2", padx=15, command=self.select_ais_file)
        ais_btn.grid(row=1, column=2, ipady=4)

        tk.Label(files_frame, text="Broker Realised P&L Account (Excel):", font=("Segoe UI", 10, "bold"),
                  fg="#f1f5f9", bg="#1e293b").grid(row=2, column=0, sticky="w", pady=(12, 5))
        self.broker_entry = tk.Entry(files_frame, font=("Segoe UI", 10), bg="#334155", fg="#ffffff",
                                       insertbackground="#ffffff", bd=1, relief="solid")
        self.broker_entry.grid(row=3, column=0, columnspan=2, sticky="ew", ipady=5, padx=(0, 10))
        broker_btn = tk.Button(files_frame, text="Browse Broker", font=("Segoe UI", 10, "bold"),
                                 bg="#6366f1", fg="#ffffff", activebackground="#4f46e5", activeforeground="#ffffff",
                                 bd=0, cursor="hand2", padx=15, command=self.select_broker_file)
        broker_btn.grid(row=3, column=2, ipady=4)

        config_frame = tk.Frame(files_frame, bg="#1e293b")
        config_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(15, 0))
        tk.Label(config_frame, text="Financial Year / Tax Rules:", font=("Segoe UI", 10, "bold"),
                  fg="#f1f5f9", bg="#1e293b").pack(side="left")
        self.fy_combo = ttk.Combobox(config_frame, values=["FY 2023-24", "FY 2024-25", "FY 2025-26"],
                                       font=("Segoe UI", 10), state="readonly", width=15)
        self.fy_combo.set("FY 2024-25")
        self.fy_combo.pack(side="left", padx=(10, 0))

        files_frame.grid_columnconfigure(0, weight=1)
        files_frame.grid_columnconfigure(1, weight=1)
        files_frame.grid_columnconfigure(2, weight=0)

        actions_frame = tk.Frame(main_container, bg="#0f172a")
        actions_frame.pack(fill="x", pady=(0, 15))
        self.start_btn = tk.Button(actions_frame, text="START RECONCILIATION", font=("Segoe UI", 11, "bold"),
                                     bg="#10b981", fg="#ffffff", activebackground="#059669", activeforeground="#ffffff",
                                     bd=0, cursor="hand2", padx=20, command=self.start_reconciliation_process)
        self.start_btn.pack(side="left", ipady=6, padx=(0, 10))
        reset_btn = tk.Button(actions_frame, text="RESET FIELDS", font=("Segoe UI", 11, "bold"),
                                bg="#475569", fg="#ffffff", activebackground="#334155", activeforeground="#ffffff",
                                bd=0, cursor="hand2", padx=15, command=self.reset_fields)
        reset_btn.pack(side="left", ipady=6, padx=(0, 10))
        exit_btn = tk.Button(actions_frame, text="EXIT APP", font=("Segoe UI", 11, "bold"),
                               bg="#ef4444", fg="#ffffff", activebackground="#dc2626", activeforeground="#ffffff",
                               bd=0, cursor="hand2", padx=15, command=self.destroy)
        exit_btn.pack(side="right", ipady=6)

        progress_frame = tk.Frame(main_container, bg="#1e293b", bd=1, relief="flat", padx=15, pady=12)
        progress_frame.pack(fill="x", pady=(0, 15))
        self.status_lbl = tk.Label(progress_frame, text="Status: Ready", font=("Segoe UI", 9, "bold"),
                                     fg="#94a3b8", bg="#1e293b")
        self.status_lbl.pack(anchor="w", pady=(0, 4))
        self.progress_bar = CanvasProgressBar(progress_frame, bg="#0f172a", fg="#6366f1")
        self.progress_bar.pack(fill="x")

        console_frame = tk.Frame(main_container, bg="#1e293b", bd=1, relief="flat", padx=15, pady=15)
        console_frame.pack(fill="both", expand=True)
        tk.Label(console_frame, text="Activity Logs:", font=("Segoe UI", 10, "bold"),
                  fg="#f1f5f9", bg="#1e293b").pack(anchor="w", pady=(0, 5))
        self.log_area = scrolledtext.ScrolledText(console_frame, font=("Consolas", 9), bg="#020617", fg="#f1f5f9",
                                                     bd=1, relief="solid", insertbackground="#ffffff")
        self.log_area.pack(fill="both", expand=True)
        self.log_area.config(state="disabled")

        self.log_area.tag_config("info", foreground="#94a3b8")
        self.log_area.tag_config("success", foreground="#34d399")
        self.log_area.tag_config("warning", foreground="#fbbf24")
        self.log_area.tag_config("error", foreground="#f87171")
        self.log_area.tag_config("bold_info", foreground="#60a5fa", font=("Consolas", 9, "bold"))

        self.write_log("Welcome to AIS BROKER RECO Engine (Master v5)! Load input Excel sheets to begin matching.", "bold_info")

    def check_dependencies(self):
        missing = []
        try:
            import pandas
        except ImportError:
            missing.append("pandas")
        try:
            import openpyxl
        except ImportError:
            missing.append("openpyxl")
        if missing:
            messagebox.showerror("Missing Dependencies",
                f"Required third-party packages are missing: {', '.join(missing)}\n\n"
                "Please run: python -m pip install pandas openpyxl")
            self.write_log("ERROR: Missing package dependencies.", "error")
        else:
            self.write_log("Base dependencies verified. 'xlrd' is only needed for legacy .xls files.", "success")

    def write_log(self, text, tag="info"):
        self.log_area.config(state="normal")
        time_str = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_area.insert("end", f"[{time_str}] {text}\n", tag)
        self.log_area.see("end")
        self.log_area.config(state="disabled")

    def set_progress(self, val, status_text=None):
        self.progress_bar.set(val)
        if status_text:
            self.status_lbl.config(text=f"Status: {status_text}", fg="#fbbf24")

    def select_ais_file(self):
        file_path = filedialog.askopenfilename(title="Select AIS Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xls *.xlsm"), ("All Files", "*.*")])
        if file_path:
            self.ais_file_path = file_path
            self.ais_entry.delete(0, "end")
            self.ais_entry.insert(0, file_path)
            self.write_log(f"AIS File Selected: {file_path}", "info")

    def select_broker_file(self):
        file_path = filedialog.askopenfilename(title="Select Broker Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xls *.xlsm"), ("All Files", "*.*")])
        if file_path:
            self.broker_file_path = file_path
            self.broker_entry.delete(0, "end")
            self.broker_entry.insert(0, file_path)
            self.write_log(f"Broker File Selected: {file_path}", "info")

    def get_safe_incremental_filename(self, output_path):
        if not os.path.exists(output_path):
            return output_path
        dir_name = os.path.dirname(output_path)
        base_name = os.path.basename(output_path)
        name, ext = os.path.splitext(base_name)
        counter = 1
        new_path = os.path.join(dir_name, f"{name}_{counter}{ext}")
        while os.path.exists(new_path):
            counter += 1
            new_path = os.path.join(dir_name, f"{name}_{counter}{ext}")
        return new_path

    def sanitize_windows_filename(self, name):
        return re.sub(r'[\\/*?:"<>|]', '_', name)

    def reset_fields(self):
        if self.is_running:
            return
        self.ais_entry.delete(0, "end")
        self.broker_entry.delete(0, "end")
        self.ais_file_path = ""
        self.broker_file_path = ""
        self.set_progress(0.0, "Ready")
        self.status_lbl.config(text="Status: Ready", fg="#94a3b8")
        self.write_log("App state reset. Ready for new files.", "bold_info")

    def start_reconciliation_process(self):
        if self.is_running:
            return
        ais = self.ais_entry.get().strip()
        broker = self.broker_entry.get().strip()
        fy = self.fy_combo.get()

        if not ais or not broker:
            messagebox.showerror("Selection Error", "Please select both the AIS and Broker Excel files before running.")
            return
        if not os.path.exists(ais):
            messagebox.showerror("File Error", f"AIS file does not exist:\n{ais}")
            return
        if not os.path.exists(broker):
            messagebox.showerror("File Error", f"Broker file does not exist:\n{broker}")
            return

        default_dir = os.path.dirname(ais)
        clean_fy = self.sanitize_windows_filename(fy.replace(" ", ""))
        default_name = f"AIS_BROKER_RECO_{clean_fy}.xlsx"

        output_file = filedialog.asksaveasfilename(initialdir=default_dir, initialfile=default_name,
            title="Save Reconciliation Results As", filetypes=[("Excel Workbook", "*.xlsx")], defaultextension=".xlsx")
        if not output_file:
            self.write_log("Operation cancelled by user (output save location not selected).", "warning")
            return

        safe_output = self.get_safe_incremental_filename(output_file)
        if safe_output != output_file:
            self.write_log(f"Output path adjusted to prevent overwrite: {os.path.basename(safe_output)}", "warning")

        self.is_running = True
        self.start_btn.config(state="disabled", bg="#475569", text="PROCESSING...")

        thread = threading.Thread(target=self.run_reco_thread, args=(ais, broker, safe_output, fy), daemon=True)
        thread.start()

    def run_reco_thread(self, ais, broker, output_path, fy):
        try:
            summary = run_reconciliation_workflow(
                ais_file=ais, broker_file=broker, output_path=output_path, fy=fy,
                progress_callback=lambda p, status: self.after(0, self.set_progress, p / 100.0, status),
                log_callback=lambda text, tag="info": self.after(0, self.write_log, text, tag)
            )
            self.after(0, self.show_success_dialog, summary, output_path)
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Processing Mismatch Error",
                f"An unexpected error halted reconciliation:\n\n{str(e)}"))
            self.after(0, self.set_progress, 1.0, "Error Encountered")
            self.after(0, lambda: self.status_lbl.config(fg="#f87171"))
        finally:
            self.after(0, self.finish_reco_state)

    def show_success_dialog(self, stats, output_path):
        self.status_lbl.config(text="Status: Completed Successfully", fg="#10b981")
        self.progress_bar.set(1.0)
        details = (
            f"RECONCILIATION COMPLETED SUCCESSFULLY!\n\n"
            f"- Total AIS rows: {stats['total_ais']}\n"
            f"- Total Broker rows: {stats['total_broker']}\n"
            f"- Matched Transactions: {stats['matched']}\n"
            f"- AIS Duplicate entries: {stats['duplicates']}\n"
            f"- Missing in AIS: {stats['missing_in_ais']}\n"
            f"- Missing in Broker: {stats['missing_in_broker']}\n"
            f"- LTCG Differences: {stats['ltcg_diffs']}\n"
            f"- STCG Differences: {stats['stcg_diffs']}\n"
            f"- Manual Review Required: {stats['review_required']}\n\n"
            f"Reconciliation Report Saved At:\n{output_path}"
        )
        messagebox.showinfo("Reconciliation Success", details)

    def finish_reco_state(self):
        self.is_running = False
        self.start_btn.config(state="normal", bg="#10b981", text="START RECONCILIATION")


if __name__ == "__main__":
    app = AISBrokerRecoApp()
    app.mainloop()

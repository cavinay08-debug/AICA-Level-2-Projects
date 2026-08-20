import os
import re
import sys
import datetime
import threading
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter import scrolledtext
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# 1. CONSTANTS & CONFIGURATION
# =====================================================================
TOLERANCE_AMOUNT = 10.0  # Configurable monetary tolerance (in Rs) for strict amount differences
SALE_CONSIDERATION_IGNORE_LIMIT = 100.0  # Minor sale value variance to ignore for brokerage/transaction charge effects
DATE_WINDOW_DAYS = 3     # Window of days for matches with minor date discrepancies (T+1/T+2)

COLUMN_SYNONYMS = {
    'isin': [
        'isin', 'isin code', 'isin_code', 'security code', 'isin number', 
        'security description isin', 'instrument isin', 'isin id',
        'security isin', 'asset isin'
    ],
    'security_name': [
        'security name', 'security description', 'security', 'scrip name', 
        'scrip', 'company name', 'description', 'stock name', 'scheme name', 
        'scheme', 'particulars', 'scrip/security description', 'instrument name',
        'name of security', 'name of share', 'name of shares', 'name of unit',
        'asset description'
    ],
    'sale_date': [
        'sale date', 'transaction date', 'trade date', 'sell date', 
        'date of sale', 'sale_date', 'date of transfer', 'transfer date', 
        'date of transaction', 'redemption date', 'sell_date', 'disposal date',
        'date of sale/transfer', 'sale transfer date'
    ],
    'purchase_date': [
        'purchase date', 'buy date', 'date of purchase', 'purchase_date', 
        'date of acquisition', 'acquisition date', 'buy_date', 'acquisition_date'
    ],
    'quantity': [
        'quantity', 'qty', 'units', 'volume', 'sale quantity', 'sell quantity', 
        'number of shares', 'no of shares', 'no. of shares', 'shares qty', 
        'sell qty', 'sale qty', 'qty sold', 'number of units', 'no of units',
        'no. of units', 'units sold'
    ],
    'sale_consideration': [
        'sale consideration', 'sales consideration', 'sale value', 'sale amount', 
        'sell value', 'sales value', 'gross sales', 'sales amount', 'consideration', 
        'redemption value', 'redemption amount', 'amount', 'value of transaction', 
        'transaction value', 'sale_value', 'consideration amount', 'sell value (gross)',
        'full value of consideration', 'sales proceeds', 'sale proceeds',
        'gross consideration', 'net consideration'
    ],
    'purchase_cost': [
        'purchase cost', 'cost of acquisition', 'acquisition cost', 'purchase value', 
        'purchase amount', 'buy value', 'buy cost', 'cost value', 'cost', 
        'cost_of_acquisition', 'purchase price/cost', 'buy value (gross)',
        'purchase price', 'indexed cost of acquisition'
    ],
    'stt': [
        'stt', 'securities transaction tax', 'stt paid', 'stt charges', 
        'stt_charges', 'stt amt', 'stt amount', 'securities transaction tax (stt)'
    ],
    'gain_class': [
        'class', 'type of gain', 'classification', 'gain type',
        'ltcg/stcg', 'capital gain type', 'nature of income', 'gain_class',
        'term', 'gain classification'
    ]
}

# =====================================================================
# 2. TAX RULE ENGINE
# =====================================================================
class TaxRuleEngine:
    """
    Isolated Indian Income Tax Capital Gains Rule Engine.
    Configurable by Financial Year.
    """
    def __init__(self, financial_year):
        self.financial_year = financial_year
        # Define historical cutoffs/rules if needed.
        # Currently covers FY 2023-24, FY 2024-25, FY 2025-26.
        
    def classify_security_type(self, isin, security_name):
        """
        Heuristic classification of security type.
        Returns: "EQUITY_SHARE", "EQUITY_MF", "DEBT_MF", "OTHER_SECURITY"
        """
        isin = str(isin).strip().upper()
        name = str(security_name).strip().upper()
        
        # Indian ISIN conventions:
        # INF -> Mutual Funds
        # INE -> Equities / Corporate Debt / Bonds
        if isin.startswith("INF"):
            # Mutual Funds
            debt_keywords = [
                "DEBT", "BOND", "LIQUID", "TREASURY", "GILT", "OVERNIGHT", "ULTRA SHORT",
                "SHORT TERM", "MEDIUM TERM", "CASH", "CONSERVATIVE", "HYBRID DEBT",
                "FIXED MATURITY", "FMP", "INCOME FUND", "DYNAMIC BOND", "SAVINGS FUND"
            ]
            # Exclude elements that have "ARBITRAGE" or "EQUITY" since they enjoy equity-oriented tax rates
            is_debt = False
            for kw in debt_keywords:
                if kw in name:
                    if "ARBITRAGE" not in name and "EQUITY" not in name:
                        is_debt = True
                        break
            if is_debt:
                return "DEBT_MF"
            else:
                return "EQUITY_MF"
                
        elif isin.startswith("INE") or isin.startswith("IN9"):
            # Shares or Corporate Bonds/NCDs
            bond_keywords = ["BOND", "NCD", "DEBENTURE", "DEB", "BND", "GOLD BOND", "SGB"]
            if any(kw in name for kw in bond_keywords):
                return "OTHER_SECURITY"
            return "EQUITY_SHARE"
        else:
            # Fallback checks based on names
            if "MUTUAL FUND" in name or "MF" in name or "FUND" in name:
                if any(kw in name for kw in ["DEBT", "LIQUID", "GILT", "BOND"]):
                    return "DEBT_MF"
                return "EQUITY_MF"
            elif any(kw in name for kw in ["BOND", "DEBENTURE", "NCD"]):
                return "OTHER_SECURITY"
            return "EQUITY_SHARE"

    def classify_transaction(self, isin, security_name, purchase_date, sale_date, stt_paid=None):
        """
        Classifies transaction into LTCG, STCG, or Review Required based on FY laws.
        Returns: (classification, holding_days, remarks)
        """
        if not purchase_date or not sale_date:
            return "Review Required", None, "Missing purchase or sale date"
            
        if not isinstance(purchase_date, (datetime.date, datetime.datetime)) or not isinstance(sale_date, (datetime.date, datetime.datetime)):
            try:
                # Try parsing if strings are passed
                purchase_date = pd.to_datetime(purchase_date).date()
                sale_date = pd.to_datetime(sale_date).date()
            except Exception:
                return "Review Required", None, "Invalid date format"
                
        if isinstance(purchase_date, datetime.datetime):
            purchase_date = purchase_date.date()
        if isinstance(sale_date, datetime.datetime):
            sale_date = sale_date.date()
            
        holding_days = (sale_date - purchase_date).days
        if holding_days < 0:
            return "Review Required", holding_days, f"Sale date ({sale_date}) is before purchase date ({purchase_date})"
            
        asset_type = self.classify_security_type(isin, security_name)
        
        # Apply provisions based on FY
        # Indian holding periods for LTCG:
        # Listed Equity & Equity MF: > 12 months (365 days)
        # Unlisted Equity: > 24 months (730 days)
        # Debt MF:
        #   - Acquired on/after April 1, 2023: Section 50AA applies (Always STCG)
        #   - Acquired before April 1, 2023: > 36 months (1095 days) for LTCG
        # Other securities (bonds/debentures): > 36 months (1095 days)
        
        if asset_type in ("EQUITY_SHARE", "EQUITY_MF"):
            # Listed Equity / Equity MF: 12 months (365 days)
            if holding_days > 365:
                return "LTCG", holding_days, f"Equity-oriented asset held > 12 months ({holding_days} days)"
            else:
                return "STCG", holding_days, f"Equity-oriented asset held <= 12 months ({holding_days} days)"
                
        elif asset_type == "DEBT_MF":
            # Debt Mutual Fund rule
            # Cut-off for Sec 50AA is April 1, 2023
            cutoff_date = datetime.date(2023, 4, 1)
            if purchase_date >= cutoff_date:
                # Section 50AA: Always treated as STCG, regardless of holding period
                return "STCG", holding_days, f"Debt MF acquired after April 1, 2023 (Sec 50AA) - strictly STCG ({holding_days} days)"
            else:
                # Pre-April 1, 2023 acquisition: 36 months (1095 days) limit for LTCG
                if holding_days > 1095:
                    return "LTCG", holding_days, f"Debt MF acquired before April 1, 2023 held > 36 months ({holding_days} days)"
                else:
                    return "STCG", holding_days, f"Debt MF acquired before April 1, 2023 held <= 36 months ({holding_days} days)"
                    
        elif asset_type == "OTHER_SECURITY":
            # Bonds/Debentures or Unlisted shares
            # General rule: unlisted is 24 months (730 days), debt bonds is 36 months (1095 days)
            # Default to 36 months for other securities if bond keywords exist
            if any(kw in security_name.upper() for kw in ["BOND", "NCD", "DEBENTURE"]):
                limit = 1095
                label = "Debt security"
            else:
                limit = 730
                label = "Other unlisted security"
                
            if holding_days > limit:
                return "LTCG", holding_days, f"{label} held > limit ({holding_days} days)"
            else:
                return "STCG", holding_days, f"{label} held <= limit ({holding_days} days)"
        else:
            return "Review Required", holding_days, f"Could not determine rules for asset type: {asset_type}"

# =====================================================================
# 3. DATA NORMALISATION & WORKBOOK INSPECTOR
# =====================================================================
class NormalizedTransaction:
    """
    Standard data object representing a transaction.
    """
    def __init__(self):
        self.match_id = None
        self.source_file = ""
        self.source_sheet = ""
        self.source_row = None
        self.isin = ""
        self.security_name = ""
        self.sale_date = None
        self.purchase_date = None
        self.quantity = 0.0
        self.sale_consideration = 0.0
        self.purchase_cost = 0.0
        self.stt = 0.0
        self.original_classification = ""
        self.calculated_classification = ""
        self.holding_days = None
        self.remarks = ""
        self.is_duplicate = False
        self.original_row_data = {}

    def to_dict(self):
        return {
            "Match ID": self.match_id,
            "Security": self.security_name,
            "ISIN": self.isin,
            "Transaction Date": self.sale_date,
            "Purchase Date": self.purchase_date,
            "Quantity": self.quantity,
            "Sale Consideration": self.sale_consideration,
            "Purchase Cost": self.purchase_cost,
            "STT": self.stt,
            "Original Classification": self.original_classification,
            "Calculated Classification": self.calculated_classification,
            "Holding Days": self.holding_days,
            "Remarks": self.remarks,
            "Source Sheet": self.source_sheet,
            "Source Row": self.source_row
        }

def clean_isin_code(val):
    if pd.isna(val) or not val:
        return ""
    cleaned = re.sub(r'[^A-Za-z0-9]', '', str(val)).strip().upper()
    return cleaned

def clean_security_name(val):
    if pd.isna(val) or not val:
        return ""
    cleaned = str(val).strip().upper()
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned

def extract_isin_from_text(val):
    if pd.isna(val) or not val:
        return ""
    match = re.search(r'\b(IN[A-Z0-9]{10})\b', str(val).upper())
    return match.group(1) if match else ""

def clean_date(val):
    if pd.isna(val) or val is None or str(val).strip() == "":
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val
    
    date_str = str(val).strip()
    if " " in date_str:
        # Separate date from time if date_str is e.g. "2023-06-15 00:00:00"
        date_str = date_str.split(" ")[0]
        
    for fmt in (
        "%d-%m-%Y", "%d/%m/%Y", "%d/%m/%y",
        "%Y-%m-%d", "%Y/%m/%d",
        "%d-%b-%Y", "%d-%B-%Y", "%d-%b-%y", "%d-%B-%y"
    ):
        try:
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    try:
        parsed = pd.to_datetime(date_str, dayfirst=True, errors="coerce")
        if pd.notna(parsed):
            return parsed.date()
    except Exception:
        pass
    return None

def clean_number(val):
    if pd.isna(val) or val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    
    val_str = str(val).strip()
    if not val_str:
        return 0.0
        
    # Remove Indian Rupee symbols, commas, spaces, currency indicators
    val_str = val_str.replace("₹", "").replace(",", "").replace("Rs", "").replace(" ", "")
    
    # Handle parentheses representation of negative numbers (e.g. "(100)" -> "-100")
    if val_str.startswith("(") and val_str.endswith(")"):
        val_str = "-" + val_str[1:-1]
        
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def normalize_header_text(val):
    if pd.isna(val) or val is None:
        return ""
    text = str(val).strip().lower()
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r'[_/\\-]+', ' ', text)
    text = re.sub(r'[^a-z0-9(). ]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def header_match_score(cell_value, synonym):
    if not cell_value or not synonym:
        return 0
    if cell_value == synonym:
        return 5
    if cell_value.startswith(synonym) or cell_value.endswith(synonym):
        return 4
    cell_tokens = set(cell_value.split())
    synonym_tokens = set(synonym.split())
    if synonym_tokens and synonym_tokens.issubset(cell_tokens):
        return 3
    if synonym in cell_value:
        return 2
    return 0

def detect_header_row(df_raw, sheet_name=""):
    """
    Finds the row containing transaction headers using synonyms list.
    Returns: (header_row_index, mapped_columns_dict, header_depth)
    """
    best_row_idx = None
    best_mapping = {}
    best_header_depth = 1
    best_score = None
    is_mf_units_sheet = "units of mf except equity fund" in str(sheet_name).strip().lower()

    # Read a larger window because AIS exports often place the real header
    # after titles, notes, or instruction rows.
    scan_limit = min(75, len(df_raw))
    for row_idx in range(scan_limit):
        row_values = [normalize_header_text(x) for x in df_raw.iloc[row_idx].values]
        candidate_rows = [(row_values, 1)]

        if row_idx + 1 < scan_limit:
            next_row_values = [normalize_header_text(x) for x in df_raw.iloc[row_idx + 1].values]
            combined_values = []
            max_len = max(len(row_values), len(next_row_values))
            for col_idx in range(max_len):
                top = row_values[col_idx] if col_idx < len(row_values) else ""
                bottom = next_row_values[col_idx] if col_idx < len(next_row_values) else ""
                combined_values.append(f"{top} {bottom}".strip())
            candidate_rows.append((combined_values, 2))

        for candidate_values, header_depth in candidate_rows:
            mapping = {}
            used_columns = set()
            quality_score = 0
            non_empty_cells = sum(1 for cell in candidate_values if cell and cell != 'nan')

            for std_col, synonyms in COLUMN_SYNONYMS.items():
                normalized_synonyms = [normalize_header_text(syn) for syn in synonyms]
                best_for_field = None
                for col_idx, cell_value in enumerate(candidate_values):
                    if col_idx in used_columns or not cell_value or cell_value == 'nan':
                        continue
                    col_score = max((header_match_score(cell_value, syn) for syn in normalized_synonyms), default=0)
                    if col_score <= 0:
                        continue
                    if best_for_field is None or col_score > best_for_field[0]:
                        best_for_field = (col_score, col_idx)

                if best_for_field is not None:
                    col_score, col_idx = best_for_field
                    mapping[std_col] = col_idx
                    used_columns.add(col_idx)
                    quality_score += col_score

            distinct_count = len(mapping)
            has_dates = 'sale_date' in mapping or 'purchase_date' in mapping
            has_values = 'sale_consideration' in mapping or 'purchase_cost' in mapping
            has_quantity_or_sheet_exception = 'quantity' in mapping or is_mf_units_sheet
            has_useful_shape = distinct_count >= 4 and has_quantity_or_sheet_exception and has_dates and has_values

            if not has_useful_shape:
                continue

            score_tuple = (distinct_count, quality_score, non_empty_cells, row_idx, header_depth)
            if best_score is None or score_tuple > best_score:
                best_score = score_tuple
                best_row_idx = row_idx
                best_mapping = mapping
                best_header_depth = header_depth

    return best_row_idx, best_mapping, best_header_depth

def resolve_excel_engine(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xls':
        try:
            import xlrd  # noqa: F401
        except ImportError as exc:
            raise ImportError(
                "Legacy .xls files require the optional 'xlrd' package. "
                "Install it with: python -m pip install xlrd"
            ) from exc
        return 'xlrd'
    return 'openpyxl'

def load_and_normalize_workbook(file_path, file_label, log_callback):
    """
    Inspects and reads transaction data from all valid sheets.
    Returns list of NormalizedTransaction objects.
    """
    log_callback(f"Inspecting workbook: {os.path.basename(file_path)}...")
    transactions = []
    
    try:
        engine = resolve_excel_engine(file_path)
        # Load excel file to inspect sheet names
        xl = pd.ExcelFile(file_path, engine=engine)
        sheet_names = xl.sheet_names
        log_callback(f"Found sheets in {file_label}: {sheet_names}")
        
        for sheet_name in sheet_names:
            # Read sheet without assuming header positions
            df_raw = xl.parse(sheet_name=sheet_name, header=None)
            if len(df_raw) == 0:
                log_callback(f"Skipping empty sheet: {sheet_name}")
                continue
                
            is_mf_units_sheet = "units of mf except equity fund" in sheet_name.strip().lower()
            header_row_idx, mapping, header_depth = detect_header_row(df_raw, sheet_name=sheet_name)
            if header_row_idx is None:
                log_callback(f"Skipping sheet '{sheet_name}': Could not identify required columns.")
                continue
                
            log_callback(
                f"Sheet '{sheet_name}' - Header detected at row {header_row_idx + 1}"
                f"{' (2-row header)' if header_depth == 2 else ''}"
            )
            log_callback(f"Mapped columns: {list(mapping.keys())}")
            
            # Slice dataframe to get values below header
            data_start_idx = header_row_idx + header_depth
            df_data = df_raw.iloc[data_start_idx:].copy()

            if header_depth == 2 and header_row_idx + 1 < len(df_raw):
                combined_headers = []
                top_headers = df_raw.iloc[header_row_idx].values
                bottom_headers = df_raw.iloc[header_row_idx + 1].values
                max_len = max(len(top_headers), len(bottom_headers))
                for col_idx in range(max_len):
                    top = normalize_header_text(top_headers[col_idx]) if col_idx < len(top_headers) else ""
                    bottom = normalize_header_text(bottom_headers[col_idx]) if col_idx < len(bottom_headers) else ""
                    combined_headers.append((f"{top} {bottom}".strip()) or f"column_{col_idx + 1}")
                df_data.columns = combined_headers[:len(df_data.columns)]
            else:
                raw_headers = df_raw.iloc[header_row_idx].values
                df_data.columns = [
                    normalize_header_text(val) or f"column_{idx + 1}"
                    for idx, val in enumerate(raw_headers[:len(df_data.columns)])
                ]
            
            # Handle merged cells: forward-fill ISIN and Security Name if available
            fill_cols = []
            if 'isin' in mapping:
                fill_cols.append(mapping['isin'])
            if 'security_name' in mapping:
                fill_cols.append(mapping['security_name'])
                
            # Convert mapping index keys to names or handle via index positions
            # We will use integer positions from mapping dict for safety
            for i in range(len(df_data)):
                raw_row = df_data.iloc[i]
                
                # Check if it's a totals or empty row
                row_str_val = " ".join([str(x).lower() for x in raw_row.values if not pd.isna(x)])
                if not row_str_val or "total" in row_str_val or "grand total" in row_str_val:
                    continue
                    
                # Build transaction
                t = NormalizedTransaction()
                t.source_file = os.path.basename(file_path)
                t.source_sheet = sheet_name
                t.source_row = int(df_data.index[i]) + 2 # Excel 1-based index (header is +1, index is +1)
                
                # Extract values using index
                if 'isin' in mapping:
                    t.isin = clean_isin_code(raw_row.iloc[mapping['isin']])
                if 'security_name' in mapping:
                    t.security_name = clean_security_name(raw_row.iloc[mapping['security_name']])
                    if not t.isin:
                        t.isin = extract_isin_from_text(t.security_name)
                if 'sale_date' in mapping:
                    t.sale_date = clean_date(raw_row.iloc[mapping['sale_date']])
                if 'purchase_date' in mapping:
                    t.purchase_date = clean_date(raw_row.iloc[mapping['purchase_date']])
                if 'quantity' in mapping:
                    t.quantity = clean_number(raw_row.iloc[mapping['quantity']])
                elif is_mf_units_sheet:
                    # Some mutual fund export layouts omit unit quantity entirely.
                    # Use a synthetic quantity so these rows can still flow through
                    # date/value based reconciliation and reporting.
                    t.quantity = 1.0
                if 'sale_consideration' in mapping:
                    t.sale_consideration = clean_number(raw_row.iloc[mapping['sale_consideration']])
                if 'purchase_cost' in mapping:
                    t.purchase_cost = clean_number(raw_row.iloc[mapping['purchase_cost']])
                if 'stt' in mapping:
                    t.stt = clean_number(raw_row.iloc[mapping['stt']])
                if 'gain_class' in mapping:
                    raw_cls = str(raw_row.iloc[mapping['gain_class']]).strip().upper()
                    if "LONG" in raw_cls or "LTCG" in raw_cls:
                        t.original_classification = "LTCG"
                    elif "SHORT" in raw_cls or "STCG" in raw_cls:
                        t.original_classification = "STCG"
                    else:
                        t.original_classification = raw_cls
                
                t.original_row_data = {str(k): v for k, v in zip(df_data.columns, raw_row.values)}
                
                # Validation checks: quantity and sale date are mandatory
                if t.quantity > 0 and t.sale_date is not None:
                    # If ISIN is missing but name exists, allow processing
                    if t.isin or t.security_name:
                        transactions.append(t)
            log_callback(
                f"Sheet '{sheet_name}' contributed "
                f"{sum(1 for t in transactions if t.source_sheet == sheet_name)} valid rows."
            )
                        
            log_callback(f"Successfully loaded {len(transactions)} transaction records from {sheet_name}.")
            
    except Exception as e:
        log_callback(f"ERROR reading workbook {file_label}: {str(e)}")
        log_callback(traceback.format_exc())
        raise e
        
    return transactions

# =====================================================================
# 4. DUPLICATE & TRANSACTION MATCHING ENGINE
# =====================================================================
def detect_ais_duplicates(ais_txns, broker_txns, log_callback):
    """
    Detects duplicates in AIS data.
    Groups AIS transactions by (ISIN, sale_date, quantity, sale_consideration).
    If AIS occurrences > Broker occurrences, the excess AIS transactions are flagged.
    """
    log_callback("Running AIS duplicate detection analysis...")
    
    # Group broker transactions by key
    broker_counts = {}
    for t in broker_txns:
        key = (t.isin, t.sale_date, round(t.quantity, 4), round(t.sale_consideration, 2))
        broker_counts[key] = broker_counts.get(key, 0) + 1
        
    # Group AIS transactions
    ais_groups = {}
    for t in ais_txns:
        key = (t.isin, t.sale_date, round(t.quantity, 4), round(t.sale_consideration, 2))
        if key not in ais_groups:
            ais_groups[key] = []
        ais_groups[key].append(t)
        
    duplicates_list = []
    exact_duplicates_count = 0
    probable_duplicates_count = 0
    
    # Process duplicate groups
    dup_group_id = 1
    for key, group in ais_groups.items():
        ais_count = len(group)
        if ais_count > 1:
            # Check corresponding broker occurrences
            broker_count = broker_counts.get(key, 0)
            
            if ais_count > broker_count:
                # Excess occurrences in AIS are duplicates
                duplicate_qty = ais_count - broker_count
                
                # Flag the duplicate entries in AIS (keep the first broker_count entries as legitimate, rest are dups)
                for idx in range(broker_count, ais_count):
                    dup_txn = group[idx]
                    dup_txn.is_duplicate = True
                    dup_txn.remarks = f"Duplicate Entry: Found {ais_count} occurrences in AIS but only {broker_count} in Broker data."
                    duplicates_list.append((dup_group_id, dup_txn, ais_count, broker_count))
                    exact_duplicates_count += 1
                
                dup_group_id += 1
                
    log_callback(f"Duplicate detection completed: Flagged {exact_duplicates_count} duplicates in AIS.")
    return duplicates_list

def match_and_reconcile(ais_txns, broker_txns, tax_engine, log_callback):
    """
    Executes the 5-pass matching engine.
    Returns: (matched_pairs, unmatched_ais, unmatched_broker)
    """
    log_callback("Starting transaction matching engine...")
    
    # Filter out duplicates from normal AIS matching pool
    matching_ais = [t for t in ais_txns if not t.is_duplicate]
    matching_broker = list(broker_txns)
    
    matched_pairs = []
    match_counter = 1
    
    # Pass 1: Strict Match (ISIN + Date + Qty + Sale Consideration)
    log_callback("Matching Pass 1: Strict Match (ISIN + Date + Qty + Consideration)...")
    unmatched_ais = []
    for a in matching_ais:
        found_idx = None
        for i, b in enumerate(matching_broker):
            if (a.isin and a.isin == b.isin) or (not a.isin and a.security_name == b.security_name):
                if a.sale_date == b.sale_date:
                    if abs(a.quantity - b.quantity) < 1e-4:
                        if abs(a.sale_consideration - b.sale_consideration) <= SALE_CONSIDERATION_IGNORE_LIMIT:
                            found_idx = i
                            break
        if found_idx is not None:
            b_match = matching_broker.pop(found_idx)
            a.match_id = f"M-{match_counter:04d}"
            b_match.match_id = a.match_id
            matched_pairs.append((a, b_match, "Strict Match"))
            match_counter += 1
        else:
            unmatched_ais.append(a)
            
    # Pass 2: Date Window Match (ISIN + Date Window +/- 3 days + Qty + Sale Consideration)
    log_callback(f"Matching Pass 2: Date Window Match (+/- {DATE_WINDOW_DAYS} days)...")
    matching_ais = unmatched_ais
    unmatched_ais = []
    for a in matching_ais:
        found_idx = None
        for i, b in enumerate(matching_broker):
            if (a.isin and a.isin == b.isin) or (not a.isin and a.security_name == b.security_name):
                if a.sale_date and b.sale_date and abs((a.sale_date - b.sale_date).days) <= DATE_WINDOW_DAYS:
                    if abs(a.quantity - b.quantity) < 1e-4:
                        if abs(a.sale_consideration - b.sale_consideration) <= SALE_CONSIDERATION_IGNORE_LIMIT:
                            found_idx = i
                            break
        if found_idx is not None:
            b_match = matching_broker.pop(found_idx)
            a.match_id = f"M-{match_counter:04d}"
            b_match.match_id = a.match_id
            matched_pairs.append((a, b_match, "Date Window Match"))
            match_counter += 1
        else:
            unmatched_ais.append(a)

    # Pass 3: Soft Description Match (Normalized name + Date + Qty + Sale Consideration)
    log_callback("Matching Pass 3: Normalized Description Match (Name + Date + Qty + Consideration)...")
    matching_ais = unmatched_ais
    unmatched_ais = []
    for a in matching_ais:
        found_idx = None
        for i, b in enumerate(matching_broker):
            # Check name similarity
            a_name = a.security_name.replace(" ", "")
            b_name = b.security_name.replace(" ", "")
            # Substring match if name is long enough
            is_name_match = (a_name in b_name) or (b_name in a_name) or (a.isin == b.isin and a.isin != "")
            if is_name_match:
                if a.sale_date == b.sale_date:
                    if abs(a.quantity - b.quantity) < 1e-4:
                        if abs(a.sale_consideration - b.sale_consideration) <= SALE_CONSIDERATION_IGNORE_LIMIT:
                            found_idx = i
                            break
        if found_idx is not None:
            b_match = matching_broker.pop(found_idx)
            a.match_id = f"M-{match_counter:04d}"
            b_match.match_id = a.match_id
            matched_pairs.append((a, b_match, "Description Match"))
            match_counter += 1
        else:
            unmatched_ais.append(a)

    # Pass 4: Soft Description & Date Window Match
    log_callback("Matching Pass 4: Description & Date Window Match...")
    matching_ais = unmatched_ais
    unmatched_ais = []
    for a in matching_ais:
        found_idx = None
        for i, b in enumerate(matching_broker):
            a_name = a.security_name.replace(" ", "")
            b_name = b.security_name.replace(" ", "")
            is_name_match = (a_name in b_name) or (b_name in a_name) or (a.isin == b.isin and a.isin != "")
            if is_name_match:
                if a.sale_date and b.sale_date and abs((a.sale_date - b.sale_date).days) <= DATE_WINDOW_DAYS:
                    if abs(a.quantity - b.quantity) < 1e-4:
                        if abs(a.sale_consideration - b.sale_consideration) <= SALE_CONSIDERATION_IGNORE_LIMIT:
                            found_idx = i
                            break
        if found_idx is not None:
            b_match = matching_broker.pop(found_idx)
            a.match_id = f"M-{match_counter:04d}"
            b_match.match_id = a.match_id
            matched_pairs.append((a, b_match, "Soft Match"))
            match_counter += 1
        else:
            unmatched_ais.append(a)

    # Pass 5: Tolerance match (ISIN + Date Window + Qty + Sale consideration within 1%)
    log_callback("Matching Pass 5: Tolerance-based Value Match (within 1% variation)...")
    matching_ais = unmatched_ais
    unmatched_ais = []
    for a in matching_ais:
        found_idx = None
        for i, b in enumerate(matching_broker):
            if (a.isin and a.isin == b.isin) or (not a.isin and a.security_name == b.security_name):
                if a.sale_date and b.sale_date and abs((a.sale_date - b.sale_date).days) <= DATE_WINDOW_DAYS:
                    if abs(a.quantity - b.quantity) < 1e-4:
                        # Allow up to 1% difference in consideration value
                        pct_diff = abs(a.sale_consideration - b.sale_consideration) / max(a.sale_consideration, 1.0)
                        if pct_diff <= 0.01:
                            found_idx = i
                            break
        if found_idx is not None:
            b_match = matching_broker.pop(found_idx)
            a.match_id = f"M-{match_counter:04d}"
            b_match.match_id = a.match_id
            matched_pairs.append((a, b_match, "Tolerance Value Match"))
            match_counter += 1
        else:
            unmatched_ais.append(a)
            
    log_callback(f"Matching finished: Reconciled {len(matched_pairs)} transactions successfully.")
    log_callback(f"Unmatched AIS records remaining: {len(unmatched_ais)}")
    log_callback(f"Unmatched Broker records remaining: {len(matching_broker)}")
    
    return matched_pairs, unmatched_ais, matching_broker

# =====================================================================
# 5. REPORT GENERATOR (OPENPYXL WRAPPER)
# =====================================================================
class RecoReportGenerator:
    """
    Handles generation and formatting of the multi-sheet reconciliation report.
    """
    def __init__(self, output_path, fy, ais_file, broker_file):
        self.output_path = output_path
        self.fy = fy
        self.ais_file = os.path.basename(ais_file)
        self.broker_file = os.path.basename(broker_file)
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Color definitions (Hex values)
        self.COLOR_INDIGO_HEADER = "1E3A8A"  # Primary dark indigo header fill
        self.COLOR_LIGHT_ZEBRA = "F8FAFC"    # Alternating row background
        self.COLOR_GREEN_MATCH = "E8F5E9"    # Green highlight for success/match
        self.COLOR_RED_DIFF = "FFEBEE"       # Red highlight for differences
        self.COLOR_YELLOW_WARN = "FFFDE7"    # Yellow highlight for warnings/review
        
        self.FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        self.FONT_BODY = Font(name="Segoe UI", size=10)
        self.FONT_BOLD = Font(name="Segoe UI", size=10, bold=True)
        
        self.ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
        self.ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
        self.ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
        
        self.BORDER_THIN = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        self.BORDER_DOUBLE_BOTTOM = Border(
            top=Side(style='thin', color='AAAAAA'),
            bottom=Side(style='double', color='1E3A8A')
        )

    def style_header(self, ws, columns):
        ws.append(columns)
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = self.FONT_HEADER
            cell.fill = PatternFill(start_color=self.COLOR_INDIGO_HEADER, end_color=self.COLOR_INDIGO_HEADER, fill_type="solid")
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_THIN
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'

    def format_rows(self, ws, start_row, end_row, num_cols):
        for row in range(start_row, end_row + 1):
            ws.row_dimensions[row].height = 20
            # Zebra striping
            fill_color = self.COLOR_LIGHT_ZEBRA if row % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.FONT_BODY
                cell.border = self.BORDER_THIN
                if cell.fill.fill_type is None:  # Only fill if not already colored
                    cell.fill = row_fill

    def autofit_columns(self, ws):
        ws.views.sheetView[0].showGridLines = True
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format and ('#,##0.00' in cell.number_format or '%' in cell.number_format):
                    # Pad length for formatted numbers
                    val += "   "
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    def write_summary_sheet(self, summary_data):
        ws = self.wb.create_sheet(title="Summary")
        ws.views.sheetView[0].showGridLines = True
        
        # Title Card
        ws.cell(row=1, column=1, value="RECONCILIATION SUMMARY REPORT").font = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
        ws.row_dimensions[1].height = 30
        
        # Meta table
        meta = [
            ("Application:", "AIS BROKER RECO (v1.0)"),
            ("Selected Financial Year:", self.fy),
            ("AIS Source Workbook:", self.ais_file),
            ("Broker Source Workbook:", self.broker_file),
            ("Reconciliation Date:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        
        row_idx = 3
        for k, v in meta:
            ws.cell(row=row_idx, column=1, value=k).font = self.FONT_BOLD
            ws.cell(row=row_idx, column=2, value=v).font = self.FONT_BODY
            row_idx += 1
            
        # Summary Statistics Headers
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="RECONCILIATION METRICS").font = Font(name="Segoe UI", size=12, bold=True, color="1E3A8A")
        row_idx += 1
        
        metrics_headers = ["Metric Description", "Record Count / Value"]
        for col, h in enumerate(metrics_headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=h)
            cell.font = self.FONT_HEADER
            cell.fill = PatternFill(start_color=self.COLOR_INDIGO_HEADER, end_color=self.COLOR_INDIGO_HEADER, fill_type="solid")
            cell.alignment = self.ALIGN_CENTER
            
        start_table_row = row_idx + 1
        
        metrics = [
            ("Total AIS Records Loaded", summary_data['total_ais']),
            ("Total Broker Records Loaded", summary_data['total_broker']),
            ("Successfully Matched Records", summary_data['matched']),
            ("Duplicate Records in AIS", summary_data['duplicates']),
            ("Missing Records in AIS (Broker only)", summary_data['missing_in_ais']),
            ("Missing Records in Broker (AIS only)", summary_data['missing_in_broker']),
            ("Long-Term Capital Gain Differences", summary_data['ltcg_diffs']),
            ("Short-Term Capital Gain Differences", summary_data['stcg_diffs']),
            ("Sales Consideration Mismatches", summary_data['sales_mismatch']),
            ("Ignored Minor Sales Variances", summary_data['ignored_sales_variances']),
            ("Other Financial Differences", summary_data['other_diffs']),
            ("Records Requiring Manual Review", summary_data['review_required']),
        ]
        
        for k, v in metrics:
            row_idx += 1
            c1 = ws.cell(row=row_idx, column=1, value=k)
            c2 = ws.cell(row=row_idx, column=2, value=v)
            c1.font = self.FONT_BODY
            c1.border = self.BORDER_THIN
            c2.font = self.FONT_BOLD
            c2.border = self.BORDER_THIN
            c2.alignment = self.ALIGN_CENTER
            
            # Format number/count
            c2.number_format = '#,##0'
            
            # Color codes
            if k in ("Successfully Matched Records"):
                c2.fill = PatternFill(start_color=self.COLOR_GREEN_MATCH, end_color=self.COLOR_GREEN_MATCH, fill_type="solid")
            elif k in ("Total AIS Records Loaded", "Total Broker Records Loaded"):
                pass
            else:
                if v > 0:
                    c2.fill = PatternFill(start_color=self.COLOR_RED_DIFF, end_color=self.COLOR_RED_DIFF, fill_type="solid")
                    
        # Style layout widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25

    def write_difference_sheet(self, title, matched_pairs_diffs):
        """
        Writes Differences sheets (LTCG / STCG Difference or Other Differences)
        """
        ws = self.wb.create_sheet(title=title)
        
        headers = [
            "Match ID", "Security", "ISIN", "Transaction Date", "Purchase Date", "Quantity", 
            "AIS Class", "Broker Class", "Calc Class", "AIS Sale Value", "Broker Sale Value", 
            "Diff Sale Value", "AIS Cost", "Broker Cost", "Diff Cost", "AIS STT", "Broker STT", 
            "Holding Days (AIS)", "Holding Days (Broker)", "Difference Type", "Remarks", "Source Sheet", "Source Row"
        ]
        
        self.style_header(ws, headers)
        
        row_num = 2
        for ais, broker, diff_type, remarks in matched_pairs_diffs:
            # Difference calculations
            diff_sale = ais.sale_consideration - broker.sale_consideration
            diff_cost = ais.purchase_cost - broker.purchase_cost
            
            row_values = [
                ais.match_id, ais.security_name, ais.isin, ais.sale_date, broker.purchase_date, ais.quantity,
                ais.original_classification, broker.original_classification, ais.calculated_classification,
                ais.sale_consideration, broker.sale_consideration, diff_sale,
                ais.purchase_cost, broker.purchase_cost, diff_cost,
                ais.stt, broker.stt,
                ais.holding_days, broker.holding_days,
                diff_type, remarks, ais.source_sheet, ais.source_row
            ]
            
            ws.append(row_values)
            
            # Format numbers & dates
            ws.cell(row=row_num, column=4).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_num, column=5).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            
            # Sale values formats
            ws.cell(row=row_num, column=10).number_format = '#,##0.00'
            ws.cell(row=row_num, column=11).number_format = '#,##0.00'
            ws.cell(row=row_num, column=12).number_format = '#,##0.00'
            
            # Cost values formats
            ws.cell(row=row_num, column=13).number_format = '#,##0.00'
            ws.cell(row=row_num, column=14).number_format = '#,##0.00'
            ws.cell(row=row_num, column=15).number_format = '#,##0.00'
            
            # STT formats
            ws.cell(row=row_num, column=16).number_format = '#,##0.00'
            ws.cell(row=row_num, column=17).number_format = '#,##0.00'
            
            # Alignments
            ws.cell(row=row_num, column=1).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=3).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=4).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=5).alignment = self.ALIGN_CENTER
            
            # Highlight differences cells
            red_fill = PatternFill(start_color=self.COLOR_RED_DIFF, end_color=self.COLOR_RED_DIFF, fill_type="solid")
            
            # Highlight classification cells if mismatch
            if ais.original_classification != broker.original_classification or ais.calculated_classification != broker.original_classification:
                ws.cell(row=row_num, column=7).fill = red_fill
                ws.cell(row=row_num, column=8).fill = red_fill
                ws.cell(row=row_num, column=9).fill = red_fill
                
            if abs(diff_sale) > TOLERANCE_AMOUNT:
                ws.cell(row=row_num, column=10).fill = red_fill
                ws.cell(row=row_num, column=11).fill = red_fill
                ws.cell(row=row_num, column=12).fill = red_fill
                
            if abs(diff_cost) > TOLERANCE_AMOUNT:
                ws.cell(row=row_num, column=13).fill = red_fill
                ws.cell(row=row_num, column=14).fill = red_fill
                ws.cell(row=row_num, column=15).fill = red_fill
                
            row_num += 1
            
        if row_num > 2:
            self.format_rows(ws, 2, row_num - 1, len(headers))
            
        self.autofit_columns(ws)

    def write_duplicates_sheet(self, duplicate_groups):
        """
        Writes 'Duplicate Entries in AIS' worksheet.
        """
        ws = self.wb.create_sheet(title="Duplicate Entries in AIS")
        headers = [
            "Dup Group Ref", "Security", "ISIN", "Transaction Date", "Quantity", 
            "Sale Consideration", "Source Row", "Source Sheet", "Remarks"
        ]
        self.style_header(ws, headers)
        
        row_num = 2
        for dup_id, t, ais_count, broker_count in duplicate_groups:
            ws.append([
                f"DUP-{dup_id:03d}", t.security_name, t.isin, t.sale_date, t.quantity,
                t.sale_consideration, t.source_row, t.source_sheet, t.remarks
            ])
            ws.cell(row=row_num, column=1).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=3).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=4).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=5).number_format = '#,##0.00'
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            
            # Fill duplicate warning
            yellow_fill = PatternFill(start_color=self.COLOR_YELLOW_WARN, end_color=self.COLOR_YELLOW_WARN, fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = yellow_fill
                
            row_num += 1
            
        if row_num > 2:
            self.format_rows(ws, 2, row_num - 1, len(headers))
        self.autofit_columns(ws)

    def write_missing_sheet(self, title, transactions_list):
        """
        Writes 'Missing Entries in AIS' or 'Missing Entries in Broker' sheets.
        """
        ws = self.wb.create_sheet(title=title)
        
        headers = [
            "Source Row", "Source Sheet", "Security", "ISIN", "Transaction Date", 
            "Quantity", "Sale Consideration", "Purchase Date", "Purchase Cost", "STT", "Reported Classification"
        ]
        self.style_header(ws, headers)
        
        row_num = 2
        for t in transactions_list:
            ws.append([
                t.source_row, t.source_sheet, t.security_name, t.isin, t.sale_date,
                t.quantity, t.sale_consideration, t.purchase_date, t.purchase_cost, t.stt, t.original_classification
            ])
            ws.cell(row=row_num, column=1).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=4).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=5).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            ws.cell(row=row_num, column=7).number_format = '#,##0.00'
            ws.cell(row=row_num, column=8).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_num, column=9).number_format = '#,##0.00'
            ws.cell(row=row_num, column=10).number_format = '#,##0.00'
            
            # Light red background for missing entries
            red_fill = PatternFill(start_color=self.COLOR_RED_DIFF, end_color=self.COLOR_RED_DIFF, fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = red_fill
                
            row_num += 1
            
        if row_num > 2:
            self.format_rows(ws, 2, row_num - 1, len(headers))
        self.autofit_columns(ws)

    def write_matched_sheet(self, matched_pairs):
        """
        Writes 'Matched Transactions' containing all matched rows side-by-side.
        """
        ws = self.wb.create_sheet(title="Matched Transactions")
        headers = [
            "Match ID", "Security", "ISIN", "Transaction Date", "Quantity", 
            "AIS Sale Value", "Broker Sale Value", "AIS Cost", "Broker Cost", 
            "AIS STT", "Broker STT", "AIS Class", "Broker Class", "Match Mode"
        ]
        self.style_header(ws, headers)
        
        row_num = 2
        for ais, broker, match_mode in matched_pairs:
            ws.append([
                ais.match_id, ais.security_name, ais.isin, ais.sale_date, ais.quantity,
                ais.sale_consideration, broker.sale_consideration,
                ais.purchase_cost, broker.purchase_cost,
                ais.stt, broker.stt,
                ais.original_classification, broker.original_classification,
                match_mode
            ])
            ws.cell(row=row_num, column=1).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=3).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=4).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=5).number_format = '#,##0.00'
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            ws.cell(row=row_num, column=7).number_format = '#,##0.00'
            ws.cell(row=row_num, column=8).number_format = '#,##0.00'
            ws.cell(row=row_num, column=9).number_format = '#,##0.00'
            ws.cell(row=row_num, column=10).number_format = '#,##0.00'
            ws.cell(row=row_num, column=11).number_format = '#,##0.00'
            
            # Color green for clean matched records (within tolerance limit)
            diff_sale = abs(ais.sale_consideration - broker.sale_consideration)
            diff_cost = abs(ais.purchase_cost - broker.purchase_cost)
            class_match = (ais.original_classification == broker.original_classification)
            
            if diff_sale <= TOLERANCE_AMOUNT and diff_cost <= TOLERANCE_AMOUNT and class_match:
                green_fill = PatternFill(start_color=self.COLOR_GREEN_MATCH, end_color=self.COLOR_GREEN_MATCH, fill_type="solid")
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = green_fill
            row_num += 1
            
        if row_num > 2:
            self.format_rows(ws, 2, row_num - 1, len(headers))
        self.autofit_columns(ws)

    def write_review_sheet(self, review_txns):
        """
        Writes 'Review Required' worksheet.
        """
        ws = self.wb.create_sheet(title="Review Required")
        headers = [
            "Source File", "Source Row", "Source Sheet", "Security", "ISIN", 
            "Transaction Date", "Quantity", "Sale Consideration", "Purchase Date", "Remarks"
        ]
        self.style_header(ws, headers)
        
        row_num = 2
        for t in review_txns:
            ws.append([
                t.source_file, t.source_row, t.source_sheet, t.security_name, t.isin,
                t.sale_date, t.quantity, t.sale_consideration, t.purchase_date, t.remarks
            ])
            ws.cell(row=row_num, column=2).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=5).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=6).alignment = self.ALIGN_CENTER
            ws.cell(row=row_num, column=7).number_format = '#,##0.00'
            ws.cell(row=row_num, column=8).number_format = '#,##0.00'
            ws.cell(row=row_num, column=9).alignment = self.ALIGN_CENTER
            
            # Fill warning color
            yellow_fill = PatternFill(start_color=self.COLOR_YELLOW_WARN, end_color=self.COLOR_YELLOW_WARN, fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = yellow_fill
            row_num += 1
            
        if row_num > 2:
            self.format_rows(ws, 2, row_num - 1, len(headers))
        self.autofit_columns(ws)

    def save(self):
        self.wb.save(self.output_path)

# =====================================================================
# 6. APP CONTROLLER / CORE RECONCILIATION PROCESSOR
# =====================================================================
def run_reconciliation_workflow(ais_file, broker_file, output_path, fy, progress_callback, log_callback):
    """
    Main background process connecting loaded tables, matching algorithms,
    tax rule validation, and openpyxl formatting.
    """
    try:
        progress_callback(5, "Initialising reconciliation run...")
        log_callback("----------------------------------------------------------------------")
        log_callback(f"Starting AIS BROKER RECO Engine for Financial Year: {fy}")
        log_callback("----------------------------------------------------------------------")
        
        # 1. Init Tax Engine
        tax_engine = TaxRuleEngine(fy)
        
        # 2. Load AIS data
        progress_callback(10, "Loading AIS data workbook...")
        ais_raw_list = load_and_normalize_workbook(ais_file, "AIS Data File", log_callback)
        
        # 3. Load Broker data
        progress_callback(25, "Loading Broker data workbook...")
        broker_raw_list = load_and_normalize_workbook(broker_file, "Broker Data File", log_callback)
        
        if not ais_raw_list:
            raise ValueError("Zero valid transaction entries extracted from AIS file. Process stopped.")
        if not broker_raw_list:
            raise ValueError("Zero valid transaction entries extracted from Broker file. Process stopped.")
            
        # 4. AIS Duplicate Detection
        progress_callback(40, "Analysing AIS duplicate entries...")
        duplicate_groups = detect_ais_duplicates(ais_raw_list, broker_raw_list, log_callback)
        
        # 5. Core matching engine
        progress_callback(55, "Matching transactions between sheets...")
        matched_pairs, unmatched_ais, unmatched_broker = match_and_reconcile(
            ais_raw_list, broker_raw_list, tax_engine, log_callback
        )
        
        # 6. Run Tax Classification checks and categorise matches
        progress_callback(70, "Evaluating Tax Classifications and differences...")
        
        ltcg_mismatches = []
        stcg_mismatches = []
        other_differences = []
        ignored_sales_variances = []
        review_required_txns = []
        
        # Check review flags in unmatched files first
        for t in unmatched_ais + unmatched_broker:
            # If transaction is missing key data, flag as review required
            if not t.isin and not t.security_name:
                t.remarks = "Empty security identifier and ISIN."
                review_required_txns.append(t)
            elif t.quantity <= 0 or t.sale_consideration <= 0:
                t.remarks = "Invalid numeric parameters (quantity/value <= 0)."
                review_required_txns.append(t)
                
        # Process matched transactions classifications & differences
        for ais, broker, match_mode in matched_pairs:
            # Calculate classification based on broker acquisition dates (or fallback to AIS dates if broker missing)
            purchase_date = broker.purchase_date if broker.purchase_date else ais.purchase_date
            
            calc_cls, h_days, tax_remarks = tax_engine.classify_transaction(
                ais.isin, ais.security_name, purchase_date, ais.sale_date
            )
            
            ais.calculated_classification = calc_cls
            broker.calculated_classification = calc_cls
            ais.holding_days = h_days
            broker.holding_days = h_days
            
            # Map original values to standardized classification
            ais_cls_norm = ais.original_classification if ais.original_classification in ("LTCG", "STCG") else None
            broker_cls_norm = broker.original_classification if broker.original_classification in ("LTCG", "STCG") else None
            
            # Determine discrepancies
            diff_remarks = []
            diff_sale = abs(ais.sale_consideration - broker.sale_consideration)
            diff_cost = abs(ais.purchase_cost - broker.purchase_cost)
            diff_stt = abs(ais.stt - broker.stt)
            
            is_class_mismatch = False
            is_amount_mismatch = False
            
            if ais_cls_norm and broker_cls_norm and ais_cls_norm != broker_cls_norm:
                is_class_mismatch = True
                diff_remarks.append(f"Classification Mismatch: AIS={ais_cls_norm}, Broker={broker_cls_norm}.")
                
            if broker_cls_norm and calc_cls != "Review Required" and broker_cls_norm != calc_cls:
                is_class_mismatch = True
                diff_remarks.append(f"Holding Period Rule Discrepancy: Broker classified as {broker_cls_norm} but rule indicates {calc_cls} ({tax_remarks}).")
                
            if calc_cls == "Review Required":
                diff_remarks.append(f"Tax calculation review: {tax_remarks}")
                review_required_txns.append(ais)
                
            if diff_sale > SALE_CONSIDERATION_IGNORE_LIMIT:
                is_amount_mismatch = True
                diff_remarks.append(
                    f"Sale Consideration Mismatch: Diff={diff_sale:.2f} "
                    f"(AIS={ais.sale_consideration:.2f}, Broker={broker.sale_consideration:.2f})."
                )
            elif diff_sale > TOLERANCE_AMOUNT:
                diff_remarks.append(
                    f"Ignored minor sale variance within Rs {SALE_CONSIDERATION_IGNORE_LIMIT:.2f}: "
                    f"Diff={diff_sale:.2f} (AIS gross={ais.sale_consideration:.2f}, "
                    f"Broker net={broker.sale_consideration:.2f})."
                )
                
            if diff_cost > TOLERANCE_AMOUNT:
                is_amount_mismatch = True
                diff_remarks.append(f"Acquisition Cost Mismatch: Diff={diff_cost:.2f} (AIS={ais.purchase_cost:.2f}, Broker={broker.purchase_cost:.2f}).")
                
            if diff_stt > TOLERANCE_AMOUNT:
                diff_remarks.append(f"STT charges mismatch: Diff={diff_stt:.2f} (AIS={ais.stt:.2f}, Broker={broker.stt:.2f}).")
                
            final_remarks = "; ".join(diff_remarks)
            ais.remarks = final_remarks
            broker.remarks = final_remarks
            
            # Place in appropriate difference sheets
            if is_class_mismatch:
                if calc_cls == "LTCG" or broker_cls_norm == "LTCG":
                    ltcg_mismatches.append((ais, broker, "Classification Mismatch", final_remarks))
                else:
                    stcg_mismatches.append((ais, broker, "Classification Mismatch", final_remarks))
            elif is_amount_mismatch:
                # Same classification, but price differences
                other_differences.append((ais, broker, "Amount Mismatch", final_remarks))
            elif diff_sale > TOLERANCE_AMOUNT:
                ignored_sales_variances.append((ais, broker, "Ignored Sales Variance", final_remarks))
            elif diff_stt > TOLERANCE_AMOUNT:
                other_differences.append((ais, broker, "STT Mismatch", final_remarks))
                
        # 7. Generate Excel Output Workbook
        progress_callback(85, "Generating Reconciliation Excel Workbook sheets...")
        
        summary_stats = {
            'total_ais': len(ais_raw_list),
            'total_broker': len(broker_raw_list),
            'matched': len(matched_pairs),
            'duplicates': len(duplicate_groups),
            'missing_in_ais': len(unmatched_broker),
            'missing_in_broker': len(unmatched_ais),
            'ltcg_diffs': len(ltcg_mismatches),
            'stcg_diffs': len(stcg_mismatches),
            'sales_mismatch': sum(
                1 for a, b, _, _ in ltcg_mismatches + stcg_mismatches + other_differences
                if abs(a.sale_consideration - b.sale_consideration) > SALE_CONSIDERATION_IGNORE_LIMIT
            ),
            'ignored_sales_variances': len(ignored_sales_variances),
            'other_diffs': len(other_differences),
            'review_required': len(review_required_txns)
        }
        
        rep = RecoReportGenerator(output_path, fy, ais_file, broker_file)
        
        rep.write_summary_sheet(summary_stats)
        rep.write_difference_sheet("Difference in Long Term", ltcg_mismatches)
        rep.write_difference_sheet("Difference in Short Term", stcg_mismatches)
        rep.write_duplicates_sheet(duplicate_groups)
        rep.write_missing_sheet("Missing Entries in AIS", unmatched_broker)
        rep.write_missing_sheet("Missing Entries in Broker", unmatched_ais)
        rep.write_difference_sheet("Other Differences", other_differences)
        rep.write_difference_sheet("Ignored Sales Variances", ignored_sales_variances)
        rep.write_matched_sheet(matched_pairs)
        rep.write_review_sheet(review_required_txns)
        
        rep.save()
        
        # Write technical log file
        log_dir = os.path.dirname(output_path)
        log_filename = os.path.splitext(os.path.basename(output_path))[0] + "_TECHNICAL_LOG.txt"
        tech_log_path = os.path.join(log_dir, log_filename)
        
        with open(tech_log_path, 'w', encoding='utf-8') as f_log:
            f_log.write(f"=== AIS BROKER RECO TECHNICAL LOG ===\n")
            f_log.write(f"Date of Reco: {datetime.datetime.now()}\n")
            f_log.write(f"Financial Year: {fy}\n")
            f_log.write(f"AIS Input: {ais_file}\n")
            f_log.write(f"Broker Input: {broker_file}\n")
            f_log.write(f"Reconciliation Output: {output_path}\n\n")
            f_log.write(f"--- Loaded Data ---\n")
            f_log.write(f"AIS rows loaded: {len(ais_raw_list)}\n")
            f_log.write(f"Broker rows loaded: {len(broker_raw_list)}\n")
            f_log.write(f"Duplicates found in AIS: {len(duplicate_groups)}\n\n")
            f_log.write(f"--- Match Stats ---\n")
            f_log.write(f"Matched pairs: {len(matched_pairs)}\n")
            f_log.write(f"Unmatched AIS: {len(unmatched_ais)}\n")
            f_log.write(f"Unmatched Broker: {len(unmatched_broker)}\n\n")
            f_log.write(f"--- Differences ---\n")
            f_log.write(f"LTCG Class mismatches: {len(ltcg_mismatches)}\n")
            f_log.write(f"STCG Class mismatches: {len(stcg_mismatches)}\n")
            f_log.write(f"Other monetary/STT mismatches: {len(other_differences)}\n")
            f_log.write(
                f"Ignored minor sales variances (<= Rs {SALE_CONSIDERATION_IGNORE_LIMIT:.2f}): "
                f"{len(ignored_sales_variances)}\n"
            )
            f_log.write(f"Review Required items: {len(review_required_txns)}\n\n")
            f_log.write(f"--- Processing Log trace ended successfully ---\n")
            
        progress_callback(100, "Processing completed! Excel file saved successfully.")
        
        log_callback("----------------------------------------------------------------------")
        log_callback(f"SUCCESS: Reconciliation workbook created at:")
        log_callback(f" > {output_path}")
        log_callback(f"Technical log saved next to it at:")
        log_callback(f" > {tech_log_path}")
        log_callback("----------------------------------------------------------------------")
        
        return summary_stats
        
    except Exception as e:
        log_callback(f"FATAL ERROR during reconciliation: {str(e)}")
        log_callback(traceback.format_exc())
        progress_callback(100, f"Error: {str(e)}")
        raise e

# =====================================================================
# 7. TKINTER USER INTERFACE (IDLE COMPATIBLE)
# =====================================================================
class CanvasProgressBar(tk.Canvas):
    """
    Polished Canvas-based progress bar matching modern dark aesthetics.
    """
    def __init__(self, parent, bg="#1e293b", fg="#6366f1", height=18, **kwargs):
        super().__init__(parent, bg=bg, height=height, highlightthickness=1, highlightbackground="#475569", bd=0, **kwargs)
        self.fg = fg
        self.progress = 0.0
        self.bind("<Configure>", self.draw)

    def set(self, val):
        self.progress = min(max(val, 0.0), 1.0)
        self.draw()

    def draw(self, event=None):
        self.delete("all")
        w = self.winfo_width()
        h = self.winfo_height()
        
        # Track background
        self.create_rectangle(2, 2, w - 2, h - 2, fill=self["bg"], width=0)
        
        # Bar fill
        if self.progress > 0:
            fill_width = int((w - 4) * self.progress)
            if fill_width > 0:
                self.create_rectangle(2, 2, 2 + fill_width, h - 2, fill=self.fg, width=0)
                
                # Draw percentage text
                pct_text = f"{int(self.progress * 100)}%"
                text_color = "#ffffff" if self.progress > 0.55 else "#a5b4fc"
                self.create_text(w / 2, h / 2, text=pct_text, fill=text_color, font=("Segoe UI", 9, "bold"))


class AISBrokerRecoApp(tk.Tk):
    """
    Main application UI interface matching Dark Slate theme.
    """
    def __init__(self):
        super().__init__()
        
        self.title("AIS BROKER RECO - Tax Capital Gains Engine")
        self.geometry("820x680")
        self.configure(bg="#0f172a") # Slate 900
        self.minsize(750, 600)
        
        # State
        self.ais_file_path = ""
        self.broker_file_path = ""
        self.is_running = False
        
        self.setup_ui()
        self.check_dependencies()

    def setup_ui(self):
        # 1. Header Card Banner
        header_frame = tk.Frame(self, bg="#1e1b4b", height=90, bd=0) # Deep Indigo
        header_frame.pack(fill="x", side="top")
        header_frame.pack_propagate(False)
        
        title_lbl = tk.Label(
            header_frame, 
            text="AIS BROKER RECO", 
            font=("Segoe UI", 22, "bold"), 
            fg="#f8fafc", # Slate 50
            bg="#1e1b4b"
        )
        title_lbl.pack(pady=(12, 1))
        
        subtitle_lbl = tk.Label(
            header_frame, 
            text="Capital Gains Reconciliation & Audit Tool (Income Tax Act, 1961)", 
            font=("Segoe UI", 9, "italic"), 
            fg="#c7d2fe", # Indigo 200
            bg="#1e1b4b"
        )
        subtitle_lbl.pack()
        
        # Main layout frame
        main_container = tk.Frame(self, bg="#0f172a")
        main_container.pack(fill="both", expand=True, padx=25, pady=20)
        
        # 2. File Selection Panel (Slate 800)
        files_frame = tk.Frame(main_container, bg="#1e293b", bd=1, relief="flat", padx=15, pady=15)
        files_frame.pack(fill="x", pady=(0, 15))
        
        # AIS Row
        ais_lbl = tk.Label(files_frame, text="AIS Data Sheet (Excel):", font=("Segoe UI", 10, "bold"), fg="#f1f5f9", bg="#1e293b")
        ais_lbl.grid(row=0, column=0, sticky="w", pady=(0, 5))
        
        self.ais_entry = tk.Entry(files_frame, font=("Segoe UI", 10), bg="#334155", fg="#ffffff", insertbackground="#ffffff", bd=1, relief="solid")
        self.ais_entry.grid(row=1, column=0, columnspan=2, sticky="ew", ipady=5, padx=(0, 10))
        
        ais_btn = tk.Button(
            files_frame, text="Browse AIS File", font=("Segoe UI", 10, "bold"), 
            bg="#6366f1", fg="#ffffff", activebackground="#4f46e5", activeforeground="#ffffff", 
            bd=0, cursor="hand2", padx=15, command=self.select_ais_file
        )
        ais_btn.grid(row=1, column=2, ipady=4)
        ais_btn.bind("<Enter>", lambda e: ais_btn.config(bg="#4f46e5"))
        ais_btn.bind("<Leave>", lambda e: ais_btn.config(bg="#6366f1"))
        
        # Broker Row
        broker_lbl = tk.Label(files_frame, text="Broker Realised P&L Account (Excel):", font=("Segoe UI", 10, "bold"), fg="#f1f5f9", bg="#1e293b")
        broker_lbl.grid(row=2, column=0, sticky="w", pady=(12, 5))
        
        self.broker_entry = tk.Entry(files_frame, font=("Segoe UI", 10), bg="#334155", fg="#ffffff", insertbackground="#ffffff", bd=1, relief="solid")
        self.broker_entry.grid(row=3, column=0, columnspan=2, sticky="ew", ipady=5, padx=(0, 10))
        
        broker_btn = tk.Button(
            files_frame, text="Browse Broker", font=("Segoe UI", 10, "bold"), 
            bg="#6366f1", fg="#ffffff", activebackground="#4f46e5", activeforeground="#ffffff", 
            bd=0, cursor="hand2", padx=15, command=self.select_broker_file
        )
        broker_btn.grid(row=3, column=2, ipady=4)
        broker_btn.bind("<Enter>", lambda e: broker_btn.config(bg="#4f46e5"))
        broker_btn.bind("<Leave>", lambda e: broker_btn.config(bg="#6366f1"))
        
        # FY Row & Configuration Info
        config_frame = tk.Frame(files_frame, bg="#1e293b")
        config_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(15, 0))
        
        fy_lbl = tk.Label(config_frame, text="Financial Year / Tax Rules:", font=("Segoe UI", 10, "bold"), fg="#f1f5f9", bg="#1e293b")
        fy_lbl.pack(side="left")
        
        self.fy_combo = ttk.Combobox(
            config_frame, values=["FY 2023-24", "FY 2024-25", "FY 2025-26"], 
            font=("Segoe UI", 10), state="readonly", width=15
        )
        self.fy_combo.set("FY 2023-24")
        self.fy_combo.pack(side="left", padx=(10, 0))
        
        # Configure Grid weight for responsiveness
        files_frame.column_dimensions = None # override pandas col settings
        files_frame.grid_columnconfigure(0, weight=1)
        files_frame.grid_columnconfigure(1, weight=1)
        files_frame.grid_columnconfigure(2, weight=0)
        
        # 3. Action Buttons Row
        actions_frame = tk.Frame(main_container, bg="#0f172a")
        actions_frame.pack(fill="x", pady=(0, 15))
        
        self.start_btn = tk.Button(
            actions_frame, text="START RECONCILIATION", font=("Segoe UI", 11, "bold"), 
            bg="#10b981", fg="#ffffff", activebackground="#059669", activeforeground="#ffffff",
            bd=0, cursor="hand2", padx=20, command=self.start_reconciliation_process
        )
        self.start_btn.pack(side="left", ipady=6, padx=(0, 10))
        self.start_btn.bind("<Enter>", lambda e: self.start_btn.config(bg="#059669") if not self.is_running else None)
        self.start_btn.bind("<Leave>", lambda e: self.start_btn.config(bg="#10b981") if not self.is_running else None)
        
        reset_btn = tk.Button(
            actions_frame, text="RESET FIELDS", font=("Segoe UI", 11, "bold"), 
            bg="#475569", fg="#ffffff", activebackground="#334155", activeforeground="#ffffff",
            bd=0, cursor="hand2", padx=15, command=self.reset_fields
        )
        reset_btn.pack(side="left", ipady=6, padx=(0, 10))
        reset_btn.bind("<Enter>", lambda e: reset_btn.config(bg="#334155"))
        reset_btn.bind("<Leave>", lambda e: reset_btn.config(bg="#475569"))
        
        exit_btn = tk.Button(
            actions_frame, text="EXIT APP", font=("Segoe UI", 11, "bold"), 
            bg="#ef4444", fg="#ffffff", activebackground="#dc2626", activeforeground="#ffffff",
            bd=0, cursor="hand2", padx=15, command=self.destroy
        )
        exit_btn.pack(side="right", ipady=6)
        exit_btn.bind("<Enter>", lambda e: exit_btn.config(bg="#dc2626"))
        exit_btn.bind("<Leave>", lambda e: exit_btn.config(bg="#ef4444"))
        
        # 4. Progress Area
        progress_frame = tk.Frame(main_container, bg="#1e293b", bd=1, relief="flat", padx=15, pady=12)
        progress_frame.pack(fill="x", pady=(0, 15))
        
        self.status_lbl = tk.Label(progress_frame, text="Status: Ready", font=("Segoe UI", 9, "bold"), fg="#94a3b8", bg="#1e293b")
        self.status_lbl.pack(anchor="w", pady=(0, 4))
        
        self.progress_bar = CanvasProgressBar(progress_frame, bg="#0f172a", fg="#6366f1")
        self.progress_bar.pack(fill="x")
        
        # 5. Log Console Area
        console_frame = tk.Frame(main_container, bg="#1e293b", bd=1, relief="flat", padx=15, pady=15)
        console_frame.pack(fill="both", expand=True)
        
        console_lbl = tk.Label(console_frame, text="Activity Logs:", font=("Segoe UI", 10, "bold"), fg="#f1f5f9", bg="#1e293b")
        console_lbl.pack(anchor="w", pady=(0, 5))
        
        self.log_area = scrolledtext.ScrolledText(
            console_frame, font=("Consolas", 9), bg="#020617", fg="#f1f5f9", 
            bd=1, relief="solid", insertbackground="#ffffff"
        )
        self.log_area.pack(fill="both", expand=True)
        self.log_area.config(state="disabled")
        
        # Colored tags
        self.log_area.tag_config("info", foreground="#94a3b8")
        self.log_area.tag_config("success", foreground="#34d399")
        self.log_area.tag_config("warning", foreground="#fbbf24")
        self.log_area.tag_config("error", foreground="#f87171")
        self.log_area.tag_config("bold_info", foreground="#60a5fa", font=("Consolas", 9, "bold"))
        
        self.write_log("Welcome to AIS BROKER RECO Engine! Load input Excel sheets to begin matching.", "bold_info")

    def check_dependencies(self):
        """Displays dialog box to install packages if missing."""
        missing = []
        try:
            import pandas
        except ImportError:
            missing.append("pandas")
        try:
            import openpyxl
        except ImportError:
            missing.append("openpyxl")
            
        if missing:
            messagebox.showerror(
                "Missing Dependencies",
                f"Required third-party packages are missing: {', '.join(missing)}\n\n"
                "Please run the following CMD command to install them:\n"
                "python -m pip install pandas openpyxl"
            )
            self.write_log("ERROR: Missing package dependencies. Re-install using CMD prompt.", "error")
        else:
            self.write_log(
                "Base dependencies verified. Note: 'xlrd' is only needed when opening legacy .xls files.",
                "success"
            )

    def write_log(self, text, tag="info"):
        """Thread-safe UI logger."""
        self.log_area.config(state="normal")
        time_str = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_area.insert("end", f"[{time_str}] {text}\n", tag)
        self.log_area.see("end")
        self.log_area.config(state="disabled")

    def set_progress(self, val, status_text=None):
        """Thread-safe progress updates."""
        self.progress_bar.set(val)
        if status_text:
            self.status_lbl.config(text=f"Status: {status_text}", fg="#fbbf24")

    def select_ais_file(self):
        file_path = filedialog.askopenfilename(
            title="Select AIS Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xls *.xlsm"), ("All Files", "*.*")]
        )
        if file_path:
            self.ais_file_path = file_path
            self.ais_entry.delete(0, "end")
            self.ais_entry.insert(0, file_path)
            self.write_log(f"AIS File Selected: {file_path}", "info")

    def select_broker_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Broker Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xls *.xlsm"), ("All Files", "*.*")]
        )
        if file_path:
            self.broker_file_path = file_path
            self.broker_entry.delete(0, "end")
            self.broker_entry.insert(0, file_path)
            self.write_log(f"Broker File Selected: {file_path}", "info")

    def get_safe_incremental_filename(self, output_path):
        """
        Creates incremented file name if file already exists in path,
        to prevent overwriting files.
        """
        if not os.path.exists(output_path):
            return output_path
            
        dir_name = os.path.dirname(output_path)
        base_name = os.path.basename(output_path)
        name, ext = os.path.splitext(base_name)
        
        counter = 1
        new_filename = f"{name}_{counter}{ext}"
        new_path = os.path.join(dir_name, new_filename)
        
        while os.path.exists(new_path):
            counter += 1
            new_filename = f"{name}_{counter}{ext}"
            new_path = os.path.join(dir_name, new_filename)
            
        return new_path

    def sanitize_windows_filename(self, name):
        """Replaces characters invalid in Windows paths."""
        return re.sub(r'[\\/*?:"<>|]', '_', name)

    def reset_fields(self):
        if self.is_running:
            return
        self.ais_entry.delete(0, "end")
        self.broker_entry.delete(0, "end")
        self.ais_file_path = ""
        self.broker_file_path = ""
        self.set_progress(0.0, "Ready")
        self.status_lbl.config(text="Status: Ready", fg="#94a3b8")
        self.write_log("App state reset. Ready for new files.", "bold_info")

    def start_reconciliation_process(self):
        if self.is_running:
            return
            
        ais = self.ais_entry.get().strip()
        broker = self.broker_entry.get().strip()
        fy = self.fy_combo.get()
        
        if not ais or not broker:
            messagebox.showerror("Selection Error", "Please select both the AIS and Broker Excel files before running.")
            return
            
        if not os.path.exists(ais):
            messagebox.showerror("File Error", f"AIS file does not exist:\n{ais}")
            return
        if not os.path.exists(broker):
            messagebox.showerror("File Error", f"Broker file does not exist:\n{broker}")
            return
            
        # Select output location beforehand
        default_dir = os.path.dirname(ais)
        clean_fy = self.sanitize_windows_filename(fy.replace(" ", ""))
        default_name = f"AIS_BROKER_RECO_{clean_fy}.xlsx"
        
        output_file = filedialog.asksaveasfilename(
            initialdir=default_dir,
            initialfile=default_name,
            title="Save Reconciliation Results As",
            filetypes=[("Excel Workbook", "*.xlsx")],
            defaultextension=".xlsx"
        )
        
        if not output_file:
            self.write_log("Operation cancelled by user (output save location not selected).", "warning")
            return
            
        # Get incremented name if file exists to prevent silent overwrite
        safe_output = self.get_safe_incremental_filename(output_file)
        if safe_output != output_file:
            self.write_log(f"Output path adjusted to prevent overwrite: {os.path.basename(safe_output)}", "warning")
            
        # Run background thread
        self.is_running = True
        self.start_btn.config(state="disabled", bg="#475569", text="PROCESSING...")
        
        thread = threading.Thread(
            target=self.run_reco_thread, 
            args=(ais, broker, safe_output, fy),
            daemon=True
        )
        thread.start()

    def run_reco_thread(self, ais, broker, output_path, fy):
        try:
            summary = run_reconciliation_workflow(
                ais_file=ais,
                broker_file=broker,
                output_path=output_path,
                fy=fy,
                progress_callback=lambda p, status: self.after(0, self.set_progress, p/100.0, status),
                log_callback=lambda text, tag="info": self.after(0, self.write_log, text, tag)
            )
            
            # Show Completion Popup Box
            self.after(0, self.show_success_dialog, summary, output_path)
            
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Processing Mismatch Error", f"An unexpected error halted reconciliation:\n\n{str(e)}"))
            self.after(0, self.set_progress, 1.0, "Error Encountered")
            self.after(0, lambda: self.status_lbl.config(fg="#f87171"))
        finally:
            self.after(0, self.finish_reco_state)

    def show_success_dialog(self, stats, output_path):
        self.status_lbl.config(text="Status: Completed Successfully", fg="#10b981")
        self.progress_bar.set(1.0)
        
        # Summary details
        details = (
            f"RECONCILIATION COMPLETED SUCCESSFULLY!\n\n"
            f"• Total AIS rows: {stats['total_ais']}\n"
            f"• Total Broker rows: {stats['total_broker']}\n"
            f"• Matched Transactions: {stats['matched']}\n"
            f"• AIS Duplicate entries: {stats['duplicates']}\n"
            f"• Missing in AIS: {stats['missing_in_ais']}\n"
            f"• Missing in Broker: {stats['missing_in_broker']}\n"
            f"• LTCG Differences: {stats['ltcg_diffs']}\n"
            f"• STCG Differences: {stats['stcg_diffs']}\n"
            f"• Manual Review Required: {stats['review_required']}\n\n"
            f"Reconciliation Report Saved At:\n{output_path}"
        )
        messagebox.showinfo("Reconciliation Success", details)

    def finish_reco_state(self):
        self.is_running = False
        self.start_btn.config(state="normal", bg="#10b981", text="START RECONCILIATION")

if __name__ == "__main__":
    app = AISBrokerRecoApp()
    app.mainloop()
